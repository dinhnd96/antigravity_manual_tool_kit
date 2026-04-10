import openpyxl

file_path = '/Users/mac/antigravity-testing-kit/Feature_02_SA_Tham_So_He_Thong/test case/SA02_Đăng xuất.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.worksheets[0]

headers = [cell.value for cell in ws[1]]
col_map = {name: idx for idx, name in enumerate(headers)}

expected_col = 'Expected' if 'Expected' in col_map else 'Expected Result'
note_col = 'Note' if 'Note' in col_map else None

updates = {
    'SA02-HAPPY-001': {
        expected_col: (
            '(i) Nghiệp vụ: Hệ thống gửi yêu cầu hủy phiên (Revoke Token) lên Server.\n'
            '(ii) UI: Hệ thống hiển thị màn hình trung gian "Đăng xuất thành công - Cảm ơn bạn đã sử dụng hệ thống ProfiX" kèm nút "Đăng nhập".\n'
            '(iii) Trạng thái/Audit: Toàn bộ thông tin phiên làm việc bị xóa sạch.\n'
            '(iv) Output: Không có.'
        )
    },
    'SA02-SECURITY-002': {
        expected_col: (
            '(i) Nghiệp vụ/Logic: Hệ thống tự động hết hạn phiên đăng nhập khi không có tương tác.\n'
            '(ii) UI: Hiển thị thông báo "Phiên làm việc đã hết hạn" và tự động chuyển về trang Đăng nhập.\n'
            '(iii) Trạng thái/Audit: Session bị hủy trên Server. Token hết hạn. Ghi log Auto-logout event.\n'
            '(iv) Output: Không có.'
        )
    },
    'SA02-NEG-003': {
        'Steps': (
            '1. Nhấn nút Đăng xuất trên Header.\n'
            '2. Đăng nhập lại và kiểm tra dữ liệu cũ.'
        ),
        expected_col: (
            '(i) Nghiệp vụ/Logic: Hệ thống không tự động lưu các tác vụ chưa hoàn tất.\n'
            '(ii) UI: Dữ liệu chưa lưu tại form cũ bị xóa sạch sau khi logout.\n'
            '(iii) Trạng thái/Audit: Không ghi nhận lưu dữ liệu nháp vào DB. Session bị hủy.\n'
            '(iv) Output: Không có.'
        ),
        note_col: 'URD không thiết kế Popup xác nhận khi Đăng xuất (kể cả có Unsaved Data). Nên luồng chỉ là 1-click logout.'
    },
    'SA02-HAPPY-005': {
        'Precondition': 'Người dùng đã đăng nhập thành công và đang ở màn hình Dashboard.',
        expected_col: (
            '(i) Nghiệp vụ/Logic: Đóng phiên làm việc hoàn toàn, xóa cache/local storage liên quan user.\n'
            '(ii) UI: Tùy thiết kế (sau màn trung gian bấm "Đăng nhập"), URL chuyển về trang đăng nhập công cộng. Giao diện hiển thị đúng form Login ban đầu.\n'
            '(iii) Trạng thái/Audit: Token hủy.\n'
            '(iv) Output: Không có.'
        )
    },
    'SA02-BOUNDARY-006': {
        expected_col: (
            '(i) Nghiệp vụ/Logic: Hệ thống vẫn nhận diện đúng mốc 1 phút để Auto-logout.\n'
            '(ii) UI: Thông báo đúng thời điểm.\n'
            '(iii) Trạng thái/Audit: Session bị terminate tại mốc 1 phút, Token hủy.\n'
            '(iv) Output: Không có.'
        )
    },
    'SA02-FLOW-007': {
        'Steps': (
            '1. Đăng nhập thành công.\n'
            '2. Mở menu Tra cứu code phí.\n'
            '3. Nhấn Đăng xuất.\n'
            '4. Thử đăng nhập lại bằng tài khoản đó.'
        ),
        expected_col: (
            '(i) Nghiệp vụ/Logic: Luồng đóng phiên và mở phiên mới diễn ra liền mạch, dữ liệu phiên cũ không bị rò rỉ sang phiên mới.\n'
            '(ii) UI: Hiển thị màn hình trung gian Đăng xuất thành công. Lần đăng nhập sau đó vào Home bình thường.\n'
            '(iii) Trạng thái/Audit: Ghi log Đăng xuất - Đăng nhập phân mảnh rõ ràng, phiên làm việc độc lập.\n'
            '(iv) Output: Không có.'
        )
    },
    'SA02-UI-008': {
        'Title': '[PENDING-BA] Hiển thị Popup xác nhận (Warning) khi Đăng xuất lúc có dữ liệu chưa lưu',
        note_col: 'TC ON HOLD — Chờ BA xác nhận có thiết kế Popup xác nhận Logout khi có unsaved data không. Nếu không có: Delete TC.'
    },
    'SA02-BOUNDARY-010': {
        expected_col: (
            '(i) Nghiệp vụ/Logic: Hệ thống phải xử lý theo default value [cần BA xác nhận default value] hoặc chặn lưu cấu hình sai từ màn quản trị.\n'
            '(ii) UI: Không xảy ra lỗi trắng trang hoặc đăng xuất ngay lập tức sau khi login.\n'
            '(iii) Trạng thái/Audit: Nếu áp dụng default value, Session sẽ Auto-logout theo thời gian mặc định đó.\n'
            '(iv) Output: Không có.'
        )
    },
    'SA02-NEGATIVE-011': {
        expected_col: (
            '(i) Nghiệp vụ/Logic: Client-side phải xóa sạch Cookie/Local Storage Token cho dù API Logout lên Server bị lỗi timeout.\n'
            '(ii) UI: Vẫn phải quay lại màn hình Login cục bộ, không treo UI chờ kết quả API.\n'
            '(iii) Trạng thái/Audit: Client-side Token/Cookie đã bị xóa khỏi LocalStorage dù API chưa phản hồi.\n'
            '(iv) Output: Không có.'
        )
    }
}

for row in ws.iter_rows(min_row=2):
    tc_id_val = row[col_map['TC_ID']].value
    if tc_id_val and tc_id_val in updates:
        for col_name, new_val in updates[tc_id_val].items():
            if col_name in col_map:
                row[col_map[col_name]].value = new_val
                print(f"Updated [{tc_id_val}] column [{col_name}]")
            else:
                print(f"WARNING: Column {col_name} not found")

wb.save(file_path)
print("\nSA02 updates applied successfully.")

import openpyxl

files = [
    'SA01_Đăng nhập.xlsx',
    'SA02_Đăng xuất.xlsx',
    'SA03_Test_Cases.xlsx',
    'SA04_Quản lý phân quyền.xlsx',
    'SA05_Ma trận phê duyệt (1).xlsx'
]

for f in files:
    try:
        wb = openpyxl.load_workbook(f'./Feature_02_SA_Tham_So_He_Thong/test case/{f}', data_only=True)
        sn = [s for s in wb.sheetnames if 'test' in s.lower() or 'case' in s.lower()][0]
        sh = wb[sn]
        first_row = list(sh.iter_rows(values_only=True))[0]
        try:
            final_status_index = [str(x).lower() for x in first_row].index('final status')
            from openpyxl.utils import get_column_letter
            col = get_column_letter(final_status_index + 1)
            print(f"{f}: Final Status is at column {col} (index {final_status_index})")
        except ValueError:
            print(f"{f}: Final Status NOT FOUND in {first_row}")
    except Exception as e:
        print(f"Error {f}: {e}")

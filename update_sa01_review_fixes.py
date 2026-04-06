import openpyxl

file_path = 'Feature_02_SA_Tham_So_He_Thong/test case/SA01_Đăng nhập.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['TestCases']

for row in range(2, ws.max_row + 1):
    tc_id = ws.cell(row=row, column=1).value
    if not tc_id:
        continue
    
    # 1. Update Expected Result format (Thêm (iv) Output)
    expected = ws.cell(row=row, column=12).value
    if expected and '(iv)' not in expected:
        expected = expected.strip() + '\n(iv) Output: Không có.'
        # Chuẩn hóa (iii) Log -> (iii) Trạng thái/Log
        if '(iii) Log:' in expected:
            expected = expected.replace('(iii) Log:', '(iii) Trạng thái/Log:')
        ws.cell(row=row, column=12).value = expected

    # 2. Fix Preconditions
    precond = ws.cell(row=row, column=10).value
    if precond == '-' or not precond or precond.strip() == '':
        if tc_id == 'SA01-NEG-004':
            ws.cell(row=row, column=10).value = 'Tài khoản tồn tại hợp lệ.'
        elif tc_id in ('SA01-HAPPY-007', 'SA01-NEG-016'):
            ws.cell(row=row, column=10).value = 'Tài khoản tồn tại, hệ thống hoạt động bình thường.'
        elif tc_id in ('SA01-UI-008', 'SA01-NEG-010', 'SA01-NEG-014'):
            ws.cell(row=row, column=10).value = 'Màn hình Login đang hiển thị.'
        elif tc_id == 'SA01-NEG-013':
            ws.cell(row=row, column=10).value = 'Điều kiện mạng bình thường.'

    # 3. Fix SA01-NEG-006 (Ambiguity)
    if tc_id == 'SA01-NEG-006':
        ws.cell(row=row, column=12).value = "(i) Nghiệp vụ: Chấp nhận phiên đăng nhập mới tại Browser B, đồng thời ngắt phiên làm việc đang tồn tại tại Browser A (Kill-sess).\n(ii) UI: Đăng nhập thành công tại Browser B. Trình duyệt A có thông báo bị đăng xuất do tài khoản đăng nhập ở nơi khác.\n(iii) Trạng thái/Log: Hệ thống ghi log lưu vết sự kiện cảnh báo đăng nhập.\n(iv) Output: Không có."
        ws.cell(row=row, column=14).value = 'Team đã thống nhất sử dụng cơ chế Kill-sess (Đá phiên cũ) thay vì chặn đăng nhập mới.'

    # 4. Update Note for SA01-NEG-004
    if tc_id == 'SA01-NEG-004':
        ws.cell(row=row, column=14).value = 'Lưu ý: Tách riêng 2 field khi thực thi Automation để pass được validation tuần tự.'

# 5. Thêm TC mới cho EntraID Network Error
new_row = ws.max_row + 1
ws.cell(row=new_row, column=1).value = 'SA01-NEG-017'
ws.cell(row=new_row, column=2).value = 'BR_04'
ws.cell(row=new_row, column=3).value = 'II.5.4.1'
ws.cell(row=new_row, column=4).value = 'SA.01'
ws.cell(row=new_row, column=5).value = 'Đăng nhập'
ws.cell(row=new_row, column=6).value = 'Xử lý lỗi khi hệ thống mất kết nối với EntraID (Network Error/503)'
ws.cell(row=new_row, column=7).value = 'Negative'
ws.cell(row=new_row, column=8).value = 'Regression'
ws.cell(row=new_row, column=9).value = 'P1'
ws.cell(row=new_row, column=10).value = 'Màn hình Login đang hiển thị.'
ws.cell(row=new_row, column=11).value = "1. Gây gián đoạn kết nối mạng phía Server tới mạng ngoài hoặc EntraID bị lỗi 503.\n2. User nhấn 'Đăng nhập qua EntraID'."
ws.cell(row=new_row, column=12).value = "(i) Nghiệp vụ: Hệ thống chặn thao tác do không kết nối được dịch vụ xác thực.\n(ii) UI: Hiển thị thông báo lỗi 'Lỗi kết nối đến hệ thống xác thực. Vui lòng thử lại sau'.\n(iii) Trạng thái/Log: Hệ thống chủ động ghi Error log (503).\n(iv) Output: Không có."
ws.cell(row=new_row, column=13).value = 'SA01-LGN-N17'

wb.save(file_path)
print(f'Done update {file_path}')

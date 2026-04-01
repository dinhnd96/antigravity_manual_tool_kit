import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation

file_path = '/Users/mac/antigravity-testing-kit/Test case/Test case_Management_ProfiX.xlsx'

def update_formulas():
    try:
        wb = openpyxl.load_workbook(file_path)
        
        # Analyze actual test case sheets
        tc_sheets = [s for s in wb.sheetnames if s not in ['📊 Dashboard', '📅 Daily Tracking', '🐞 Bug Tracker', 'Coverage', 'Dedup_Log']]
        
        dashboard = wb['📊 Dashboard']
        tracking = wb['📅 Daily Tracking']
        bug_tracker = wb['🐞 Bug Tracker']
        
        # --- 1. Dashboard Formulas ---
        # Clear old data
        for i in range(3, dashboard.max_row + 1):
            for j in range(1, 9):
                dashboard.cell(row=i, column=j).value = None
                
        row_idx = 3
        for idx, sheet_name in enumerate(tc_sheets, 1):
            sheet = wb[sheet_name]
            # Find Final Status column, fallback to U
            status_col_letter = 'U'
            for col in range(1, sheet.max_column + 1):
                val = sheet.cell(row=2, column=col).value
                if val and 'Final Status' in str(val):
                    status_col_letter = get_column_letter(col)
                    break
                    
            dashboard.cell(row=row_idx, column=1).value = idx
            dashboard.cell(row=row_idx, column=2).value = sheet_name
            
            # Total TC: =COUNTA('Sheet'!A:A) - 2 (subtract headers)
            dashboard.cell(row=row_idx, column=3).value = f"=COUNTA('{sheet_name}'!A:A)-2"
            
            # Passed / Failed / N/A
            dashboard.cell(row=row_idx, column=4).value = f'=COUNTIF(\'{sheet_name}\'!{status_col_letter}:{status_col_letter}, "Pass")'
            dashboard.cell(row=row_idx, column=5).value = f'=COUNTIF(\'{sheet_name}\'!{status_col_letter}:{status_col_letter}, "Fail")'
            dashboard.cell(row=row_idx, column=6).value = f'=COUNTIF(\'{sheet_name}\'!{status_col_letter}:{status_col_letter}, "N/A")'
            
            # Execution % = (Passed + Failed) / Total TC
            dashboard.cell(row=row_idx, column=7).value = f"=IF(C{row_idx}>0, (D{row_idx}+E{row_idx})/C{row_idx}, 0)"
            dashboard.cell(row=row_idx, column=7).number_format = '0.00%'
            
            # Pass Rate % = Passed / Total TC
            dashboard.cell(row=row_idx, column=8).value = f"=IF(C{row_idx}>0, D{row_idx}/C{row_idx}, 0)"
            dashboard.cell(row=row_idx, column=8).number_format = '0.00%'
            
            row_idx += 1
            
        # Total Row
        dashboard.cell(row=row_idx, column=2).value = "TOTAL"
        dashboard.cell(row=row_idx, column=3).value = f"=SUM(C3:C{row_idx-1})"
        dashboard.cell(row=row_idx, column=4).value = f"=SUM(D3:D{row_idx-1})"
        dashboard.cell(row=row_idx, column=5).value = f"=SUM(E3:E{row_idx-1})"
        dashboard.cell(row=row_idx, column=6).value = f"=SUM(F3:F{row_idx-1})"
        dashboard.cell(row=row_idx, column=7).value = f"=IF(C{row_idx}>0, (D{row_idx}+E{row_idx})/C{row_idx}, 0)"
        dashboard.cell(row=row_idx, column=7).number_format = '0.00%'
        dashboard.cell(row=row_idx, column=8).value = f"=IF(C{row_idx}>0, D{row_idx}/C{row_idx}, 0)"
        dashboard.cell(row=row_idx, column=8).number_format = '0.00%'

        for j in range(2, 9):
            dashboard.cell(row=row_idx, column=j).font = Font(bold=True)

        # --- 2. Daily Tracking Formulas ---
        for i in range(3, 30):
            # Defects Found = COUNTIFS of Detected Date (Col G) in Bug Tracker matching Date (Col A) in Tracking
            tracking.cell(row=i, column=4).value = f"=COUNTIF('🐞 Bug Tracker'!G:G, A{i})"
            # Critical Issues
            tracking.cell(row=i, column=5).value = f'=COUNTIFS(\'🐞 Bug Tracker\'!G:G, A{i}, \'🐞 Bug Tracker\'!C:C, "Critical")'
            # Blockers
            tracking.cell(row=i, column=6).value = f'=COUNTIFS(\'🐞 Bug Tracker\'!G:G, A{i}, \'🐞 Bug Tracker\'!C:C, "Blocker")'

        # --- 3. Bug Tracker Data Validation ---
        dv_severity = DataValidation(type="list", formula1='"Minor,Major,Critical,Blocker"', allow_blank=True)
        dv_priority = DataValidation(type="list", formula1='"Low,Medium,High,Urgent"', allow_blank=True)
        dv_status = DataValidation(type="list", formula1='"New,Assigned,In Progress,Fixed,Re-test,Closed,Rejected"', allow_blank=True)
        
        bug_tracker.add_data_validation(dv_severity)
        bug_tracker.add_data_validation(dv_priority)
        bug_tracker.add_data_validation(dv_status)
        
        dv_severity.add('C3:C1000')
        dv_priority.add('D3:D1000')
        dv_status.add('H3:H1000')

        wb.save(file_path)
        print("Formulas updated successfully.")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == '__main__':
    update_formulas()

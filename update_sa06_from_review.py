import openpyxl
import shutil

# Đường dẫn file
source_file = 'Feature_02_SA_Tham_So_He_Thong/test case/SA06 - reviewed.xlsx'
dest_file = 'Feature_02_SA_Tham_So_He_Thong/test case/SA06 - updated_final.xlsx'

# Copy file trước khi thao tác
shutil.copy(source_file, dest_file)

wb = openpyxl.load_workbook(dest_file)
ws = wb.active

# Lấy index của header
headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column+1)}

col_title = headers['Title']
col_steps = headers['Steps']
col_exp = headers['Expected Result']
col_pre = headers['Precondition']
col_feat = headers['Feature']

for r in range(2, ws.max_row + 1):
    tc_id = ws.cell(r, headers['TC_ID']).value
    if not tc_id: continue
    
    # SA06-BR-NEG-002: Lỗi check trùng tên
    if tc_id == "SA06-BR-NEG-002":
        ws.cell(r, col_title).value = "Kiểm tra hệ thống báo lỗi khi nhập trùng Mã điều kiện đã tồn tại"
        ws.cell(r, col_steps).value = "1. Nhập Mã điều kiện đã tồn tại trong hệ thống (Tên điều kiện nhập hợp lệ hoặc trùng đều được).\n2. Nhấn Xác nhận."
        ws.cell(r, col_exp).value = "(i) Nghiệp vụ/Logic: DB kiểm tra constraint (Unique) và chặn tạo mới bản ghi.\n(ii) UI: Hiển thị thông báo lỗi 'Mã điều kiện đã tồn tại'. Form thiết lập không tắt.\n(iii) Trạng thái/Audit: Không tạo thay đổi, không sinh log.\n(iv) Output: Không có."
        
    # SA06-BR-HAP-006: Thiếu field khi ETL
    elif tc_id == "SA06-BR-HAP-006":
        ws.cell(r, col_steps).value = "1. Tại dropdown 'Nguồn dữ liệu', lần lượt thử các trường hợp:\n1a. Chọn 'API'.\n1b. Chọn 'ETL'.\n2. Nhập thông tin cấu hình tương ứng:\n- Nếu API: Nhập ô 'Mapping note Msg'.\n- Nếu ETL: Chọn list từ 'Bảng dữ liệu' và sau đó tiếp tục chọn 'Trường dữ liệu' tương ứng với bảng đó.\n3. Nhấn Xác nhận."

    # SA06-BR-HAP-005: Thêm thiết lập CTUD
    elif tc_id == "SA06-BR-HAP-005":
        ws.cell(r, col_title).value = "Kiểm tra chỉ hiển thị danh sách điều kiện tính phí có trạng thái = 'Hoạt động' tại màn tạo Code phí / CTƯĐ"
        ws.cell(r, col_pre).value = "1. 'COND-ACT' trạng thái Hoạt động, 'COND-INACT' trạng thái Không hoạt động.\n2. Mở màn hình tạo/chỉnh sửa Code phí HOẶC mở màn hình tạo/chỉnh sửa CTƯĐ."

    # SA06-BR-NEG-003: Update error message
    elif tc_id == "SA06-BR-NEG-003":
        ws.cell(r, col_exp).value = "(i) Nghiệp vụ/Logic: Phát hiện ràng buộc tham chiếu, chặn update.\n(ii) UI: Hệ thống bật popup/toast cảnh báo 'Không thể sửa trạng thái điều kiện tính phí. Vui lòng đảm bảo CTƯĐ đã hết hiệu lực và Code phí đã huỷ trước khi thực hiện'.\n(iii) Trạng thái/Audit: Không tạo thay đổi.\n(iv) Output: Không có."

    # SA06-BR-HAP-004: Update precondition về trạng thái
    elif tc_id == "SA06-BR-HAP-004":
        ws.cell(r, col_pre).value = "1. Hệ thống có sẵn Điều kiện tính phí thuộc 1 trong 2 tình huống sau:\n- Trạng thái 'Không hoạt động'.\n- Trạng thái 'Hoạt động', NHƯNG chưa từng gán vào đâu, hoặc gán vào CTƯĐ/Code phí đã hết hiệu lực/đã hủy."

    # SA06-BR-HAP-001: Chi tiết field
    elif tc_id == "SA06-BR-HAP-001":
        ws.cell(r, col_steps).value = "1. Tại màn 'Danh mục điều kiện tính phí', chọn nút 'Thêm mới'.\n2. Nhập đầy đủ, hợp lệ các box của nhóm trường bắt buộc (Mã, Tên, Kiểu dữ liệu, Trạng thái).\n3. Tại Nguồn dữ liệu, chọn API (và nhập Mapping note Msg) HOẶC chọn ETL (và cấu hình Bảng, Trường dữ liệu).\n4. Nhấn 'Xác nhận'."

    # SA06-UI-001: Navigation và tên sai
    elif tc_id == "SA06-UI-001":
        ws.cell(r, col_title).value = "Kiểm tra giao diện màn hình Thêm mới Điều kiện tính phí"
        ws.cell(r, col_feat).value = "Thêm mới Điều kiện tính phí"
        ws.cell(r, col_steps).value = "1. Tại trang Danh sách Điều kiện tính phí, click nút 'Thêm mới'.\n2. Quan sát màn hình/popup Thêm mới bật lên."
        
    # SA06-UI-004: Ambiguity tự động tải
    elif tc_id == "SA06-UI-004":
        ws.cell(r, col_steps).value = "1. Thực hiện nhập điều kiện tìm kiếm đa dạng:\n1a. Lọc ra kết quả có Tồn tại Data thỏa mãn.\n1b. Lọc ra kết quả Empty (Không có data thỏa mãn).\n2. Nhấn nút Tìm kiếm (hoặc gõ Enter nếu hỗ trợ).\n3. Click nút 'Xóa bộ lọc / Clear' sau khi test xong."

    # SA06-UI-005: Xuất file
    elif tc_id == "SA06-UI-005":
        ws.cell(r, col_exp).value = "(i) Nghiệp vụ/Logic: Trigger Export API lấy dữ liệu theo đúng Query lúc user nhấn xuất.\n(ii) UI: Không đổi trang, hiển thị loading.\n(iii) Trạng thái/Audit: Ghi log Export.\n(iv) Output: File excel có đủ header, dữ liệu các hàng/cột bên trong file đảm bảo khớp 100% với danh sách dữ liệu đang hiển thị trên lưới đã filter."

wb.save(dest_file)
print("Updated Test Cases successfully. Saved to:", dest_file)

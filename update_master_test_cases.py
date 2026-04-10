import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import os

master_path = './Feature_02_SA_Tham_So_He_Thong/test case/ProfiX_Master_Test_Cases.xlsx'
wb = openpyxl.load_workbook(master_path)

files_to_sync = [
    ('SA01', 'SA01_Đăng nhập.xlsx'),
    ('SA02', 'SA02_Đăng xuất.xlsx'),
    ('SA03', 'SA03_Test_Cases.xlsx'),
    ('SA04', 'SA04_Quản lý phân quyền.xlsx'),
    ('SA05', 'SA05_Ma trận phê duyệt (1).xlsx')
]

std_headers = ['TC_ID', 'BR_Ref', 'URD_Ref', 'Module', 'Feature', 'Title', 'Type', 'Category', 
               'Priority', 'Precondition', 'Steps', 'Expected Result', 'Trace_ID', 'Note', 
               'Status R1', 'Tester R1', 'Date R1', 'Status R2', 'Tester R2', 'Date R2', 'Final Status']

for module_name, filename in files_to_sync:
    file_path = f"./Feature_02_SA_Tham_So_He_Thong/test case/{filename}"
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        continue
    
    src_wb = openpyxl.load_workbook(file_path, data_only=True)
    src_sheet_name = [s for s in src_wb.sheetnames if 'test' in s.lower() or 'case' in s.lower()][0]
    src_sh = src_wb[src_sheet_name]
    
    if module_name in wb.sheetnames:
        del wb[module_name]
    dest_sh = wb.create_sheet(module_name)
    
    # write std headers
    dest_sh.append(std_headers)
    
    # parse source
    rows = list(src_sh.iter_rows(values_only=True))
    src_header = [str(x).lower().strip() if x else '' for x in rows[0]]
    
    col_map = {}
    for i, h in enumerate(std_headers):
        h_l = h.lower()
        if 'expected' in h_l: h_l = 'expected'
        for j, sh in enumerate(src_header):
            if h_l in sh or sh in h_l:
                col_map[i] = j
                break
                
    for row in rows[1:]:
        if not any(row): continue
        dest_row = [''] * 21
        for i in range(21):
            if i in col_map and col_map[i] < len(row):
                dest_row[i] = row[col_map[i]]
        
        # clean-up logic
        title = str(dest_row[5]).lower() if dest_row[5] else ''
        if 'xem' in title or 'view' in title or 'chi tiết' in title:
            dest_row[11] = "Form hiển thị dữ liệu read-only. Không làm thay đổi Database."
        if 'đóng' in title or 'hủy' in title or 'cancel' in title:
            dest_row[11] = "Đóng form, quay về dữ liệu màn trước. Không sinh bản ghi Nháp/Chờ duyệt."
            
        dest_sh.append(dest_row)
        
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    fill_blue = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    dest_sh.column_dimensions['F'].width = 35 # Title
    dest_sh.column_dimensions['J'].width = 35 # Precondition
    dest_sh.column_dimensions['K'].width = 45 # Steps
    dest_sh.column_dimensions['L'].width = 45 # Expected
    
    for row_idx, row in enumerate(dest_sh.iter_rows()):
        for col_idx, cell in enumerate(row):
            cell.border = border
            if row_idx == 0:
                cell.font = Font(bold=True)
                cell.fill = fill_blue
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if 14 <= col_idx <= 20: 
                    cell.fill = fill_yellow
                    cell.alignment = Alignment(horizontal='center', vertical='top')

# Refresh Dashboard
dash = wb['📊 Dashboard']
dash.delete_rows(3, dash.max_row)

sa_sheets = [s for s in wb.sheetnames if s.startswith('SA')]
sa_sheets.sort() 

start_row = 3
thin = Side(border_style="thin", color="000000")
border = Border(top=thin, left=thin, right=thin, bottom=thin)

for i, module in enumerate(sa_sheets):
    r = start_row + i
    
    # find exact final status column letter
    sh = wb[module]
    hr = list(sh.iter_rows(max_row=1, values_only=True))[0]
    final_col_idx = -1
    for j, h in enumerate(hr):
        if h and 'final status' in str(h).lower():
            final_col_idx = j
            break
            
    col_letter = get_column_letter(final_col_idx + 1) if final_col_idx != -1 else 'U'
    
    dash.cell(row=r, column=1, value=i+1).border = border
    dash.cell(row=r, column=2, value=module).border = border
    dash.cell(row=r, column=3, value=f"=COUNTA('{module}'!A:A)-1").border = border
    
    dash.cell(row=r, column=4, value=f'=COUNTIF(\'{module}\'!{col_letter}:{col_letter},"Pass")').border = border
    dash.cell(row=r, column=5, value=f'=COUNTIF(\'{module}\'!{col_letter}:{col_letter},"Fail")').border = border
    dash.cell(row=r, column=6, value=f'=COUNTIF(\'{module}\'!{col_letter}:{col_letter},"Blocked")').border = border
    dash.cell(row=r, column=7, value=f'=COUNTIF(\'{module}\'!{col_letter}:{col_letter},"N/A")').border = border
    
    dash.cell(row=r, column=8, value=f'=IF(C{r}>0,(D{r}+E{r}+F{r})/C{r},0)').border = border
    dash.cell(row=r, column=8).number_format = '0.00%'
    
    dash.cell(row=r, column=9, value=f'=IF(C{r}>0,D{r}/C{r},0)').border = border
    dash.cell(row=r, column=9).number_format = '0.00%'

total_r = start_row + len(sa_sheets)
dash.cell(row=total_r, column=2, value='TOTAL').font = Font(bold=True)
for c, l in enumerate(['C','D','E','F','G']):
    cell = dash.cell(row=total_r, column=c+3, value=f"=SUM({l}3:{l}{total_r-1})")
    cell.font = Font(bold=True)
    cell.border = border

dash.cell(row=total_r, column=8, value=f"=IF(C{total_r}>0,(D{total_r}+E{total_r}+F{total_r})/C{total_r},0)").font = Font(bold=True)
dash.cell(row=total_r, column=8).border = border
dash.cell(row=total_r, column=8).number_format = '0.00%'

dash.cell(row=total_r, column=9, value=f"=IF(C{total_r}>0,D{total_r}/C{total_r},0)").font = Font(bold=True)
dash.cell(row=total_r, column=9).border = border
dash.cell(row=total_r, column=9).number_format = '0.00%'

wb.save(master_path)
print("SUCCESS: Sync SA01-05 and Dashboard update completed.")

import openpyxl

wb = openpyxl.load_workbook('./Feature_02_SA_Tham_So_He_Thong/test case/ProfiX_Master_Test_Cases.xlsx', data_only=True)
print("Dashboard Rows:")
dash = wb['📊 Dashboard']
for i, row in enumerate(dash.iter_rows(values_only=True)):
    if i < 15:
        print(row)

print("\nSA06 Headers:")
sa06 = wb['SA06']
for i, row in enumerate(sa06.iter_rows(values_only=True)):
    if i < 2:
        print(row)

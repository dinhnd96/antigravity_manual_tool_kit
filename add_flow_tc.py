import openpyxl
from copy import copy

file_path = '/Users/mac/antigravity-testing-kit/Feature_02_SA_Tham_So_He_Thong/test case/Test_Cases_SA09_Final_Standard.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['TestCases']

# Find col indexes
header_row = 1
col_map = {}
for idx, cell in enumerate(ws[header_row], start=1):
    if cell.value:
        col_map[str(cell.value).strip()] = idx

def get_col(name):
    return col_map.get(name, -1)

new_flow_case = {
    'TC_ID': 'SA09-FLOW-001',
    'BR_Ref': 'FLOW.01',
    'URD_Ref': 'I.1.1.1',
    'Module': 'SA.09',
    'Feature': 'Luồng E2E',
    'Title': 'Kiểm tra luồng liên hoàn (Life-cycle): Thêm mới, Tìm kiếm, Sửa và Xóa bản ghi Nhóm code phí',
    'Type': 'Happy',
    'Category': 'Smoke',
    'Priority': 'P1',
    'Precondition': '1. User Maker được phân quyền đầy đủ chức năng Thêm mới, Sửa, Xóa, Tìm kiếm.\n2. Hệ thống chứa sẵn dữ liệu Code phí hợp lệ chưa được gán nhóm.',
    'Steps': '1. [Thêm mới] Tạo 1 Nhóm code phí hợp lệ và nhấn Xác nhận.\n2. [Tìm kiếm] Lọc theo Mã nhóm vừa tạo.\n3. [Sửa] Nhấn icon Sửa, cập nhật "Tên nhóm" và "Mức độ ưu tiên", nhấn Xác nhận.\n4. [Xóa] Tại lưới, nhấn icon Xóa dòng vừa sửa.',
    'Expected': '(i) Logic: Dữ liệu luân chuyển đúng luồng (CRUD), liên kết dữ liệu không bị hỏng (DB Reference). Cuối cùng dữ liệu bị đánh dấu Xóa.\n(ii) UI: Sau mỗi step đều báo Toast xanh thành công. Lưới co giãn cập nhật đúng data sau Thêm/Sửa và biến mất sau Xóa.\n(iii) Trạng thái/Audit: Sinh trạng thái nháp ban đầu, sau đó cập nhật thông tin và tiến tới trạng thái Xóa. Audit Log ghi nhận đầy đủ 3 actions gắn với Maker.\n(iv) File/Email: (Không có).',
    'Trace_ID': 'E2E-LIFECYCLE-001',
    'Note': 'TC Luồng E2E nhằm kiểm tra độ liền mạch và tính toàn vẹn dữ liệu khi đi ngang qua hết các button của màn hình.'
}

target_row = ws.max_row + 1
source_row = 2

for col_name, val in new_flow_case.items():
    c_idx = get_col(col_name)
    if c_idx != -1:
        new_cell = ws.cell(row=target_row, column=c_idx)
        new_cell.value = val
        source_cell = ws.cell(row=source_row, column=c_idx)
        if source_cell.has_style:
            new_cell.font = copy(source_cell.font)
            new_cell.border = copy(source_cell.border)
            new_cell.fill = copy(source_cell.fill)
            new_cell.number_format = copy(source_cell.number_format)
            new_cell.protection = copy(source_cell.protection)
            new_cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')

wb.save(file_path)
print("Đã bổ sung Test Case FLOW thành công.")

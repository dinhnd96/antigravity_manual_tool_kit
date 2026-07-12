"""US31 - Batch 5: NFR (SC-36→SC-38) + Cross-check"""
import openpyxl
from openpyxl.styles import Alignment, Border, Side
import os

FEATURE = "US31 - Báo cáo tổng doanh thu phí"

TCS = [
    # TC-041 (SC-36)
    ("US31-TC-041", "SC-36",
     "Mục \"Yêu cầu nghiệp vụ\", Navigation – *\"Báo cáo >> Báo cáo tổng doanh thu phí\"* + QTC-10",
     FEATURE, "Phân quyền",
     "NFR: User không có quyền truy cập — Menu bị ẩn hoặc thông báo không có quyền",
     "NFR", "High",
     "1. User đã đăng nhập hệ thống ProfiX.\n2. User KHÔNG có quyền truy cập menu Báo cáo >> Báo cáo tổng doanh thu phí.",
     "1. Truy cập menu Báo cáo.\n"
     "2. Quan sát danh sách sub-menu.",
     "(i) Nghiệp vụ/Logic: Hệ thống không cho phép truy cập tính năng Báo cáo tổng doanh thu phí.\n"
     "(ii) UI: Menu 'Báo cáo tổng doanh thu phí' bị ẩn trong danh sách sub-menu, "
     "hoặc khi truy cập URL trực tiếp hiển thị thông báo 'Bạn không có quyền truy cập'.",
     "[Theo QTC-10]"),

    # TC-042 (SC-37)
    ("US31-TC-042", "SC-37",
     "Bảng mô tả nút STT 1 (Tra cứu) – Best practice NFR: Chống spam click, chỉ xử lý 1 request/click",
     FEATURE, "Tra cứu báo cáo",
     "NFR: Spam click nút Tra cứu — Hệ thống chỉ xử lý 1 request",
     "NFR", "Medium",
     "1. User đã đăng nhập hệ thống ProfiX.\n"
     "2. User có quyền truy cập Báo cáo tổng doanh thu phí.\n"
     "3. Đã nhập Từ ngày = '01/01/2025', Đến ngày = '31/03/2025'.",
     "1. Nhấn nút 'Tra cứu' liên tiếp 3-5 lần nhanh (double/triple click).\n"
     "2. Quan sát network request (F12 > Network tab).",
     "(i) Nghiệp vụ/Logic: Hệ thống chỉ gửi DUY NHẤT 1 request lên BE. "
     "Không gửi request trùng lặp gây overload server.\n"
     "(ii) UI: Nút 'Tra cứu' bị disabled tạm thời sau click đầu tiên (hoặc hiển thị loading), "
     "sau đó enable lại khi có response. Lưới hiển thị kết quả 1 lần duy nhất.",
     ""),

    # TC-043 (SC-38)
    ("US31-TC-043", "SC-38",
     "Bảng mô tả nút STT 3 (Tải xuống) – Best practice NFR: Chống spam click, chỉ xử lý 1 request/click",
     FEATURE, "Tải xuống báo cáo",
     "NFR: Spam click nút Tải xuống — Hệ thống chỉ tải 1 file",
     "NFR", "Medium",
     "1. User đã đăng nhập hệ thống ProfiX.\n"
     "2. Đã tra cứu thành công, lưới đang hiển thị kết quả.",
     "1. Nhấn nút 'Tải xuống' → chọn 'Excel'.\n"
     "2. Ngay lập tức nhấn 'Tải xuống' → chọn 'Excel' lần 2 (trong vòng 1 giây).\n"
     "3. Kiểm tra số file đã tải về.",
     "(i) Nghiệp vụ/Logic: Hệ thống chỉ tải DUY NHẤT 1 file. "
     "Không tạo nhiều request tải file trùng lặp.\n"
     "(ii) UI: Nút 'Tải xuống' bị disabled tạm thời sau click đầu tiên. "
     "Chỉ 1 file .xlsx được tải về thư mục Downloads.",
     ""),
]

fpath = os.path.join(os.path.dirname(os.path.dirname(__file__)), "US31_TestCases.xlsx")
wb = openpyxl.load_workbook(fpath)
ws = wb["Test Cases"]
start_row = ws.max_row + 1
data_align = Alignment(vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))

for ri, tc in enumerate(TCS, start_row):
    for ci, val in enumerate(tc, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.alignment = data_align
        cell.border = thin_border

ws.auto_filter.ref = f"A1:S{ws.max_row}"
wb.save(fpath)

total = ws.max_row - 1
print(f"✅ Batch 5 done: {len(TCS)} TC (NFR)")
print(f"📄 Total: {total} TC")

# Cross-check: verify all 43 SC covered
all_sc_refs = set()
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
    for cell in row:
        if cell.value:
            all_sc_refs.add(cell.value.strip())

expected_scs = set(f"SC-{i:02d}" for i in range(1, 44))
missing = expected_scs - all_sc_refs
extra = all_sc_refs - expected_scs

print(f"\n🔍 Cross-check:")
print(f"   SC covered: {len(all_sc_refs)}/43")
if missing:
    print(f"   ❌ MISSING: {sorted(missing)}")
else:
    print(f"   ✅ Không sót SC nào!")
if extra:
    print(f"   ➕ Extra: {sorted(extra)}")

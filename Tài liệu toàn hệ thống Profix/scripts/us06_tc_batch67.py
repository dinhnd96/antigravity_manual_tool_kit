#!/usr/bin/env python3
"""US06 TC Batch 6+7: Data Integrity + NFR (FINAL)"""
import openpyxl
from openpyxl.styles import Alignment, Border, Side
import os

FPATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "US06_TestCases.xlsx")
wb = openpyxl.load_workbook(FPATH)
ws = wb["Test Cases"]
wrap = Alignment(wrap_text=True, vertical="top")
thin = Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
FEAT = "Khai báo Biểu phí"

data = [
    # ===== DATA INTEGRITY =====
    ("US06-TC-066","SC-50",'BA QA-03.1 – "khoảng thời gian hiệu lực không trùng"',FEAT,"Toàn vẹn dữ liệu",
     "1 Code phí gán nhiều Biểu phí: khoảng HL KHÔNG trùng → thành công","Data Integrity","High",
     "1. User đăng nhập quyền Maker\n2. CP001 đã gán cho BP-A (HL: 01/01-30/06/2026)\n3. Đang Thêm mới BP-B",
     "1. Thêm mới BP-B với HL=01/07/2026, HHL=31/12/2026\n2. Gán CP001\n3. Click 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Khoảng HL không trùng → cho phép gán. CP001 thuộc cả BP-A và BP-B.\n(ii) UI: Lưu thành công.","[BA QA-03.1]"),

    ("US06-TC-067","SC-51",'BA QA-03.1 – "khoảng thời gian hiệu lực trùng"',FEAT,"Toàn vẹn dữ liệu",
     "1 Code phí gán nhiều Biểu phí: khoảng HL TRÙNG → BE chặn","Data Integrity","High",
     "1. User đăng nhập quyền Maker\n2. CP001 đã gán cho BP-A (HL: 01/01-30/06/2026)\n3. Đang Thêm mới BP-C",
     "1. Thêm mới BP-C với HL=01/03/2026, HHL=31/08/2026 (trùng 01/03-30/06)\n2. Gán CP001\n3. Click 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Khoảng HL trùng → BE chặn, không cho phép lưu.\n(ii) UI: Hiển thị thông báo lỗi trùng khoảng hiệu lực.","[BA QA-03.1]"),

    ("US06-TC-068","SC-52",'Mục "Thêm mới" P37-P40 – "thay đổi chỉ trong scope Biểu phí"',FEAT,"Toàn vẹn dữ liệu",
     "Sửa Code phí gốc tại Biểu phí: Code phí gốc tại Danh mục SPDV bị ảnh hưởng","Data Integrity","High",
     "1. User đăng nhập quyền Maker\n2. BP-030 có CP001 (Giá trị ĐK gốc = '20000')",
     "1. Chỉnh sửa BP-030, sửa CP001: Giá trị ĐK = '50000'\n2. Xác nhận → Checker duyệt\n3. Mở Danh mục SPDV → Xem CP001",
     "(i) Nghiệp vụ/Logic: CP001 tại BP-030 có Giá trị = 50000. CP001 tại Danh mục SPDV cũng thay đổi = 50000.\n(ii) UI: Cả 2 màn hình hiển thị giá trị đồng bộ (50000).",""),

    ("US06-TC-069","SC-53",'Mục "Thêm mới" P25 – "thay thế tất cả"',FEAT,"Toàn vẹn dữ liệu",
     "Upload thay thế: sau upload, lưới chỉ hiển thị Code phí từ file","Data Integrity","High",
     "1. User đăng nhập quyền Maker\n2. Đang Thêm mới, đã tích CP001, CP002\n3. File Excel chứa CP003",
     "1. Upload file Excel chứa CP003\n2. Kiểm tra lưới Thông tin chi tiết",
     "(i) Nghiệp vụ/Logic: CP001, CP002 bị xóa hoàn toàn. Chỉ còn CP003 từ file.\n(ii) UI: Lưới hiển thị 1 dòng CP003. Cây SPDV cập nhật tương ứng.",""),

    ("US06-TC-070","SC-54",'Mục "Chỉnh sửa" P61',FEAT,"Toàn vẹn dữ liệu",
     "Chỉnh sửa Chưa hiệu lực: bỏ gắn Code phí → không còn liên kết sau duyệt","Data Integrity","Medium",
     "1. User đăng nhập quyền Maker\n2. BP-031 Chưa hiệu lực, có CP001, CP002",
     "1. Chỉnh sửa BP-031\n2. Xóa CP002 khỏi lưới\n3. Xác nhận → Checker duyệt\n4. Mở BP-031",
     "(i) Nghiệp vụ/Logic: Sau duyệt, BP-031 chỉ còn CP001. CP002 không liên kết.\n(ii) UI: Lưới Thông tin chi tiết BP-031 hiển thị 1 dòng CP001.",""),

    ("US06-TC-071","SC-78",'Mục "Thêm mới" P46 + "Chỉnh sửa" P67',FEAT,"Toàn vẹn dữ liệu",
     "Sau Checker duyệt: Code phí đồng bộ trên lưới Danh mục SPDV và Biểu phí","Data Integrity","Medium",
     "1. Checker vừa duyệt tác vụ Thêm mới BP-032 có CP001",
     "1. Mở lưới Danh mục Biểu phí → xem BP-032 → kiểm tra Code phí\n2. Mở Danh mục SPDV → tìm CP001 → kiểm tra liên kết Biểu phí",
     "(i) Nghiệp vụ/Logic: CP001 hiển thị đồng bộ tại cả 2 màn hình.\n(ii) UI: BP-032 hiển thị CP001. Danh mục SPDV ghi nhận CP001 thuộc BP-032.",""),

    ("US06-TC-072","SC-79",'QTC-08 – "Lịch sử tác động"',FEAT,"Toàn vẹn dữ liệu",
     "Xem Lịch sử tác động: hiển thị đúng 4 trường theo QTC-08","Data Integrity","Medium",
     "1. BP-033 đã qua Thêm mới + Chỉnh sửa (đã duyệt)",
     "1. Mở chi tiết BP-033\n2. Click tab/link 'Lịch sử tác động'",
     "(i) Nghiệp vụ/Logic: Ghi nhận 2 bản ghi: Thêm mới và Chỉnh sửa. Đúng 4 cột.\n(ii) UI: Hiển thị bảng: Ngày cập nhật | Tác động | Người cập nhật | Người phê duyệt.","[Theo QTC-08]"),

    # ===== NFR =====
    ("US06-TC-073","SC-55",'QTC-12 + NFR spam-click Thêm mới',FEAT,"NFR – Spam click",
     "Double-click Xác nhận khi Thêm mới → hệ thống chỉ xử lý 1 lần","NFR","High",
     "1. User đăng nhập quyền Maker\n2. Đã nhập đủ thông tin Thêm mới hợp lệ + gán Code phí",
     "1. Click nhanh 2 lần liên tiếp nút 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Chỉ tạo 1 bản ghi Chờ duyệt. Không tạo trùng.\n(ii) UI: Toast thành công 1 lần. Nút disable sau click đầu tiên.",""),

    ("US06-TC-074","SC-56",'QTC-12 + NFR spam-click Chỉnh sửa',FEAT,"NFR – Spam click",
     "Double-click Xác nhận khi Chỉnh sửa → chỉ xử lý 1 lần","NFR","High",
     "1. User đăng nhập quyền Maker\n2. Đang chỉnh sửa BP-034",
     "1. Sửa Tên = 'Updated'\n2. Click nhanh 2 lần 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Chỉ 1 tác vụ Chỉnh sửa Chờ duyệt.\n(ii) UI: Nút disable sau click đầu.",""),

    ("US06-TC-075","SC-57",'NFR spam-click Xóa Code phí',FEAT,"NFR – Spam click",
     "Double-click Xóa Code phí khỏi lưới → chỉ xóa 1 lần","NFR","Medium",
     "1. Đang Thêm mới, lưới có CP001, CP002",
     "1. Click nhanh 2 lần nút 'Xóa' tại CP001",
     "(i) Nghiệp vụ/Logic: Chỉ xóa CP001 1 lần, không ảnh hưởng CP002.\n(ii) UI: CP001 biến mất. CP002 vẫn hiển thị.",""),

    ("US06-TC-076","SC-58",'QTC-12 + NFR spam-click Phê duyệt',FEAT,"NFR – Spam click",
     "Double-click Phê duyệt tại Tác vụ chờ duyệt → chỉ xử lý 1 lần","NFR","High",
     "1. User đăng nhập quyền Checker\n2. Tác vụ BP-035 Chờ duyệt",
     "1. Click nhanh 2 lần nút 'Phê duyệt'",
     "(i) Nghiệp vụ/Logic: Chỉ phê duyệt 1 lần.\n(ii) UI: Nút disable. Toast 'Phê duyệt thành công' hiển thị 1 lần.",""),

    ("US06-TC-077","SC-59",'QTC-12 + NFR spam-click Từ chối',FEAT,"NFR – Spam click",
     "Double-click Từ chối tại Tác vụ chờ duyệt → chỉ xử lý 1 lần","NFR","Medium",
     "1. User đăng nhập quyền Checker\n2. Tác vụ BP-036 Chờ duyệt",
     "1. Click nhanh 2 lần nút 'Từ chối'\n2. Nhập lý do = 'Sai thông tin'",
     "(i) Nghiệp vụ/Logic: Chỉ từ chối 1 lần.\n(ii) UI: Nút disable. Toast 'Từ chối thành công' hiển thị 1 lần.",""),

    ("US06-TC-078","SC-60",'QTC-12 + NFR spam-click Xóa Pending',FEAT,"NFR – Spam click",
     "Double-click Xóa tại Tác vụ Pending → chỉ xóa 1 lần","NFR","Medium",
     "1. User đăng nhập quyền Maker\n2. Tác vụ BP-037 ở Từ chối duyệt tại Tác vụ Pending",
     "1. Click nhanh 2 lần nút 'Xóa'",
     "(i) Nghiệp vụ/Logic: Chỉ xóa 1 lần.\n(ii) UI: Bản ghi biến mất khỏi Tác vụ Pending.",""),

    ("US06-TC-079","SC-72",'QTC-10 – "Ma trận dữ liệu theo Khối KHCN"',FEAT,"NFR – Phân quyền",
     "User Khối KHCN chỉ thấy Biểu phí có Code phí Loại KH = KHCN/DNSN/CBNV","NFR","High",
     "1. User đăng nhập thuộc Khối KHCN\n2. Hệ thống có BP-A (Code phí KHCN), BP-B (Code phí KHTC)",
     "1. Mở Danh mục Biểu phí\n2. Kiểm tra danh sách hiển thị",
     "(i) Nghiệp vụ/Logic: Chỉ hiển thị BP-A. BP-B (KHTC) không xuất hiện.\n(ii) UI: Lưới chỉ hiển thị Biểu phí thuộc phạm vi Khối KHCN.","[Theo QTC-10]"),

    ("US06-TC-080","SC-73",'QTC-10 – "Ma trận dữ liệu theo Khối KHDN"',FEAT,"NFR – Phân quyền",
     "User Khối KHDN chỉ thấy Biểu phí có Code phí Loại KH = KHTC","NFR","High",
     "1. User đăng nhập thuộc Khối KHDN\n2. Hệ thống có BP-A (KHCN), BP-B (KHTC)",
     "1. Mở Danh mục Biểu phí\n2. Kiểm tra danh sách",
     "(i) Nghiệp vụ/Logic: Chỉ hiển thị BP-B (KHTC). BP-A không xuất hiện.\n(ii) UI: Lưới chỉ hiển thị Biểu phí thuộc phạm vi Khối KHDN.","[Theo QTC-10]"),
]

start = ws.max_row + 1
for i, row in enumerate(data):
    for j, val in enumerate(row):
        c = ws.cell(row=start+i, column=j+1, value=val)
        c.alignment = wrap; c.border = thin
wb.save(FPATH)

# Cross-check
all_sc = set()
for r in range(2, ws.max_row+1):
    sc = ws.cell(row=r, column=2).value
    if sc: all_sc.add(sc)
print(f"✅ Batch 6+7 (Data Integrity + NFR): {len(data)} TC → {FPATH}")
print(f"📊 TỔNG: {ws.max_row-1} TC | SC duy nhất: {len(all_sc)}")
print(f"📋 SC list: {sorted(all_sc)}")

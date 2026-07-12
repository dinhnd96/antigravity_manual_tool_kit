"""US31 - Batch 3: UI/UX & Field Validation (SC-19→SC-27, SC-39, SC-41) → Append"""
import openpyxl
from openpyxl.styles import Alignment, Border, Side
import os

FEATURE = "US31 - Báo cáo tổng doanh thu phí"
MOD_DK = "Điều kiện tìm kiếm"
MOD_LUOI = "Lưới kết quả"
MOD_TONG = "Tổng cộng"
MOD_MH = "Màn hình"

PRECON = (
    "1. User đã đăng nhập hệ thống ProfiX.\n"
    "2. User có quyền truy cập menu Báo cáo >> Báo cáo tổng doanh thu phí.\n"
    "3. Đã truy cập màn hình Báo cáo tổng doanh thu phí."
)
PRECON_DATA = PRECON + "\n4. Đã tra cứu thành công, lưới đang hiển thị kết quả."

TCS = [
    # TC-019 (SC-19)
    ("US31-TC-019", "SC-19",
     "Bảng mô tả trường STT 3 – *\"Người dùng nhấn vào Biểu phí, hệ thống hiển thị danh sách biểu phí (format: Mã biểu phí - Tên biểu phí)\"* + QTC-01.1",
     FEATURE, MOD_DK,
     "Combobox Biểu phí — Hiển thị đúng format, hỗ trợ search-as-you-type",
     "UI/UX & Field Validation", "Medium",
     PRECON,
     "1. Nhấn vào trường Biểu phí.\n"
     "2. Quan sát danh sách dropdown.\n"
     "3. Gõ 'BP001' vào ô tìm kiếm.\n"
     "4. Quan sát kết quả lọc.",
     "(i) Nghiệp vụ/Logic: Hệ thống hiển thị danh sách biểu phí với format 'Mã biểu phí - Tên biểu phí'. "
     "Khi gõ 'BP001', chỉ hiển thị các biểu phí có mã chứa 'BP001'.\n"
     "(ii) UI: Combobox hiển thị dropdown với danh sách. Kết quả lọc realtime khi gõ text.",
     "[Theo QTC-01.1]"),

    # TC-020 (SC-20)
    ("US31-TC-020", "SC-20",
     "Bảng mô tả trường STT 4 – *\"Người dùng nhấn vào Code phí, hệ thống hiển thị danh sách Code phí (format: Mã code phí - Tên code phí)\"* + QTC-01.1",
     FEATURE, MOD_DK,
     "Combobox Code phí — Hiển thị đúng format, hỗ trợ search-as-you-type",
     "UI/UX & Field Validation", "Medium",
     PRECON,
     "1. Nhấn vào trường Code phí.\n"
     "2. Quan sát danh sách dropdown.\n"
     "3. Gõ 'Phí chuyển' vào ô tìm kiếm.\n"
     "4. Quan sát kết quả lọc.",
     "(i) Nghiệp vụ/Logic: Hệ thống hiển thị danh sách code phí với format 'Mã code phí - Tên code phí'. "
     "Khi gõ 'Phí chuyển', chỉ hiển thị các code phí có tên chứa 'Phí chuyển'.\n"
     "(ii) UI: Combobox hiển thị dropdown. Kết quả lọc realtime.",
     "[Theo QTC-01.1]"),

    # TC-021 (SC-21)
    ("US31-TC-021", "SC-21",
     "Bảng mô tả trường STT 5 – *\"Cho phép người dùng chọn loại tính phí: Theo giao dịch/Theo định kỳ\"* + QTC-01.2",
     FEATURE, MOD_DK,
     "Dropdown Loại tính phí — Hiển thị đúng 2 giá trị, chỉ chọn 1",
     "UI/UX & Field Validation", "Medium",
     PRECON,
     "1. Nhấn vào trường Loại tính phí.\n"
     "2. Quan sát dropdown.",
     "(i) Nghiệp vụ/Logic: Dropdown hiển thị đúng 2 giá trị: 'Theo giao dịch' và 'Theo định kỳ'. Chỉ cho phép chọn 1.\n"
     "(ii) UI: Dropdown list mở ra, hiển thị 2 lựa chọn. Không cho phép nhập text tự do.",
     "[Theo QTC-01.2]"),

    # TC-022 (SC-22)
    ("US31-TC-022", "SC-22",
     "Bảng mô tả trường STT 8-19 + BA trả lời QA-04.2 – *\"Ưu tiên áp dụng theo mô tả chi tiết\"*",
     FEATURE, MOD_LUOI,
     "Lưới hiển thị đúng 12 cột theo mô tả chi tiết",
     "UI/UX & Field Validation", "High",
     PRECON_DATA,
     "1. Quan sát header các cột trên lưới danh sách.",
     "(i) Nghiệp vụ/Logic: Lưới hiển thị đúng 12 cột theo thứ tự mô tả: STT, Mã Chi nhánh, Tên chi nhánh, "
     "Loại tiền, Loại tính phí, Biểu phí, Code phí, Tên phí, Doanh thu phí (Nguyên tệ), "
     "Doanh thu phí (VND), Từ ngày, Đến ngày.\n"
     "(ii) UI: Tất cả 12 cột hiển thị đúng tên header. Không thừa, không thiếu cột.",
     ""),

    # TC-023 (SC-23)
    ("US31-TC-023", "SC-23",
     "Bảng mô tả trường STT 16-17 (Number) + QTC-01.4 – *\"Phân cách hàng nghìn, VND/JPY không thập phân\"*",
     FEATURE, MOD_LUOI,
     "Cột Doanh thu phí hiển thị đúng format Number — Phân cách hàng nghìn, ngoại lệ VND/JPY",
     "UI/UX & Field Validation", "High",
     PRECON_DATA + "\n5. Lưới có dữ liệu gồm Loại tiền VND, USD, JPY.",
     "1. Quan sát cột Doanh thu phí (Nguyên tệ) dòng Loại tiền = 'USD' (VD: 12345.67).\n"
     "2. Quan sát cột Doanh thu phí (VND) dòng Loại tiền = 'VND' (VD: 5427167).\n"
     "3. Quan sát cột Doanh thu phí (Nguyên tệ) dòng Loại tiền = 'JPY' (VD: 98765).",
     "(i) Nghiệp vụ/Logic: USD hiển thị 2 số thập phân (12,345.67). VND không có thập phân (5,427,167). "
     "JPY không có thập phân (98,765).\n"
     "(ii) UI: Giá trị Number có phân cách hàng nghìn bằng dấu phẩy. Thập phân bằng dấu chấm (nếu có).",
     "[Theo QTC-01.4]"),

    # TC-024 (SC-24)
    ("US31-TC-024", "SC-24",
     "Bảng mô tả trường STT 18-19 (Date) + BA trả lời QA-04.4 – *\"Ưu tiên áp dụng theo mô tả chi tiết\"* + QTC-01.5",
     FEATURE, MOD_LUOI,
     "Cột Từ ngày/Đến ngày trên lưới hiển thị đúng format dd/mm/yyyy (không có giờ)",
     "UI/UX & Field Validation", "Medium",
     PRECON_DATA,
     "1. Quan sát cột Từ ngày và Đến ngày trên lưới kết quả.",
     "(i) Nghiệp vụ/Logic: Cột hiển thị ngày đúng format dd/mm/yyyy, không bao gồm giờ phút giây.\n"
     "(ii) UI: VD: '01/01/2025', '31/03/2025'. Không hiển thị '00:00:00 - 01/01/2025'.",
     "[Theo QTC-01.5]"),

    # TC-025 (SC-25)
    ("US31-TC-025", "SC-25",
     "QTC-01.7 – *\"Trường hợp các trường dữ liệu không có giá trị, hệ thống hiển thị blank (rỗng)\"*",
     FEATURE, MOD_LUOI,
     "Trường rỗng trên lưới hiển thị blank (không hiển thị -, N/A, Null)",
     "UI/UX & Field Validation", "Low",
     PRECON_DATA + "\n5. Lưới có bản ghi mà một số trường (VD: Tên phí) không có giá trị.",
     "1. Quan sát các ô trống trên lưới (VD: cột Tên phí không có dữ liệu).",
     "(i) Nghiệp vụ/Logic: Trường không có giá trị hiển thị rỗng (blank).\n"
     "(ii) UI: Ô trống trên lưới KHÔNG hiển thị ký tự '-', 'N/A', hoặc 'Null'. Chỉ để trống.",
     "[Theo QTC-01.7]"),

    # TC-026 (SC-26)
    ("US31-TC-026", "SC-26",
     "Bảng mô tả trường STT 20 (Tổng cộng, mục 'Thông tin khác') + BA trả lời QA-01.5 + QA-04.3",
     FEATURE, MOD_TONG,
     "Tổng cộng hiển thị ở vùng riêng (Thông tin khác), sticky footer, tính toàn bộ DL",
     "UI/UX & Field Validation", "High",
     PRECON_DATA + "\n5. Kết quả có >50 bản ghi (2+ trang).",
     "1. Quan sát vị trí dòng Tổng cộng.\n"
     "2. Cuộn lưới xuống cuối.\n"
     "3. Chuyển sang trang 2.\n"
     "4. Quan sát lại dòng Tổng cộng.",
     "(i) Nghiệp vụ/Logic: Tổng cộng = tổng Doanh thu phí VND của TOÀN BỘ dữ liệu (xuyên suốt mọi trang), "
     "không chỉ trang hiện tại.\n"
     "(ii) UI: Dòng Tổng cộng hiển thị ở vùng riêng (Thông tin khác) ngoài lưới, dạng footer cố định (sticky). "
     "Giá trị không thay đổi khi chuyển trang. Format Number theo QTC-01.4.",
     "[Theo QTC-01.4]"),

    # TC-027 (SC-27)
    ("US31-TC-027", "SC-27",
     "QTC-04 – *\"Mặc định không hiển thị danh sách kết quả khi chưa nhập điều kiện tìm kiếm\"*",
     FEATURE, MOD_MH,
     "Truy cập màn hình lần đầu — Lưới mặc định rỗng, chờ nhấn Tra cứu",
     "UI/UX & Field Validation", "Medium",
     "1. User đã đăng nhập hệ thống ProfiX.\n2. User có quyền truy cập Báo cáo tổng doanh thu phí.",
     "1. Truy cập menu Báo cáo >> Báo cáo tổng doanh thu phí.\n"
     "2. Quan sát lưới danh sách.",
     "(i) Nghiệp vụ/Logic: Lưới mặc định không hiển thị kết quả. Chờ user nhập điều kiện và nhấn 'Tra cứu'.\n"
     "(ii) UI: Lưới rỗng (có thể hiển thị message 'Không có dữ liệu' hoặc trống). "
     "Các trường điều kiện ở trạng thái mặc định.",
     "[Theo QTC-04]"),

    # TC-028 (SC-39) — CIF removed regression
    ("US31-TC-028", "SC-39",
     "BA xác nhận QA-01.1 – *\"Trường dư, BA đã cập nhật US v2 — xóa nút Tra cứu CIF\"*",
     FEATURE, MOD_MH,
     "Regression: Nút 'Tra cứu CIF' đã bị xóa khỏi màn hình v2",
     "UI/UX & Field Validation", "Medium",
     "1. User đã đăng nhập hệ thống ProfiX.\n2. Hệ thống đang chạy version v2 (US31 updated).",
     "1. Truy cập menu Báo cáo >> Báo cáo tổng doanh thu phí.\n"
     "2. Quan sát toàn bộ vùng Điều kiện tìm kiếm và các nút chức năng.",
     "(i) Nghiệp vụ/Logic: Nút 'Tra cứu CIF' KHÔNG còn tồn tại trên màn hình. "
     "Không có trường Mã CIF trong vùng điều kiện tìm kiếm.\n"
     "(ii) UI: Chỉ có 3 nút chức năng: Tra cứu, Xóa tra cứu, Tải xuống. Không có nút Tra cứu CIF.",
     "Regression test sau khi BA loại bỏ tính năng CIF khỏi US31 v2"),

    # TC-029 (SC-41) — Combobox Chi nhánh
    ("US31-TC-029", "SC-41",
     "Bảng mô tả trường STT 2 – *\"Cho phép người dùng tìm kiếm và chọn chi nhánh\"* + QTC-01.1",
     FEATURE, MOD_DK,
     "Combobox Mã Chi nhánh — Hỗ trợ tìm kiếm theo mã hoặc tên, chỉ chọn 1",
     "UI/UX & Field Validation", "Medium",
     PRECON,
     "1. Nhấn vào trường Mã Chi nhánh.\n"
     "2. Gõ 'Hà Nội' vào ô tìm kiếm.\n"
     "3. Quan sát kết quả lọc.\n"
     "4. Chọn '001 - CN Hà Nội'.",
     "(i) Nghiệp vụ/Logic: Hệ thống lọc và hiển thị chi nhánh chứa 'Hà Nội'. Chỉ cho chọn 1 chi nhánh.\n"
     "(ii) UI: Combobox hiển thị dropdown. Khi gõ text, kết quả lọc realtime. Sau khi chọn, hiển thị '001 - CN Hà Nội'.",
     "[Theo QTC-01.1]"),
]

fpath = os.path.join(os.path.dirname(os.path.dirname(__file__)), "US31_TestCases.xlsx")
wb = openpyxl.load_workbook(fpath)
ws = wb["Test Cases"]
start_row = ws.max_row + 1
data_align = Alignment(vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))

for ri, tc in enumerate(TCS, start_row):
    for ci, val in enumerate(tc, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.alignment = data_align
        cell.border = thin_border

ws.auto_filter.ref = f"A1:S{ws.max_row}"
wb.save(fpath)
print(f"✅ Batch 3 done: {len(TCS)} TC (UI/UX)")
print(f"📄 Total: {ws.max_row - 1} TC")

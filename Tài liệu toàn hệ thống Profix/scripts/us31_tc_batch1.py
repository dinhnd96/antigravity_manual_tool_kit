"""US31 - Batch 1: Happy Path (SC-01 → SC-09) → Create Excel"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

FEATURE = "US31 - Báo cáo tổng doanh thu phí"
MOD_TRA = "Tra cứu báo cáo"
MOD_TAI = "Tải xuống báo cáo"

COLS = ["TC_ID","SC_Ref","Reference","Feature","Module","Title","Type","Priority",
        "Precondition","Steps","Expected","Note",
        "Tester","Ngày Test","Kết quả","Bug ID","Retest","Retest Result","Ghi chú QA"]

PRECON_BASE = (
    "1. User đã đăng nhập hệ thống ProfiX.\n"
    "2. User có quyền truy cập menu Báo cáo >> Báo cáo tổng doanh thu phí.\n"
    "3. Hệ thống có dữ liệu giao dịch thu phí đã ghi nhận."
)

TCS = [
    # TC-001 (SC-01)
    ("US31-TC-001", "SC-01",
     "Mục \"Yêu cầu nghiệp vụ\" – *\"Hiển thị tổng doanh thu phí nguyên tệ đã thu được của mỗi nhóm Chi nhánh + Loại tiền + Loại tính phí + Biểu phí + Code phí\"* + BA xác nhận QA-03.2",
     FEATURE, MOD_TRA,
     "Tra cứu đầy đủ điều kiện hợp lệ — Lưới hiển thị kết quả group by đúng tổ hợp",
     "Happy Path", "High",
     PRECON_BASE + "\n4. User không thuộc Khối KHCN/KHDN/KHDNL (Combobox Khối cho phép chọn).",
     "1. Truy cập menu Báo cáo >> Báo cáo tổng doanh thu phí.\n"
     "2. Chọn Khối = 'KHCN'.\n"
     "3. Chọn Mã Chi nhánh = '001 - CN Hà Nội'.\n"
     "4. Chọn Biểu phí = 'BP001 - Biểu phí chuẩn'.\n"
     "5. Chọn Code phí = 'FEE001 - Phí chuyển tiền'.\n"
     "6. Chọn Loại tính phí = 'Theo giao dịch'.\n"
     "7. Chọn Từ ngày = '01/01/2025', Đến ngày = '31/03/2025'.\n"
     "8. Nhấn nút 'Tra cứu'.",
     "(i) Nghiệp vụ/Logic: Hệ thống tìm kiếm và trả kết quả đúng theo tất cả điều kiện đã chọn. "
     "Mỗi dòng trên lưới đại diện cho 1 tổ hợp duy nhất (Chi nhánh + Loại tiền + Loại tính phí + Biểu phí + Code phí). "
     "Doanh thu phí = tổng các giao dịch thuộc tổ hợp đó.\n"
     "(ii) UI: Lưới danh sách hiển thị kết quả. Dòng Tổng cộng cập nhật giá trị VND. "
     "Phân trang hiển thị nếu >50 bản ghi.",
     "[Theo QTC-04, QTC-06]"),

    # TC-002 (SC-02)
    ("US31-TC-002", "SC-02",
     "QTC-04 – *\"Nếu không nhập điều kiện tìm kiếm không bắt buộc, hệ thống tìm kiếm tất cả\"*",
     FEATURE, MOD_TRA,
     "Tra cứu chỉ với Từ ngày + Đến ngày (bỏ trống điều kiện không bắt buộc) — Hệ thống trả toàn bộ dữ liệu",
     "Happy Path", "High",
     PRECON_BASE,
     "1. Truy cập menu Báo cáo >> Báo cáo tổng doanh thu phí.\n"
     "2. Để trống Khối, Mã Chi nhánh, Biểu phí, Code phí, Loại tính phí.\n"
     "3. Chọn Từ ngày = '01/01/2025', Đến ngày = '31/03/2025'.\n"
     "4. Nhấn nút 'Tra cứu'.",
     "(i) Nghiệp vụ/Logic: Hệ thống trả về tất cả dữ liệu doanh thu phí trong khoảng 01/01/2025 – 31/03/2025 "
     "không phân biệt Khối, Chi nhánh, Biểu phí, Code phí, Loại tính phí.\n"
     "(ii) UI: Lưới hiển thị toàn bộ kết quả. Dòng Tổng cộng hiển thị tổng VND toàn bộ.",
     "[Theo QTC-04]"),

    # TC-003 (SC-03)
    ("US31-TC-003", "SC-03",
     "Bảng mô tả trường STT 1 (Khối) – *\"Nếu người dùng thuộc khối KHCN [...] hiển thị mặc định khối của user đó và không cho phép sửa\"* + BA xác nhận Combobox disabled",
     FEATURE, MOD_TRA,
     "User thuộc Khối KHCN — Combobox Khối hiển thị KHCN và disabled",
     "Happy Path", "High",
     "1. User đã đăng nhập, thuộc Khối KHCN.\n2. User có quyền truy cập Báo cáo tổng doanh thu phí.",
     "1. Truy cập menu Báo cáo >> Báo cáo tổng doanh thu phí.\n"
     "2. Quan sát trường Khối.\n"
     "3. Chọn Từ ngày = '01/01/2025', Đến ngày = '31/03/2025'.\n"
     "4. Nhấn nút 'Tra cứu'.",
     "(i) Nghiệp vụ/Logic: Hệ thống chỉ trả về dữ liệu thuộc Khối KHCN. "
     "Không thể tra cứu dữ liệu Khối khác.\n"
     "(ii) UI: Combobox Khối hiển thị 'KHCN', trạng thái disabled (greyed out), không cho thao tác.",
     "[Theo QTC-10]"),

    # TC-004 (SC-04)
    ("US31-TC-004", "SC-04",
     "Bảng mô tả trường STT 1 (Khối) – *\"Nếu người dùng không thuộc [...] hệ thống hiển thị các khối và cho phép người dùng chọn\"*",
     FEATURE, MOD_TRA,
     "User không thuộc KHCN/KHDN/KHDNL — Combobox Khối cho phép chọn tự do, để trống = tất cả",
     "Happy Path", "High",
     "1. User đã đăng nhập, không thuộc Khối KHCN/KHDN/KHDNL (ví dụ: user Hội sở).\n2. User có quyền truy cập Báo cáo tổng doanh thu phí.",
     "1. Truy cập menu Báo cáo >> Báo cáo tổng doanh thu phí.\n"
     "2. Quan sát trường Khối — dropdown cho phép chọn.\n"
     "3. Để trống Khối.\n"
     "4. Chọn Từ ngày = '01/01/2025', Đến ngày = '31/03/2025'.\n"
     "5. Nhấn nút 'Tra cứu'.",
     "(i) Nghiệp vụ/Logic: Hệ thống trả về dữ liệu của tất cả các Khối (KHCN, KHDN, KHDNL).\n"
     "(ii) UI: Combobox Khối ở trạng thái enabled, cho phép chọn hoặc để trống. "
     "Lưới hiển thị kết quả từ mọi Khối.",
     ""),

    # TC-005 (SC-05)
    ("US31-TC-005", "SC-05",
     "Bảng mô tả nút STT 2 (Xoá tra cứu) – *\"Người dùng nhấn vào button Xoá tra cứu, hệ thống xoá các điều kiện tìm kiếm đang có\"* + QTC-04",
     FEATURE, MOD_TRA,
     "Xóa tra cứu — Hệ thống xóa toàn bộ điều kiện, lưới về mặc định",
     "Happy Path", "Medium",
     PRECON_BASE + "\n4. Đã thực hiện tra cứu thành công, lưới đang hiển thị kết quả.",
     "1. Nhấn nút 'Xóa tra cứu'.",
     "(i) Nghiệp vụ/Logic: Hệ thống xóa toàn bộ điều kiện tìm kiếm đang có. "
     "Lưới danh sách trở về trạng thái mặc định (không hiển thị kết quả).\n"
     "(ii) UI: Tất cả các trường Khối, Mã Chi nhánh, Biểu phí, Code phí, Loại tính phí, "
     "Từ ngày, Đến ngày đều được reset về trống/giá trị mặc định. Lưới rỗng.",
     "[Theo QTC-04]"),

    # TC-006 (SC-06)
    ("US31-TC-006", "SC-06",
     "Bảng mô tả trường STT 20 – *\"trường hợp lưới danh sách tổng doanh thu phí không có dữ liệu thì Tổng cộng hiển thị = 0\"*",
     FEATURE, MOD_TRA,
     "Tra cứu không có kết quả — Lưới rỗng, Tổng cộng = 0",
     "Happy Path", "Medium",
     PRECON_BASE,
     "1. Truy cập menu Báo cáo >> Báo cáo tổng doanh thu phí.\n"
     "2. Chọn Biểu phí = 'BP999 - Biểu phí không tồn tại dữ liệu'.\n"
     "3. Chọn Từ ngày = '01/01/2020', Đến ngày = '01/01/2020'.\n"
     "4. Nhấn nút 'Tra cứu'.",
     "(i) Nghiệp vụ/Logic: Hệ thống không tìm thấy bản ghi nào thỏa mãn điều kiện.\n"
     "(ii) UI: Lưới danh sách hiển thị trống (blank). "
     "Dòng Tổng cộng hiển thị = 0.",
     ""),

    # TC-007 (SC-07)
    ("US31-TC-007", "SC-07",
     "Bảng mô tả nút STT 3 (Tải xuống) – *\"hệ thống hiện dropdownlist Excel/PDF [...] tải xuống báo cáo\"* + QTC-05",
     FEATURE, MOD_TAI,
     "Tải xuống định dạng Excel — File .xlsx tải thành công",
     "Happy Path", "High",
     PRECON_BASE + "\n4. Đã tra cứu thành công, lưới đang hiển thị kết quả.",
     "1. Nhấn nút 'Tải xuống'.\n"
     "2. Dropdown hiển thị 2 lựa chọn: Excel / PDF.\n"
     "3. Chọn 'Excel'.",
     "(i) Nghiệp vụ/Logic: Hệ thống tải file .xlsx chứa đúng dữ liệu đang hiển thị trên lưới.\n"
     "(ii) UI: File tải xuống thành công. Tên file = 'Báo cáo tổng doanh thu phí - yyyymmddhhmmss.xlsx'.",
     "[Theo QTC-05]"),

    # TC-008 (SC-08)
    ("US31-TC-008", "SC-08",
     "Bảng mô tả nút STT 3 (Tải xuống) – *\"hệ thống hiện dropdownlist Excel/PDF [...] tải xuống báo cáo\"* + QTC-05",
     FEATURE, MOD_TAI,
     "Tải xuống định dạng PDF — File .pdf tải thành công",
     "Happy Path", "High",
     PRECON_BASE + "\n4. Đã tra cứu thành công, lưới đang hiển thị kết quả.",
     "1. Nhấn nút 'Tải xuống'.\n"
     "2. Dropdown hiển thị 2 lựa chọn: Excel / PDF.\n"
     "3. Chọn 'PDF'.",
     "(i) Nghiệp vụ/Logic: Hệ thống tải file .pdf chứa đúng dữ liệu đang hiển thị trên lưới.\n"
     "(ii) UI: File tải xuống thành công. Tên file = 'Báo cáo tổng doanh thu phí - yyyymmddhhmmss.pdf'.",
     "[Theo QTC-05]"),

    # TC-009 (SC-09)
    ("US31-TC-009", "SC-09",
     "QTC-05 – *\"Luôn cho phép tải xuống dù có dữ liệu hay không\"*",
     FEATURE, MOD_TAI,
     "Tải xuống khi lưới rỗng — Hệ thống vẫn cho phép tải file",
     "Happy Path", "Medium",
     PRECON_BASE + "\n4. Đã tra cứu nhưng không có kết quả (lưới rỗng).",
     "1. Nhấn nút 'Tải xuống'.\n"
     "2. Chọn 'Excel'.",
     "(i) Nghiệp vụ/Logic: Hệ thống vẫn cho phép tải file. File chứa header (tên các cột) nhưng không có dữ liệu.\n"
     "(ii) UI: File .xlsx tải xuống thành công, không hiển thị lỗi.",
     "[Theo QTC-05]"),
]

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Cases"

# Header style
header_font = Font(bold=True, size=10, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))

# Write headers
for ci, col_name in enumerate(COLS, 1):
    cell = ws.cell(row=1, column=ci, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Write data
data_align = Alignment(vertical="top", wrap_text=True)
for ri, tc in enumerate(TCS, 2):
    for ci, val in enumerate(tc, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.alignment = data_align
        cell.border = thin_border

# Column widths
COL_WIDTHS = [12,8,40,20,18,45,15,10,40,50,55,25,10,12,10,10,10,12,15]
for ci, w in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

ws.auto_filter.ref = f"A1:S{len(TCS)+1}"
ws.freeze_panes = "A2"

out = os.path.join(os.path.dirname(os.path.dirname(__file__)), "US31_TestCases.xlsx")
wb.save(out)
print(f"✅ Batch 1 done: {len(TCS)} TC (SC-01→SC-09 Happy Path)")
print(f"📄 File: {out}")

#!/usr/bin/env python3
"""US06 TC Batch 4+5: Field Validation + Business Logic"""
import openpyxl
from openpyxl.styles import Alignment, Border, Side
import os

FPATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "US06_TestCases.xlsx")
wb = openpyxl.load_workbook(FPATH)
ws = wb["Test Cases"]
wrap = Alignment(wrap_text=True, vertical="top")
thin = Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
FEAT = "Khai báo Biểu phí"
PRE = "1. User đăng nhập quyền Maker tại Danh mục Biểu phí"

data = [
    # ===== FIELD VALIDATION =====
    ("US06-TC-049","SC-36",'Mockup Thêm mới vs Chỉnh sửa',FEAT,"Thêm mới – UI",
     "Thêm mới KHÔNG có trường Trạng thái; Chỉnh sửa CÓ Trạng thái readonly","Field Validation","Medium",PRE,
     "1. Mở màn hình Thêm mới → kiểm tra có trường Trạng thái không\n2. Mở màn hình Chỉnh sửa BP-001 → kiểm tra trường Trạng thái",
     "(i) Nghiệp vụ/Logic: Thêm mới không hiển thị Trạng thái (chưa tồn tại). Chỉnh sửa hiển thị readonly.\n(ii) UI: Thêm mới không có field. Chỉnh sửa có field Trạng thái dạng readonly.",""),

    ("US06-TC-050","SC-37",'Bảng mô tả trường R13 – "Search bar tìm kiếm nhanh và gần đúng"',FEAT,"Thêm mới – UI",
     "Cây SPDV: Search bar tìm kiếm gần đúng theo Mã/Tên SPDV/Code phí","Field Validation","Medium",
     PRE+"\n2. Đang ở Thêm mới, click trường SPDV/Code phí",
     "1. Nhập 'chuyen tien' vào Search bar (không dấu)\n2. Kiểm tra kết quả lọc trên cây",
     "(i) Nghiệp vụ/Logic: Hệ thống tìm kiếm gần đúng, bỏ dấu, case-insensitive.\n(ii) UI: Cây SPDV lọc hiển thị các node chứa 'chuyển tiền'.","[Theo QTC-02]"),

    ("US06-TC-051","SC-40",'Bảng mô tả trường R22-R30 – "dấu ◎ readonly"',FEAT,"Thêm mới – UI",
     "Lưới Thông tin chi tiết: tất cả trường readonly, chỉ có nút hành động","Field Validation","Medium",
     PRE+"\n2. Đã gán Code phí CP001 vào lưới",
     "1. Kiểm tra các cột: Mã phí, Tên phí, SPDV, Loại tiền tệ, Đối tượng thu phí, Loại KH, Khai báo theo nhóm KH, Quy tắc tính phí\n2. Thử click vào ô dữ liệu",
     "(i) Nghiệp vụ/Logic: Dữ liệu Code phí là readonly, không cho sửa trực tiếp trên lưới.\n(ii) UI: Tất cả ô dữ liệu không editable. Chỉ có nút Chỉnh sửa/Xem/Xóa ở cột Hành động.",""),

    ("US06-TC-052","SC-41",'QTC-11 – "FE-First Error Handling"',FEAT,"Thêm mới – UI",
     "Nhập ký tự đặc biệt XSS/SQL Injection → FE sanitize hoặc BE chặn","Field Validation","Medium",PRE,
     "1. Nhập Mã = '<script>alert(1)</script>'\n2. Nhập Tên = 'DROP TABLE; --'\n3. Click 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Hệ thống sanitize hoặc từ chối ký tự nguy hiểm.\n(ii) UI: FE chặn hoặc BE trả lỗi. Không thực thi mã độc.","[Theo QTC-11]"),

    ("US06-TC-053","SC-42",'Mục "Chỉnh sửa" P62 + BA QA-01.6',FEAT,"Chỉnh sửa – UI",
     "Đang hiệu lực: chỉ Ngày HHL enabled, các trường khác disabled","Field Validation","High",
     "1. User đăng nhập quyền Maker\n2. BP-015 Đang hiệu lực",
     "1. Click 'Chỉnh sửa' tại BP-015\n2. Kiểm tra trạng thái tất cả trường: Mã, Tên, Ngày BH, Ngày HL, Số VB, Tên VB, Link iDoc, SPDV/Code phí\n3. Kiểm tra trường Ngày HHL",
     "(i) Nghiệp vụ/Logic: Chỉ Ngày HHL cho phép sửa. Tất cả trường khác và Code phí không tác động được.\n(ii) UI: Tất cả trường disabled trừ Ngày hết hiệu lực. Nút Chỉnh sửa/Xóa Code phí ẩn hoặc disabled.","[BA QA-01.6]"),

    ("US06-TC-054","SC-43",'QTC-14.1 – "No-Change Guard"',FEAT,"Chỉnh sửa – UI",
     "Chỉnh sửa không thay đổi gì → FE disable Xác nhận hoặc cảnh báo","Field Validation","Medium",
     "1. User đăng nhập quyền Maker\n2. BP-016 Chưa hiệu lực",
     "1. Click 'Chỉnh sửa' tại BP-016\n2. KHÔNG thay đổi bất kỳ thông tin nào\n3. Kiểm tra nút 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Hệ thống không cho lưu khi không có thay đổi.\n(ii) UI: Nút 'Xác nhận' disabled hoặc hiển thị cảnh báo 'Không có thay đổi để lưu'.","[Theo QTC-14.1]"),

    ("US06-TC-055","SC-68",'Mockup Thêm mới, Bảng mô tả trường R5-R12',FEAT,"Thêm mới – UI",
     "Trường bắt buộc (★) hiển thị dấu (*) đúng; Link iDoc không có dấu","Field Validation","Low",PRE,
     "1. Mở màn hình Thêm mới\n2. Kiểm tra dấu (*) tại: Mã, Tên, Ngày BH, Ngày HL, Ngày HHL, Số VB, Tên VB\n3. Kiểm tra trường Link iDoc",
     "(i) Nghiệp vụ/Logic: 7 trường bắt buộc có dấu (*). Link iDoc không bắt buộc.\n(ii) UI: Dấu (*) hiển thị đúng tại 7 trường. Link iDoc không có dấu (*).",""),

    ("US06-TC-056","SC-69",'Mockup Thêm mới vs Chỉnh sửa – "nút Tải xuống"',FEAT,"Thêm mới – UI",
     "Thêm mới: không hiển thị Tải xuống. Chỉnh sửa: hiển thị cả Tải lên + Tải xuống","Field Validation","Low",PRE,
     "1. Mở Thêm mới → kiểm tra có nút 'Tải xuống' không\n2. Mở Chỉnh sửa BP-001 → kiểm tra nút 'Chọn tệp' và 'Tải xuống'",
     "(i) Nghiệp vụ/Logic: Thêm mới chưa có dữ liệu nên không cần Tải xuống.\n(ii) UI: Thêm mới chỉ có 'Chọn tệp'. Chỉnh sửa có cả 'Chọn tệp' và 'Tải xuống'.",""),

    ("US06-TC-057","SC-70",'QTC-01.5 – "Date format dd/mm/yyyy"',FEAT,"Thêm mới – UI",
     "Các trường ngày hiển thị đúng format dd/mm/yyyy","Field Validation","Low",PRE,
     "1. Chọn Ngày ban hành, Ngày hiệu lực, Ngày hết hiệu lực\n2. Kiểm tra format hiển thị",
     "(i) Nghiệp vụ/Logic: Đúng format chuẩn hệ thống.\n(ii) UI: Tất cả trường ngày hiển thị dd/mm/yyyy (VD: 06/05/2026).","[Theo QTC-01.5]"),

    # ===== BUSINESS LOGIC =====
    ("US06-TC-058","SC-38",'Bảng mô tả trường R13 – "cascade tích/bỏ tích"',FEAT,"Thêm mới – Cascade SPDV",
     "Tích chọn SPDV cha → tự động tích con + Code phí. Bỏ tích cha → bỏ tích con","Business Logic","High",
     PRE+"\n2. Cây SPDV có: SPDV-A (cha) → SPDV-A1 (con) → CP-A1-01 (Code phí)",
     "1. Tích chọn SPDV-A → kiểm tra SPDV-A1 và CP-A1-01\n2. Bỏ tích SPDV-A → kiểm tra SPDV-A1 và CP-A1-01",
     "(i) Nghiệp vụ/Logic: Tích cha → tự động tích tất cả con + Code phí. Bỏ tích cha → bỏ tất cả.\n(ii) UI: Checkbox cascade đúng. Lưới cập nhật tương ứng.",""),

    ("US06-TC-059","SC-39",'Bảng mô tả trường R13 – "bỏ tích SPDV cấp cha liền trước"',FEAT,"Thêm mới – Cascade SPDV",
     "Bỏ tích tất cả con đồng cấp → tự động bỏ tích cha liền trước","Business Logic","High",
     PRE+"\n2. SPDV-A có 2 con: SPDV-A1, SPDV-A2. Cả 2 đang được tích chọn",
     "1. Bỏ tích SPDV-A1\n2. Kiểm tra SPDV-A (cha vẫn tích vì còn A2)\n3. Bỏ tích SPDV-A2\n4. Kiểm tra SPDV-A",
     "(i) Nghiệp vụ/Logic: Khi không còn con đồng cấp nào được tích → cha tự động bỏ tích.\n(ii) UI: Sau bước 3, SPDV-A tự động unchecked.",""),

    ("US06-TC-060","SC-44",'Mục "Thêm mới" P47-P49 + BA QA-03.3 – "realtime"',FEAT,"Trạng thái Biểu phí",
     "Trạng thái tự động chuyển Chưa hiệu lực → Đang hiệu lực (realtime)","Business Logic","High",
     "1. Biểu phí BP-020 đã duyệt, Ngày HL = ngày hệ thống hôm nay",
     "1. Mở lưới Danh mục Biểu phí\n2. Tìm BP-020\n3. Kiểm tra cột Trạng thái",
     "(i) Nghiệp vụ/Logic: Ngày HT >= Ngày HL → trạng thái = Đang hiệu lực (tính realtime).\n(ii) UI: Cột Trạng thái hiển thị 'Đang hiệu lực'.","[BA QA-03.3]"),

    ("US06-TC-061","SC-45",'Mục "Thêm mới" P50 + BA QA-03.3',FEAT,"Trạng thái Biểu phí",
     "Trạng thái tự động chuyển Đang hiệu lực → Hết hiệu lực","Business Logic","High",
     "1. BP-021 có Ngày HHL = hôm qua (đã qua)",
     "1. Mở lưới Danh mục\n2. Tìm BP-021\n3. Kiểm tra Trạng thái",
     "(i) Nghiệp vụ/Logic: Ngày HT > Ngày HHL → Hết hiệu lực.\n(ii) UI: Cột Trạng thái = 'Hết hiệu lực'.","[BA QA-03.3]"),

    ("US06-TC-062","SC-46",'Mục "Thêm mới" P46 + QTC-12',FEAT,"Maker-Checker",
     "Checker phê duyệt Thêm mới → Biểu phí hiển thị trên lưới chính thức","Business Logic","High",
     "1. User đăng nhập quyền Checker\n2. Tác vụ Thêm mới BP-022 đang Chờ duyệt",
     "1. Mở Tác vụ chờ duyệt\n2. Xem chi tiết BP-022\n3. Click 'Phê duyệt'",
     "(i) Nghiệp vụ/Logic: Biểu phí được lưu chính thức. Trạng thái = Đã duyệt.\n(ii) UI: Toast 'Phê duyệt thành công'. Bản ghi cập nhật trạng thái Đã duyệt. Biểu phí hiển thị trên lưới Danh mục.","[Theo QTC-12]"),

    ("US06-TC-063","SC-47",'QTC-12 mục 2 – "Tác vụ pending của tôi"',FEAT,"Maker-Checker",
     "Checker từ chối Thêm mới → Maker nhận bản ghi tại Tác vụ Pending","Business Logic","High",
     "1. User đăng nhập quyền Checker\n2. Tác vụ Thêm mới BP-023 đang Chờ duyệt",
     "1. Mở tác vụ BP-023\n2. Click 'Từ chối'\n3. Nhập lý do = 'Sai thông tin ngày'\n4. Đăng nhập Maker, kiểm tra Tác vụ Pending",
     "(i) Nghiệp vụ/Logic: Bản ghi ở trạng thái Từ chối duyệt. Maker có thể sửa lại tại Tác vụ Pending.\n(ii) UI: Bản ghi tại Tác vụ chờ duyệt cập nhật trạng thái Từ chối. Maker thấy bản ghi tại Tác vụ Pending.","[Theo QTC-12]"),

    ("US06-TC-064","SC-49",'BA xác nhận QA-01.1',FEAT,"Thêm mới – Code phí Ngừng HĐ",
     "Code phí Ngừng hoạt động dưới SPDV Hoạt động VẪN được chọn","Business Logic","Medium",
     PRE+"\n2. SPDV-B Hoạt động có Code phí CP-B1 (Ngừng hoạt động)",
     "1. Mở cây SPDV\n2. Tìm SPDV-B → kiểm tra CP-B1\n3. Tích chọn CP-B1",
     "(i) Nghiệp vụ/Logic: Code phí Ngừng HĐ vẫn được phép chọn vào Biểu phí.\n(ii) UI: CP-B1 hiển thị trên cây và cho phép tích chọn.","[BA QA-01.1]"),

    ("US06-TC-065","SC-71",'QTC-12 mục 2 – "Xóa: Chờ duyệt hoặc Từ chối duyệt"',FEAT,"Maker-Checker",
     "Maker xóa tác vụ Chờ duyệt/Từ chối tại Tác vụ Pending","Business Logic","Medium",
     "1. User đăng nhập quyền Maker\n2. Tồn tại tác vụ BP-024 ở trạng thái Từ chối duyệt tại Tác vụ Pending",
     "1. Tại Tác vụ Pending, tìm BP-024\n2. Click 'Xóa'",
     "(i) Nghiệp vụ/Logic: Tác vụ bị xóa. Biểu phí (nếu Thêm mới) không tồn tại trên hệ thống.\n(ii) UI: Bản ghi biến mất khỏi Tác vụ Pending.","[Theo QTC-12]"),
]

start = ws.max_row + 1
for i, row in enumerate(data):
    for j, val in enumerate(row):
        c = ws.cell(row=start+i, column=j+1, value=val)
        c.alignment = wrap; c.border = thin
wb.save(FPATH)
print(f"✅ Batch 4+5 (Field Validation + Business Logic): {len(data)} TC → {FPATH}")

# -*- coding: utf-8 -*-
"""US35 – Sinh file Excel Test Case Suite từ 4 batch data"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from us35_tc_batch1 import TC_BATCH1
from us35_tc_batch2 import TC_BATCH2
from us35_tc_batch3 import TC_BATCH3
from us35_tc_batch4 import TC_BATCH4
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

ALL_TC = TC_BATCH1 + TC_BATCH2 + TC_BATCH3 + TC_BATCH4
OUT = os.path.join(os.path.dirname(os.path.dirname(__file__)), "US35_TestCase_Suite.xlsx")

wb = Workbook()
ws = wb.active
ws.title = "Test Cases"

HEADERS = ["TC_ID","SC_Ref","Reference","Feature","Module","Title","Type","Priority",
           "Precondition","Steps","Expected","Note",
           "Tester","Ngày Test","Kết quả","Bug ID","Retest","Ngày Retest","Ghi chú QA"]
COL_WIDTHS = [14,8,45,16,20,55,16,10,40,50,55,30,12,12,10,10,10,12,15]

# Header style
hdr_font = Font(name="Times New Roman", bold=True, color="FFFFFF", size=10)
hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"))

for col_idx, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = hdr_align
    cell.border = thin_border

# Type colors
TYPE_FILLS = {
    "Happy Path": PatternFill(start_color="E2EFDA", fill_type="solid"),
    "Negative Path": PatternFill(start_color="FCE4EC", fill_type="solid"),
    "Boundary Value": PatternFill(start_color="FFF3E0", fill_type="solid"),
    "Field Validation": PatternFill(start_color="E3F2FD", fill_type="solid"),
    "Business Logic": PatternFill(start_color="F3E5F5", fill_type="solid"),
    "Data Integrity": PatternFill(start_color="FFF9C4", fill_type="solid"),
    "NFR": PatternFill(start_color="E0F7FA", fill_type="solid"),
}
data_font = Font(name="Times New Roman", size=10)
data_align = Alignment(vertical="top", wrap_text=True)

for row_idx, tc in enumerate(ALL_TC, 2):
    tc_id, sc_ref, ref, feature, module, title, tc_type, priority, precon, steps, expected, note = tc
    values = [tc_id, sc_ref, ref, feature, module, title, tc_type, priority,
              precon, steps, expected, note, "", "", "", "", "", "", ""]
    fill = TYPE_FILLS.get(tc_type, None)
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = data_font
        cell.alignment = data_align
        cell.border = thin_border
        if fill and col_idx <= 8:
            cell.fill = fill

# Column widths
for i, w in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[chr(64+i) if i <= 26 else None].width = w
# Fix columns > Z
import openpyxl.utils
for i, w in enumerate(COL_WIDTHS):
    col_letter = openpyxl.utils.get_column_letter(i+1)
    ws.column_dimensions[col_letter].width = w

# Freeze pane
ws.freeze_panes = "A2"
# Auto filter
ws.auto_filter.ref = f"A1:S{len(ALL_TC)+1}"

ws.sheet_properties.pageSetUpPr = None
wb.save(OUT)

print(f"✅ Đã tạo: {OUT}")
print(f"📊 Tổng TC: {len(ALL_TC)}")

# Cross-check SC coverage
from collections import Counter
sc_refs = [tc[1] for tc in ALL_TC]
sc_unique = sorted(set(sc_refs), key=lambda x: int(x.split("-")[1]))
print(f"🔍 SC Coverage: {len(sc_unique)} SC unique")
expected_scs = [f"SC-{str(i).zfill(2)}" for i in range(1, 68)]
missing = [sc for sc in expected_scs if sc not in sc_unique]
if missing:
    print(f"⚠️ THIẾU SC: {missing}")
else:
    print("✅ 100% SC coverage (SC-01 → SC-67)")

type_counts = Counter(tc[6] for tc in ALL_TC)
for t, c in sorted(type_counts.items()):
    print(f"   {t}: {c}")

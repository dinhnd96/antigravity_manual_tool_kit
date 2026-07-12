"""US31 - Batch 2: Negative Path + BVA (SC-10 → SC-18) → Append Excel"""
import openpyxl
from openpyxl.styles import Alignment, Border, Side
import os

FEATURE = "US31 - Báo cáo tổng doanh thu phí"
MOD_TRA = "Tra cứu báo cáo"
MOD_LUOI = "Lưới kết quả"

PRECON = (
    "1. User đã đăng nhập hệ thống ProfiX.\n"
    "2. User có quyền truy cập menu Báo cáo >> Báo cáo tổng doanh thu phí.\n"
    "3. Đã truy cập màn hình Báo cáo tổng doanh thu phí."
)

TCS = [
    # TC-010 (SC-10)
    ("US31-TC-010", "SC-10",
     "Bảng mô tả trường STT 6 (Từ ngày ★) – *\"là field bắt buộc\"* + QTC-14.5",
     FEATURE, MOD_TRA,
     "Bỏ trống Từ ngày (bắt buộc) — FE chặn, hiển thị lỗi validation",
     "Negative Path", "High",
     PRECON,
     "1. Để trống Từ ngày.\n"
     "2. Chọn Đến ngày = '31/03/2025'.\n"
     "3. Nhấn nút 'Tra cứu'.",
     "(i) Nghiệp vụ/Logic: Hệ thống KHÔNG gọi BE. FE chặn lại do trường bắt buộc bị bỏ trống.\n"
     "(ii) UI: Hiển thị thông báo lỗi 'Trường này bắt buộc' dưới field Từ ngày. Field Từ ngày có viền đỏ/highlight.",
     "[Theo QTC-14.5]"),

    # TC-011 (SC-11)
    ("US31-TC-011", "SC-11",
     "Bảng mô tả trường STT 7 (Đến ngày ★) – *\"là field bắt buộc\"* + QTC-14.5",
     FEATURE, MOD_TRA,
     "Bỏ trống Đến ngày (bắt buộc) — FE chặn, hiển thị lỗi validation",
     "Negative Path", "High",
     PRECON,
     "1. Chọn Từ ngày = '01/01/2025'.\n"
     "2. Để trống Đến ngày.\n"
     "3. Nhấn nút 'Tra cứu'.",
     "(i) Nghiệp vụ/Logic: Hệ thống KHÔNG gọi BE. FE chặn lại do trường bắt buộc bị bỏ trống.\n"
     "(ii) UI: Hiển thị thông báo lỗi 'Trường này bắt buộc' dưới field Đến ngày. Field Đến ngày có viền đỏ/highlight.",
     "[Theo QTC-14.5]"),

    # TC-012 (SC-12)
    ("US31-TC-012", "SC-12",
     "Bảng mô tả trường STT 6 – *\"Không chọn ngày lớn hơn field Đến ngày\"*",
     FEATURE, MOD_TRA,
     "Từ ngày > Đến ngày — FE không cho phép chọn",
     "Negative Path", "High",
     PRECON,
     "1. Chọn Từ ngày = '15/04/2025'.\n"
     "2. Chọn Đến ngày = '01/03/2025' (trước Từ ngày).\n"
     "3. Nhấn nút 'Tra cứu'.",
     "(i) Nghiệp vụ/Logic: Hệ thống KHÔNG cho phép chọn Từ ngày > Đến ngày hoặc chặn khi nhấn Tra cứu.\n"
     "(ii) UI: FE hiển thị lỗi validation hoặc không cho phép chọn ngày không hợp lệ tại datepicker.",
     ""),

    # TC-013 (SC-13)
    ("US31-TC-013", "SC-13",
     "Bảng mô tả trường STT 6-7 (Date, dd/mm/yyyy) + QTC-01.5",
     FEATURE, MOD_TRA,
     "Nhập ngày không hợp lệ (sai format) — FE chặn",
     "Negative Path", "Medium",
     PRECON,
     "1. Tại field Từ ngày, thử nhập giá trị '32/13/2025' (ngày/tháng không tồn tại).\n"
     "2. Quan sát phản hồi FE.",
     "(i) Nghiệp vụ/Logic: Hệ thống không chấp nhận giá trị ngày không hợp lệ.\n"
     "(ii) UI: FE không cho nhập (datepicker chặn) hoặc hiển thị lỗi format ngày.",
     "[Theo QTC-01.5]"),

    # TC-014 (SC-14)
    ("US31-TC-014", "SC-14",
     "Bảng mô tả trường STT 6-7 + QTC-01.5 – *\"Từ ngày: 00:00:00.000 [...] Đến ngày: 23:59:59.999\"*",
     FEATURE, MOD_TRA,
     "BVA: Từ ngày = Đến ngày (cùng 1 ngày) — Hệ thống trả kết quả đúng 1 ngày",
     "Boundary Value", "High",
     PRECON + "\n4. Hệ thống có giao dịch thu phí vào ngày 15/03/2025.",
     "1. Chọn Từ ngày = '15/03/2025'.\n"
     "2. Chọn Đến ngày = '15/03/2025'.\n"
     "3. Nhấn nút 'Tra cứu'.",
     "(i) Nghiệp vụ/Logic: Hệ thống trả kết quả giao dịch trong đúng 1 ngày (00:00:00.000 – 23:59:59.999 của 15/03/2025).\n"
     "(ii) UI: Lưới hiển thị dữ liệu. Cột Từ ngày = '15/03/2025', Đến ngày = '15/03/2025'.",
     "[Theo QTC-01.5]"),

    # TC-015 (SC-15)
    ("US31-TC-015", "SC-15",
     "BA trả lời QA-02.1 – *\"Đã có quy tắc chung về việc không giới hạn dữ liệu\"*",
     FEATURE, MOD_TRA,
     "BVA: Khoảng thời gian rất lớn (>5 năm) — Hệ thống vẫn xử lý",
     "Boundary Value", "Medium",
     PRECON + "\n4. Hệ thống có dữ liệu giao dịch từ 2020 đến 2025.",
     "1. Chọn Từ ngày = '01/01/2020'.\n"
     "2. Chọn Đến ngày = '31/12/2025' (khoảng 6 năm).\n"
     "3. Nhấn nút 'Tra cứu'.",
     "(i) Nghiệp vụ/Logic: Hệ thống vẫn xử lý và trả kết quả bình thường, không timeout, không chặn.\n"
     "(ii) UI: Lưới hiển thị dữ liệu. Không có cảnh báo giới hạn thời gian.",
     "BA xác nhận không giới hạn khoảng thời gian tra cứu"),

    # TC-016 (SC-16)
    ("US31-TC-016", "SC-16",
     "BA trả lời QA-02.2 – *\"không giới hạn dữ liệu\"* + QTC-01.4",
     FEATURE, MOD_LUOI,
     "BVA: Doanh thu phí VND rất lớn (hàng nghìn tỷ) — Hiển thị đúng format, không overflow",
     "Boundary Value", "Medium",
     PRECON + "\n4. Hệ thống có tổ hợp có doanh thu VND = 1,234,567,890,123.45 (hàng nghìn tỷ).",
     "1. Tra cứu điều kiện có dữ liệu doanh thu lớn.\n"
     "2. Quan sát cột Doanh thu phí (VND) và dòng Tổng cộng.",
     "(i) Nghiệp vụ/Logic: Hệ thống hiển thị chính xác giá trị, không bị tràn hay cắt số.\n"
     "(ii) UI: Cột Doanh thu phí (VND) hiển thị đúng format: '1,234,567,890,123' (phân cách hàng nghìn, VND không có thập phân). "
     "Tổng cộng cũng hiển thị đúng format.",
     "[Theo QTC-01.4]"),

    # TC-017 (SC-17)
    ("US31-TC-017", "SC-17",
     "QTC-06 – *\"Mặc định hiển thị 50 bản ghi/trang\"*",
     FEATURE, MOD_LUOI,
     "BVA: Kết quả đúng 50 bản ghi — Hiển thị 1 trang, phân trang disabled",
     "Boundary Value", "Medium",
     PRECON + "\n4. Điều kiện tra cứu trả về đúng 50 bản ghi.",
     "1. Tra cứu với điều kiện trả về đúng 50 bản ghi.\n"
     "2. Quan sát phân trang.",
     "(i) Nghiệp vụ/Logic: Hệ thống hiển thị toàn bộ 50 bản ghi trên 1 trang.\n"
     "(ii) UI: Phân trang hiển thị '1 - 50 của 50 bản ghi'. Nút Next/Prev disabled hoặc không hiển thị.",
     "[Theo QTC-06]"),

    # TC-018 (SC-18)
    ("US31-TC-018", "SC-18",
     "QTC-06 – *\"Mặc định hiển thị 50 bản ghi/trang\"*",
     FEATURE, MOD_LUOI,
     "BVA: Kết quả 51 bản ghi — Hiển thị 2 trang",
     "Boundary Value", "Medium",
     PRECON + "\n4. Điều kiện tra cứu trả về 51 bản ghi.",
     "1. Tra cứu với điều kiện trả về 51 bản ghi.\n"
     "2. Quan sát phân trang.\n"
     "3. Nhấn nút Next (trang 2).",
     "(i) Nghiệp vụ/Logic: Trang 1 hiển thị 50 bản ghi, trang 2 hiển thị 1 bản ghi.\n"
     "(ii) UI: Phân trang hiển thị '1 - 50 của 51 bản ghi'. Nút Next enabled. Trang 2 hiển thị bản ghi thứ 51.",
     "[Theo QTC-06]"),
]

# Append to existing file
fpath = os.path.join(os.path.dirname(os.path.dirname(__file__)), "US31_TestCases.xlsx")
wb = openpyxl.load_workbook(fpath)
ws = wb["Test Cases"]
start_row = ws.max_row + 1
data_align = Alignment(vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))

for ri, tc in enumerate(TCS, start_row):
    for ci, val in enumerate(tc, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.alignment = data_align
        cell.border = thin_border

ws.auto_filter.ref = f"A1:S{ws.max_row}"
wb.save(fpath)
print(f"✅ Batch 2 done: {len(TCS)} TC (SC-10→SC-18 Negative+BVA)")
print(f"📄 Total: {ws.max_row - 1} TC")

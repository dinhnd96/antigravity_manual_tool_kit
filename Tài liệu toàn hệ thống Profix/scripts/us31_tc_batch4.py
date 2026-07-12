"""US31 - Batch 4: Business Logic + Data Integrity (SC-28→SC-35, SC-40, SC-42, SC-43)"""
import openpyxl
from openpyxl.styles import Alignment, Border, Side
import os

FEATURE = "US31 - Báo cáo tổng doanh thu phí"
MOD_TRA = "Tra cứu báo cáo"
MOD_LUOI = "Lưới kết quả"
MOD_TONG = "Tổng cộng"
MOD_TAI = "Tải xuống báo cáo"

PRECON = (
    "1. User đã đăng nhập hệ thống ProfiX.\n"
    "2. User có quyền truy cập menu Báo cáo >> Báo cáo tổng doanh thu phí.\n"
    "3. Đã truy cập màn hình Báo cáo tổng doanh thu phí."
)
PRECON_DATA = PRECON + "\n4. Đã tra cứu thành công, lưới đang hiển thị kết quả."

TCS = [
    # TC-030 (SC-28) KHDNL
    ("US31-TC-030", "SC-28",
     "Bảng mô tả trường STT 1 (Khối) + BA trả lời QA-04.5 – *\"KHDNL là khối Khách hàng doanh nghiệp lớn [...] 2 Khối trên là khác nhau\"*",
     FEATURE, MOD_TRA,
     "User thuộc Khối KHDNL — Combobox disabled, lưới chỉ trả dữ liệu Khối KHDNL",
     "Business Logic", "High",
     "1. User đã đăng nhập, thuộc Khối KHDNL.\n2. User có quyền truy cập Báo cáo tổng doanh thu phí.\n3. Hệ thống có dữ liệu thuộc Khối KHDNL.",
     "1. Truy cập menu Báo cáo >> Báo cáo tổng doanh thu phí.\n"
     "2. Quan sát trường Khối.\n"
     "3. Chọn Từ ngày = '01/01/2025', Đến ngày = '31/03/2025'.\n"
     "4. Nhấn nút 'Tra cứu'.",
     "(i) Nghiệp vụ/Logic: Hệ thống chỉ trả về dữ liệu thuộc Khối KHDNL. Không thể tra cứu dữ liệu Khối khác.\n"
     "(ii) UI: Combobox Khối hiển thị 'KHDNL', trạng thái disabled. Lưới chỉ chứa dữ liệu KHDNL.",
     "BA xác nhận KHDNL là Khối riêng biệt, khác KHDN"),

    # TC-031 (SC-29) KHDN
    ("US31-TC-031", "SC-29",
     "Bảng mô tả trường STT 1 (Khối) + QTC-10",
     FEATURE, MOD_TRA,
     "User thuộc Khối KHDN — Combobox disabled, lưới chỉ trả dữ liệu Khối KHDN",
     "Business Logic", "High",
     "1. User đã đăng nhập, thuộc Khối KHDN.\n2. User có quyền truy cập Báo cáo tổng doanh thu phí.\n3. Hệ thống có dữ liệu thuộc Khối KHDN.",
     "1. Truy cập menu Báo cáo >> Báo cáo tổng doanh thu phí.\n"
     "2. Quan sát trường Khối.\n"
     "3. Chọn Từ ngày = '01/01/2025', Đến ngày = '31/03/2025'.\n"
     "4. Nhấn nút 'Tra cứu'.",
     "(i) Nghiệp vụ/Logic: Hệ thống chỉ trả về dữ liệu thuộc Khối KHDN (vừa và nhỏ). "
     "Không lẫn dữ liệu KHDNL hoặc KHCN.\n"
     "(ii) UI: Combobox Khối hiển thị 'KHDN', trạng thái disabled.",
     "[Theo QTC-10]"),

    # TC-032 (SC-30)
    ("US31-TC-032", "SC-30",
     "BA trả lời QA-01.4 – *\"Khi Loại tiền = VND, 2 cột cùng giá trị\"*",
     FEATURE, MOD_LUOI,
     "Loại tiền = VND — Cột Nguyên tệ và VND hiển thị cùng giá trị",
     "Business Logic", "High",
     PRECON_DATA + "\n5. Lưới có bản ghi với Loại tiền = 'VND', Doanh thu = 5,427,167.",
     "1. Quan sát dòng có Loại tiền = 'VND'.\n"
     "2. So sánh giá trị cột Doanh thu phí (Nguyên tệ) và Doanh thu phí (VND).",
     "(i) Nghiệp vụ/Logic: Khi Loại tiền = VND, cột Doanh thu phí (Nguyên tệ) = Doanh thu phí (VND) = 5,427,167.\n"
     "(ii) UI: Cả 2 cột hiển thị cùng giá trị '5,427,167'. Không có thập phân (VND).",
     ""),

    # TC-033 (SC-31)
    ("US31-TC-033", "SC-31",
     "BA trả lời QA-03.1 – *\"Dữ liệu đã được quy đổi và lưu từ thời điểm phát sinh giao dịch và không tính lại khi truy xuất báo cáo\"*",
     FEATURE, MOD_LUOI,
     "Tỷ giá VND — Giá trị đã lưu từ thời điểm GD, không tính lại khi tra cứu",
     "Business Logic", "High",
     PRECON + "\n4. Hệ thống có giao dịch ngoại tệ (USD) thu ngày 15/01/2025 với tỷ giá lúc đó = 24,500 VND/USD.\n"
     "5. Tỷ giá hiện tại (ngày tra cứu) = 25,000 VND/USD.",
     "1. Tra cứu với Từ ngày = '01/01/2025', Đến ngày = '31/01/2025'.\n"
     "2. Quan sát dòng có giao dịch USD ngày 15/01/2025.\n"
     "3. Kiểm tra cột Doanh thu phí (VND).",
     "(i) Nghiệp vụ/Logic: Cột Doanh thu phí (VND) hiển thị giá trị đã quy đổi theo tỷ giá 24,500 (thời điểm phát sinh GD), "
     "KHÔNG phải 25,000 (tỷ giá hiện tại). Giá trị không thay đổi dù tra cứu ở thời điểm khác.\n"
     "(ii) UI: Giá trị VND cố định, không đổi giữa các lần tra cứu.",
     ""),

    # TC-034 (SC-32)
    ("US31-TC-034", "SC-32",
     "QTC-06 – *\"Sắp xếp mặc định theo ngày update, ngày tạo giảm dần\"*",
     FEATURE, MOD_LUOI,
     "Lưới sắp xếp mặc định theo ngày update/tạo giảm dần",
     "Business Logic", "Medium",
     PRECON_DATA,
     "1. Quan sát thứ tự các dòng trên lưới kết quả.",
     "(i) Nghiệp vụ/Logic: Dữ liệu sắp xếp mặc định theo ngày update/ngày tạo giảm dần (gần nhất lên trước).\n"
     "(ii) UI: Dòng đầu tiên là bản ghi có ngày cập nhật gần nhất.",
     "[Theo QTC-06]"),

    # TC-035 (SC-42)
    ("US31-TC-035", "SC-42",
     "Bảng mô tả trường STT 5 – *\"Cho phép người dùng chọn loại tính phí: Theo giao dịch/Theo định kỳ\"*",
     FEATURE, MOD_TRA,
     "Filter Loại tính phí = 'Theo giao dịch' — Lưới chỉ hiển thị dữ liệu Theo giao dịch",
     "Business Logic", "Medium",
     PRECON + "\n4. Hệ thống có dữ liệu cả 2 loại: Theo giao dịch và Theo định kỳ.",
     "1. Chọn Loại tính phí = 'Theo giao dịch'.\n"
     "2. Chọn Từ ngày = '01/01/2025', Đến ngày = '31/03/2025'.\n"
     "3. Nhấn nút 'Tra cứu'.\n"
     "4. Quan sát cột Loại tính phí trên lưới.",
     "(i) Nghiệp vụ/Logic: Lưới chỉ hiển thị dữ liệu có Loại tính phí = 'Theo giao dịch'. "
     "Không lẫn dữ liệu 'Theo định kỳ'.\n"
     "(ii) UI: Tất cả dòng trên lưới có cột Loại tính phí = 'Theo giao dịch'.",
     ""),

    # TC-036 (SC-33) Data Integrity
    ("US31-TC-036", "SC-33",
     "Mục \"Yêu cầu nghiệp vụ\" v2 – *\"tổng doanh thu phí [...] của mỗi nhóm Chi nhánh + Loại tiền + Loại tính phí + Biểu phí + Code phí\"* + BA xác nhận QA-03.2",
     FEATURE, MOD_LUOI,
     "Data Integrity: Group by đúng tổ hợp — Mỗi dòng là 1 tổ hợp duy nhất",
     "Data Integrity", "High",
     PRECON_DATA + "\n5. Hệ thống có nhiều giao dịch thuộc cùng tổ hợp (CN Hà Nội + VND + Theo GD + BP001 + FEE001).",
     "1. Tra cứu với Mã Chi nhánh = '001 - CN Hà Nội', Từ ngày = '01/01/2025', Đến ngày = '31/03/2025'.\n"
     "2. Kiểm tra dòng có tổ hợp (001 + VND + Theo giao dịch + BP001 + FEE001) trên lưới.\n"
     "3. Kiểm tra số lượng dòng có cùng tổ hợp.",
     "(i) Nghiệp vụ/Logic: Chỉ có DUY NHẤT 1 dòng cho tổ hợp (001 + VND + Theo giao dịch + BP001 + FEE001). "
     "Doanh thu = tổng tất cả giao dịch thuộc tổ hợp đó.\n"
     "(ii) UI: Không có 2 dòng trùng tổ hợp trên lưới.",
     ""),

    # TC-037 (SC-34) File export
    ("US31-TC-037", "SC-34",
     "QTC-05 – *\"Template = fields trên lưới, tải xuống theo điều kiện tìm kiếm\"*",
     FEATURE, MOD_TAI,
     "Data Integrity: File Excel chứa đúng dữ liệu đang hiển thị trên lưới",
     "Data Integrity", "High",
     PRECON_DATA,
     "1. Nhấn nút 'Tải xuống' → chọn 'Excel'.\n"
     "2. Mở file .xlsx đã tải.\n"
     "3. So sánh dữ liệu trong file với dữ liệu đang hiển thị trên lưới.",
     "(i) Nghiệp vụ/Logic: File Excel chứa đúng dữ liệu theo điều kiện tra cứu đang áp dụng. "
     "Tất cả 12 cột có trên lưới đều xuất hiện trong file. Số lượng bản ghi khớp.\n"
     "(ii) UI: File mở thành công, header cột khớp với lưới, dữ liệu khớp 100%.",
     "[Theo QTC-05]"),

    # TC-038 (SC-35) Tổng cộng
    ("US31-TC-038", "SC-35",
     "Bảng mô tả trường STT 20 – *\"Hiển thị tổng doanh thu phí VND trong khoảng thời gian từ ngày – đến ngày\"* + BA trả lời QA-01.5",
     FEATURE, MOD_TONG,
     "Data Integrity: Tổng cộng = Tổng VND toàn bộ bản ghi (xuyên suốt các trang)",
     "Data Integrity", "High",
     PRECON_DATA + "\n5. Kết quả có >50 bản ghi (2+ trang).",
     "1. Cộng thủ công cột Doanh thu phí (VND) của tất cả bản ghi (trang 1 + trang 2).\n"
     "2. So sánh với giá trị dòng Tổng cộng trên màn hình.",
     "(i) Nghiệp vụ/Logic: Tổng cộng = Tổng cột Doanh thu phí (VND) của TOÀN BỘ bản ghi. "
     "Không chỉ trang hiện tại.\n"
     "(ii) UI: Giá trị Tổng cộng khớp chính xác với phép tính thủ công.",
     ""),

    # TC-039 (SC-40) Ngày giao dịch
    ("US31-TC-039", "SC-40",
     "BA trả lời QA-01.6 – *\"BA đã cập nhật US dùng thống nhất trường Ngày giao dịch\"* + Bảng mô tả trường STT 6-7",
     FEATURE, MOD_TRA,
     "Data Integrity: Hệ thống lọc chính xác theo trường 'Ngày giao dịch' trong lịch sử thu phí",
     "Data Integrity", "High",
     PRECON + "\n4. Hệ thống có giao dịch thu phí với Ngày giao dịch = 15/02/2025.\n"
     "5. Giao dịch này có Ngày tạo bản ghi = 16/02/2025 (khác Ngày giao dịch).",
     "1. Chọn Từ ngày = '15/02/2025', Đến ngày = '15/02/2025'.\n"
     "2. Nhấn nút 'Tra cứu'.\n"
     "3. Kiểm tra giao dịch có xuất hiện.\n"
     "4. Chọn Từ ngày = '16/02/2025', Đến ngày = '16/02/2025'.\n"
     "5. Nhấn nút 'Tra cứu'.\n"
     "6. Kiểm tra giao dịch có xuất hiện.",
     "(i) Nghiệp vụ/Logic: Ở bước 3, giao dịch PHẢI xuất hiện (Ngày giao dịch = 15/02). "
     "Ở bước 6, giao dịch KHÔNG xuất hiện (Ngày giao dịch ≠ 16/02, đó là Ngày tạo). "
     "Hệ thống lọc theo 'Ngày giao dịch', không phải 'Ngày tạo bản ghi'.\n"
     "(ii) UI: Lưới hiển thị/không hiển thị đúng theo logic lọc.",
     ""),

    # TC-040 (SC-43) Tổng cộng in export
    ("US31-TC-040", "SC-43",
     "Bảng mô tả trường STT 20 – *\"Hiển thị tổng doanh thu phí VND\"* + QTC-05",
     FEATURE, MOD_TAI,
     "Data Integrity: File export bao gồm dòng Tổng cộng doanh thu phí VND",
     "Data Integrity", "Medium",
     PRECON_DATA,
     "1. Nhấn nút 'Tải xuống' → chọn 'Excel'.\n"
     "2. Mở file .xlsx.\n"
     "3. Kiểm tra cuối file có dòng Tổng cộng.",
     "(i) Nghiệp vụ/Logic: File Excel bao gồm dòng Tổng cộng doanh thu phí VND ở cuối dữ liệu.\n"
     "(ii) UI: Dòng Tổng cộng hiển thị trong file, giá trị khớp với Tổng cộng trên màn hình.",
     "[Theo QTC-05]"),
]

fpath = os.path.join(os.path.dirname(os.path.dirname(__file__)), "US31_TestCases.xlsx")
wb = openpyxl.load_workbook(fpath)
ws = wb["Test Cases"]
start_row = ws.max_row + 1
data_align = Alignment(vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))

for ri, tc in enumerate(TCS, start_row):
    for ci, val in enumerate(tc, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.alignment = data_align
        cell.border = thin_border

ws.auto_filter.ref = f"A1:S{ws.max_row}"
wb.save(fpath)
print(f"✅ Batch 4 done: {len(TCS)} TC (BL+DI)")
print(f"📄 Total: {ws.max_row - 1} TC")

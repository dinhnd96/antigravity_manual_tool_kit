#!/usr/bin/env python3
"""US06 TC Generator - Batch 1: Happy Path (SC-01 to SC-11, SC-05b, SC-19, SC-61)"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Cases"

# Headers
HEADERS = ["TC_ID","SC_Ref","Reference","Feature","Module","Title","Type","Priority",
           "Precondition","Steps","Expected","Note","Tester","Ngày test","Kết quả",
           "Bug ID","Retest","Retest Result","Ghi chú QA"]

hdr_font = Font(bold=True, size=10)
hdr_fill = PatternFill("solid", fgColor="D9E2F3")
wrap = Alignment(wrap_text=True, vertical="top")
thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))

for i, h in enumerate(HEADERS, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.border = thin

# Column widths
widths = [12,8,40,18,25,50,15,8,40,55,55,30]+[12]*7
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64+i) if i<=26 else 'A'+chr(64+i-26)].width = w

FEAT = "Khai báo Biểu phí"
PRE_MAKER = "1. User đăng nhập với quyền Maker tại chức năng Danh mục Biểu phí\n2. Đã có SPDV và Code phí ở trạng thái Hoạt động trong hệ thống"

data = [
    # SC-01
    ("US06-TC-001", "SC-01",
     'Mục "Thêm mới Biểu phí", Flowchart Bước 1→7c – "Maker khai báo đầy đủ thông tin hợp lệ"',
     FEAT, "Thêm mới Biểu phí",
     "Thêm mới Biểu phí thành công với đầy đủ thông tin hợp lệ (E2E)",
     "Happy Path", "High",
     PRE_MAKER,
     "1. Tại Danh mục Biểu phí, click nút 'Thêm mới'\n"
     "2. Nhập Mã = 'BP-2026-001'\n"
     "3. Nhập Tên Biểu phí = 'Biểu phí chuyển tiền nội địa Q3/2026'\n"
     "4. Chọn Ngày ban hành = 01/05/2026\n"
     "5. Chọn Ngày hiệu lực = 01/07/2026\n"
     "6. Chọn Ngày hết hiệu lực = 31/12/2026\n"
     "7. Nhập Số văn bản = 'QĐ-123/2026'\n"
     "8. Nhập Tên văn bản = 'QĐ ban hành biểu phí CTNĐ'\n"
     "9. Nhập Link iDoc = 'https://idoc.bank.vn/qd123'\n"
     "10. Tích chọn Code phí 'CP001 - Phí chuyển tiền' từ cây SPDV\n"
     "11. Click nút 'Xác nhận'",
     "--- TRƯỚC KHI DUYỆT (MAKER) ---\n"
     "(i) Nghiệp vụ/Logic: Hệ thống ghi nhận thành công. Bản ghi Biểu phí ở trạng thái Chờ duyệt, Hành động = 'Thêm mới'.\n"
     "(ii) UI: Toast 'Thêm mới thành công'. Bản ghi hiển thị tại Tác vụ Pending của tôi.\n\n"
     "--- SAU KHI LAST CHECKER DUYỆT ---\n"
     "(i) Nghiệp vụ/Logic: Hệ thống lưu dữ liệu chính thức. Bản ghi cập nhật trạng thái Đã duyệt.\n"
     "(ii) UI: Toast 'Phê duyệt thành công'. Biểu phí hiển thị trên lưới Danh mục Biểu phí với trạng thái Chưa hiệu lực.",
     ""),

    # SC-02
    ("US06-TC-002", "SC-02",
     'Mục "Thêm mới", đoạn tích chọn tại trường SPDV/Code phí – "tích chọn SPDV cấp cha → tự động tích con"',
     FEAT, "Thêm mới – Gán Code phí",
     "Gán Code phí bằng tích chọn cây SPDV – tích chọn SPDV cấp cha tự động tích con và Code phí",
     "Happy Path", "High",
     PRE_MAKER + "\n3. Đang ở màn hình Thêm mới Biểu phí, đã nhập đủ thông tin chung",
     "1. Click vào trường 'Sản phẩm/Dịch vụ/Code phí'\n"
     "2. Cây SPDV hiển thị các SPDV có trạng thái = Hoạt động\n"
     "3. Tích chọn SPDV cấp 1 'SPDV001 - Chuyển tiền'\n"
     "4. Kiểm tra các SPDV cấp con và Code phí trực thuộc",
     "(i) Nghiệp vụ/Logic: Hệ thống tự động tích chọn tất cả SPDV cấp con và Code phí trực thuộc SPDV cấp 1 'Chuyển tiền'.\n"
     "(ii) UI: Checkbox SPDV cấp 1 và tất cả con hiển thị trạng thái checked. Lưới Thông tin chi tiết hiển thị danh sách Code phí vừa được gán.",
     ""),

    # SC-03
    ("US06-TC-003", "SC-03",
     'Mục "Thêm mới", đoạn Upload file danh sách code phí – "Tham chiếu tới US07"',
     FEAT, "Thêm mới – Gán Code phí",
     "Gán Code phí bằng Upload file Excel thành công (tham chiếu US07)",
     "Happy Path", "High",
     PRE_MAKER + "\n3. Đã chuẩn bị file Excel template hợp lệ chứa Code phí CP001, CP002",
     "1. Tại màn hình Thêm mới Biểu phí, click nút 'Chọn tệp'\n"
     "2. Chọn file 'DanhSachCodePhi.xlsx'\n"
     "3. Kiểm tra lưới Thông tin chi tiết",
     "(i) Nghiệp vụ/Logic: Hệ thống đọc file và gán Code phí CP001, CP002 vào Biểu phí. Tự động tích chọn tương ứng trên cây SPDV.\n"
     "(ii) UI: Lưới Thông tin chi tiết hiển thị 2 dòng Code phí CP001, CP002 với đầy đủ thông tin.",
     "[Theo US07]"),

    # SC-04
    ("US06-TC-004", "SC-04",
     'Mục "Thêm mới", đoạn Hành động Chỉnh sửa Code phí – "Điều kiện tính phí: Mô tả, Giá trị"',
     FEAT, "Thêm mới – Sửa Code phí",
     "Sửa Điều kiện tính phí và Quy tắc tính phí của Code phí trong Biểu phí thành công",
     "Happy Path", "High",
     PRE_MAKER + "\n3. Đang ở Thêm mới Biểu phí, đã gán Code phí CP001 vào lưới",
     "1. Tại lưới Thông tin chi tiết, click nút 'Chỉnh sửa' tại dòng CP001\n"
     "2. Popup sửa Code phí hiển thị\n"
     "3. Sửa Điều kiện tính phí: Mô tả = 'Phí áp dụng KH VIP', Giá trị = '50000'\n"
     "4. Sửa Quy tắc tính phí: Giá trị số = '0.5', Tối thiểu = '10000', Tối đa = '500000'\n"
     "5. Xác nhận thay đổi tại popup",
     "(i) Nghiệp vụ/Logic: Hệ thống ghi nhận thay đổi Điều kiện tính phí và Quy tắc tính phí cho Code phí CP001 trong phạm vi Biểu phí hiện tại.\n"
     "(ii) UI: Popup đóng. Lưới Thông tin chi tiết cập nhật thông tin Code phí CP001 đã sửa.",
     ""),

    # SC-05
    ("US06-TC-005", "SC-05",
     'Mục "Thêm mới" P41 – "Xem chi tiết Code phí giống màn hình Xem tại Danh mục SPDV"',
     FEAT, "Thêm mới – Xem Code phí",
     "Xem chi tiết Code phí trong luồng Thêm mới – hiển thị đúng như Danh mục SPDV",
     "Happy Path", "Medium",
     PRE_MAKER + "\n3. Đang ở Thêm mới Biểu phí, đã gán Code phí CP001",
     "1. Tại lưới Thông tin chi tiết, click nút 'Xem' tại dòng CP001\n"
     "2. Kiểm tra màn hình Xem chi tiết Code phí",
     "(i) Nghiệp vụ/Logic: Hệ thống hiển thị đầy đủ thông tin chi tiết Code phí CP001 (readonly). Không thay đổi dữ liệu.\n"
     "(ii) UI: Popup/màn hình Xem hiển thị giống màn hình Xem code phí tại Danh mục SPDV.",
     ""),

    # SC-05b
    ("US06-TC-006", "SC-05b",
     'Mục "Thêm mới" P42 – "Bỏ gắn Code phí ra khỏi Biểu phí"',
     FEAT, "Thêm mới – Bỏ gắn Code phí",
     "Bỏ gắn (Xóa) Code phí khỏi Biểu phí trong luồng Thêm mới",
     "Happy Path", "High",
     PRE_MAKER + "\n3. Đang ở Thêm mới, đã gán Code phí CP001, CP002 vào lưới (2 dòng)",
     "1. Tại lưới Thông tin chi tiết, click nút 'Xóa' tại dòng CP002\n"
     "2. Kiểm tra lưới Thông tin chi tiết\n"
     "3. Kiểm tra trường SPDV/Code phí trên cây",
     "(i) Nghiệp vụ/Logic: Code phí CP002 bị bỏ gắn khỏi Biểu phí. Cây SPDV tự động bỏ tích CP002.\n"
     "(ii) UI: Lưới Thông tin chi tiết chỉ còn 1 dòng CP001. CP002 biến mất khỏi lưới.",
     ""),

    # SC-06
    ("US06-TC-007", "SC-06",
     'Mục "Thêm mới" P43-P44 – "Hành động = Thêm mới"',
     FEAT, "Thêm mới – Tác vụ chờ duyệt",
     "Maker khai báo không sửa Code phí → Hành động tác vụ = 'Thêm mới'",
     "Happy Path", "High",
     PRE_MAKER + "\n3. Đã gán Code phí CP001 nhưng KHÔNG chỉnh sửa thông tin chi tiết Code phí",
     "1. Nhập đầy đủ thông tin chung hợp lệ (Mã='BP-TM-001', Tên='BP Test TM'...)\n"
     "2. Gán Code phí CP001 (không sửa gì)\n"
     "3. Click nút 'Xác nhận'\n"
     "4. Kiểm tra tác vụ tại Tác vụ Pending",
     "(i) Nghiệp vụ/Logic: Bản ghi Biểu phí ở trạng thái Chờ duyệt. Hành động = 'Thêm mới'.\n"
     "(ii) UI: Toast thành công. Bản ghi hiển thị tại Tác vụ Pending với Hành động = 'Thêm mới'.",
     ""),

    # SC-07
    ("US06-TC-008", "SC-07",
     'Mục "Thêm mới" P45 – "Hành động = Thêm mới – Sửa Code phí"',
     FEAT, "Thêm mới – Tác vụ chờ duyệt",
     "Maker khai báo + sửa Code phí → Hành động = 'Thêm mới – Sửa Code phí'",
     "Happy Path", "High",
     PRE_MAKER + "\n3. Đã gán Code phí CP001",
     "1. Nhập đầy đủ thông tin chung hợp lệ (Mã='BP-TM-002'...)\n"
     "2. Gán Code phí CP001\n"
     "3. Click 'Chỉnh sửa' tại dòng CP001, sửa Giá trị Điều kiện = '30000'\n"
     "4. Click nút 'Xác nhận'\n"
     "5. Kiểm tra tác vụ tại Tác vụ Pending",
     "(i) Nghiệp vụ/Logic: Bản ghi ở trạng thái Chờ duyệt. Hành động = 'Thêm mới – Sửa Code phí'.\n"
     "(ii) UI: Bản ghi tại Tác vụ Pending hiển thị Hành động = 'Thêm mới – Sửa Code phí'.",
     ""),

    # SC-08
    ("US06-TC-009", "SC-08",
     'Mục "Chỉnh sửa" P55-P60 – "Chưa hiệu lực: sửa toàn bộ trừ Mã + sửa Code phí"',
     FEAT, "Chỉnh sửa – Chưa hiệu lực",
     "Chỉnh sửa Biểu phí Chưa hiệu lực: sửa thông tin chung + sửa Code phí thành công",
     "Happy Path", "High",
     "1. User đăng nhập quyền Maker tại Danh mục Biểu phí\n"
     "2. Tồn tại Biểu phí 'BP-001' trạng thái Chưa hiệu lực (Ngày HL tương lai)\n"
     "3. Không có tác vụ Chỉnh sửa nào đang Chờ duyệt cho BP-001",
     "1. Click nút 'Chỉnh sửa' tại dòng BP-001\n"
     "2. Kiểm tra trường Mã Biểu phí = disabled\n"
     "3. Sửa Tên = 'BP Chuyển tiền Q4 (updated)'\n"
     "4. Sửa Ngày hết hiệu lực = 31/03/2027\n"
     "5. Click 'Chỉnh sửa' tại Code phí CP001, sửa Tối thiểu = '15000'\n"
     "6. Click nút 'Xác nhận'",
     "--- TRƯỚC KHI DUYỆT (MAKER) ---\n"
     "(i) Nghiệp vụ/Logic: Hệ thống ghi nhận chỉnh sửa. Mã BP-001 giữ nguyên. Bản ghi Chờ duyệt, Hành động = 'Chỉnh sửa – Sửa Code phí'.\n"
     "(ii) UI: Toast 'Chỉnh sửa thành công'. Bản ghi hiển thị tại Tác vụ Pending.\n\n"
     "--- SAU KHI LAST CHECKER DUYỆT ---\n"
     "(i) Nghiệp vụ/Logic: Hệ thống cập nhật dữ liệu chính thức. Mã BP-001 không đổi. Trạng thái Đã duyệt.\n"
     "(ii) UI: Toast 'Phê duyệt thành công'. Lưới Danh mục hiển thị thông tin mới (Tên, Ngày HHL).",
     ""),

    # SC-09
    ("US06-TC-010", "SC-09",
     'Mục "Chỉnh sửa" P62 – "Đang hiệu lực: chỉ sửa Ngày HHL"',
     FEAT, "Chỉnh sửa – Đang hiệu lực",
     "Chỉnh sửa Biểu phí Đang hiệu lực: sửa Ngày hết hiệu lực hợp lệ",
     "Happy Path", "High",
     "1. User đăng nhập quyền Maker\n"
     "2. Biểu phí 'BP-002' trạng thái Đang hiệu lực (HL=01/01/2026, HHL=30/06/2026)\n"
     "3. Ngày hệ thống = 06/05/2026",
     "1. Click 'Chỉnh sửa' tại BP-002\n"
     "2. Kiểm tra tất cả trường disabled trừ Ngày hết hiệu lực\n"
     "3. Sửa Ngày hết hiệu lực = 31/12/2026 (> 06/05/2026 và > 01/01/2026)\n"
     "4. Click nút 'Xác nhận'",
     "--- TRƯỚC KHI DUYỆT (MAKER) ---\n"
     "(i) Nghiệp vụ/Logic: Hệ thống ghi nhận. Mã BP-002 không đổi. Hành động = 'Chỉnh sửa'.\n"
     "(ii) UI: Toast 'Chỉnh sửa thành công'. Bản ghi tại Tác vụ Pending.\n\n"
     "--- SAU KHI LAST CHECKER DUYỆT ---\n"
     "(i) Nghiệp vụ/Logic: Ngày HHL cập nhật = 31/12/2026. Trạng thái = Đã duyệt.\n"
     "(ii) UI: Lưới Danh mục hiển thị Ngày HHL mới = 31/12/2026.",
     "[BA QA-01.6] Đang hiệu lực chỉ sửa Ngày HHL"),

    # SC-10
    ("US06-TC-011", "SC-10",
     'Mục "Chỉnh sửa" P63 – "Hết hiệu lực: chuyển đổi code phí sang Biểu phí mới"',
     FEAT, "Chỉnh sửa – Hết hiệu lực",
     "Biểu phí Hết hiệu lực: chuyển đổi Code phí sang Biểu phí mới (tham chiếu US09)",
     "Happy Path", "Medium",
     "1. User đăng nhập quyền Maker\n"
     "2. Biểu phí 'BP-003' trạng thái Hết hiệu lực",
     "1. Click 'Chỉnh sửa' tại BP-003\n"
     "2. Kiểm tra hệ thống cho phép chuyển đổi Code phí sang Biểu phí mới\n"
     "3. Thực hiện chuyển đổi theo US09",
     "(i) Nghiệp vụ/Logic: Hệ thống cho phép chuyển đổi Code phí từ BP-003 sang Biểu phí mới có trạng thái = Đang hiệu lực.\n"
     "(ii) UI: Hiển thị chức năng chuyển đổi biểu phí. Tham chiếu US09.",
     "[Theo US09]"),

    # SC-11
    ("US06-TC-012", "SC-11",
     'Mục "Chỉnh sửa" P64-P66 – "Hành động Chỉnh sửa hoặc Chỉnh sửa – Sửa Code phí"',
     FEAT, "Chỉnh sửa – Tác vụ chờ duyệt",
     "Sau Maker chỉnh sửa không sửa Code phí → Hành động = 'Chỉnh sửa'",
     "Happy Path", "High",
     "1. User đăng nhập quyền Maker\n"
     "2. Biểu phí 'BP-004' trạng thái Chưa hiệu lực",
     "1. Click 'Chỉnh sửa' tại BP-004\n"
     "2. Sửa Tên = 'BP Updated Name'\n"
     "3. KHÔNG sửa Code phí\n"
     "4. Click nút 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Bản ghi Chờ duyệt. Hành động = 'Chỉnh sửa' (không phải 'Chỉnh sửa – Sửa Code phí').\n"
     "(ii) UI: Tác vụ Pending hiển thị Hành động = 'Chỉnh sửa'.",
     ""),

    # SC-19
    ("US06-TC-013", "SC-19",
     'Mục "Thêm mới" P25 – "thay thế tất cả code phí đã chọn trước đó"',
     FEAT, "Thêm mới – Gán Code phí",
     "Tích chọn trước → Upload file sau → hệ thống THAY THẾ toàn bộ Code phí",
     "Happy Path", "High",
     PRE_MAKER + "\n3. File Excel chứa Code phí CP003, CP004",
     "1. Tại Thêm mới, tích chọn Code phí CP001, CP002 trên cây SPDV\n"
     "2. Lưới hiển thị 2 dòng: CP001, CP002\n"
     "3. Click nút 'Chọn tệp', upload file chứa CP003, CP004\n"
     "4. Kiểm tra lưới Thông tin chi tiết",
     "(i) Nghiệp vụ/Logic: Hệ thống thay thế toàn bộ. CP001, CP002 bị xóa. Chỉ còn CP003, CP004 từ file.\n"
     "(ii) UI: Lưới Thông tin chi tiết hiển thị 2 dòng CP003, CP004. Cây SPDV cập nhật tích chọn tương ứng.",
     ""),

    # SC-61
    ("US06-TC-014", "SC-61",
     'Mục "Chỉnh sửa" P67 + QTC-12 – "Checker duyệt Chỉnh sửa thành công"',
     FEAT, "Chỉnh sửa – Maker-Checker",
     "Checker phê duyệt thành công tác vụ Chỉnh sửa → Biểu phí cập nhật",
     "Happy Path", "High",
     "1. User đăng nhập quyền Checker\n"
     "2. Tồn tại tác vụ Chỉnh sửa Biểu phí BP-004 ở trạng thái Chờ duyệt",
     "1. Tại Tác vụ chờ duyệt, click xem tác vụ Chỉnh sửa BP-004\n"
     "2. Kiểm tra thông tin thay đổi\n"
     "3. Click nút 'Phê duyệt'",
     "(i) Nghiệp vụ/Logic: Hệ thống cập nhật dữ liệu Biểu phí BP-004 chính thức. Trạng thái = Đã duyệt.\n"
     "(ii) UI: Toast 'Phê duyệt thành công'. Bản ghi tại Tác vụ chờ duyệt cập nhật trạng thái Đã duyệt. Lưới Danh mục hiển thị thông tin mới.",
     "[Theo QTC-12]"),
]

for row_idx, row in enumerate(data, 2):
    for col_idx, val in enumerate(row, 1):
        c = ws.cell(row=row_idx, column=col_idx, value=val)
        c.alignment = wrap
        c.border = thin

out = os.path.join(os.path.dirname(os.path.dirname(__file__)), "US06_TestCases.xlsx")
wb.save(out)
print(f"✅ Batch 1 (Happy Path): {len(data)} TC → {out}")

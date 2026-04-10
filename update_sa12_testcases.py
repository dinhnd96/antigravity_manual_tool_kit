import openpyxl

file_path = '/Users/mac/antigravity-testing-kit/Feature_02_SA_Tham_So_He_Thong/test case/SA12_Quan_Ly_Bieu_Mau_Final.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['TestCases']

headers = [cell.value for cell in ws[1]]
col_map = {name: idx for idx, name in enumerate(headers)}

for row in ws.iter_rows(min_row=2):
    tc_id = row[col_map['TC_ID']].value
    if not tc_id:
        continue
    
    # 1. Update Maker -> Người dùng
    for col_name in ['Precondition', 'Steps', 'Expected']:
        val = row[col_map[col_name]].value
        if val:
            val = val.replace('Đăng nhập Maker có quyền', 'Người dùng có quyền')
            val = val.replace('Đăng nhập Maker.', 'Đăng nhập vào hệ thống.')
            val = val.replace('Maker', 'Người dùng')
            row[col_map[col_name]].value = val

    # 2. Update SA12-BR01-HAP-001
    if tc_id == 'SA12-BR01-HAP-001':
        expected = row[col_map['Expected']].value
        if expected:
            # Delete " hoặc [Chờ duyệt]"
            expected = expected.replace(' hoặc [Chờ duyệt]', '')
            row[col_map['Expected']].value = expected

    # 3. Update SA12-BR01-NEG-001 Validation fields
    elif tc_id == 'SA12-BR01-NEG-001':
        steps = row[col_map['Steps']].value
        if steps:
            steps = steps.replace(
                '[Bảng nguồn dữ liệu], [Tiêu đề]',
                '[Bảng nguồn dữ liệu], [Tiêu đề], [Nội dung]'
            )
            row[col_map['Steps']].value = steps
            
        note = row[col_map['Note']].value
        if note:
            note = note.replace(
                'Mã, Tên, Tiêu đề, Nguồn',
                'Mã, Tên, Tiêu đề, Nguồn, Nội dung'
            )
            row[col_map['Note']].value = note

wb.save(file_path)
print("SA12 test cases updated successfully.")

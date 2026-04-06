import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation
import os
import datetime

source_dir = 'Feature_02_SA_Tham_So_He_Thong/test case/'
files_to_sync = sorted([f for f in os.listdir(source_dir) if f.startswith('SA') and f.endswith('.xlsx')])
modules = [f[:4] for f in files_to_sync] # SA06, SA07...

wb_master = openpyxl.Workbook()
wb_master.remove(wb_master.active) # Remove default sheet

# Thêm sheet Dashboard
ws_dash = wb_master.create_sheet('📊 Dashboard')
ws_dash.merge_cells('A1:I1')
cell_title = ws_dash.cell(1, 1, 'PROFIX — QUALITY DASHBOARD SUMMARY')
cell_title.font = Font(bold=True, size=14)
cell_title.alignment = Alignment(horizontal='center')

headers_dash = ['STT', 'Module', 'Total TC', 'Passed', 'Failed', 'Blocked', 'N/A', 'Execution %', 'Pass Rate %']
for col, val in enumerate(headers_dash, 1):
    cell = ws_dash.cell(2, col, val)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='D9EAD3', fill_type='solid')

# Thêm sheet Daily Tracking
ws_daily = wb_master.create_sheet('📅 Daily Tracking')
ws_daily.merge_cells('A1:J1')
cell_daily_title = ws_daily.cell(1, 1, 'DAILY EXECUTION TRACKING — TEAM PROFIX')
cell_daily_title.font = Font(bold=True, size=14)
cell_daily_title.alignment = Alignment(horizontal='center')

members = ['Định', 'Vân', 'Vân Anh', 'Thanh', 'Hiền', 'Thủy', 'Thương']
headers_daily = ['Date'] + members + ['Total/Day', 'Tổng lũy kế']
for col, val in enumerate(headers_daily, 1):
    cell = ws_daily.cell(2, col, val)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='D9EAD3', fill_type='solid')

# Styles
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='top')
header_fill = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
exec_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

base_headers = ['TC_ID', 'BR_Ref', 'URD_Ref', 'Module', 'Feature', 'Title', 'Type', 'Category', 'Priority', 'Precondition', 'Steps', 'Expected Result', 'Trace_ID', 'Note']
exec_headers = ['Status R1', 'Tester R1', 'Date R1', 'Status R2', 'Tester R2', 'Date R2', 'Final Status']

def apply_anti_pattern(title, steps, expected):
    title_lower = str(title).lower()
    
    if 'xem' in title_lower or 'read-only' in title_lower:
        if steps and ('xác nhận' in str(steps).lower() or 'hoàn tất' in str(steps).lower()):
            clean_steps = [s for s in str(steps).split('\n') if 'xác nhận' not in s.lower() and 'hoàn tất' not in s.lower() and 'chờ duyệt' not in s.lower()]
            steps = '\n'.join(clean_steps)
        expected = "Form hiển thị dữ liệu read-only. Không làm thay đổi Database."
        
    if 'đóng' in title_lower or 'hủy' in title_lower or 'cancel' in title_lower:
        expected = "Đóng form, quay về dữ liệu màn trước. Không sinh bản ghi Nháp/Chờ duyệt."
        
    return steps, expected

# Xử lý các module
total_files = 0
total_anti_patterns = 0

for file_name in files_to_sync:
    total_files += 1
    module_name = file_name[:4]
    file_path = os.path.join(source_dir, file_name)
    wb_source = openpyxl.load_workbook(file_path, data_only=True)
    
    # Tìm sheet Testcase
    ws_source = None
    for s in wb_source.sheetnames:
        if 'TestCase' in s or 'Testcase' in s or 'Test Cases' in s:
            ws_source = wb_source[s]
            break
    if not ws_source:
        ws_source = wb_source.active
        
    # Tạo sheet
    ws_dest = wb_master.create_sheet(module_name)
    
    # Validation Dropdowns
    dv_status1 = DataValidation(type="list", formula1='"Pass,Fail,Blocked,N/A"', allow_blank=True)
    dv_status2 = DataValidation(type="list", formula1='"Pass,Fail,Blocked,N/A"', allow_blank=True)
    dv_status_final = DataValidation(type="list", formula1='"Pass,Fail,Blocked,N/A"', allow_blank=True)
    dv_tester1 = DataValidation(type="list", formula1=f'"{",".join(members)}"', allow_blank=True)
    dv_tester2 = DataValidation(type="list", formula1=f'"{",".join(members)}"', allow_blank=True)
    
    ws_dest.add_data_validation(dv_status1)
    ws_dest.add_data_validation(dv_status2)
    ws_dest.add_data_validation(dv_status_final)
    ws_dest.add_data_validation(dv_tester1)
    ws_dest.add_data_validation(dv_tester2)

    # Ghi Headers
    all_headers = base_headers + exec_headers
    for col, val in enumerate(all_headers, 1):
        cell = ws_dest.cell(1, col, val)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.fill = header_fill if col <= 14 else PatternFill(start_color='FCE4D6', fill_type='solid')
        cell.alignment = center_align

    # Cột định dạng width
    for col in [6, 10, 11, 12]: # F, J, K, L
        ws_dest.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 40

    row_dest = 2
    for r in range(2, ws_source.max_row + 1):
        tc_id = ws_source.cell(r, 1).value
        if not tc_id:
            continue
            
        data = [ws_source.cell(r, c).value for c in range(1, 15)]
        # Nếu data ngắn hơn 14 thì pad
        while len(data) < 14:
            data.append('')
            
        # Anti-pattern check
        orig_steps, orig_expected = data[10], data[11]
        data[10], data[11] = apply_anti_pattern(data[5], data[10], data[11])
        if data[10] != orig_steps or data[11] != orig_expected:
            total_anti_patterns += 1

        for c, val in enumerate(data, 1):
            cell = ws_dest.cell(row_dest, c, val)
            cell.border = thin_border
            if c in [6, 10, 11, 12]:
                cell.alignment = wrap_align
                
        # Format exec columns
        for c in range(15, 22):
            cell = ws_dest.cell(row_dest, c)
            cell.border = thin_border
            cell.fill = exec_fill
            cell.alignment = center_align
            
            # Apply DV
            if c == 15: dv_status1.add(cell)
            elif c == 18: dv_status2.add(cell)
            elif c == 21: dv_status_final.add(cell)
            elif c == 16: dv_tester1.add(cell)
            elif c == 19: dv_tester2.add(cell)
            elif c == 17 or c == 20: 
                cell.number_format = 'DD/MM/YYYY'
                
        row_dest += 1
        
    ws_dest.freeze_panes = 'A2'

# Populate Dashboard formulas
for idx, mod in enumerate(modules, 0):
    row = idx + 3
    ws_dash.cell(row, 1, idx + 1).border = thin_border
    ws_dash.cell(row, 2, mod).border = thin_border
    
    ws_dash.cell(row, 3, f"=COUNTA('{mod}'!A:A)-1").border = thin_border
    ws_dash.cell(row, 4, f'=COUNTIF(\'{mod}\'!U:U,"Pass")').border = thin_border
    ws_dash.cell(row, 5, f'=COUNTIF(\'{mod}\'!U:U,"Fail")').border = thin_border
    ws_dash.cell(row, 6, f'=COUNTIF(\'{mod}\'!U:U,"Blocked")').border = thin_border
    ws_dash.cell(row, 7, f'=COUNTIF(\'{mod}\'!U:U,"N/A")').border = thin_border
    
    cell_exec = ws_dash.cell(row, 8, f'=IF(C{row}>0,(D{row}+E{row}+F{row})/C{row},0)')
    cell_exec.number_format = '0.00%'
    cell_exec.border = thin_border
    
    cell_pass = ws_dash.cell(row, 9, f'=IF(C{row}>0,D{row}/C{row},0)')
    cell_pass.number_format = '0.00%'
    cell_pass.border = thin_border

total_row = len(modules) + 3
ws_dash.cell(total_row, 2, 'TOTAL').font = Font(bold=True)
for c in range(3, 10):
    col_let = openpyxl.utils.get_column_letter(c)
    cell = ws_dash.cell(total_row, c, f"=SUM({col_let}3:{col_let}{total_row-1})")
    cell.font = Font(bold=True)
    cell.border = thin_border

# Populate Daily formulas
start_date = datetime.date.today()
for i in range(20):
    row = i + 3
    date_val = start_date + datetime.timedelta(days=i)
    cell_date = ws_daily.cell(row, 1, date_val)
    cell_date.number_format = 'DD/MM/YYYY'
    cell_date.border = thin_border

    for m_idx, member in enumerate(members, 2):
        formula_parts = []
        for mod in modules:
            col_letter = openpyxl.utils.get_column_letter(m_idx)
            # R1 (Tester: P, Date: Q) and R2 (Tester: S, Date: T)
            formula_parts.append(f"COUNTIFS('{mod}'!$P:$P,{col_letter}$2,'{mod}'!$Q:$Q,$A{row})")
            formula_parts.append(f"COUNTIFS('{mod}'!$S:$S,{col_letter}$2,'{mod}'!$T:$T,$A{row})")
        
        full_form = "=" + " + ".join(formula_parts)
        cell_m = ws_daily.cell(row, m_idx, full_form)
        cell_m.border = thin_border
        
    cell_total = ws_daily.cell(row, 9, f"=SUM(B{row}:H{row})")
    cell_total.border = thin_border
    cell_cumu = ws_daily.cell(row, 10, f"=SUM(I$3:I{row})")
    cell_cumu.border = thin_border

out_path = 'Feature_02_SA_Tham_So_He_Thong/test case/ProfiX_Master_Test_Cases.xlsx'
wb_master.save(out_path)

print(f"REPORT: ")
print(f"- Số file Module đã duyệt: {total_files}")
print(f"- Số vị trí Anti-pattern được clean-up: {total_anti_patterns}")
print(f"- Đã lưu Master file tại: {out_path}")

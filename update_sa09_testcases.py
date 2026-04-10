import openpyxl

file_path = '/Users/mac/antigravity-testing-kit/Feature_02_SA_Tham_So_He_Thong/test case/SA09.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['TestCases']

# Create mapping of headers
headers = []
for cell in ws[1]:
    headers.append(cell.value)

col_map = {name: idx for idx, name in enumerate(headers)}

def extract_layers(text):
    return text if text else ""

# 1. Update existing rules
for row in ws.iter_rows(min_row=2):
    tc_id = row[col_map['TC_ID']].value
    if not tc_id:
        continue
        
    if tc_id == 'SA09-BR-HAP-001':
        old_steps = row[col_map['Steps']].value
        # Thay Mã nhóm, Tên nhóm, vv bằng đúng label
        row[col_map['Steps']].value = "1. Chọn Thêm mới.\n2. Nhập: Mã (*), Tên nhóm code phí (*), Thứ tự ưu tiên (*), Tần suất, Ngày thu.\n3. Chọn các Code phí định kỳ từ danh sách.\n4. Nhấn Xác nhận."
        
    elif tc_id == 'SA09-BR-NEG-001':
        row[col_map['Precondition']].value = "1. User Maker được phân quyền chức năng Thêm mới.\n2. Đang ở màn hình Thêm mới Nhóm code phí."
        row[col_map['Steps']].value = "1. Để trống một trong các trường: [Mã], [Tên nhóm code phí], [Thứ tự ưu tiên], [Tần suất], [Ngày thu].\n2. Nhập các trường còn lại.\n3. Nhấn Xác nhận."
        row[col_map['Note']].value = "Đã gộp các variants: 1a: Bỏ trống Mã, 1b: Bỏ trống Tên nhóm code phí, 1c: Bỏ trống Thứ tự ưu tiên, 1d: Bỏ trống Tần suất, 1e: Bỏ trống Ngày thu."
        
    elif tc_id == 'SA09-UI-01-ADD-001':
        expected = row[col_map['Expected']].value
        expected = expected.replace(
            "(ii) UI: Form hiển thị, đầy đủ các field: Mã nhóm, Tên nhóm, Loại tính phí, Mức độ ưu tiên... dạng Edit.",
            "(ii) UI: Form Thêm mới hiển thị: Mã (*), Tên nhóm code phí (*), Thứ tự ưu tiên (*), Tần suất (Hàng tháng / Hàng năm), Ngày thu (radio + input), Ngày thu theo dữ liệu (radio), Bảng dữ liệu (dropdown), Trường dữ liệu (dropdown), Code phí (dropdown + nút Kiểm tra). Grid 'Danh sách code phí': Mã phí, Tên phí, Loại tiền tệ, Mã hạch toán, VAT, Công thức tính phí. Nút: Xác nhận, Đóng."
        )
        row[col_map['Expected']].value = expected
        
    elif tc_id == 'SA09-BR-HAP-002':
        # Split into A
        row[col_map['TC_ID']].value = 'SA09-BR-HAP-002A'
        row[col_map['Steps']].value = '1. Tại màn hình Thêm mới, chọn Tần suất = "Hàng tháng".\n2. Mở Cây SPDV để chọn mã phí.'
        row[col_map['Note']].value = '' # remove duplicate note
        
    elif tc_id == 'SA09-BR-HAP-003':
        row[col_map['Precondition']].value = "1. Đăng nhập Maker.\n2. Hệ thống cấu hình 2 Nhóm phí (Nhóm A Priority=1, Nhóm B Priority=2).\n3. Hai nhóm phí có cùng Ngày thu trên cùng Tài khoản.\n4. Cả 2 Nhóm phí đã ở trạng thái Đã duyệt."
        exp = row[col_map['Expected']].value
        if exp:
            exp = exp.replace(
                "(iv) Output: Message Format đúng định dạng Json yêu cầu.",
                "(iv) Output: Message trên Topic Kafka: Nhóm A (priority=1) có offset thấp hơn Nhóm B (priority=2). JSON có trường priority đúng giá trị."
            )
            row[col_map['Expected']].value = exp
            
    elif tc_id == 'SA09-UI-05-DELETE-001':
        row[col_map['Precondition']].value = "1. Bản ghi Nhóm code phí đang ở trạng thái Nháp.\n2. User Maker được phân quyền Xóa."
        
    elif tc_id == 'SA09-FLOW-001':
        steps = row[col_map['Steps']].value
        steps = steps.replace(
            "3. [Sửa] Nhấn icon Sửa",
            "2.5. [Checker Duyệt] Đăng nhập Checker, duyệt bản ghi vừa tạo -> trạng thái 'Đã duyệt'.\n3. [Sửa] Đăng nhập Maker, nhấn icon Sửa"
        )
        row[col_map['Steps']].value = steps
        
        exp = row[col_map['Expected']].value
        if exp:
            exp = exp.replace(
                "Sinh trạng thái nháp ban đầu, sau đó cập nhật thông tin và tiến tới trạng thái Xóa.",
                "Sinh trạng thái Chờ duyệt ban đầu -> Đã duyệt -> Cập nhật thành Chờ duyệt -> Xóa."
            )
            row[col_map['Expected']].value = exp
            
    elif tc_id == 'SA09-UI-07-EXCEL-001':
        exp = row[col_map['Expected']].value
        if exp:
            exp = exp.replace("User A", "Maker")
            row[col_map['Expected']].value = exp


# Build new rows to append
new_rows_data = [
    {
        'TC_ID': 'SA09-BR-HAP-002B',
        'BR_Ref': 'BR_04',
        'URD_Ref': 'I.1.1.1',
        'Module': 'SA.09',
        'Feature': 'Thêm mới - Cây SPDV',
        'Title': 'Kiểm tra logic lọc dữ liệu Code phí trên cây SPDV theo Tần suất = Hàng năm',
        'Type': 'Happy',
        'Category': 'Regression',
        'Priority': 'P2',
        'Precondition': '1. Đăng nhập Maker.\n2. Hệ thống có đa dạng Code phí (Loại tính phí định kỳ/một lần, tần suất năm).',
        'Steps': '1. Tại màn hình Thêm mới, chọn Tần suất = "Hàng năm".\n2. Mở Cây SPDV để chọn mã phí.',
        'Expected': '(i) Nghiệp vụ/Logic: Chỉ những mã phí có Loại = "Định kỳ" và Tần suất = Hàng năm mới hiển thị.\n(ii) UI: Cây SPDV hiển thị phân cấp đúng các mã phí hợp lệ.\n(iii) Trạng thái/Audit: (N/A).\n(iv) Output: (Không có).',
        'Trace_ID': 'BR04-TREE-FILTER',
        'Note': ''
    },
    {
        'TC_ID': 'SA09-BR-HAP-002C',
        'BR_Ref': 'BR_04',
        'URD_Ref': 'I.1.1.1',
        'Module': 'SA.09',
        'Feature': 'Thêm mới - Cây SPDV',
        'Title': 'Kiểm tra logic lọc dữ liệu Code phí trên cây SPDV theo Ngày thu theo dữ liệu',
        'Type': 'Happy',
        'Category': 'Regression',
        'Priority': 'P2',
        'Precondition': '1. Đăng nhập Maker.\n2. Hệ thống có đa dạng Code phí.',
        'Steps': '1. Tại màn hình Thêm mới, chọn radio "Ngày thu theo dữ liệu".\n2. Chọn Bảng dữ liệu và Trường dữ liệu.\n3. Nhấn Kiểm tra / Mở Cây SPDV.',
        'Expected': '(i) Nghiệp vụ/Logic: Chỉ những mã phí phù hợp điều kiện thu theo dữ liệu hiển thị.\n(ii) UI: Trả về danh sách phí hợp lệ.\n(iii) Trạng thái/Audit: (N/A).\n(iv) Output: (Không có).',
        'Trace_ID': 'BR04-TREE-FILTER',
        'Note': ''
    },
    {
        'TC_ID': 'SA09-UI-05-DELETE-003',
        'BR_Ref': 'UI-FUNC.05',
        'URD_Ref': 'I.1.1.1',
        'Module': 'SA.09',
        'Feature': 'Xóa',
        'Title': 'Kiểm tra hệ thống hiển thị popup Confirm và Xóa thành công bản ghi ở trạng thái Chờ duyệt',
        'Type': 'Happy',
        'Category': 'Regression',
        'Priority': 'P2',
        'Precondition': '1. Bản ghi Nhóm code phí đang ở trạng thái Chờ duyệt.\n2. User Maker được phân quyền Xóa.',
        'Steps': '1. Tại màn hình danh sách, nhấn icon Xóa tại dòng bản ghi.\n2. Hệ thống hiển thị popup confirm.\n3. Nhấn Xác nhận.',
        'Expected': '(i) Nghiệp vụ/Logic: Bản ghi bị đánh dấu Xóa.\n(ii) UI: Toast "Xóa thành công", bản ghi biến mất khỏi lưới.\n(iii) Trạng thái/Audit: Ghi nhận log Xóa bởi Maker vào ngày T.\n(iv) Output: (Không có).',
        'Trace_ID': 'UI05-DEL-HAP',
        'Note': ''
    },
    {
        'TC_ID': 'SA09-UI-KIEMTRA-001',
        'BR_Ref': 'UI-FUNC.01',
        'URD_Ref': 'I.1.1.1',
        'Module': 'SA.09',
        'Feature': 'Thêm mới',
        'Title': 'Kiểm tra hoạt động của nút "Kiểm tra" khi nhập Code phí hợp lệ',
        'Type': 'Happy',
        'Category': 'Regression',
        'Priority': 'P2',
        'Precondition': '1. Ở màn hình Thêm mới.\n2. Chọn 1 code phí đang tồn tại chưa thuộc nhóm nào.',
        'Steps': '1. Trỏ vào dropdown Code phí, chọn mã phí.\n2. Nhấn nút "Kiểm tra".',
        'Expected': '(i) Nghiệp vụ/Logic: Mã phí được đánh giá là hợp lệ có thể add vào nhóm.\n(ii) UI: Mã phí được đẩy xuống lưới "Danh sách code phí".\n(iii) Trạng thái/Audit: Không đổi.\n(iv) Output: (Không có).',
        'Trace_ID': 'UI-CHECK-BTN-HAP',
        'Note': ''
    },
    {
        'TC_ID': 'SA09-UI-KIEMTRA-002',
        'BR_Ref': 'UI-FUNC.01',
        'URD_Ref': 'I.1.1.1',
        'Module': 'SA.09',
        'Feature': 'Thêm mới',
        'Title': 'Kiểm tra hành vi bấm nút "Kiểm tra" báo lỗi khi nhập Code phí đã được gán nhóm khác',
        'Type': 'Negative',
        'Category': 'Regression',
        'Priority': 'P2',
        'Precondition': '1. Ở màn hình Thêm mới.\n2. Biết 1 mã phí đã thuộc nhóm khác.',
        'Steps': '1. Gõ thẳng mã code phí đã qua sử dụng.\n2. Nhấn nút "Kiểm tra".',
        'Expected': '(i) Nghiệp vụ/Logic: Hệ thống chặn thêm do trùng BR05.\n(ii) UI: Báo popup "Code phí đã tồn tại trong nhóm khác". Lưới không update.\n(iii) Trạng thái/Audit: Tương tự.\n(iv) Output: Không có.',
        'Trace_ID': 'UI-CHECK-BTN-NEG',
        'Note': ''
    },
    {
        'TC_ID': 'SA09-FLOW-CHECKER-001',
        'BR_Ref': 'FLOW.01',
        'URD_Ref': 'I.1.1.1',
        'Module': 'SA.09',
        'Feature': 'Phê duyệt',
        'Title': 'Kiểm tra chức năng Checker Duyệt bản ghi Nhóm code phí',
        'Type': 'Happy',
        'Category': 'Smoke',
        'Priority': 'P1',
        'Precondition': '1. Đăng nhập Checker.\n2. Có 1 bản ghi đang ở trạng thái "Chờ duyệt" do Maker tạo.',
        'Steps': '1. Chọn bản ghi Chờ duyệt.\n2. Nhấn icon Duyệt.\n3. Xác nhận pop-up.',
        'Expected': '(i) Nghiệp vụ/Logic: Bản ghi đổi trạng thái thành công.\n(ii) UI: Toast "Duyệt thành công".\n(iii) Trạng thái/Audit: Record status = Đã duyệt. Log ghi user Checker phê duyệt.\n(iv) Output: Có thể có Kafka message ghi đè xuống Core.',
        'Trace_ID': 'E2E-CHECKER-APP',
        'Note': ''
    },
    {
        'TC_ID': 'SA09-FLOW-CHECKER-002',
        'BR_Ref': 'FLOW.01',
        'URD_Ref': 'I.1.1.1',
        'Module': 'SA.09',
        'Feature': 'Phê duyệt',
        'Title': 'Kiểm tra chức năng Checker Từ chối bản ghi Nhóm code phí',
        'Type': 'Negative',
        'Category': 'Regression',
        'Priority': 'P2',
        'Precondition': '1. Đăng nhập Checker.\n2. Có 1 bản ghi Chờ duyệt.',
        'Steps': '1. Nhấn nút Từ chối.\n2. Nhập lý do.\n3. Xác nhận.',
        'Expected': '(i) Nghiệp vụ/Logic: Request bị cancel, lưu lý do từ chối.\n(ii) UI: Phản hồi Từ chối thành công.\n(iii) Trạng thái/Audit: Trạng thái đổi thành "Từ chối".\n(iv) Output: Push thông báo cho Maker.',
        'Trace_ID': 'E2E-CHECKER-REJ',
        'Note': ''
    },
    {
        'TC_ID': 'SA09-UI-PAGE-001',
        'BR_Ref': 'UI-FUNC.06',
        'URD_Ref': 'I.1.1.1',
        'Module': 'SA.09',
        'Feature': 'Danh sách',
        'Title': 'Kiểm tra tính năng phân trang trên lưới Danh sách',
        'Type': 'UI',
        'Category': 'Regression',
        'Priority': 'P3',
        'Precondition': 'Có lớn hơn 10 bản ghi.',
        'Steps': '1. Mở màn hình quản trị.\n2. Chuyển trang 1 sang 2.\n3. Đổi số bản ghi hiển thị 10 sang 20.',
        'Expected': '(i) Nghiệp vụ/Logic: Load đủ số bản ghi và đúng offset.\n(ii) UI: Grid update ngay lập tức.\n(iii) Trạng thái: NA.\n(iv) Output: NA.',
        'Trace_ID': 'UI-LIST-PAGE',
        'Note': ''
    }
]

for item in new_rows_data:
    new_row = []
    for h in headers:
        new_row.append(item.get(h, ''))
    ws.append(new_row)

print("TestCases sheet updated")

# 2. Update Coverage Sheet
ws_cov = wb['Coverage']
ws_cov.append(['TC chức năng UI', 'UI-FUNC.01', 'COVERED', 'TC cho nút Kiểm Tra'])
ws_cov.append(['TC chức năng UI', 'UI-FUNC.06', 'COVERED', 'Phân trang'])
ws_cov.append(['Luồng Nghiệp vụ', 'FLOW.01', 'COVERED', 'Kịch bản Checker (Approve/Reject)'])

for row in ws_cov.iter_rows(min_row=2):
    if row[1].value == 'BR_04':
        row[3].value = 'Cover đủ Hàng tháng, Hàng năm, Theo Dữ liệu'
    elif row[1].value == 'BR_06':
        row[3].value = 'Cập nhật đủ output format Priority Topic'

print("Coverage sheet updated")
wb.save(file_path)

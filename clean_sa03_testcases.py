import openpyxl

file_path = '/Users/mac/antigravity-testing-kit/Feature_02_SA_Tham_So_He_Thong/test case/SA03_Quản lý người dùng.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['TestCases'] if 'TestCases' in wb.sheetnames else wb.worksheets[0]

headers = [cell.value for cell in ws[1]]
col_map = {name: idx for idx, name in enumerate(headers)}

updates = {
    'SA03-BOUNDARY-009': {
        'Title': '[PENDING-BA] Tìm kiếm với từ khóa chứa ký tự đặc biệt hoặc độ dài tối đa',
        'Note': 'ANTI-PATTERN: URD chưa quy định giới hạn độ dài hay chặn ký tự DB tìm kiếm. Cần BA xác nhận spec trước khi test, nếu không Dev sẽ reject bug.'
    },
    'SA03-FLOW-010': {
        'Title': '[MARK FOR DELETE] Luồng đăng nhập bị chặn khi Inactive',
        'Note': 'DUPLICATE/OUT OF SCOPE: Việc Inactive không đăng nhập được thuộc phạm vi của màn SA.01 (Đăng nhập). Ở module SA.03 chỉ cần verify Admin lưu trạng thái Inactive thành công trên DB/Lưới UI là đủ.'
    },
    'SA03-SECURITY-011': {
        'Title': '[PENDING-BA] Security: Chặn Admin tự khóa tài khoản của chính mình (Self-Lockout)',
        'Note': 'ANTI-PATTERN CRITICAL: URD không có business rule cấm Admin tự Active/Inactive bản thân. Cần BA confirm bổ sung chặn rule này, nếu BA không bổ sung thì phải Delete TC này.'
    }
}

for row in ws.iter_rows(min_row=2):
    tc_id_val = row[col_map['TC_ID']].value
    if tc_id_val and tc_id_val in updates:
        for col_name, new_val in updates[tc_id_val].items():
            if col_name in col_map:
                row[col_map[col_name]].value = new_val
                print(f"Updated [{tc_id_val}] column [{col_name}]")

wb.save(file_path)
print("\nSA03 cleanup applied successfully.")

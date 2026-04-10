import openpyxl

wb = openpyxl.load_workbook('./Feature_02_SA_Tham_So_He_Thong/test case/SA01_Đăng nhập.xlsx', data_only=True)
print("SA01 Sheets:", wb.sheetnames)
if 'Test Cases' in wb.sheetnames:
    tc = wb['Test Cases']
    for i, row in enumerate(tc.iter_rows(values_only=True)):
        if i < 1:
            print("Row 1:", row)
            break

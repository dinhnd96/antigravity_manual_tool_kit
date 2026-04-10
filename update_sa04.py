import openpyxl

wb = openpyxl.load_workbook('Feature_02_SA_Tham_So_He_Thong/SA04_Test_Cases.xlsx')

# 1. Update Coverage Sheet
ws_cov = wb['Coverage']
# Clear existing rows (keep header)
ws_cov.delete_rows(2, ws_cov.max_row)

rows = [
    ('Business Rule', 'BR_01', 'Covered', 'Tham chiếu bởi SA04-BR-HAP-001, SA04-BR-HAP-002, SA04-BR-NEG-001'),
    ('Business Rule', 'BR_02', 'Covered', 'Tham chiếu bởi SA04-BR-HAP-003, SA04-UI-NEG-002b (đề xuất)'),
    ('Business Rule', 'BR_03', 'Covered', 'Tham chiếu bởi SA04-BR-NEG-002, SA04-BR-NEG-003, SA04-BR-NEG-004 (đề xuất)'),
    ('Business Rule', 'BR_04', 'Covered', 'Tham chiếu bởi SA04-BR-HAP-001, SA04-BR-HAP-002'),
    ('Business Rule', 'BR_05', 'Covered', 'Tham chiếu bởi SA04-BR-HAP-001'),
    ('UI Function', 'UI-FUNC.01', 'Covered', 'Tham chiếu bởi SA04-UI-001'),
    ('UI Function', 'UI-FUNC.02', 'Covered', 'Tham chiếu bởi SA04-UI-002, SA04-UI-NEG-002a, SA04-UI-NEG-002b'),
    ('UI Function', 'UI-FUNC.03', 'Covered', 'Tham chiếu bởi SA04-UI-003'),
    ('UI Function', 'UI-FUNC.04', 'Covered', 'Tham chiếu bởi SA04-UI-004, SA04-UI-NEG-004 (đề xuất)'),
    ('UI Function', 'UI-FUNC.05', 'Covered', 'Tham chiếu bởi SA04-UI-005, SA04-UI-NEG-005a (đề xuất)'),
    ('UI Function', 'UI-FUNC.06', 'Covered', 'Tham chiếu bởi SA04-UI-006'),
    ('E2E Flow', 'E2E', 'Covered', 'Tham chiếu bởi SA04-E2E-HAP-001'),
    ('UI List', 'UI-FUNC.00', 'Pending', 'Cần đề xuất SA04-UI-LIST-001')
]
for row in rows:
    ws_cov.append(row)

# 2. Update Test Cases Sheet
ws_tc = wb['Test Cases']

updates = {
    'SA04-BR-HAP-001': {
        'Expected': "(i) Nghiệp vụ/Logic: Lưu bản ghi nhóm quyền thành công, không qua duyệt (BR_05 - quản trị hệ thống không yêu cầu duyệt).\n(ii) UI: Form đóng thông báo 'Thêm mới Nhóm quyền thành công'.\n(iii) Trạng thái/Audit: Bản ghi lưu đổi trạng thái hiển thị trên lưới là Hoạt động. Sinh log Audit.\n(iv) Output: Không sinh message/file."
    },
    'SA04-BR-HAP-002': {
        'Expected': "(i) Nghiệp vụ/Logic: Lưu bản ghi nhóm quyền thành công và ghi nhận toàn bộ các quyền đã chọn.\n(ii) UI: Form đóng thông báo 'Thêm mới Nhóm quyền thành công' màu xanh.\n(iii) Trạng thái/Audit: Bản ghi Nhóm quyền hiển thị trên lưới với trạng thái Hoạt động. Sinh log Audit thêm mới.\n(iv) Output: Không."
    },
    'SA04-BR-HAP-003': {
        'Type': 'Exception'
    },
    'SA04-UI-001': {
        'Expected': "(i) Nghiệp vụ/Logic: Gọi API lấy dữ liệu form thành công.\n(ii) UI: Hiển thị màn hình Thêm mới Nhóm quyền có đầy đủ các field: ID nhóm quyền (*), Tên nhóm quyền (*), danh sách Quyền, nút Xác nhận, Đóng.\n(iii) Trạng thái/Audit: Không thay đổi.\n(iv) Output: Không."
    },
    'SA04-UI-005': {
        'Steps': "1. Click nút Tìm kiếm.\n2. Nhập tiêu chí tìm kiếm hợp lệ (3a: Theo ID nhóm quyền, 3b: Theo Tên nhóm quyền).\n3. Click xác nhận tìm."
    },
    'SA04-E2E-HAP-001': {
        'Steps': "1. Ở tính năng quản lý phân quyền, click Thêm mới Nhóm quyền.\n2. Nhập thông tin bắt buộc, chọn quyền và Xác nhận.\n3. Nhập ID vừa tạo để Tìm kiếm trên danh sách.\n4. Click nút Xem để đối chiếu, sau đó Đóng.\n5. Click nút Sửa, cấp quyền mới, Xác nhận.\n6. Click Tải xuống.\n7. Xóa bản ghi vừa tạo, xác nhận popup. Verify dòng biến mất khỏi grid."
    }
}

for i, row in enumerate(ws_tc.iter_rows(values_only=False)):
    if i == 0:
        continue
    tc_id = row[0].value
    if tc_id in updates:
        if 'Expected' in updates[tc_id]:
            row[11].value = updates[tc_id]['Expected']
        if 'Type' in updates[tc_id]:
            row[6].value = updates[tc_id]['Type']
        if 'Steps' in updates[tc_id]:
            row[10].value = updates[tc_id]['Steps']

# Append new missing TCs at the end
new_tcs = [
    ('SA04-BR-NEG-004', 'BR_03, UI-FUNC.02', 'SA.04', 'SA04', 'Sửa', 'Kiểm tra hệ thống báo lỗi khi Sửa Tên nhóm quyền trùng với bản ghi khác', 'Negative', 'Regression', 'P1', 
        '1. Tồn tại nhóm quyền có Tên "Nhóm báo cáo".\n2. Người quản trị ở màn hình danh sách Quản lý phân quyền.',
        '1. Nhấn nút Sửa tại dòng bản ghi khác "Nhóm báo cáo".\n2. Nhập Tên nhóm quyền thành "Nhóm báo cáo".\n3. Nhấn Xác nhận.', 
        "(i) Nghiệp vụ/Logic: Hệ thống chặn lưu vì Tên nhóm quyền bị trùng.\n(ii) UI: Màn hình giữ nguyên, hiển thị popup/text báo lỗi 'Tên nhóm quyền đã tồn tại'.\n(iii) Trạng thái/Audit: Không tạo thay đổi, không sinh log.\n(iv) Output: Không.",
        'BR03-DUP-NAME-EDIT', 'Bổ sung theo Báo cáo Gap (#2)', None, None, None, None, None, None, None),
    ('SA04-UI-NEG-004', 'UI-FUNC.04', 'SA.04', 'SA04', 'Xóa', 'Kiểm tra hệ thống chặn Xóa khi nhóm quyền đang được gán cho user', 'Negative', 'Smoke', 'P1', 
        '1. Tồn tại nhóm quyền đang được gán cho ít nhất 1 người dùng.\n2. Người quản trị ở màn hình danh sách Quản lý phân quyền.',
        '1. Nhấn nút Xóa tại dòng bản ghi đang được gán user.\n2. Xác nhận xóa trên popup hệ thống.', 
        "(i) Nghiệp vụ/Logic: Chặn Xóa do vi phạm toàn vẹn dữ liệu.\n(ii) UI: Hiển thị popup cảnh báo Nhóm quyền đang được sử dụng.\n(iii) Trạng thái/Audit: Không tạo thay đổi, không sinh log.\n(iv) Output: Không.",
        'UI-04-DEL-DENY', 'Bổ sung theo Báo cáo Gap (#3)', None, None, None, None, None, None, None),
    ('SA04-UI-NEG-002a', 'UI-FUNC.02, BR_01', 'SA.04', 'SA04', 'Sửa', 'Kiểm tra hệ thống chặn lưu khi bỏ trống trường bắt buộc lúc Sửa', 'Negative', 'Regression', 'P2', 
        '1. Tồn tại bản ghi Nhóm quyền.\n2. Người quản trị ở màn hình Quản lý phân quyền.',
        '1. Nhấn nút Sửa tại dòng bản ghi.\n2. Xóa trắng trường bắt buộc (Tên nhóm quyền).\n3. Nhấn Xác nhận.', 
        "(i) Nghiệp vụ/Logic: Hệ thống chặn không cho phép lưu bản ghi do vi phạm ràng buộc trống.\n(ii) UI: Form không tắt, focus và bôi đỏ viền ô lỗi.\n(iii) Trạng thái/Audit: Không tạo thay đổi, không sinh log.\n(iv) Output: Không.",
        'UI-02-REQ-EDIT', 'Bổ sung theo Báo cáo Gap (#10)', None, None, None, None, None, None, None),    
    ('SA04-UI-NEG-002b', 'UI-FUNC.02, BR_02', 'SA.04', 'SA04', 'Sửa', 'Kiểm tra hệ thống hủy lưu dữ liệu khi nhấn Đóng lúc Sửa', 'Exception', 'Regression', 'P2', 
        '1. Tồn tại bản ghi Nhóm quyền.\n2. Người quản trị ở form Sửa bản ghi.',
        '1. Thay đổi thông tin bất kỳ.\n2. Chọn nút Đóng.', 
        "(i) Nghiệp vụ/Logic: Hủy thao tác Sửa, hệ thống không lưu thông tin vừa nhập.\n(ii) UI: Đóng form Sửa, trở về màn hình danh sách Nhóm quyền.\n(iii) Trạng thái/Audit: Không cập nhật bản ghi, không sinh log.\n(iv) Output: Không.",
        'UI-02-CLOSE-EDIT', 'Bổ sung theo Báo cáo Gap (#10)', None, None, None, None, None, None, None), 
    ('SA04-UI-NEG-005a', 'UI-FUNC.05', 'SA.04', 'SA04', 'Tìm kiếm', 'Kiểm tra kết quả danh sách rỗng khi Tìm kiếm với giá trị không tồn tại', 'Negative', 'Regression', 'P2', 
        '1. Đang ở màn hình danh sách Quản lý phân quyền.',
        '1. Tìm kiếm theo ID/Tên nhóm quyền bằng giá trị không tồn tại.\n2. Chọn nút Tìm kiếm.', 
        "(i) Nghiệp vụ/Logic: Truy vấn database không lấy ra bản ghi nào.\n(ii) UI: Grid hiển thị rỗng với thông báo 'Không tìm thấy kết quả'.\n(iii) Trạng thái/Audit: Không tạo thay đổi.\n(iv) Output: Không.",
        'UI-05-EMPTY-SRC', 'Bổ sung theo Báo cáo Gap (#4)', None, None, None, None, None, None, None), 
    ('SA04-UI-LIST-001', 'UI-FUNC.00', 'SA.04', 'SA04', 'Xem', 'Verify thông tin cột ở lưới danh sách Nhóm quyền theo chuẩn Mockup', 'Happy', 'Regression', 'P3', 
        '1. Đang ở màn hình danh sách Quản lý phân quyền.',
        '1. Xem lưới danh sách.', 
        "(i) Nghiệp vụ/Logic: Trả về field hiển thị trên lưới.\n(ii) UI: Lưới hiển thị 12 cột: Mã nhóm, Tên nhóm, Loại nhóm, Trạng thái, Mô tả, Kênh, Ngày tạo, Người tạo, Ngày sửa, Người sửa, Hành động (với nút sửa, xem, xóa).\n(iii) Trạng thái/Audit: Không tạo thay đổi.\n(iv) Output: Không.",
        'UI-LST-COLS', 'Bổ sung theo Báo cáo Gap (#12)', None, None, None, None, None, None, None)          
]

for new_tc in new_tcs:
    ws_tc.append(new_tc)

wb.save('Feature_02_SA_Tham_So_He_Thong/SA04_Test_Cases_Updated.xlsx')
print('Update complete: SA04_Test_Cases_Updated.xlsx created')

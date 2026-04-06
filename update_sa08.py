import openpyxl

file_path = 'Feature_02_SA_Tham_So_He_Thong/test case/SA08_Test_Cases.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active

# Column mapping (1-based)
c_TC_ID = 1
c_BR_Ref = 2
c_URD_Ref = 3
c_Type = 7
c_Category = 8
c_Precondition = 10
c_Steps = 11
c_Note = 14

for row in range(1, ws.max_row + 1):
    tc_id = ws.cell(row, c_TC_ID).value
    
    if not tc_id:
        continue
    
    # 1. SA08-HPY-001: Đổi Type -> Negative
    if tc_id == 'SA08-HPY-001':
        ws.cell(row, c_Type).value = 'Negative'
        precond = ws.cell(row, c_Precondition).value
        if precond and 'Người dùng đã đăng nhập hệ thống' in precond:
            ws.cell(row, c_Precondition).value = precond.replace('Người dùng đã đăng nhập hệ thống', 'Đăng nhập hệ thống bằng tài khoản Maker')
            
    # 2. SA08-AUTH-001: Update ref, type, category, note
    elif tc_id == 'SA08-AUTH-001':
        ws.cell(row, c_BR_Ref).value = 'SEC-01'
        ws.cell(row, c_Type).value = 'Security'
        ws.cell(row, c_Category).value = 'Smoke'
        note = str(ws.cell(row, c_Note).value or '')
        ws.cell(row, c_Note).value = (note + "\nAssumption - Security check ngoài phạm vi URD module").strip()

    # 3. SA08-CALC-001: URD_Ref, Note
    elif tc_id == 'SA08-CALC-001':
        ws.cell(row, c_URD_Ref).value = 'SA.08/BR_03'
        note = str(ws.cell(row, c_Note).value or '')
        ws.cell(row, c_Note).value = (note + "\nCần automation hoặc tool đo phiên thực tế (VD: cấu hình N=1 phút để test nhanh)").strip()

    # 4. SA08-CALC-002: URD_Ref
    elif tc_id == 'SA08-CALC-002':
        ws.cell(row, c_URD_Ref).value = 'SA.08/BR_03'

    # 5. SA08-E2E-001: Thêm verify grid vào steps
    elif tc_id == 'SA08-E2E-001':
        steps = str(ws.cell(row, c_Steps).value or '')
        if '4. Nhấn nút Xem' in steps:
            steps_new = steps.replace('4. Nhấn nút Xem', '4. Nhấn nút Xem\n4b. Quan sát lưới danh sách, xác nhận cột Giá trị hiển thị đúng giá trị vừa sửa')
            ws.cell(row, c_Steps).value = steps_new

    # 6. SA08-UI-VIEW-001: Fix typo "nút Xét" -> "nút Xem"
    elif tc_id == 'SA08-UI-VIEW-001':
        steps = str(ws.cell(row, c_Steps).value or '')
        if 'nút Xét' in steps:
            ws.cell(row, c_Steps).value = steps.replace('nút Xét', 'nút Xem')

    # 7. SA08-UI-EDIT-001: Fix Note reference
    elif tc_id == 'SA08-UI-EDIT-001':
        note = str(ws.cell(row, c_Note).value or '')
        if 'ĐÃ COVER' in note and 'TC_ID=' not in note:
            ws.cell(row, c_Note).value = "ĐÃ COVER ở TC_ID=SA08-CALC-001 và TC_ID=SA08-NEG-001"

    # 8. SA08-NEG-001: Add note about mandatory fields
    elif tc_id == 'SA08-NEG-001':
        note = str(ws.cell(row, c_Note).value or '')
        ws.cell(row, c_Note).value = (note + "\nCần BA xác nhận danh sách trường bắt buộc trong SA.08").strip()

wb.save(file_path)
print("Updated Test Cases in " + file_path)

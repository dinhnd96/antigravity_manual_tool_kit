import openpyxl

file_path = '/Users/mac/antigravity-testing-kit/Feature_02_SA_Tham_So_He_Thong/test case/SA10_Quan_Ly_Job_Phi_Final_Standard.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['TestCases']

headers = [cell.value for cell in ws[1]]
col_map = {name: idx for idx, name in enumerate(headers)}

# Update existing TCs
for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
    tc_id = row[col_map['TC_ID']].value
    if not tc_id:
        continue
    
    # 1. Role replace in Precondition, Steps, Expected
    for col_name in ['Precondition', 'Steps', 'Expected']:
        val = row[col_map[col_name]].value
        if val:
            val = val.replace('Đăng nhập Maker.', 'Đăng nhập vào hệ thống.')
            val = val.replace('User Maker có quyền.', 'Người dùng có quyền.')
            val = val.replace('User Maker', 'Người dùng')
            val = val.replace('Maker', 'Người dùng')
            row[col_map[col_name]].value = val

    # 2. TC SA10-BR01-HAP-001
    if tc_id == 'SA10-BR01-HAP-001':
        steps = row[col_map['Steps']].value
        steps = steps.replace('Mã Job, Thứ tự chạy, Nhóm code phí', 'Mã số job, Thứ chạy job, Nhóm code phí')
        row[col_map['Steps']].value = steps

    # 3. TC SA10-BR01-NEG-002
    elif tc_id == 'SA10-BR01-NEG-002':
        row[col_map['Steps']].value = '1. Để trống một trong các trường: [Mã số job], [Thứ chạy job], [Tên job], [Nhóm code phí], [Lệnh thực thi].\n2. Nhập các trường còn lại.\n3. Nhấn Xác nhận.'
        row[col_map['Note']].value = 'Gộp các trường hợp: Bỏ trống từng field bắt buộc của Job theo variants 1a, 1b...'

    # 4. TC SA10-UI-01-ADD-001
    elif tc_id == 'SA10-UI-01-ADD-001':
        expected = row[col_map['Expected']].value
        expected = expected.replace(
            'Mã Job, Thứ tự, Nhóm code phí (dropdown), Tần suất... Các field rỗng hợp lệ.',
            'Mã số job (*), Thứ chạy job (*), Tên job (*), Nhóm code phí (*), Mô tả job, Kiểu job, Từ thời điểm, Đến thời điểm, Tần suất, Ngày thu, Ngày thực thi tiếp theo, Lệnh thực thi (*). Các field rỗng hợp lệ.'
        )
        row[col_map['Expected']].value = expected


# 5. Add New TC: SA10-LOG-NEG-001
new_tc = {
    'TC_ID': 'SA10-LOG-NEG-001',
    'BR_Ref': 'General',
    'URD_Ref': 'I.1.1.1',
    'Module': 'SA.10',
    'Feature': 'Thay thế TK',
    'Title': 'Kiểm tra ngoại lệ xử lý trích thu khi tất cả tài khoản thay thế đều không đủ số dư',
    'Type': 'Negative',
    'Category': 'Regression',
    'Priority': 'P2',
    'Precondition': '1. Đăng nhập vào hệ thống.\n2. Cả TK mặc định và mọi TK phụ (thay thế) của Khách hàng đều có 0 VNĐ.',
    'Steps': '1. Đợi hoặc kích hoạt Chạy thủ công Job thu phí đối với Khách hàng này.\n2. Kiểm tra log thực thi.',
    'Expected': '(i) Nghiệp vụ/Logic: Job ghi nhận log Failed, trigger bắn Email thông báo lỗi theo BR_04 do không trích thu được.\n(ii) UI: (N/A).\n(iii) Trạng thái/Audit: Job log báo [Lỗi] - "Không có tài khoản đủ số dư".\n(iv) File/Email: Nhận được email cấu hình.',
    'Trace_ID': 'LOG-ACCOUNT-FAILALL',
    'Note': 'Bổ sung theo rà soát luồng biên.'
}

new_row = []
for h in headers:
    new_row.append(new_tc.get(h, ''))
ws.append(new_row)

# 6. Update Coverage Sheet
ws_cov = wb['Coverage']
ws_cov.append(['Quy tắc chung', 'General', 'COVERED', 'Testing sweep account fallback khi tất cả accounts không đủ số dư'])

wb.save(file_path)
print("Updated SA10 perfectly.")

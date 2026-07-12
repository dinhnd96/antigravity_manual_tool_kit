import openpyxl
from copy import copy

# Load workbook
wb = openpyxl.load_workbook('Daily_Checklist_Timeline.xlsx')

def copy_cell_style_to_notes(sheet, r):
    src_cell = sheet.cell(row=r, column=5)
    
    # Thiết lập alignment cho ô chính E
    src_cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top', horizontal='left')
    
    # Copy style sang F, G, H, I
    for col in range(6, 10):
        dst_cell = sheet.cell(row=r, column=col)
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.border = copy(src_cell.border)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top', horizontal='left')

for day in ['T2', 'T3', 'T4', 'T5', 'T6', 'T7', 'CN']:
    sheet = wb[day]
    
    # 1. Điều chỉnh độ rộng cột F, G, H, I thành 13.0 để phần ghi chú nới rộng thêm 4 ô rất đẹp
    for col_letter in ['F', 'G', 'H', 'I']:
        sheet.column_dimensions[col_letter].width = 13.0
        
    # 2. Quét các merged cells hiện tại để tìm các dòng tiêu đề gộp A:E
    ranges = list(sheet.merged_cells.ranges)
    header_rows = set()
    
    # Unmerge và merge lại từ A đến I cho các dòng tiêu đề gộp A:E (kể cả hàng 1)
    for r_range in ranges:
        if r_range.min_col == 1 and r_range.max_col == 5:
            # Lưu lại hàng tiêu đề
            for r in range(r_range.min_row, r_range.max_row + 1):
                header_rows.add(r)
            try:
                sheet.unmerge_cells(start_row=r_range.min_row, start_column=r_range.min_col,
                                    end_row=r_range.max_row, end_column=r_range.max_col)
                sheet.merge_cells(start_row=r_range.min_row, start_column=1,
                                  end_row=r_range.max_row, end_column=9)
                print(f"{day}: Merged header rows {r_range.min_row} to {r_range.max_row} from A to I")
            except Exception as e:
                print(f"Error merging header in {day}: {e}")
                
    # 3. Gộp E2:I2 cho dòng Header (hàng 2)
    copy_cell_style_to_notes(sheet, 2)
    sheet.merge_cells(start_row=2, start_column=5, end_row=2, end_column=9)
    sheet.cell(row=2, column=5).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    
    # 4. Quét tất cả các dòng từ dòng 3 trở đi để giãn dòng và gộp ghi chú E:I
    for r in range(3, sheet.max_row + 1):
        # Nếu dòng r là dòng tiêu đề gộp (đã được xử lý ở trên)
        if r in header_rows:
            sheet.row_dimensions[r].height = 26
            continue
            
        # Nếu dòng r là dòng chi tiết thường:
        # Giãn dòng hợp lý dựa vào nội dung
        val_c = sheet.cell(row=r, column=3).value
        val_c_str = str(val_c) if val_c else ""
        
        if 'Nhật ký sáng' in val_c_str:
            sheet.row_dimensions[r].height = 110
        elif 'Nhật ký tối' in val_c_str:
            sheet.row_dimensions[r].height = 170
        else:
            sheet.row_dimensions[r].height = 22
            
        # Copy style và merge cột E đến I (nới thêm 4 ô ghi chú)
        copy_cell_style_to_notes(sheet, r)
        try:
            sheet.merge_cells(start_row=r, start_column=5, end_row=r, end_column=9)
        except Exception as e:
            # Đề phòng nếu ô đã bị merge
            print(f"Warning merge row {r} in {day}: {e}")

print("Successfully updated all routine sheets with notes column extension and row height adjustments!")
wb.save('Daily_Checklist_Timeline.xlsx')

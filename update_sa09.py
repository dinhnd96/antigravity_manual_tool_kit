import openpyxl

file_path = 'Feature_02_SA_Tham_So_He_Thong/test case/SA09.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active

c_TC_ID = 1
c_BR_Ref = 2
c_URD_Ref = 3
c_Module = 4
c_Feature = 5
c_Title = 6
c_Type = 7
c_Category = 8
c_Priority = 9
c_Precondition = 10
c_Steps = 11
c_Expected = 12
c_Trace_ID = 13
c_Note = 14

seen_tcs = set()
rows_to_delete = []

# Phân tách duplicate
for r in range(2, ws.max_row + 1):
    tc_id = ws.cell(r, c_TC_ID).value
    if tc_id in seen_tcs and tc_id is not None:
        rows_to_delete.append(r)
    elif tc_id is not None:
        seen_tcs.add(tc_id)

for r in reversed(rows_to_delete):
    ws.delete_rows(r)

# Phân tách update chi tiết
for row in range(2, ws.max_row + 1):
    tc_id = ws.cell(row, c_TC_ID).value
    if not tc_id:
        continue
        
    expected_val = str(ws.cell(row, c_Expected).value or '')
    if '(i) Logic:' in expected_val:
        expected_val = expected_val.replace('(i) Logic:', '(i) Nghiệp vụ/Logic:')
    if '(iv) File/Email:' in expected_val:
        expected_val = expected_val.replace('(iv) File/Email:', '(iv) Output:')
    ws.cell(row, c_Expected).value = expected_val

    precond_val = str(ws.cell(row, c_Precondition).value or '')

    if tc_id == 'SA09-UI-001':
        ws.cell(row, c_TC_ID).value = 'SA09-BR-HAP-004'
        steps_val = str(ws.cell(row, c_Steps).value or '')
        if '3.' not in steps_val:
            ws.cell(row, c_Steps).value = steps_val.strip() + "\n3. Mở lại lưới danh sách, xác nhận không có bản ghi vừa nhập."

    elif tc_id == 'SA09-BR-HAP-003':
        ws.cell(row, c_Category).value = 'Smoke'
        ws.cell(row, c_Type).value = 'Integration'
        if '1. Đăng nhập Maker.' not in precond_val: # Sửa precond cũ để nối Maker lên trên cùng
            if '1. ' in precond_val:
                precond_val = precond_val.replace('1. ', '2. ').replace('2. ', '3. ')
            ws.cell(row, c_Precondition).value = "1. Đăng nhập Maker.\n" + precond_val
            
    elif tc_id in ['SA09-BR-HAP-002', 'SA09-BR-NEG-003', 'SA09-UI-06-SEARCH-001', 'SA09-UI-04-VIEW-001', 'SA09-UI-07-EXCEL-001']:
        if not precond_val.startswith('1. Đăng nhập Maker') and not 'Maker' in precond_val:
            new_precond = "1. Đăng nhập Maker.\n2. " + precond_val.replace('1. ', '').replace('2. ', '\n3. ')
            ws.cell(row, c_Precondition).value = new_precond
            
    if tc_id == 'SA09-UI-03-EDIT-001':
        steps_val = str(ws.cell(row, c_Steps).value or '')
        if '1b.' not in steps_val:
            ws.cell(row, c_Steps).value = steps_val.replace('\n2.', '\n1b. Kiểm tra form load đúng thông tin hiện tại của bản ghi.\n2.')
            
    if tc_id == 'SA09-BR-HAP-002':
        ws.cell(row, c_Feature).value = 'Thêm mới - Cây SPDV'

# Bổ sung các TC bị GAP
new_tcs = [
    ['SA09-UI-01-ADD-001', 'UI-FUNC.01', 'I.1.1.1', 'SA.09', 'Thêm mới', 'Kiểm tra giao diện form Thêm mới Nhóm code phí', 'UI', 'Regression', 'P3', '1. Đăng nhập Maker.\n2. Đang ở màn hình danh sách Quản lý nhóm code phí.', '1. Nhấn nút Thêm mới.\n2. Kiểm tra giao diện form.', '(i) Nghiệp vụ/Logic: Bật màn hình theo điều kiện phân quyền.\n(ii) UI: Form hiển thị, đầy đủ các field: Mã nhóm, Tên nhóm, Loại tính phí, Mức độ ưu tiên... dạng Edit.\n(iii) Trạng thái/Audit: Không thay đổi.\n(iv) Output: Không sinh message.', 'UI-ADD-001', 'Gap UI Added'],
    ['SA09-UI-05-DELETE-002', 'UI-FUNC.05', 'I.1.1.1', 'SA.09', 'Xóa', 'Kiểm tra chặn xóa bản ghi khi Nhóm code phí đang có trạng thái Đã duyệt', 'Negative', 'Regression', 'P2', '1. Đăng nhập Maker.\n2. Hệ thống có bản ghi Nhóm code phí Đã duyệt.', '1. Nhấn icon Xóa tại dòng dữ liệu.\n2. Confirm xóa.', '(i) Nghiệp vụ/Logic: Chặn do trạng thái không phải Nháp/Chờ duyệt.\n(ii) UI: Báo popup cảnh báo "Không thể xóa bản ghi do đang ở trạng thái Hoạt động/Đã duyệt".\n(iii) Trạng thái/Audit: Không tạo thay đổi.\n(iv) Output: Không sinh message.', 'UI-DEL-NEG', 'Gap Negative Added']
]

for new_tc in new_tcs:
    # pad out empty fields
    while len(new_tc) < ws.max_column:
        new_tc.append('')
    ws.append(new_tc)

wb.save(file_path)
print("Updated successfully")

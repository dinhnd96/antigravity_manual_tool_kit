import openpyxl

file_path = 'Feature_02_SA_Tham_So_He_Thong/test case/SA07_Quan_Ly_Nhom_Khach_Hang_Updated.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active

# Header mapping
headers = {}
for i, cell in enumerate(ws[1]):
    headers[cell.value] = i

for row in range(2, ws.max_row + 1):
    tc_id = ws.cell(row, headers['TC_ID'] + 1).value
    
    if not tc_id:
        continue
        
    precond = str(ws.cell(row, headers['Precondition'] + 1).value or '')
    
    # M-05 to M-10 Fix Precondition format
    if not precond.startswith('1.') and precond.strip() and precond != 'None':
        ws.cell(row, headers['Precondition'] + 1).value = "1. Đăng nhập Maker.\n2. " + precond
    elif precond == 'Maker.' or precond == 'Maker, màn hình Thêm mới.' or precond == 'Màn hình Thêm mới.':
        ws.cell(row, headers['Precondition'] + 1).value = "1. Đăng nhập Maker.\n2. Ở màn hình Thêm mới."
    
    # C-01, C-03
    if tc_id == 'TC_SA07_012':
        ws.cell(row, headers['Trace_ID'] + 1).value = 'BR02-CLOSE-BTN'
    # C-02, C-04
    elif tc_id == 'TC_SA07_013':
        ws.cell(row, headers['Trace_ID'] + 1).value = 'BR03-UNASSIGNED-FREE'
    # H-01
    elif tc_id == 'TC_SA07_010':
        ws.cell(row, headers['Type'] + 1).value = 'Integration'
        ws.cell(row, headers['Precondition'] + 1).value = "1. Đăng nhập Maker.\n2. Hệ thống chưa có bản ghi Nhóm KH 'NEW_GROUP'."
        ws.cell(row, headers['Steps'] + 1).value = "1. Tạo Nhóm KH 'NEW_GROUP'.\n2. Gán NEW_GROUP vào Code phí đang Hoạt động.\n3. Thử đổi 'NEW_GROUP' thành Không hoạt động (Verify chặn).\n4. Quay lại sửa Code phí thành Hủy.\n5. Đổi trạng thái 'NEW_GROUP' sang Không hoạt động (Verify thành công)."
        ws.cell(row, headers['Expected'] + 1).value = "(i) Nghiệp vụ/Logic: Ràng buộc cascade liên bảng DB thành công.\n(ii) UI: View map chuẩn từng Step, luồng end-to-end hợp lệ.\n(iii) Trạng thái/Audit: CSDL thay đổi chuỗi trạng thái liên hoàn.\n(iv) Output: E2E check passed."
    # H-02
    elif tc_id == 'TC_SA07_003':
        ws.cell(row, headers['Note'] + 1).value = "Assumption - Cần BA xác nhận quy tắc xử lý dấu phẩy với operator ="
    # H-03
    elif tc_id == 'TC_SA07_004':
        ws.cell(row, headers['Note'] + 1).value = "Assumption - Cần BA xác nhận rule TRIM"
    # M-01
    elif tc_id == 'TC_SA07_002':
        ws.cell(row, headers['Expected'] + 1).value = "(i) Nghiệp vụ/Logic: Parsing List Option Data.\n(ii) UI: Field nhận đa giá trị cách nhau dấu phẩy, form không báo lỗi format.\n(iii) Trạng thái/Audit: Data view only.\n(iv) Output: Không xuất file."
    # M-02
    elif tc_id == 'TC_SA07_001':
        ws.cell(row, headers['Title'] + 1).value = "Kiểm tra danh sách 'Operator' hiển thị [=, IN] và 'Trạng thái' hiển thị [Hoạt động, Không hoạt động]"
        ws.cell(row, headers['Steps'] + 1).value = "1. Tại form Thêm mới, mở dropdown 'Operator'.\n2. Mở dropdown 'Trạng thái'."
        ws.cell(row, headers['Note'] + 1).value = "Đã gộp kiểm tra 2 dropdown BR_01"
    # M-03
    elif tc_id == 'TC_SA07_008':
        ws.cell(row, headers['Steps'] + 1).value = "1. Nhấn nút 'Thêm mới' tại màn danh sách.\n2. Quan sát layer popup vừa mở."
        ws.cell(row, headers['Expected'] + 1).value = "(i) Nghiệp vụ/Logic: Route sang trang mới.\n(ii) UI: Bật Layer popup Thêm Mới với các controls rỗng sẵn sàng nhập, có nút lưu/đóng đầy đủ.\n(iii) Trạng thái/Audit: Không.\n(iv) Output: Load view."
    # M-04
    elif tc_id == 'TC_SA07_009':
        ws.cell(row, headers['Steps'] + 1).value = "1. Nhấn icon 'Xem' dòng dữ liệu bất kỳ.\n2. Kiểm tra trạng thái các input."
        ws.cell(row, headers['Expected'] + 1).value = "(i) Nghiệp vụ/Logic: Lấy DTO show.\n(ii) UI: Popup bật. Tất cả các field (Mã nhóm, Tên nhóm, Operator, Trạng thái) đều ở chế độ Read-only vô hiệu hóa, không có nút Xác nhận.\n(iii) Trạng thái/Audit: DB Readonly.\n(iv) Output: none."

# Merged 011 into 001, so we delete 011
rows_to_delete = []
for row in range(2, ws.max_row + 1):
    tc_id = ws.cell(row, headers['TC_ID'] + 1).value
    if tc_id == 'TC_SA07_011':
        rows_to_delete.append(row)

for row in reversed(rows_to_delete):
    ws.delete_rows(row)

# H-04 Add GAP UI-FUNC.02
new_row = [
    'TC_SA07_014', 'UI-FUNC.02', 'I.1.1.1', 'SA.07', 'Quản lý nhóm khách hàng',
    'Kiểm tra màn hình Sửa (Edit) load đầy đủ dữ liệu cũ',
    'Happy', 'Regression', 'P2',
    '1. Đăng nhập Maker.\n2. Tồn tại bản ghi Nhóm KH hợp lệ.',
    '1. Nhấn icon Sửa trên màn hình danh sách.\n2. Kiểm tra giao diện form.',
    '(i) Nghiệp vụ/Logic: Fetch API chi tiết data.\n(ii) UI: Form load chuẩn xác data cũ của bản ghi vào các control tương ứng, các trường chỉnh sửa được enabled.\n(iii) Trạng thái/Audit: Không.\n(iv) Output: Không.',
    'UI-EDIT-FORM', 'Bổ sung Gap'
]
# padding
while len(new_row) < len(headers):
    new_row.append('')

ws.append(new_row)

wb.save(file_path)
print("Updated Test Cases in " + file_path)

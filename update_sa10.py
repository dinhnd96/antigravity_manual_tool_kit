import openpyxl

file_path = 'Feature_02_SA_Tham_So_He_Thong/test case/SA10_Quan_Ly_Job_Phi_Final_Standard.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['TestCases'] if 'TestCases' in wb.sheetnames else wb.active

c_TC_ID = 1
c_BR_Ref = 2
c_URD_Ref = 3
c_Module = 4
c_Feature = 5
c_Title = 6
c_Type = 7
c_Category = 8
c_Priority = 9
c_Precondition = 10
c_Steps = 11
c_Expected = 12
c_Trace_ID = 13
c_Note = 14

for row in range(2, ws.max_row + 1):
    tc_id = ws.cell(row, c_TC_ID).value
    if not tc_id:
        continue
        
    expected_val = str(ws.cell(row, c_Expected).value or '')
    if '(i) Logic:' in expected_val:
        expected_val = expected_val.replace('(i) Logic:', '(i) Nghiệp vụ/Logic:')
    if '(iii) Trạng thái:' in expected_val:
        expected_val = expected_val.replace('(iii) Trạng thái:', '(iii) Trạng thái/Audit:')
    ws.cell(row, c_Expected).value = expected_val

    precond_val = str(ws.cell(row, c_Precondition).value or '')
    if precond_val and 'Đăng nhập Maker' not in precond_val and 'Maker' not in precond_val:
        if '1. ' in precond_val:
            precond_val = precond_val.replace('1. ', '2. ').replace('2. ', '3. ')
            ws.cell(row, c_Precondition).value = "1. Đăng nhập Maker.\n" + precond_val
        else:
            ws.cell(row, c_Precondition).value = "1. Đăng nhập Maker.\n2. " + precond_val

    if tc_id == 'SA10-UI-04-VIEW-001':
        ws.cell(row, c_TC_ID).value = 'SA10-UI-02-VIEW-001'
        ws.cell(row, c_BR_Ref).value = 'UI-FUNC.02'
        if ws.cell(row, c_Type).value == 'UI':
            ws.cell(row, c_Type).value = 'Happy'

    elif tc_id == 'SA10-UI-FUNC-001':
        steps_val = str(ws.cell(row, c_Steps).value or '')
        if '3.' not in steps_val:
            ws.cell(row, c_Steps).value = steps_val.strip() + '\n3. Nhấn nút [Ngưng đình chỉ].\n4. Chờ lịch chạy tiếp theo, xác nhận Job tự chạy thành công.'
        if ws.cell(row, c_Type).value == 'UI':
            ws.cell(row, c_Type).value = 'Happy'

    elif tc_id == 'SA10-BR02-UI-001':
        ws.cell(row, c_Type).value = 'Negative'

    elif tc_id == 'SA10-UI-FILTER-001':
        if ws.cell(row, c_Type).value == 'UI':
            ws.cell(row, c_Type).value = 'Happy'

    elif tc_id == 'SA10-UI-EXPORT-001':
        steps_val = str(ws.cell(row, c_Steps).value or '')
        if 'Kiêm tra' in steps_val:
            ws.cell(row, c_Steps).value = steps_val.replace('Kiêm tra', 'Kiểm tra')

new_tcs = [
    ['SA10-UI-01-ADD-001', 'UI-FUNC.01', 'I.1.1.1', 'SA.10', 'Thêm mới', 'Kiểm tra giao diện form Thêm mới Job phí định kỳ', 'UI', 'Regression', 'P3', '1. Đăng nhập Maker.\n2. Đang ở màn hình danh sách Job.', '1. Nhấn nút Thêm mới.\n2. Kiểm tra giao diện popup/form hiển thị.', '(i) Nghiệp vụ/Logic: Không lưu vào DB.\n(ii) UI: Form hiển thị với đầy đủ các field: Mã Job, Thứ tự, Nhóm code phí (dropdown), Tần suất... Các field rỗng hợp lệ.\n(iii) Trạng thái/Audit: Không.\n(iv) Output: Không.', 'UI-ADD-001', 'Gap UI Added'],
    ['SA10-UI-03-EDIT-001', 'UI-FUNC.03', 'I.1.1.1', 'SA.10', 'Sửa', 'Kiểm tra chức năng giao diện tải dữ liệu lên form Sửa Job', 'Happy', 'Regression', 'P2', '1. Đăng nhập Maker.\n2. Có sẵn bản ghi Job đang tồn tại.', '1. Tại lưới danh sách, nhấn icon Sửa.\n2. Kiểm tra dữ liệu được bind lên form.', '(i) Nghiệp vụ/Logic: Load data từ DB theo ID.\n(ii) UI: Form Sửa bật lên, các ô input được điền sẵn đúng giá trị của Job.\n(iii) Trạng thái/Audit: Không.\n(iv) Output: Không.', 'UI-EDIT-001', 'Gap UI Added'],
    ['SA10-BR05-HAP-001', 'BR_05', 'I.1.1.1', 'SA.10', 'Tận thu', 'Kiểm tra thêm mới thành công khi chọn Tận thu và điền đủ thông tin nguồn thu', 'Happy', 'Smoke', 'P1', '1. Đăng nhập Maker.\n2. Màn hình Thêm mới.', '1. Nhập các thông tin bắt buộc chung.\n2. Tích chọn [Có Tận thu].\n3. Chọn [Thu vào số dư tối thiểu] hoặc nhập [Thu từ TK của KH].\n4. Nhấn Xác nhận.', '(i) Nghiệp vụ/Logic: Hệ thống ghi nhận job lưu thành công cấu hình Tận thu.\n(ii) UI: Toast "Thêm mới job thành công".\n(iii) Trạng thái/Audit: Sinh bản ghi Job mới lưu đầy đủ dữ liệu nguồn tận thu, Audit log có Maker.\n(iv) Output: Không.', 'BR05-TAN-THU-HAP', 'Happy Case Added']
]

for new_tc in new_tcs:
    while len(new_tc) < ws.max_column:
        new_tc.append('')
    ws.append(new_tc)

wb.save(file_path)
print("Updated successfully")

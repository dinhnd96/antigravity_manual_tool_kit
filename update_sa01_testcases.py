import openpyxl

file_path = '/Users/mac/antigravity-testing-kit/Feature_02_SA_Tham_So_He_Thong/test case/SA01_Đăng nhập.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.worksheets[0]

headers = [cell.value for cell in ws[1]]
col_map = {name: idx for idx, name in enumerate(headers)}

# Detect expected column name (could be 'Expected' or 'Expected Result')
expected_col = 'Expected' if 'Expected' in col_map else 'Expected Result'
note_col = 'Note' if 'Note' in col_map else None

print("Headers:", headers)
print("Expected col:", expected_col)
print("Note col:", note_col)

updates = {
    # Finding 1: SA01-NEG-002 - thiếu layer (iii)(iv)
    'SA01-NEG-002': {
        expected_col: (
            '(i) Nghiệp vụ/Logic: Hệ thống chặn truy cập, không tạo phiên làm việc.\n'
            '(ii) UI: Hiển thị thông báo "Tên đăng nhập hoặc mật khẩu không chính xác".\n'
            '(iii) Trạng thái/Audit: Không tạo Session. Không ghi Audit log đăng nhập thành công.\n'
            '(iv) Output: Không có.'
        )
    },
    # Finding 2: SA01-NEG-003 - thiếu layer (iii)(iv)
    'SA01-NEG-003': {
        expected_col: (
            '(i) Nghiệp vụ/Logic: Hệ thống xác thực thất bại, không cấp quyền truy cập.\n'
            '(ii) UI: Hiển thị thông báo lỗi "Tên đăng nhập hoặc mật khẩu không chính xác".\n'
            '(iii) Trạng thái/Audit: Không tạo Session. Ghi nhận attempt login sai (dùng cho cơ chế lockout).\n'
            '(iv) Output: Không có.'
        )
    },
    # Finding 3: SA01-NEG-004 - thiếu layer (i)(iii)(iv)
    'SA01-NEG-004': {
        expected_col: (
            '(i) Nghiệp vụ/Logic: Hệ thống chặn submit form do thiếu thông tin bắt buộc.\n'
            '(ii) UI: Báo lỗi Validation đỏ tại trường bị bỏ trống: "Vui lòng nhập <Tên trường>".\n'
            '(iii) Trạng thái/Audit: Không gọi API xác thực. Không sinh Session hay log đăng nhập.\n'
            '(iv) Output: Không có.'
        ),
        'Steps': (
            '1. Để trống Tên đăng nhập, nhập Mật khẩu hợp lệ. Nhấn "Đăng nhập".\n'
            '2. Nhập Tên đăng nhập hợp lệ, để trống Mật khẩu. Nhấn "Đăng nhập".\n'
            '3. Để trống cả Tên đăng nhập lẫn Mật khẩu. Nhấn "Đăng nhập".'
        )
    },
    # Finding 4: SA01-NEG-006 - thiếu layer (iii)(iv)
    'SA01-NEG-006': {
        expected_col: (
            '(i) Nghiệp vụ/Logic: Chặn truy cập phiên thứ hai hoặc tự động ngắt phiên cũ (tùy thiết kế).\n'
            '(ii) UI: Thông báo trên Browser B: Tài khoản đang được sử dụng ở thiết bị khác. '
            'Trên Browser A (nếu bị kick): thông báo Phiên đăng nhập đã hết hạn, vui lòng đăng nhập lại.\n'
            '(iii) Trạng thái/Audit: Phiên cũ của Browser A bị terminate/invalidate. Log ghi nhận sự kiện đăng nhập trùng lặp.\n'
            '(iv) Output: Không có.'
        ),
        note_col: 'Assumption: Cần BA xác nhận hành vi: Ngắt phiên cũ (Browser A bị kick) hay Chặn phiên mới (Browser B không vào được).'
    },
    # Finding 5: SA01-HAPPY-007 - thiếu layer + mockup mismatch
    'SA01-HAPPY-007': {
        expected_col: (
            '(i) Nghiệp vụ/Logic: Hệ thống lưu thông tin đăng nhập vào Cookie/LocalStorage để tự động điền ở lần sau.\n'
            '(ii) UI: Sau khi mở lại trang, Tên đăng nhập được tự động điền vào ô (hoặc tự động đăng nhập tùy config).\n'
            '(iii) Trạng thái/Audit: Cookie "Remember Me" được ghi với TTL tương ứng.\n'
            '(iv) Output: Không có.'
        ),
        note_col: 'Assumption: Mockup màn đăng nhập hiện tại KHÔNG hiển thị Checkbox "Ghi nhớ". Cần BA/Designer xác nhận có thêm element này không trước khi Execute TC.'
    },
    # Finding 6: SA01-UI-008 - thiếu layer (i)(iii)(iv)
    'SA01-UI-008': {
        expected_col: (
            '(i) Nghiệp vụ/Logic: Không có API nào được gọi. Đây là behavior thuần UI client-side.\n'
            '(ii) UI: Mật khẩu hiển thị dạng text khi click icon Mắt (bật), dạng dấu chấm/sao khi click lại (tắt). Icon thay đổi trạng thái theo.\n'
            '(iii) Trạng thái/Audit: Không thay đổi. Không sinh log.\n'
            '(iv) Output: Không có.'
        )
    },
    # Finding 7: SA01-HAPPY-009 - thiếu layer (iii)(iv)
    'SA01-HAPPY-009': {
        expected_col: (
            '(i) Nghiệp vụ/Logic: Hệ thống nhận Token hợp lệ từ EntraID và thiết lập phiên làm việc.\n'
            '(ii) UI: Chuyển hướng về trang chủ ProfiX với quyền/tài khoản tương ứng. Menu hiển thị đúng theo nhóm quyền được gán.\n'
            '(iii) Trạng thái/Audit: Token EntraID được lưu vào Session/Cookie. Log ghi nhận đăng nhập thành công qua EntraID.\n'
            '(iv) Output: Không có.'
        )
    },
    # Finding 8: SA01-SECURITY-012 - Anti-pattern "30 phút"
    'SA01-SECURITY-012': {
        expected_col: (
            '(i) Nghiệp vụ/Logic: Hệ thống ghi nhận số lần đăng nhập sai. Sau lần thứ 5, tài khoản bị chuyển sang trạng thái Inactive.\n'
            '(ii) UI: Từ lần thứ 5 trở đi hiển thị thông báo "Tài khoản đã bị khóa, vui lòng liên hệ quản trị viên để được hỗ trợ".\n'
            '(iii) Trạng thái/Audit: Tài khoản User_A chuyển sang trạng thái Inactive trong DB. Log Security ghi nhận sự kiện Lockout/Brute-force Detected.\n'
            '(iv) Output: Không có.'
        ),
        note_col: 'Assumption: URD không quy định thời gian khóa cụ thể (30 phút). Hành vi tài khoản bị Inactive phải được Admin vào SA.03 mở lại (theo quy trình).'
    },
    # Finding 9: SA01-NEG-014 - Anti-pattern "50 ký tự"
    'SA01-NEG-014': {
        expected_col: (
            '(i) Nghiệp vụ/Logic: Hệ thống chặn input không an toàn hoặc vượt quá giới hạn chiều dài trước khi gửi lên Server.\n'
            '(ii) UI: Báo lỗi "Tên đăng nhập không hợp lệ" hoặc "Vượt quá độ dài cho phép [cần BA xác nhận giới hạn ]".\n'
            '(iii) Trạng thái/Audit: Không gọi API xác thực. Không sinh log.\n'
            '(iv) Output: Không có.'
        ),
        note_col: 'Assumption: URD không quy định giới hạn độ dài cụ thể của trường Tên đăng nhập. Cần BA xác nhận max-length rule trước khi chốt con số.'
    },
    # Finding 10: SA01-FLOW-011 - TC quá mơ hồ
    'SA01-FLOW-011': {
        'Steps': (
            '1. Truy cập link hệ thống ProfiX.\n'
            '2. Nhập Tên đăng nhập và Mật khẩu hợp lệ. Nhấn "Đăng nhập".\n'
            '3. Kiểm tra màn hình Home Dashboard: Widget "Phí định kỳ 2025", "Tổng phí tháng", "Doanh thu từ chương trình ưu đãi" hiển thị đúng.\n'
            '4. Kiểm tra Menu trái hiển thị đủ nhóm: Tham số, Tra cứu, Các chức năng khác.\n'
            '5. Giữ nguyên trạng thái trong ít nhất 5 phút, không thực hiện thao tác.'
        ),
        expected_col: (
            '(i) Nghiệp vụ/Logic: Phiên làm việc được khởi tạo thành công và giữ ổn định cho đến khi Logout hoặc Timeout.\n'
            '(ii) UI: Dashboard hiển thị đúng các widget theo thiết kế (image2). Menu hiển thị đầy đủ nhóm chức năng theo quyền của tài khoản.\n'
            '(iii) Trạng thái/Audit: Session hợp lệ, không bị expire trong thời gian thao tác.\n'
            '(iv) Output: Không có.'
        )
    },
}

# Apply all updates
for row in ws.iter_rows(min_row=2):
    tc_id_val = row[col_map['TC_ID']].value
    if tc_id_val and tc_id_val in updates:
        for col_name, new_val in updates[tc_id_val].items():
            if col_name in col_map:
                row[col_map[col_name]].value = new_val
                print(f"  Updated [{tc_id_val}] column [{col_name}]")
            else:
                print(f"  WARNING: Column '{col_name}' not found for TC {tc_id_val}")

wb.save(file_path)
print("\nSA01 update completed successfully.")

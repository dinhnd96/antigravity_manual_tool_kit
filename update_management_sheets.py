"""
Script cập nhật công thức và định dạng cho 3 Sheet: Dashboard, Daily Tracking, Bug Tracker
Theo chuẩn test_case_management_sync prompt - Tối ưu hóa cho dự án ProfiX
"""
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

# Đường dẫn file Master
FILE_PATH = '/Users/mac/antigravity-testing-kit/Test case/Test case_Management_ProfiX.xlsx'

SHEET_DASHBOARD   = '📊 Dashboard'
SHEET_TRACKING    = '📅 Daily Tracking'
SHEET_BUG_TRACKER = '🐞 Bug Tracker'
MGMT_SHEETS       = {SHEET_DASHBOARD, SHEET_TRACKING, SHEET_BUG_TRACKER, 'Coverage', 'Dedup_Log'}
MEMBERS           = ['Định', 'Vân', 'Vân Anh', 'Thanh', 'Hiền', 'Thủy']
DATA_ROWS         = range(3, 62)   # hàng 3 → 61 (~2 tháng tracking)

# ── Style helpers ────────────────────────────────────────────────────────────
thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def h1_fill(): return PatternFill('solid', fgColor='1F4E79')      # Xanh Navy Đậm
def h2_fill(): return PatternFill('solid', fgColor='D9E1F2')      # Xanh Pastel Nhạt (Header)
def total_fill(): return PatternFill('solid', fgColor='FFF2CC')    # Vàng Nhạt (Total)
def error_fill(): return PatternFill('solid', fgColor='FFC7CE')    # Đỏ Nhạt (Fail/Bug)

def h1_font(size=14): return Font(bold=True, color='FFFFFF', size=size)
def h2_font(): return Font(bold=True, size=11)
def body_font(): return Font(size=10)
def bold_font(): return Font(bold=True, size=10)

def pct_fmt(ws, row, col):
    ws.cell(row=row, column=col).number_format = '0.00%'

def style_header1(ws, merge_range, value):
    """Hàng 1 – Tiêu đề lớn, chuyên nghiệp"""
    try:
        ws.merge_cells(merge_range)
    except Exception:
        pass
    cell = ws.cell(row=1, column=1)
    cell.value = value
    cell.font = h1_font()
    cell.fill = h1_fill()
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

def style_cell(cell, fill=None, font=None, center=False, wrap=True, border=BORDER):
    if border: cell.border = border
    if fill:   cell.fill   = fill
    if font:   cell.font   = font
    cell.alignment = Alignment(
        horizontal='center' if center else 'left',
        vertical='top',
        wrap_text=wrap
    )

# ── 1. DASHBOARD ────────────────────────────────────────────────────────────
def update_dashboard(wb):
    if SHEET_DASHBOARD not in wb.sheetnames:
        wb.create_sheet(SHEET_DASHBOARD)
    ws = wb[SHEET_DASHBOARD]
    
    # Lấy danh sách Sheet chứa Test Case (trừ các sheet quản lý)
    tc_sheets = [s for s in wb.sheetnames if s not in MGMT_SHEETS]

    # Unmerge all existing merged cells
    merged_cells_ranges = list(ws.merged_cells.ranges)
    for m_range in merged_cells_ranges:
        ws.unmerge_cells(str(m_range))

    # Reset nội dung
    for row in ws.iter_rows(min_row=1, max_row=max(ws.max_row, 100)):
        for cell in row:
            cell.value = None
            cell.style = 'Normal'

    # Hàng 1: Title
    style_header1(ws, 'A1:I1', 'PROFIX — QUALITY DASHBOARD SUMMARY')

    # Hàng 2: Header
    headers = ['STT', 'Module (Feature)', 'Total TC', 'Passed', 'Failed', 'Blocked', 'N/A', 'Execution %', 'Pass Rate %']
    widths = [6, 40, 12, 10, 10, 10, 8, 15, 15]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=c, value=h)
        style_cell(cell, fill=h2_fill(), font=h2_font(), center=True)
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[2].height = 22

    # Dữ liệu từng Module
    for idx, s_name in enumerate(tc_sheets, 1):
        r = idx + 2
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=s_name)
        ws.cell(row=r, column=3, value=f"=COUNTA('{s_name}'!A:A)-2")
        ws.cell(row=r, column=4, value=f"=COUNTIF('{s_name}'!U:U,\"Pass\")")
        ws.cell(row=r, column=5, value=f"=COUNTIF('{s_name}'!U:U,\"Fail\")")
        ws.cell(row=r, column=6, value=f"=COUNTIF('{s_name}'!U:U,\"Blocked\")")
        ws.cell(row=r, column=7, value=f"=COUNTIF('{s_name}'!U:U,\"N/A\")")
        
        # Execution % = (P + F + B) / Total
        ws.cell(row=r, column=8, value=f"=IF(C{r}>0,(D{r}+E{r}+F{r})/C{r},0)")
        pct_fmt(ws, r, 8)
        
        # Pass Rate % = P / Total
        ws.cell(row=r, column=9, value=f"=IF(C{r}>0,D{r}/C{r},0)")
        pct_fmt(ws, r, 9)

        for c in range(1, 10):
            style_cell(ws.cell(row=r, column=c), font=body_font(), center=(c in [1, 3, 4, 5, 6, 7, 8, 9]))

    # Hàng kết quả tổng (TOTAL)
    total_row = len(tc_sheets) + 3
    ws.cell(row=total_row, column=2, value="TOTAL SUMMARY")
    for c in range(3, 8):
        col_letter = get_column_letter(c)
        ws.cell(row=total_row, column=c, value=f"=SUM({col_letter}3:{col_letter}{total_row-1})")
    
    ws.cell(row=total_row, column=8, value=f"=IF(C{total_row}>0,(D{total_row}+E{total_row}+F{total_row})/C{total_row},0)")
    pct_fmt(ws, total_row, 8)
    ws.cell(row=total_row, column=9, value=f"=IF(C{total_row}>0,D{total_row}/C{total_row},0)")
    pct_fmt(ws, total_row, 9)

    for c in range(1, 10):
        style_cell(ws.cell(row=total_row, column=c), fill=total_fill(), font=bold_font(), center=(c != 2))
    
    print(f"  [v] Sheet {SHEET_DASHBOARD} updated.")

# ── 2. DAILY TRACKING ───────────────────────────────────────────────────────
def update_daily_tracking(wb):
    if SHEET_TRACKING not in wb.sheetnames:
        wb.create_sheet(SHEET_TRACKING)
    ws = wb[SHEET_TRACKING]
    tc_sheets = [s for s in wb.sheetnames if s not in MGMT_SHEETS]

    # Unmerge all existing merged cells
    merged_cells_ranges = list(ws.merged_cells.ranges)
    for m_range in merged_cells_ranges:
        ws.unmerge_cells(str(m_range))

    # Reset nội dung
    for row in ws.iter_rows(min_row=1, max_row=max(ws.max_row, 70)):
        for cell in row:
            cell.value = None
            cell.style = 'Normal'

    last_col_idx = 2 + len(MEMBERS) + 2
    style_header1(ws, f'A1:{get_column_letter(last_col_idx)}1', 'DAILY EXECUTION TRACKING — TEAM PROFIX')

    # Headers
    headers = ['Ngày (Date)'] + MEMBERS + ['Total/Day', 'Lũy kế']
    widths = [15] + [12]*len(MEMBERS) + [13, 14]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=c, value=h)
        style_cell(cell, fill=h2_fill(), font=h2_font(), center=True)
        ws.column_dimensions[get_column_letter(c)].width = w

    # Công thức hàng ngày
    for idx, r_idx in enumerate(DATA_ROWS, 1):
        r = r_idx
        # Cột A: Để user nhập ngày hoặc dùng công thức auto date nếu cần
        ws.cell(row=r, column=1).number_format = 'DD/MM/YYYY'
        style_cell(ws.cell(row=r, column=1), font=body_font(), center=True)

        # Công thức cho từng member
        for m_idx, member in enumerate(MEMBERS, 2):
            parts = [f"COUNTIFS('{s}'!R:R,\"{member}\",'{s}'!T:T,A{r})" for s in tc_sheets]
            formula = "=" + "+".join(parts) if parts else "=0"
            style_cell(ws.cell(row=r, column=m_idx), font=body_font(), center=True)
            ws.cell(row=r, column=m_idx).value = formula

        # Cột Total/Day
        sum_range = f"{get_column_letter(2)}{r}:{get_column_letter(1+len(MEMBERS))}{r}"
        ws.cell(row=r, column=len(MEMBERS)+2, value=f"=SUM({sum_range})")
        style_cell(ws.cell(row=r, column=len(MEMBERS)+2), font=bold_font(), center=True)

        # Cột Lũy kế
        total_day_col = get_column_letter(len(MEMBERS)+2)
        ws.cell(row=r, column=len(MEMBERS)+3, value=f"=SUM({total_day_col}$3:{total_day_col}{r})")
        style_cell(ws.cell(row=r, column=len(MEMBERS)+3), font=bold_font(), center=True, fill=total_fill())

    print(f"  [v] Sheet {SHEET_TRACKING} updated (Rows 3-61).")

# ── 3. BUG TRACKER ──────────────────────────────────────────────────────────
def update_bug_tracker(wb):
    if SHEET_BUG_TRACKER not in wb.sheetnames:
        wb.create_sheet(SHEET_BUG_TRACKER)
    ws = wb[SHEET_BUG_TRACKER]

    # KHÔNG xóa dữ liệu, chỉ cập nhật Header và Styling
    style_header1(ws, f'A1:L1', 'PROFIX — DEFECT TRACKING SYSTEM (BUG TRACKER)')

    headers = ['Bug ID', 'Related TC', 'Summary / Title', 'Module', 'Severity', 'Priority', 'Env', 'Steps to Reproduce', 'Expected', 'Actual', 'Assignee', 'Status']
    widths = [12, 15, 45, 12, 10, 10, 15, 45, 45, 45, 15, 12]
    
    # Update Header Row (Row 2)
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=c, value=h)
        style_cell(cell, fill=h2_fill(), font=h2_font(), center=True)
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[2].height = 25

    # Định dạng các dòng dữ liệu hiện có (nếu có)
    max_r = ws.max_row
    if max_r < 3: max_r = 10 # Default format cho 10 dòng đầu nếu trống

    for r in range(3, max_r + 1):
        for c in range(1, 13):
            cell = ws.cell(row=r, column=c)
            # Giữ nguyên value
            style_cell(cell, font=body_font(), wrap=True)
            # Special formatting cho Severity S1/S2 hoặc status
            val = str(cell.value)
            if val in ['S1', 'S2', 'Fatal', 'Critical', 'High']:
                cell.font = Font(bold=True, color='9C0006')
                cell.fill = error_fill()
            if val in ['Open', 'New', 'Reopen']:
                cell.font = Font(bold=True)

    print(f"  [v] Sheet {SHEET_BUG_TRACKER} formatted (Preserved data).")

# ── MAIN ────────────────────────────────────────────────────────────────────
def main():
    if not os.path.exists(FILE_PATH):
        print(f"❌ File not found: {FILE_PATH}")
        return

    print(f"🚀 Updating Master File: {os.path.basename(FILE_PATH)}")
    wb = openpyxl.load_workbook(FILE_PATH)

    update_dashboard(wb)
    update_daily_tracking(wb)
    update_bug_tracker(wb)

    wb.save(FILE_PATH)
    print(f"\n✅ SUCCESS! All management sheets updated and styled.")
    print(f"   - File: {FILE_PATH}")

if __name__ == '__main__':
    main()

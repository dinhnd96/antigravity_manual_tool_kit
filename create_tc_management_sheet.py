from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

OUTPUT_PATH = "/Users/mac/antigravity-testing-kit/TC_Management_ProfiX.xlsx"

wb = Workbook()

# ---- Color Palette ----
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "D6E4F7"
LIGHT_GREEN = "E2EFDA"
HEADER_GRAY = "D9D9D9"
WHITE       = "FFFFFF"
PASS_GREEN  = "70AD47"
FAIL_RED    = "FF0000"
BLOCK_ORG   = "ED7D31"
INPROG_YEL  = "FFD966"

# ---- Helper ----
def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header_row(ws, row_num, col_count, bg_color=MID_BLUE, font_color=WHITE):
    fill = PatternFill("solid", fgColor=bg_color)
    font = Font(bold=True, color=font_color, size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = thin_border()

# ================================================================
# SHEET 1: Test Plan Overview
# ================================================================
ws_plan = wb.active
ws_plan.title = "📋 Test Plan"

# Title
ws_plan.merge_cells("A1:H1")
title_cell = ws_plan["A1"]
title_cell.value = "TEST PLAN — HỆ THỐNG PROFIX PHASE 1 (SIT)"
title_cell.font = Font(bold=True, size=16, color=WHITE)
title_cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws_plan.row_dimensions[1].height = 35

# Info block
info = [
    ["Dự án:", "ProfiX – Hệ thống Quản lý Phí Tập Trung Phase 1", "", "Giai đoạn:", "System Integration Testing (SIT)"],
    ["QA Lead:", "[Tên của bạn]", "", "Ngày lập:", "19/03/2026"],
    ["URD Ref:", "PVCB_URD_ProfiX Phase 1_ver 0.7", "", "Deadline SIT:", "31/03/2026"],
]
for i, row in enumerate(info, start=2):
    ws_plan.cell(row=i, column=1, value=row[0]).font = Font(bold=True)
    ws_plan.cell(row=i, column=2, value=row[1])
    ws_plan.cell(row=i, column=4, value=row[3]).font = Font(bold=True)
    ws_plan.cell(row=i, column=5, value=row[4])

# Scope table
scope_header = ["STT", "Nhóm Chức Năng", "Các Module Chi Tiết", "PIC Test", "Trạng Thái"]
ws_plan.append([])
ws_plan.append(scope_header)
style_header_row(ws_plan, 7, len(scope_header))

scope_data = [
    [1, "Khai báo Tham số (PR)", "PR01: SPDV & Code phí KHCN/KHDN\nPR02: Công thức tính phí", "Định", ""],
    [2, "Tra cứu (SE)", "SE01→SE05: Xem cây, Lịch sử, CTƯĐ", "Vân Anh", ""],
    [3, "Tác vụ Phê duyệt (OT)", "OT01: Duyệt / Từ chối", "Thanh + Vân Anh", ""],
    [4, "Quản trị Hệ thống (SA)", "SA01-09: Đăng nhập, User, Phân quyền", "Hiệp + Thanh", ""],
    [5, "Báo cáo (RP)", "RP01→RP05: Sao kê, Nợ phí, Dashboard", "Vân Anh", ""],
]
for i, row in enumerate(scope_data, start=8):
    for j, val in enumerate(row, start=1):
        c = ws_plan.cell(row=i, column=j, value=val)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = thin_border()
        c.fill = PatternFill("solid", fgColor=LIGHT_BLUE if i % 2 == 0 else WHITE)
    ws_plan.row_dimensions[i].height = 45

ws_plan.column_dimensions["A"].width = 6
ws_plan.column_dimensions["B"].width = 25
ws_plan.column_dimensions["C"].width = 50

# ================================================================
# TC Sheet Construction
# ================================================================
TC_HEADERS = [
    "TC_ID", "Module", "Feature", "Title", "Type", "Category", "Priority", 
    "Precondition", "Steps", "Expected", "URD_Ref", "BR_Ref", "Trace_ID", "Note",
    "Status R1", "Tester R1", "Date R1", 
    "Status R2", "Tester R2", "Date R2", 
    "Final Status"
]
# Col mapping: 
# O=15(R1), P=16(T1), Q=17(D1)
# R=18(R2), S=19(T2), T=20(D2)
# U=21(Final)

STATUS_OPTIONS   = '"Pass,Fail,Blocked,In Progress,N/A"'
PRIORITY_OPTIONS = '"P1,P2,P3"'
TYPE_OPTIONS     = '"Happy,Negative,Boundary,Smoke,Regression"'
CATEGORY_OPTIONS = '"Functional,Security,UI/UX,Integration,Performance"'
TESTER_OPTIONS   = '"Định,Thanh,Vân Anh,Hiệp"'

def create_tc_sheet(wb, sheet_name, module_name, sample_rows=[], tab_color="2E75B6"):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = tab_color

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(TC_HEADERS))}1")
    ws["A1"].value = f"TEST CASES — {module_name}"
    ws["A1"].font = Font(bold=True, size=13, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Header row
    ws.append(TC_HEADERS)
    style_header_row(ws, 2, len(TC_HEADERS))
    ws.row_dimensions[2].height = 25

    # Sample rows
    for i, row_data in enumerate(sample_rows, start=3):
        # Truncate or pad row_data to match TC_HEADERS
        # Original sample had 14 cols, now we have 21
        clean_row = list(row_data) + [""] * (len(TC_HEADERS) - len(row_data))
        for j, val in enumerate(clean_row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.border = thin_border()
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.fill = PatternFill("solid", fgColor=LIGHT_BLUE if i % 2 == 0 else WHITE)
        ws.row_dimensions[i].height = 40

    # Auto-formula for Final Status (Column U)
    # Formula: =IF(R3<>"", R3, O3)
    max_row = 100
    for r in range(3, max_row + 1):
        ws.cell(row=r, column=21, value=f'=IF(R{r}<>"", R{r}, O{r})')
        # Style empty rows below samples
        if r > 3 + len(sample_rows):
            for col in range(1, len(TC_HEADERS) + 1):
                if col == 21: continue # skip since we put formula
                c = ws.cell(row=r, column=col)
                c.border = thin_border()
                c.fill = PatternFill("solid", fgColor=WHITE)

    # Data Validations
    dv_status   = DataValidation(type="list", formula1=STATUS_OPTIONS,   allow_blank=True)
    dv_tester   = DataValidation(type="list", formula1=TESTER_OPTIONS,   allow_blank=True)
    dv_type     = DataValidation(type="list", formula1=TYPE_OPTIONS,     allow_blank=True)
    dv_category = DataValidation(type="list", formula1=CATEGORY_OPTIONS, allow_blank=True)
    dv_prio     = DataValidation(type="list", formula1=PRIORITY_OPTIONS, allow_blank=True)

    dv_type.sqref     = f"E3:E{max_row}"
    dv_category.sqref = f"F3:F{max_row}"
    dv_prio.sqref     = f"G3:G{max_row}"
    dv_status.sqref   = f"O3:O{max_row} R3:R{max_row}" # Applied to R1 and R2
    dv_tester.sqref   = f"P3:P{max_row} S3:S{max_row}" # Applied to R1 and R2

    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_category)
    ws.add_data_validation(dv_prio)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_tester)

    # Date formatting for Q and T
    for r in range(3, max_row + 1):
        ws.cell(row=r, column=17).number_format = "DD/MM/YYYY" # Q
        ws.cell(row=r, column=20).number_format = "DD/MM/YYYY" # T

    # Conditional Formatting
    green = PatternFill("solid", fgColor=PASS_GREEN)
    red   = PatternFill("solid", fgColor=FAIL_RED)
    org   = PatternFill("solid", fgColor=BLOCK_ORG)
    yel   = PatternFill("solid", fgColor=INPROG_YEL)

    # Apply to R1, R2 and Final (O, R, U)
    for col_let in ["O", "R", "U"]:
        ws.conditional_formatting.add(f"{col_let}3:{col_let}{max_row}", CellIsRule(operator="equal", formula=['"Pass"'], fill=green))
        ws.conditional_formatting.add(f"{col_let}3:{col_let}{max_row}", CellIsRule(operator="equal", formula=['"Fail"'], fill=red))
        ws.conditional_formatting.add(f"{col_let}3:{col_let}{max_row}", CellIsRule(operator="equal", formula=['"Blocked"'], fill=org))
        ws.conditional_formatting.add(f"{col_let}3:{col_let}{max_row}", CellIsRule(operator="equal", formula=['"In Progress"'], fill=yel))

    # Type colors (E)
    ws.conditional_formatting.add(f"E3:E{max_row}", CellIsRule(operator="equal", formula=['"Happy"'], fill=green))
    ws.conditional_formatting.add(f"E3:E{max_row}", CellIsRule(operator="equal", formula=['"Negative"'], fill=red))

    # Column widths
    widths = [14, 18, 25, 38, 12, 18, 9, 30, 42, 40, 14, 14, 14, 25, 14, 14, 14, 14, 14, 14, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"
    return ws

# Create Sheets
SHEETS = ["'🧪 TC_PR01'", "'🧪 TC_PR02'", "'🧪 TC_OT01'", "'🧪 TC_SA'", "'🧪 TC_SE'", "'🧪 TC_RP'"]
# (Simplified sample data for brevity)
create_tc_sheet(wb, "🧪 TC_PR01", "PR01", sample_rows=[
    ["PR01-001", "PR01", "Feature", "Sample Happy Case", "Happy", "Functional", "P1", "Precon", "Steps", "Expected", "URD-01", "BR-01", "ID-01", "Note", "Pass", "Định", "18/03", "", "", "", ""],
    ["PR01-002", "PR01", "Feature", "Sample Re-test Case", "Negative", "Functional", "P1", "Precon", "Steps", "Expected", "URD-01", "BR-01", "ID-02", "Note", "Fail", "Định", "18/03", "Pass", "Thanh", "19/03", ""]
])
create_tc_sheet(wb, "🧪 TC_PR02", "PR02", sample_rows=[])
create_tc_sheet(wb, "🧪 TC_OT01", "OT01", sample_rows=[])
create_tc_sheet(wb, "🧪 TC_SA", "SA", sample_rows=[])
create_tc_sheet(wb, "🧪 TC_SE", "SE", sample_rows=[])
create_tc_sheet(wb, "🧪 TC_RP", "RP", sample_rows=[])

# ================================================================
# Bug Tracker
# ================================================================
ws_bug = wb.create_sheet("🐞 Bug Tracker")
ws_bug.append(["Bug ID", "TC ID", "Title", "Module", "Severity", "Priority", "Env", "Steps", "Expected", "Actual", "Owner", "Status"])
style_header_row(ws_bug, 1, 12, bg_color="C00000")

# ================================================================
# Dashboard
# ================================================================
ws_dash = wb.create_sheet("📊 Dashboard")
ws_dash.merge_cells("A1:F1")
ws_dash["A1"].value = "📊 QA DASHBOARD — PROFIX PHASE 1"
ws_dash["A1"].font = Font(bold=True, size=14, color=WHITE)
ws_dash["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
ws_dash["A1"].alignment = Alignment(horizontal="center")

# Summary on Column U (Final Status)
def countif_all(status):
    return "+".join([f'COUNTIF({s}!U:U,\"{status}\")' for s in SHEETS])

dash_labels = [
    ["", "Chỉ Số", "Giá Trị", "Diễn Giải"],
    ["", "Total Designed",  f"="+"+".join([f'COUNTA({s}!A3:A100)' for s in SHEETS]),  "Tổng TC"],
    ["", "✅ Passed",       f"={countif_all('Pass')}",        "Kết quả cuối"],
    ["", "❌ Failed",       f"={countif_all('Fail')}",        "Kết quả cuối"],
    ["", "🟠 Blocked",      f"={countif_all('Blocked')}",     "Kết quả cuối"],
    ["", "Pass Rate (%)",   f"=IFERROR(ROUND({countif_all('Pass')}/(" + "+".join([f'COUNTA({s}!A3:A100)' for s in SHEETS]) + ")*100,1),0)", "% thành công"],
]
for i, row in enumerate(dash_labels, start=3):
    for j, val in enumerate(row, start=1):
        c = ws_dash.cell(row=i, column=j, value=val)
        c.border = thin_border()
        if j > 1: c.alignment = Alignment(horizontal="center")

# ================================================================
# Daily Tracking
# ================================================================
ws_daily = wb.create_sheet("📅 Daily Tracking")
TEAM = ["Định", "Thanh", "Vân Anh", "Hiệp"]
DATES = ["18/03", "19/03", "20/03", "21/03", "24/03"]

ws_daily.merge_cells(f"A1:{get_column_letter(2+len(TEAM))}1")
ws_daily["A1"].value = "📅 DAILY TEST PROGRESS (R1 + R2 Combined)"
ws_daily["A1"].font = Font(bold=True, size=14, color=WHITE)
ws_daily["A1"].fill = PatternFill("solid", fgColor="B8860B")

daily_headers = ["Ngày", "Tổng"] + TEAM
for j, h in enumerate(daily_headers, start=1):
    c = ws_daily.cell(row=2, column=j, value=h)
    c.fill = PatternFill("solid", fgColor="1F3864")
    c.font = Font(bold=True, color=WHITE)
    c.border = thin_border()

# Row per date
for i, d in enumerate(DATES, start=3):
    ws_daily.cell(row=i, column=1, value=d).border = thin_border()
    
    # Total combined executions (R1 date + R2 date)
    total_parts = []
    for s in SHEETS:
        total_parts.append(f'COUNTIF({s}!Q:Q,"{d}") + COUNTIF({s}!T:T,"{d}")')
    ws_daily.cell(row=i, column=2, value="=" + "+".join(total_parts)).border = thin_border()

    # Per Tester (R1 + R2)
    for t_idx, tester in enumerate(TEAM):
        parts = []
        for s in SHEETS:
            # Tester R1 @ Date R1 OR Tester R2 @ Date R2
            parts.append(f'COUNTIFS({s}!P:P,"{tester}",{s}!Q:Q,"{d}") + COUNTIFS({s}!S:S,"{tester}",{s}!T:T,"{d}")')
        ws_daily.cell(row=i, column=3 + t_idx, value="=" + "+".join(parts)).border = thin_border()

wb.save(OUTPUT_PATH)
print(f"✅ Enhanced file created: {OUTPUT_PATH}")

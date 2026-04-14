import openpyxl

file_path = '/Users/mac/antigravity-testing-kit/Feature_02_SA_Tham_So_He_Thong/test case/ProfiX_Master_Test_Cases_Final.xlsx'
wb = openpyxl.load_workbook(file_path)

ws = wb['📅 Daily Tracking']

tc_sheets = [s for s in wb.sheetnames if s.startswith('SA')]

headers = ['Date', 'Định', 'Vân', 'Vân Anh', 'Thanh', 'Hiền', 'Thủy', 'Thương', 'Total/Day', 'Tổng lũy kế']
members = headers[1:8]

for row in range(3, 61):
    date_cell = f'A{row}'
    
    for col_idx, member in enumerate(members, 2):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        
        # Real columns: Q = Tester R1, R = Date R1
        countifs_parts = []
        for sheet in tc_sheets:
            sheet_name_escaped = sheet.replace("'", "''")
            countifs_parts.append(f"COUNTIFS('{sheet_name_escaped}'!Q:Q, \"{member}\", '{sheet_name_escaped}'!R:R, {date_cell})")
            
        formula = "=" + " + ".join(countifs_parts)
        ws[f'{col_letter}{row}'].value = formula

wb.save(file_path)
print("Updated Daily Tracking formulas successfully to use Q:Q (Tester) and R:R (Date).")

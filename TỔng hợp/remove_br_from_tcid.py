import openpyxl
import re

file_path = '/Users/mac/antigravity-testing-kit/Feature_02_SA_Tham_So_He_Thong/test case/Test_Cases_SA09_Final_Standard.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['TestCases']

# Find TC_ID column
tc_id_col = None
for idx, cell in enumerate(ws[1], start=1):
    if cell.value and str(cell.value).strip() == 'TC_ID':
        tc_id_col = idx
        break

counters = {'HAP': 1, 'NEG': 1, 'UI': 1}

for row_idx in range(2, ws.max_row + 1):
    cell = ws.cell(row=row_idx, column=tc_id_col)
    tc_id = cell.value
    if tc_id and isinstance(tc_id, str):
        # Find format like SA09-BR01-HAP-001
        match = re.match(r'SA09-BR0\d+-([A-Z]+)-\d+', tc_id)
        if match:
            type_str = match.group(1)
            if type_str not in counters:
                counters[type_str] = 1
            
            new_tc_id = f"SA09-{type_str}-{counters[type_str]:03d}"
            counters[type_str] += 1
            
            cell.value = new_tc_id
            print(f"Changed {tc_id} -> {new_tc_id}")

wb.save(file_path)
print("Bỏ tiền tố BR trong TC_ID thành công!")

import openpyxl
import sys
import re

def update_precondition(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb["TestCases"]

    col_map = {cell.value: cell.column for cell in ws[1]}
    col_pre = col_map.get("Precondition")
    col_note = col_map.get("Note")
    
    # Tên tính năng to theo tài liệu URD
    main_feature_name = "Quản lý nhóm code phí định kỳ"

    for row in range(2, ws.max_row + 1):
        precondition = ws.cell(row=row, column=col_pre).value or ""
        note = ws.cell(row=row, column=col_note).value or ""

        if precondition:
            new_pre = []
            updated = False
            for line in str(precondition).split('\n'):
                # Catch both the generic one and the specific sub-feature one we added earlier
                # Look for "Được phân quyền vào chức năng [something]"
                if "Được phân quyền vào chức năng" in line:
                    num_match = re.match(r'^(\d+\.\s*)', line)
                    prefix = num_match.group(1) if num_match else "1. "
                    new_line = f"{prefix}Được phân quyền vào chức năng {main_feature_name}."
                    
                    if line != new_line:
                        line = new_line
                        updated = True
                new_pre.append(line)
                
            if updated:
                ws.cell(row=row, column=col_pre).value = "\n".join(new_pre)
                
                # Check mapping in note
                note_lines = note.split('\n')
                new_notes = []
                for n_line in note_lines:
                    if "từ 'Đăng nhập chung chung' sang chức năng" in n_line:
                        pass # remove old note
                    elif "Đã cập nhật tên tính năng to" not in n_line and n_line.strip():
                        new_notes.append(n_line)
                
                final_note = "\n".join(new_notes)
                new_review_note = "📝 REVIEW: Đã cập nhật Pre-condition sang tên tính năng to (Quản lý nhóm code phí định kỳ)."
                if final_note:
                    ws.cell(row=row, column=col_note).value = final_note + "\n\n" + new_review_note
                else:
                    ws.cell(row=row, column=col_note).value = new_review_note

    wb.save(file_path)
    print("Done applying preconditions update!")

if __name__ == "__main__":
    update_precondition(sys.argv[1])

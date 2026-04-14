import openpyxl

file_path = '/Users/mac/antigravity-testing-kit/Feature_02_SA_Tham_So_He_Thong/test case/Test_Cases_SA09_Final_Standard.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['TestCases']

# Find TC_ID column
tc_id_col = None
for idx, cell in enumerate(ws[1], start=1):
    if cell.value and str(cell.value).strip() == 'TC_ID':
        tc_id_col = idx
        break

for row_idx in range(2, ws.max_row + 1):
    cell = ws.cell(row=row_idx, column=tc_id_col)
    tc_id = cell.value
    if tc_id and isinstance(tc_id, str):
        # We need to change SA09-HAP-001 -> SA09-BR-HAP-001
        # And SA09-NEG-001 -> SA09-BR-NEG-001
        # But NOT SA09-UI-001
        if tc_id.startswith('SA09-HAP-') or tc_id.startswith('SA09-NEG-'):
            new_tc_id = tc_id.replace('SA09-', 'SA09-BR-', 1)
            cell.value = new_tc_id
            print(f"Fixed: {tc_id} -> {new_tc_id}")

wb.save(file_path)
print("Đã thêm chữ 'BR' (và bỏ cụm số 0x) vào TC_ID thành công!")

import openpyxl

file_path = 'Feature_02_SA_Tham_So_He_Thong/test case/SA12_Quan_Ly_Bieu_Mau_Final.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['TestCases'] if 'TestCases' in wb.sheetnames else wb.active

c_TC_ID = 1
c_Type = 7
c_Precondition = 10
c_Expected = 12
c_Note = 14

for row in range(2, ws.max_row + 1):
    tc_id = ws.cell(row, c_TC_ID).value
    if not tc_id:
        continue
        
    expected_val = str(ws.cell(row, c_Expected).value or '')
    if '(i) Logic:' in expected_val:
        expected_val = expected_val.replace('(i) Logic:', '(i) Nghiệp vụ/Logic:')
    if '(iii) Trạng thái:' in expected_val:
        expected_val = expected_val.replace('(iii) Trạng thái:', '(iii) Trạng thái/Audit:')
    ws.cell(row, c_Expected).value = expected_val

    precond_val = str(ws.cell(row, c_Precondition).value or '')
    if precond_val and 'Maker' not in precond_val:
        if '1. ' in precond_val:
            precond_val = precond_val.replace('1. ', '2. ').replace('2. ', '3. ')
            ws.cell(row, c_Precondition).value = "1. Đăng nhập Maker.\n" + precond_val
        else:
            ws.cell(row, c_Precondition).value = "1. Đăng nhập Maker.\n2. " + precond_val

    if tc_id == 'SA12-BR02-UI-001':
        ws.cell(row, c_Type).value = 'Negative'

    elif tc_id == 'SA12-UI-01-SEARCH-HAP':
        ws.cell(row, c_Type).value = 'Happy'

    elif tc_id == 'SA12-UI-04-VIEW-HAP':
        ws.cell(row, c_Type).value = 'Happy'

    elif tc_id == 'SA12-LOG-01-INTEGRATION':
        note = str(ws.cell(row, c_Note).value or '')
        ws.cell(row, c_Note).value = note + '\nAssumption: Cần BA xác nhận hệ thống có thực hiện validate placeholder theo Bảng nguồn hay chỉ lưu nguyên text.'

new_tcs = [
    ['SA12-UI-02-EDIT-001', 'UI-FUNC.02', 'I.1.1.1', 'SA.12', 'Sửa', 'Kiểm tra chức năng giao diện load dữ liệu lên form Sửa biểu mẫu', 'Happy', 'Regression', 'P2', '1. Đăng nhập Maker có quyền Quản lý biểu mẫu.\n2. Lưới danh sách có biểu mẫu đang hoạt động.', '1. Tại lưới danh sách, nhấn icon Sửa trên 1 dòng.\n2. Xác nhận dữ liệu được điền vào form.\n3. Sửa nhẹ nội dung và chọn Xác nhận.', '(i) Nghiệp vụ/Logic: Bản ghi được lưu lại thành công.\n(ii) UI: Form hiển thị ở chế độ chỉnh sửa, data cũ load đúng xuống các fields. Sau khi Xác nhận, có Toast thành công.\n(iii) Trạng thái/Audit: Audit log cập nhật Maker thay đổi nội dung.\n(iv) Output: N/A.', 'UI-EDIT-HAP', 'Gap UI Added']
]

for new_tc in new_tcs:
    while len(new_tc) < ws.max_column:
        new_tc.append('')
    ws.append(new_tc)

wb.save(file_path)
print("Updated successfully")

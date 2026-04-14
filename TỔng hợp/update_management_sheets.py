"""
Script cập nhật chuyên sâu cho Master File Test Case ProfiX
Hỗ trợ: Dashboard, Daily Tracking (7 members), Bug Tracker & TC Sheets formatting (Module 6)
"""
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import os

FILE_PATH = '/Users/mac/antigravity-testing-kit/Test case/Test case_Management_ProfiX.xlsx'

SHEET_DASHBOARD   = '📊 Dashboard'
SHEET_TRACKING    = '📅 Daily Tracking'
SHEET_BUG_TRACKER = '🐞 Bug Tracker'
# Các sheet quản lý (không phải sheet chứa Test Case)
MGMT_SHEETS       = {SHEET_DASHBOARD, SHEET_TRACKING, SHEET_BUG_TRACKER, 'Coverage', 'Dedup_Log'}
# Danh sách thành viên (Đã thêm Thương)
MEMBERS           = ['Định', 'Vân', 'Vân Anh', 'Thanh', 'Hiền', 'Thủy', 'Thương']
DATA_ROWS         = range(3, 100) # Mở rộng lên 100 hàng cho Tracking (~3 tháng)

# Status & Tester lists for Data Validation
STATUS_LIST = '"Pass,Fail,Blocked,N/A"'
TESTER_LIST = f'"{", ".join(MEMBERS)}"'

# ── Style helpers ────────────────────────────────────────────────────────────
thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def h1_fill(): return PatternFill('solid', fgColor='1F4E79')      # Xanh Navy
def h2_fill(): return PatternFill('solid', fgColor='D9E1F2')      # Xanh Pastel Header
def exec_fill(): return PatternFill('solid', fgColor='FCE4D6')    # Cam nhạt cho Exec columns
def total_fill(): return PatternFill('solid', fgColor='FFF2CC')    # Vàng Nhạt

def h1_font(size=14): return Font(bold=True, color='FFFFFF', size=size)
def h2_font(): return Font(bold=True, size=11)
def body_font(): return Font(size=10)
def bold_font(): return Font(bold=True, size=10)

def style_cell(cell, fill=None, font=None, center=False, vtop=True, wrap=True, border=BORDER):
    if border: cell.border = border
    if fill:   cell.fill   = fill
    if font:   cell.font   = font
    cell.alignment = Alignment(
        horizontal='center' if center else 'left',
        vertical='top' if vtop else 'center',
        wrap_text=wrap
    )

def style_header1(ws, merge_range, value):
    try:
        ws.merge_cells(merge_range)
    except: pass
    cell = ws.cell(row=1, column=1)
    cell.value = value
    cell.font = h1_font()
    cell.fill = h1_fill()
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

# ── 1. DASHBOARD ────────────────────────────────────────────────────────────
def update_dashboard(wb):
    if SHEET_DASHBOARD not in wb.sheetnames: wb.create_sheet(SHEET_DASHBOARD)
    ws = wb[SHEET_DASHBOARD]
    tc_sheets = [s for s in wb.sheetnames if s not in MGMT_SHEETS]

    # Unmerge & Clean
    for m in list(ws.merged_cells.ranges): ws.unmerge_cells(str(m))
    for r in ws.iter_rows(max_row=max(ws.max_row, 100), max_col=15):
        for c in r: c.value = None; c.style = 'Normal'

    style_header1(ws, 'A1:I1', 'PROFIX — QUALITY DASHBOARD SUMMARY')
    headers = ['STT', 'Module (Feature)', 'Total TC', 'Passed', 'Failed', 'Blocked', 'N/A', 'Execution %', 'Pass Rate %']
    widths = [6, 45, 12, 11, 11, 11, 9, 15, 15]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=c, value=h)
        style_cell(cell, fill=h2_fill(), font=h2_font(), center=True)
        ws.column_dimensions[get_column_letter(c)].width = w

    for idx, s in enumerate(tc_sheets, 1):
        r = idx + 2
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=s)
        # Sử dụng cột U (Final Status) để đếm cho Dashboard
        ws.cell(row=r, column=3, value=f"=COUNTA('{s}'!A:A)-2")
        ws.cell(row=r, column=4, value=f"=COUNTIF('{s}'!U:U,\"Pass\")")
        ws.cell(row=r, column=5, value=f"=COUNTIF('{s}'!U:U,\"Fail\")")
        ws.cell(row=r, column=6, value=f"=COUNTIF('{s}'!U:U,\"Blocked\")")
        ws.cell(row=r, column=7, value=f"=COUNTIF('{s}'!U:U,\"N/A\")")
        ws.cell(row=r, column=8, value=f"=IF(C{r}>0,(D{r}+E{r}+F{r})/C{r},0)")
        ws.cell(row=r, column=9, value=f"=IF(C{r}>0,D{r}/C{r},0)")
        ws.cell(row=r, column=8).number_format = ws.cell(row=r, column=9).number_format = '0.00%'
        for c in range(1, 10): style_cell(ws.cell(row=r, column=c), font=body_font(), center=(c != 2))

    t_r = len(tc_sheets) + 3
    ws.cell(row=t_r, column=2, value="TOTAL")
    for c in range(3, 8):
        ws.cell(row=t_r, column=c, value=f"=SUM({get_column_letter(c)}3:{get_column_letter(c)}{t_r-1})")
    ws.cell(row=t_r, column=8, value=f"=IF(C{t_r}>0,(D{t_r}+E{t_r}+F{t_r})/C{t_r},0)")
    ws.cell(row=t_r, column=9, value=f"=IF(C{t_r}>0,D{t_r}/C{t_r},0)")
    ws.cell(row=t_r, column=8).number_format = ws.cell(row=t_r, column=9).number_format = '0.00%'
    for c in range(1, 10): style_cell(ws.cell(row=t_r, column=c), fill=total_fill(), font=bold_font(), center=(c != 2))
    print(f"  [v] {SHEET_DASHBOARD} updated.")

# ── 2. DAILY TRACKING ───────────────────────────────────────────────────────
def update_daily_tracking(wb):
    if SHEET_TRACKING not in wb.sheetnames: wb.create_sheet(SHEET_TRACKING)
    ws = wb[SHEET_TRACKING]
    tc_sheets = [s for s in wb.sheetnames if s not in MGMT_SHEETS]

    # Unmerge & Clean
    for m in list(ws.merged_cells.ranges): ws.unmerge_cells(str(m))
    for r in ws.iter_rows(max_row=max(ws.max_row, 120), max_col=15):
        for c in r: c.value = None; c.style = 'Normal'

    last_col = 1 + len(MEMBERS) + 2 # Date + Members + Total/Day + Lũy kế
    style_header1(ws, f'A1:{get_column_letter(last_col)}1', 'DAILY EXECUTION TRACKING — TEAM PROFIX')
    headers = ['Date'] + MEMBERS + ['Total/Day', 'Lũy kế']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        style_cell(cell, fill=h2_fill(), font=h2_font(), center=True)
        ws.column_dimensions[get_column_letter(c)].width = 15 if c == 1 else 12

    for r in DATA_ROWS:
        ws.cell(row=r, column=1).number_format = 'DD/MM/YYYY'
        style_cell(ws.cell(row=r, column=1), center=True)
        # sum counting across all TC sheets (column P = Tester R1, column Q = Date R1)
        for m_idx, member in enumerate(MEMBERS, 2):
            parts = [f"COUNTIFS('{s}'!P:P,\"{member}\",'{s}'!Q:Q,A{r})" for s in tc_sheets]
            ws.cell(row=r, column=m_idx, value="=" + "+".join(parts) if parts else "=0")
            style_cell(ws.cell(row=r, column=m_idx), center=True)
        # Total/Day
        col_end = 1 + len(MEMBERS)
        ws.cell(row=r, column=col_end+1, value=f"=SUM(B{r}:{get_column_letter(col_end)}{r})")
        style_cell(ws.cell(row=r, column=col_end+1), font=bold_font(), center=True)
        # Lũy kế
        t_col = get_column_letter(col_end+1)
        ws.cell(row=r, column=col_end+2, value=f"=SUM({t_col}$3:{t_col}{r})")
        style_cell(ws.cell(row=r, column=col_end+2), font=bold_font(), center=True, fill=total_fill())
    print(f"  [v] {SHEET_TRACKING} updated.")

# ── 3. TC SHEETS (MODULE 6) ─────────────────────────────────────────────────
def format_all_tc_sheets(wb):
    tc_sheets = [s for s in wb.sheetnames if s not in MGMT_SHEETS]
    
    # 3.1. Define Data Validations
    dv_status = DataValidation(type="list", formula1=STATUS_LIST, allow_blank=True)
    dv_tester = DataValidation(type="list", formula1=TESTER_LIST, allow_blank=True)
    
    for s_name in tc_sheets:
        ws = wb[s_name]
        # Thêm validation vào sheet
        if dv_status not in ws.data_validations: ws.add_data_validation(dv_status)
        if dv_tester not in ws.data_validations: ws.add_data_validation(dv_tester)
        
        # O: Status R1, P: Tester R1, Q: Date R1, R: Status R2, S: Tester R2, T: Date R2, U: Final Status
        exec_cols = [('O','Status R1',dv_status), ('P','Tester R1',dv_tester), ('Q','Date R1',None), 
                     ('R','Status R2',dv_status), ('S','Tester R2',dv_tester), ('T','Date R2',None), 
                     ('U','Final Status',dv_status)]
        
        for col_let, h_val, dv in exec_cols:
            # Header
            cell = ws[f"{col_let}2"]
            cell.value = h_val
            style_cell(cell, fill=h2_fill(), font=h2_font(), center=True)
            ws.column_dimensions[col_let].width = 15
            
            # Data cells formatting (Row 3 -> 500)
            target_range = f"{col_let}3:{col_let}500"
            for r in range(3, 501):
                c = ws.cell(row=r, column=ord(col_let)-64)
                style_cell(c, fill=exec_fill(), center=True)
                if h_val.startswith('Date'): c.number_format = 'DD/MM/YYYY'
            
            # Apply Validation
            if dv: dv.add(f"{col_let}3:{col_let}500")

    print(f"  [v] Formatted {len(tc_sheets)} Test Case sheets (Module 6).")

# ── 4. BUG TRACKER ──────────────────────────────────────────────────────────
def update_bug_tracker(wb):
    if SHEET_BUG_TRACKER not in wb.sheetnames: wb.create_sheet(SHEET_BUG_TRACKER)
    ws = wb[SHEET_BUG_TRACKER]
    style_header1(ws, 'A1:L1', 'PROFIX — BUG TRACKER')
    headers = ['Bug ID', 'Related TC', 'Title', 'Module', 'Severity', 'Priority', 'Env', 'Steps', 'Expected', 'Actual', 'Assignee', 'Status']
    widths = [12, 15, 45, 12, 10, 10, 15, 45, 45, 45, 15, 12]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=c, value=h)
        style_cell(cell, fill=h2_fill(), font=h2_font(), center=True)
        ws.column_dimensions[get_column_letter(c)].width = w
    
    # Simple check for critical bugs highlighting
    for r in range(3, ws.max_row + 1):
        for c in range(1, 13):
            cell = ws.cell(row=r, column=c)
            style_cell(cell, wrap=True)
            if str(cell.value) in ['S1', 'S2', 'Fatal', 'Critical']:
                cell.font = Font(bold=True, color='9C0006')
                cell.fill = PatternFill('solid', fgColor='FFC7CE')
    print(f"  [v] {SHEET_BUG_TRACKER} styled.")

# ── MAIN ────────────────────────────────────────────────────────────────────
def main():
    if not os.path.exists(FILE_PATH): return print(f"❌ File not found: {FILE_PATH}")
    print(f"🚀 Updating Master: {os.path.basename(FILE_PATH)}")
    wb = openpyxl.load_workbook(FILE_PATH)
    
    update_dashboard(wb)
    update_daily_tracking(wb)
    update_bug_tracker(wb)
    format_all_tc_sheets(wb)
    
    wb.save(FILE_PATH)
    print(f"\n✅ SUCCESS! All modules standardized.")

if __name__ == '__main__':
    main()

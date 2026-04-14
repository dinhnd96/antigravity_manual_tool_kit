import openpyxl

file_path = '/Users/mac/antigravity-testing-kit/Feature_02_SA_Tham_So_He_Thong/test case/SA09.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['TestCases']

# Header mapping
headers = [cell.value for cell in ws[1]]
col_map = {name: idx for idx, name in enumerate(headers)}

rows_to_delete = []

for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
    tc_id = row[col_map['TC_ID']].value
    if not tc_id:
        continue
        
    # Mark checker flows and duplicate deletes for deletion
    if tc_id in ('SA09-FLOW-CHECKER-001', 'SA09-FLOW-CHECKER-002', 'SA09-UI-05-DELETE-002', 'SA09-UI-05-DELETE-003'):
        rows_to_delete.append(row_idx)
        continue

    # Update Precondition to remove "Maker" -> "Người dùng"
    precond = row[col_map['Precondition']].value
    if precond:
        precond = precond.replace('Đăng nhập Maker', 'Đăng nhập vào hệ thống')
        precond = precond.replace('User Maker', 'Người dùng')
        precond = precond.replace('Maker', 'Người dùng')
        precond = precond.replace('Chờ duyệt', 'Hoạt động')
        precond = precond.replace('Nháp', 'Hoạt động')
        precond = precond.replace('Đã duyệt', 'Hoạt động')
        row[col_map['Precondition']].value = precond

    # Update Steps to remove Maker
    steps = row[col_map['Steps']].value
    if steps:
        steps = steps.replace('Maker', 'Người dùng')
        # fix flow 001
        if tc_id == 'SA09-FLOW-001':
            lines = steps.split('\n')
            new_lines = [l for l in lines if 'Checker Duyệt' not in l]
            row[col_map['Steps']].value = '\n'.join(new_lines)
        else:
            row[col_map['Steps']].value = steps
            
    # Update Expected to remove Checker/Maker/Chờ duyệt
    expected = row[col_map['Expected']].value
    if expected:
        expected = expected.replace('Maker', 'Người dùng')
        expected = expected.replace('[Chờ duyệt]', '[Hoạt động]')
        expected = expected.replace('Chờ duyệt', 'Hoạt động')
        expected = expected.replace('Nháp', 'Hoạt động')
        expected = expected.replace('Đã duyệt', 'Hoạt động')
        
        if tc_id == 'SA09-FLOW-001':
            expected = expected.replace(
                "Sinh trạng thái Hoạt động ban đầu -> Hoạt động -> Cập nhật thành Hoạt động -> Xóa.",
                "Bản ghi được thêm mới -> Cập nhật thông tin -> Xóa thành công."
            )
        row[col_map['Expected']].value = expected

    # Update Title
    title = row[col_map['Title']].value
    if title:
        title = title.replace('trạng thái Chờ duyệt', 'thành công')
        title = title.replace('bản ghi nháp', 'bản ghi')
        row[col_map['Title']].value = title

# Delete rows in reverse order
for row_idx in sorted(rows_to_delete, reverse=True):
    ws.delete_rows(row_idx)

# Update Coverage
ws_cov = wb['Coverage']
cov_rows_to_delete = []
for row_idx, row in enumerate(ws_cov.iter_rows(min_row=2), 2):
    ref = row[1].value
    note = row[3].value
    if note and 'Checker' in note:
        cov_rows_to_delete.append(row_idx)
        
for row_idx in sorted(cov_rows_to_delete, reverse=True):
    ws_cov.delete_rows(row_idx)

wb.save(file_path)
print("Updated SA09 test cases to remove Maker/Checker and approval flow.")

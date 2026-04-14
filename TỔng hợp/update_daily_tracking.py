import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

file_path = '/Users/mac/antigravity-testing-kit/Feature_02_SA_Tham_So_He_Thong/test case/ProfiX_Master_Test_Cases_Final.xlsx'
wb = openpyxl.load_workbook(file_path)

if '📅 Daily Tracking' in wb.sheetnames:
    ws = wb['📅 Daily Tracking']
else:
    ws = wb.create_sheet('📅 Daily Tracking')

# Setup Header if not exists
ws.merge_cells('A1:J1')
ws['A1'] = 'DAILY EXECUTION TRACKING — TEAM PROFIX'
ws['A1'].font = Font(bold=True, size=14)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

headers = ['Date', 'Định', 'Vân', 'Vân Anh', 'Thanh', 'Hiền', 'Thủy', 'Thương', 'Total/Day', 'Tổng lũy kế']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col_idx)
    cell.value = header
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Find all test case sheets (starting with SA)
tc_sheets = [s for s in wb.sheetnames if s.startswith('SA')]

# Write formulas from row 3 to 60
for row in range(3, 61):
    date_cell = f'A{row}'
    
    # Columns B to H correspond to members
    members = headers[1:8]
    for col_idx, member in enumerate(members, 2):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        
        # Build formula summing COUNTIFS for all TC sheets
        # P = Tester R1, Q = Date R1
        countifs_parts = []
        for sheet in tc_sheets:
            # We must wrap sheet name in single quotes
            sheet_name_escaped = sheet.replace("'", "''")
            countifs_parts.append(f"COUNTIFS('{sheet_name_escaped}'!P:P, \"{member}\", '{sheet_name_escaped}'!Q:Q, {date_cell})")
            
        formula = "=" + " + ".join(countifs_parts)
        cell = ws[f'{col_letter}{row}']
        cell.value = formula
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        
    # Total/Day (I)
    i_cell = f'I{row}'
    ws[i_cell].value = f"=SUM(B{row}:H{row})"
    ws[i_cell].alignment = Alignment(horizontal='center')
    ws[i_cell].border = thin_border
    
    # Tổng lũy kế (J)
    j_cell = f'J{row}'
    ws[j_cell].value = f"=SUM(I$3:I{row})"
    ws[j_cell].alignment = Alignment(horizontal='center')
    ws[j_cell].border = thin_border
    
    # Apply border to A cell as well
    ws[f'A{row}'].border = thin_border
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    ws[f'A{row}'].number_format = 'DD/MM/YYYY'

wb.save(file_path)
print("Updated Daily Tracking formulas successfully across 60 rows.")

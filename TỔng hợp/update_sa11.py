import openpyxl

file_path = 'Feature_02_SA_Tham_So_He_Thong/test case/SA11_Tra_Cuu_Lich_Su_Final.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['TestCases'] if 'TestCases' in wb.sheetnames else wb.active

c_TC_ID = 1
c_Feature = 5
c_Type = 7
c_Precondition = 10
c_Steps = 11
c_Expected = 12

for row in range(2, ws.max_row + 1):
    tc_id = ws.cell(row, c_TC_ID).value
    if not tc_id:
        continue
        
    expected_val = str(ws.cell(row, c_Expected).value or '')
    if '(i) Logic:' in expected_val:
        expected_val = expected_val.replace('(i) Logic:', '(i) Nghiệp vụ/Logic:')
    
    if tc_id == 'SA11-UI-04-VIEW-HAP':
        if 'băt' in expected_val:
            expected_val = expected_val.replace('Thời gian băt đầu', 'Thời gian bắt đầu')
            
    ws.cell(row, c_Expected).value = expected_val

    precond_val = str(ws.cell(row, c_Precondition).value or '')
    if precond_val and 'Đăng nhập Maker' not in precond_val and 'Maker' not in precond_val:
        if '1. ' in precond_val:
            precond_val = precond_val.replace('1. ', '2. ').replace('2. ', '3. ')
            ws.cell(row, c_Precondition).value = "1. Đăng nhập Maker.\n" + precond_val
        else:
            ws.cell(row, c_Precondition).value = "1. Đăng nhập Maker.\n2. " + precond_val

    if tc_id == 'SA11-UI-06-PAGING':
        ws.cell(row, c_Type).value = 'Happy'
        ws.cell(row, c_Feature).value = 'Phân trang'

wb.save(file_path)
print("Updated successfully")

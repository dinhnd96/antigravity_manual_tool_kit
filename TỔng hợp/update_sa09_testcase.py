import openpyxl

file_path = '/Users/mac/antigravity-testing-kit/Feature_02_SA_Tham_So_He_Thong/test case/Test_Cases_SA09_Final_Standard.xlsx'

# Open workbook
wb = openpyxl.load_workbook(file_path)
ws = wb['TestCases']
# Find column indexes
header_row = 1
col_map = {}
for idx, cell in enumerate(ws[header_row], start=1):
    if cell.value:
        col_map[cell.value.strip()] = idx

def get_col(name):
    return col_map.get(name, -1)

# Helper for cell modification
def update_row(r_idx, updates):
    for col_name, val in updates.items():
        c_idx = get_col(col_name)
        if c_idx != -1:
            ws.cell(row=r_idx, column=c_idx).value = val
            ws.cell(row=r_idx, column=c_idx).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')

tc_id_col = get_col('TC_ID')

for row_idx in range(header_row + 1, ws.max_row + 1):
    tc_id_cell = ws.cell(row=row_idx, column=tc_id_col)
    if not tc_id_cell.value:
        continue
    
    tc_id = str(tc_id_cell.value).strip()

    # 1. Update UI-FUNC.06
    if tc_id == 'SA09-UI-01-BASIC-001':
        update_row(row_idx, {
            'TC_ID': 'SA09-UI-06-SEARCH-001',
            'BR_Ref': 'UI-FUNC.06',
            'Title': 'Kiểm tra chức năng Tìm kiếm và Làm mới bộ lọc trên lưới danh sách'
        })
    
    # 2. Update BR_06 (Logic Error)
    elif tc_id == 'SA09-BR06-HAP-001':
        update_row(row_idx, {
            'Precondition': '1. Hệ thống cấu hình 2 Nhóm phí (Nhóm A có mức ưu tiên = 1, Nhóm B có mức ưu tiên = 2).\n2. Hai nhóm phí có cùng Ngày thu trên cùng một Tài khoản/Khách hàng.',
            'Steps': '1. Duyệt cấu hình 2 Nhóm phí.\n2. Chạy Job thu phí.\n3. Kiểm tra message được đẩy sang Topic của Kafka/ProfiX.',
            'Expected': '(i) Logic: Dữ liệu được ghi vào Topic theo thứ tự ưu tiên cấu hình.\n(ii) UI: (N/A).\n(iii) Trạng thái/Audit: Bản ghi của Nhóm A (Priority 1) được đẩy xuống Topic TRƯỚC bản ghi của Nhóm B (Priority 2). Job Log ghi nhận thành công.\n(iv) File/Email: Message Format đúng định dạng Json yêu cầu.'
        })

    # 3. Format error BR_01
    elif tc_id == 'SA09-BR01-HAP-001':
        update_row(row_idx, {
            'Precondition': '1. User Maker được phân quyền chức năng Thêm mới.\n2. Có sẵn Code phí định kỳ chưa thuộc nhóm nào.'
        })

    # 4. Format error BR_03
    elif tc_id == 'SA09-BR03-NEG-001':
        update_row(row_idx, {
            'Precondition': '1. Trong hệ thống đã tồn tại Mã nhóm "MONTH_FEE_01".\n2. User đang ở màn hình Thêm mới.'
        })
    
    # 5. Add note to BR_04
    elif tc_id == 'SA09-BR04-HAP-001':
        note_val = ws.cell(row=row_idx, column=get_col('Note')).value
        new_note = (note_val + "\n" if note_val else "") + "Cần test bổ sung luồng lọc theo 'Ngày thu theo dữ liệu' (Variant)."
        update_row(row_idx, {'Note': new_note})

# Constructing new test cases to append
new_cases = [
    {
        'TC_ID': 'SA09-UI-03-EDIT-001',
        'BR_Ref': 'UI-FUNC.03',
        'URD_Ref': 'I.1.1.1',
        'Module': 'SA.09',
        'Feature': 'Sửa',
        'Title': 'Kiểm tra chức năng Sửa bản ghi Nhóm code phí định kỳ khi ở trạng thái Chờ duyệt',
        'Type': 'Happy',
        'Category': 'Regression',
        'Priority': 'P2',
        'Precondition': '1. Bản ghi Nhóm code phí đang ở trạng thái Chờ duyệt.\n2. User Maker được phân quyền Sửa.',
        'Steps': '1. Tại màn hình danh sách, chọn bản ghi [Chờ duyệt] và nhấn icon Sửa.\n2. Sửa thông tin (VD: Tên nhóm).\n3. Nhấn Xác nhận.',
        'Expected': '(i) Logic: Bản ghi nháp được cập nhật nội dung.\n(ii) UI: Hiển thị Toast thông báo "Cập nhật thành công". Quay về màn hình danh sách.\n(iii) Trạng thái/Audit: Trạng thái bản ghi vẫn là [Chờ duyệt], ghi nhận log Sửa.\n(iv) File/Email: (Không có).',
        'Trace_ID': 'UI03-EDIT-HAP',
        'Note': ''
    },
    {
        'TC_ID': 'SA09-UI-05-DELETE-001',
        'BR_Ref': 'UI-FUNC.05',
        'URD_Ref': 'I.1.1.1',
        'Module': 'SA.09',
        'Feature': 'Xóa',
        'Title': 'Kiểm tra hệ thống hiển thị popup Confirm và Xóa thành công bản ghi nháp',
        'Type': 'Happy',
        'Category': 'Regression',
        'Priority': 'P2',
        'Precondition': '1. Bản ghi Nhóm code phí đang ở trạng thái Nháp/Chờ duyệt.\n2. User Maker được phân quyền Xóa.',
        'Steps': '1. Tại màn hình danh sách, nhấn icon Xóa tại dòng bản ghi.\n2. Hệ thống hiển thị popup confirm "Bạn có chắc chắn muốn xóa?".\n3. Nhấn Xác nhận.',
        'Expected': '(i) Logic: Bản ghi bị đánh dấu Xóa (Deleted) trong hệ thống.\n(ii) UI: Có popup Confirm. Sau khi xác nhận, hiện Toast "Xóa thành công", bản ghi biến mất khỏi lưới.\n(iii) Trạng thái/Audit: Ghi nhận log Xóa bởi Maker vào ngày T.\n(iv) File/Email: (Không có).',
        'Trace_ID': 'UI05-DEL-HAP',
        'Note': ''
    }
]

# Append new test cases
target_row = ws.max_row + 1
from copy import copy

# Copy styling from the last valid row (e.g. row 2)
source_row = 2

for case in new_cases:
    for col_name, val in case.items():
        c_idx = get_col(col_name)
        if c_idx != -1:
            new_cell = ws.cell(row=target_row, column=c_idx)
            new_cell.value = val
            source_cell = ws.cell(row=source_row, column=c_idx)
            if source_cell.has_style:
                new_cell.font = copy(source_cell.font)
                new_cell.border = copy(source_cell.border)
                new_cell.fill = copy(source_cell.fill)
                new_cell.number_format = copy(source_cell.number_format)
                new_cell.protection = copy(source_cell.protection)
                new_cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
    target_row += 1

wb.save(file_path)
print("Updated Excel file successfully.")

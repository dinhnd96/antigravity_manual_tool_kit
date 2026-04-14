import openpyxl
import shutil
import warnings
import re

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

source_file = 'Feature_02_SA_Tham_So_He_Thong/test case/SA01_Đăng nhập.xlsx'
dest_file = 'Feature_02_SA_Tham_So_He_Thong/test case/SA01_Đăng nhập - updated_final.xlsx'

shutil.copy(source_file, dest_file)

wb = openpyxl.load_workbook(dest_file)
ws = wb.active

headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]

# Tìm cột dựa trên headers hiện tại
col_id = headers.index('TC_ID') + 1 if 'TC_ID' in headers else 1
col_title = headers.index('Title') + 1 if 'Title' in headers else headers.index('Test Case Name') + 1 if 'Test Case Name' in headers else 6
col_exp = None
for name in ['Expected', 'Expected Result']:
    if name in headers:
        col_exp = headers.index(name) + 1
        break

for r in range(2, ws.max_row + 1):
    tc_id = ws.cell(r, col_id).value
    if not tc_id: continue
    
    # 1. Update TC_ID Format
    new_id = str(tc_id)
    new_id = new_id.replace('SA01-HAPPY-', 'SA01-BR-HAP-')
    new_id = new_id.replace('SA01-SECURITY-', 'SA01-BR-SEC-')
    new_id = new_id.replace('SA01-NEGATIVE-', 'SA01-BR-NEG-')
    new_id = new_id.replace('SA01-NEG-', 'SA01-BR-NEG-')
    new_id = new_id.replace('SA01-BOUNDARY-', 'SA01-BR-BOU-')
    new_id = new_id.replace('SA01-UI-', 'SA01-UI-')
    ws.cell(r, col_id).value = new_id
    
    title = str(ws.cell(r, col_title).value) if ws.cell(r, col_title).value else ""
    exp = str(ws.cell(r, col_exp).value).strip() if col_exp and ws.cell(r, col_exp).value else ""
    
    # 2. Đảm bảo cấu trúc 4 Lớp (i-ii-iii-iv)
    if '(i)' in exp and '(ii)' in exp:
        if '(iii)' not in exp:
            # Chèn (iii) đúng lúc trước chữ (iv) nếu có, hoặc cuối dòng
            if '(iv)' in exp:
                exp = exp.replace('(iv)', '(iii) Trạng thái/Audit: Hệ thống không ghi nhận Log thay đổi dữ liệu.\n(iv)')
            else:
                exp += '\n(iii) Trạng thái/Audit: Hệ thống không ghi nhận Log thay đổi trạng thái.'
        if '(iv)' not in exp:
            exp += '\n(iv) Output: Không có Output.'
    
    # 3. Gắn cờ cảnh báo BA cho các luồng Ghi nhớ (007, 016) và EntraID (009, 010, 013, 015, 017)
    needs_flag = False
    if any(x in new_id for x in ['007', '016', '009', '010', '013', '015', '017']):
        needs_flag = True
        
    if needs_flag and not title.startswith("[CHỜ BA CONFIRM MOCKUP]"):
        title = "[CHỜ BA CONFIRM MOCKUP] " + title

    ws.cell(r, col_title).value = title
    if col_exp:
        ws.cell(r, col_exp).value = exp

wb.save(dest_file)
print(f"Updated Test Cases successfully. Saved to: {dest_file}")

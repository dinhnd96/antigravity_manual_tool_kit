import openpyxl
import shutil
import warnings
import re

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

source_file = 'Feature_02_SA_Tham_So_He_Thong/test case/SA03_Quản lý người dùng.xlsx'
dest_file = 'Feature_02_SA_Tham_So_He_Thong/test case/SA03_Quản lý người dùng - updated_final.xlsx'

shutil.copy(source_file, dest_file)

wb = openpyxl.load_workbook(dest_file)
ws = wb.active

headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]

col_id = headers.index('TC_ID') + 1 if 'TC_ID' in headers else headers.index('Test Case ID') + 1 if 'Test Case ID' in headers else 1
col_title = headers.index('Title') + 1 if 'Title' in headers else 6
col_steps = headers.index('Steps') + 1 if 'Steps' in headers else 11
col_exp = None
for name in ['Expected', 'Expected Result', 'Expected results']:
    if name in headers:
        col_exp = headers.index(name) + 1
        break

# Mảng các tên Header chuẩn để mapping cho dòng add mới
header_names = headers

for r in range(2, ws.max_row + 1):
    tc_id = ws.cell(r, col_id).value
    if not tc_id or str(tc_id) == 'TC_ID' or str(tc_id) == 'Test Case ID': continue
    
    # 1. Format TC_ID chuẩn
    new_id = str(tc_id)
    new_id = new_id.replace('SA03-HAPPY-', 'SA03-BR-HAP-')
    new_id = new_id.replace('SA03-SECURITY-', 'SA03-BR-SEC-')
    new_id = new_id.replace('SA03-NEG-', 'SA03-BR-NEG-')
    new_id = new_id.replace('SA03-BOUNDARY-', 'SA03-BR-BOU-')
    new_id = new_id.replace('SA03-UI-', 'SA03-UI-')
    ws.cell(r, col_id).value = new_id
    
    title = str(ws.cell(r, col_title).value) if ws.cell(r, col_title).value else ""
    exp = str(ws.cell(r, col_exp).value).strip() if col_exp and ws.cell(r, col_exp).value else ""
    steps = str(ws.cell(r, col_steps).value).strip() if ws.cell(r, col_steps).value else ""
    
    # 2. Cắm cờ BA
    if new_id == 'SA03-BR-HAP-002' and not title.startswith('[CẦN BA'):
        title = "[CẦN BA CONFIRM ICON XEM MOCKUP] " + title
        
    # 3. Clarify UI-007
    if new_id == 'SA03-UI-007':
        steps = "1. Tại Grid Quản lý người dùng, click vào biểu tượng Bánh răng (Sửa).\n2. Tại Popup 'Sửa nhóm quyền', chọn thử một Nhóm quyền bất kỳ ở Combobox.\n3. Quan sát các ô Checkbox trong lưới ma trận Tham số bên dưới."
        ws.cell(r, col_steps).value = steps
        
    # 4. Fill Auto Audit Layers
    if '(i)' in exp and '(ii)' in exp:
        if '(iii)' not in exp:
            exp += '\n(iii) Trạng thái/Audit: Hệ thống không tạo mới dòng log nào đặc biệt (Trừ log hệ thống chung).'
        if '(iv)' not in exp:
            exp += '\n(iv) Output: Không sinh Message Broker hay Xuất File.'

    ws.cell(r, col_title).value = title
    if col_exp:
        ws.cell(r, col_exp).value = exp

# 5. Append TC Bị Gap (Thêm mới, Tải lên)
new_tcs = [
    {
        'id': 'SA03-BR-HAP-017', 
        'title': 'Thêm mới người dùng thành công (Nhập đầy đủ thông tin vào Form)',
        'steps': '1. Click nút [ + Thêm mới ] tại màn hình chính.\n2. Nhập đầy đủ thông tin: User, Tên hiển thị, Khối, Phòng ban, Enum Nhóm quyền và Trạng thái.\n3. Click [Xác nhận].',
        'exp': '(i) Nghiệp vụ: Hệ thống validate OK và Lưu trữ User mới vào Database thông suốt, không yêu cầu phân luồng Maker/Checker.\n(ii) UI: Đóng pop-up Thêm mới, có Toast xanh báo thành công. Grid Danh sách tự động reload và show User mới thêm.\n(iii) Trạng thái/Audit: Log hành động Tạo người dùng.\n(iv) Output: Không có.'
    },
    {
        'id': 'SA03-BR-NEG-018', 
        'title': 'Hiển thị cảnh báo Validation khi bỏ trống field bắt buộc tại form Thêm mới',
        'steps': '1. Click nút [ + Thêm mới ].\n2. Xoá trắng hoặc không nhập các field bắt buộc, ví dụ: "User" và "Nhóm quyền".\n3. Click [Đóng] hoặc [Xác nhận].',
        'exp': '(i) Nghiệp vụ: Chặn truy vấn đẩy xuống DB, validate Failed.\n(ii) UI: Hệ thống không tắt Popup. Các TextBox rỗng bị bôi đỏ (Focus state), hiện dòng Error tool-tip "Trường bắt buộc".\n(iii) Trạng thái/Audit: Không tạo User.\n(iv) Output: Không có.'
    },
    {
        'id': 'SA03-BR-HAP-019', 
        'title': 'Tải lên (Import) hàng loạt tài khoản người dùng thành công bằng File mẫu',
        'steps': '1. Click nút [ ⭱ Tải lên ].\n2. Tải File_Template.xlsx đính kèm và dán Data 10 User chuẩn format.\n3. Chèn file lên form Upload và Click [Xác nhận].',
        'exp': '(i) Nghiệp vụ: Backend parse file Excel thành công, insert Batch 10 dòng vào Database.\n(ii) UI: Có Animation Spinner/Loading tiến trình tải. Báo Upload 100% OK. Grid Danh sách reload hiển thị đủ User đã import.\n(iii) Trạng thái/Audit: Ghi log Import hàng loạt.\n(iv) Output: Gửi in-app Notification/Email báo cáo kết quả (Nếu hệ thống có thiết kế Async).'
    },
    {
        'id': 'SA03-BR-NEG-020', 
        'title': 'Nhận diện lỗi Validation khi [Tải lên] file sai mở rộng dạng (PDF, DOC)',
        'steps': '1. Click nút [ ⭱ Tải lên ].\n2. Chọn 1 file document.pdf (hoặc text.txt) thay vì file định dạng EXCEL (.xls/.xlsx).',
        'exp': '(i) Nghiệp vụ: Cổng chặn trực tiếp đuôi file phi chuẩn.\n(ii) UI: Bật popup/toast báo đỏ "Định dạng file không hỗ trợ, vui lòng chọn file .xlsx".\n(iii) Trạng thái/Audit: Không.\n(iv) Output: Không có.'
    }
]

cur_row = ws.max_row + 1
for tc in new_tcs:
    for c, col_name in enumerate(header_names, 1):
        if not col_name: continue
        val = ''
        if col_name == 'TC_ID' or col_name == 'Test Case ID' or col_name == 'Test case ID':
            val = tc['id']
        elif col_name == 'Title':
            val = tc['title']
        elif col_name == 'Steps':
            val = tc['steps']
        elif col_name in ['Expected', 'Expected Result', 'Expected results']:
            val = tc['exp']
        elif col_name == 'Module':
            val = 'SA.03'
        elif col_name == 'Precondition':
            val = '1. User Admin được Login vào mạng Admin thành công.'
        elif col_name == 'Type':
            if 'Tải lên' in tc['title']: val = 'Integration'
            elif 'thành công' in tc['title']: val = 'Happy'
            else: val = 'Negative'
            
        ws.cell(cur_row, c).value = val
    cur_row += 1

wb.save(dest_file)
print(f"Updated Test Cases and Appended Missing TC successfully. Saved to: {dest_file}")

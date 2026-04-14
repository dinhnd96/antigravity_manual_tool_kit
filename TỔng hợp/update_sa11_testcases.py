import openpyxl

file_path = '/Users/mac/antigravity-testing-kit/Feature_02_SA_Tham_So_He_Thong/test case/SA11_Tra_Cuu_Lich_Su_Final.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['TestCases']

headers = [cell.value for cell in ws[1]]
col_map = {name: idx for idx, name in enumerate(headers)}

for row in ws.iter_rows(min_row=2):
    tc_id = row[col_map['TC_ID']].value
    if not tc_id:
        continue
    
    # 1. Update Maker -> Người dùng
    for col_name in ['Precondition', 'Steps', 'Expected']:
        val = row[col_map[col_name]].value
        if val:
            val = val.replace('Đăng nhập Maker.', 'Đăng nhập vào hệ thống.')
            val = val.replace('Maker', 'Người dùng')
            row[col_map[col_name]].value = val

    # 2. Update Column names in SA11-UI-01-SEARCH-HAP
    if tc_id == 'SA11-UI-01-SEARCH-HAP':
        expected = row[col_map['Expected']].value
        # Replace '(ii) UI: Lưới reload nhanh, dữ liệu hiển thị đúng các cột: Thời gian chạy, Tên job, Kết quả.'
        expected = expected.replace(
            'Các cột: Thời gian chạy, Tên job, Kết quả',
            'Các cột lấy từ spec: Ngày chạy job, Mã job, Số thứ tự job, Tên job, Nhóm Code phí, Mô tả, Kiểu job, Từ thời điểm, Đến thời điểm, Lệnh thực thi, Ngày thực thi, Trạng thái vận hành, Lần thực thi, Thời gian thực hiện, Người thực hiện'
        )
        # fallback replace if text was slightly different
        expected = expected.replace(
            'dữ liệu hiển thị đúng các cột: Thời gian chạy, Tên job, Kết quả.',
            'dữ liệu hiển thị đúng các cột như Mockup: Ngày chạy job, Trạng thái vận hành, Tên job...'
        )
        # Let's forcefully set line (ii)
        lines = expected.split('\n')
        for i, line in enumerate(lines):
            if line.startswith('(ii) UI:'):
                lines[i] = '(ii) UI: Lưới reload nhanh, dữ liệu hiển thị đủ 15 cột: Ngày chạy job, Mã job, Số thứ tự job, Tên job, Nhóm Code phí, Mô tả, Kiểu job, Từ thời điểm, Đến thời điểm, Lệnh thực thi, Ngày thực thi, Trạng thái vận hành, Lần thực thi, Thời gian thực hiện, Người thực hiện.'
        row[col_map['Expected']].value = '\n'.join(lines)

    # 3. Update File Export column names in SA11-UI-05-EXPORT-HAP
    elif tc_id == 'SA11-UI-05-EXPORT-HAP':
        expected = row[col_map['Expected']].value
        lines = expected.split('\n')
        for i, line in enumerate(lines):
            if line.startswith('(iv)'):
                lines[i] = '(iv) File/Email: File .xlsx tải về thành công, hiển thị chính xác các cột như trên UI (Ngày chạy job, Tên job, Trạng thái vận hành...).'
        row[col_map['Expected']].value = '\n'.join(lines)

wb.save(file_path)
print("SA11 updated successfully.")

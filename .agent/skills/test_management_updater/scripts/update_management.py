"""
ProfiX Master Test Case Management — Sync, Dashboard & Daily Tracking
Engine: openpyxl (preserves formatting, formulas, validations)
"""

import os
import sys
import traceback
from datetime import datetime, timedelta

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─── Configuration ───────────────────────────────────────────────
TARGET_FILE = "ProfiX_Master_Test_Cases.xlsx"

# Team members for Daily Tracking
TESTERS = ["Định", "Vân", "Vân Anh", "Thương"]

# Dropdown options
STATUS_OPTIONS = '"Pass,Fail,Blocked,Doing,N/A"'
TESTER_OPTIONS = '"' + ",".join(TESTERS) + '"'

# 19 columns: shared 12 + 7 tracking
SHARED_HEADERS = [
    "TC_ID", "SC_Ref", "Reference", "Feature", "Module", "Title",
    "Type", "Priority", "Precondition", "Steps", "Expected", "Note",
]
ROUND_TRACKING = [
    "Status R1", "Tester R1", "Date R1",
]
SINGLE_TRACKING = [
    "Tester", "Ngày test", "Kết quả",
    "Bug ID", "Retest", "Retest Result", "Ghi chú TL",
]

# ─── Colors ──────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F7"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F2F2F2"

# ─── Style helpers ───────────────────────────────────────────────
def _thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _style_header(ws, row, col_count, bg=MID_BLUE, fg=WHITE):
    fill = PatternFill("solid", fgColor=bg)
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = Font(bold=True, color=fg, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()

def _set_col_widths(ws, headers):
    wide = {"Steps", "Expected", "Precondition"}
    mid  = {"Title", "Reference"}
    for j, h in enumerate(headers, 1):
        if h in wide:
            ws.column_dimensions[get_column_letter(j)].width = 50
        elif h in mid:
            ws.column_dimensions[get_column_letter(j)].width = 40
        else:
            ws.column_dimensions[get_column_letter(j)].width = 18

def _detect_format(ws):
    """Detect if sheet uses Round-based (A) or Single-test (B) format."""
    h13 = ws.cell(row=1, column=13).value
    if h13 and "Status" in str(h13):
        return "round"
    return "single"

def _status_col_letter(fmt):
    """Return the column letter of the primary status field."""
    if fmt == "round":
        return "M"  # Status R1 = col 13
    return "O"      # Kết quả = col 15

def _tester_date_refs(fmt, sheet_name):
    """Return list of (tester_col, date_col) for COUNTIFS formulas."""
    if fmt == "round":
        return [
            (f"'{sheet_name}'!N:N", f"'{sheet_name}'!O:O"),  # Tester R1, Date R1
        ]
    return [
        (f"'{sheet_name}'!M:M", f"'{sheet_name}'!N:N"),
    ]


# ═══════════════════════════════════════════════════════════════════
#  BUILD DASHBOARD
# ═══════════════════════════════════════════════════════════════════
def build_dashboard(wb, us_sheets):
    """Create or overwrite the 📊 Dashboard sheet."""
    if "📊 Dashboard" in wb.sheetnames:
        del wb["📊 Dashboard"]
    ws = wb.create_sheet("📊 Dashboard", 0)

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"].value = "PROFIX — QUALITY DASHBOARD SUMMARY"
    ws["A1"].font = Font(bold=True, size=14, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Headers at row 2
    headers = ["STT", "Module", "Total TC", "Passed", "Failed", "Blocked", "N/A", "Execution %", "Pass Rate %"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=2, column=j, value=h)
    _style_header(ws, 2, len(headers))

    # One row per US sheet
    for idx, (sname, fmt) in enumerate(us_sheets):
        r = idx + 3
        scol = _status_col_letter(fmt)

        ws.cell(row=r, column=1, value=idx + 1).border = _thin_border()
        ws.cell(row=r, column=2, value=sname).border = _thin_border()
        ws.cell(row=r, column=3, value=f"=COUNTA('{sname}'!A:A)-1").border = _thin_border()
        ws.cell(row=r, column=4, value=f'=COUNTIF(\'{sname}\'!{scol}:{scol},"Pass")').border = _thin_border()
        ws.cell(row=r, column=5, value=f'=COUNTIF(\'{sname}\'!{scol}:{scol},"Fail")').border = _thin_border()
        ws.cell(row=r, column=6, value=f'=COUNTIF(\'{sname}\'!{scol}:{scol},"Blocked")').border = _thin_border()
        ws.cell(row=r, column=7, value=f'=COUNTIF(\'{sname}\'!{scol}:{scol},"N/A")').border = _thin_border()
        ws.cell(row=r, column=8, value=f"=IF(C{r}>0,(D{r}+E{r}+F{r})/C{r},0)").border = _thin_border()
        ws.cell(row=r, column=9, value=f"=IF(C{r}>0,D{r}/C{r},0)").border = _thin_border()

        # Format percentage columns
        ws.cell(row=r, column=8).number_format = '0.0%'
        ws.cell(row=r, column=9).number_format = '0.0%'

    # TOTAL row
    total_r = len(us_sheets) + 3
    ws.cell(row=total_r, column=2, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_r, column=2).border = _thin_border()
    for c in range(3, 8):
        ws.cell(row=total_r, column=c, value=f"=SUM({get_column_letter(c)}3:{get_column_letter(c)}{total_r-1})").border = _thin_border()
    ws.cell(row=total_r, column=8, value=f"=IF(C{total_r}>0,(D{total_r}+E{total_r}+F{total_r})/C{total_r},0)").border = _thin_border()
    ws.cell(row=total_r, column=9, value=f"=IF(C{total_r}>0,D{total_r}/C{total_r},0)").border = _thin_border()
    ws.cell(row=total_r, column=8).number_format = '0.0%'
    ws.cell(row=total_r, column=9).number_format = '0.0%'

    # Column widths
    for j, w in enumerate([6, 12, 12, 10, 10, 10, 8, 14, 14], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    print(f"  ✅ Dashboard: {len(us_sheets)} modules")


# ═══════════════════════════════════════════════════════════════════
#  BUILD DAILY TRACKING
# ═══════════════════════════════════════════════════════════════════
def build_daily_tracking(wb, us_sheets, start_date=None, num_days=30):
    """Create or overwrite the 📅 Daily Tracking sheet."""
    if "📅 Daily Tracking" in wb.sheetnames:
        del wb["📅 Daily Tracking"]
    ws = wb.create_sheet("📅 Daily Tracking", 1)

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"].value = "DAILY EXECUTION TRACKING — TEAM PROFIX"
    ws["A1"].font = Font(bold=True, size=13, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Headers row 2
    headers = ["Date"] + TESTERS + ["Total/Day", "Tổng lũy kế"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=2, column=j, value=h)
    _style_header(ws, 2, len(headers))

    # Determine start date
    if start_date is None:
        start_date = datetime(2026, 4, 7)

    # Build COUNTIFS formula parts for each tester
    for day_idx in range(num_days):
        r = day_idx + 3
        dt = start_date + timedelta(days=day_idx)
        ws.cell(row=r, column=1, value=dt).border = _thin_border()
        ws.cell(row=r, column=1).number_format = "DD/MM/YYYY"

        # For each tester (columns B-E)
        for t_idx, tester in enumerate(TESTERS):
            col = t_idx + 2  # B=2, C=3, D=4, E=5
            parts = []
            for sname, fmt in us_sheets:
                for tester_ref, date_ref in _tester_date_refs(fmt, sname):
                    parts.append(f'COUNTIFS({tester_ref},{get_column_letter(col)}$2,{date_ref},$A{r})')
            formula = "=" + " + ".join(parts) if parts else "=0"
            ws.cell(row=r, column=col, value=formula).border = _thin_border()

        # Total/Day = SUM(B:E)
        ws.cell(row=r, column=6, value=f"=SUM(B{r}:E{r})").border = _thin_border()
        # Cumulative
        ws.cell(row=r, column=7, value=f"=SUM(F$3:F{r})").border = _thin_border()

    # Column widths
    ws.column_dimensions["A"].width = 14
    for j in range(2, 8):
        ws.column_dimensions[get_column_letter(j)].width = 14

    print(f"  ✅ Daily Tracking: {num_days} days, {len(TESTERS)} testers")


# ═══════════════════════════════════════════════════════════════════
#  FORMAT US SHEET
# ═══════════════════════════════════════════════════════════════════
def format_us_sheet(ws, fmt):
    """Apply formatting, validations, and formulas to a US sheet."""
    headers = SHARED_HEADERS + (ROUND_TRACKING if fmt == "round" else SINGLE_TRACKING)
    max_r = ws.max_row

    # Header styling
    _style_header(ws, 1, len(headers))
    _set_col_widths(ws, headers)

    # Data rows formatting
    for r in range(2, max_r + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            # Zebra striping
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    # Data Validations
    if fmt == "round":
        dv_status = DataValidation(type="list", formula1=STATUS_OPTIONS, allow_blank=True)
        dv_tester = DataValidation(type="list", formula1=TESTER_OPTIONS, allow_blank=True)
        dv_status.sqref = f"M2:M{max_r + 50}"
        dv_tester.sqref = f"N2:N{max_r + 50}"
        ws.add_data_validation(dv_status)
        ws.add_data_validation(dv_tester)
    else:
        dv_status = DataValidation(type="list", formula1=STATUS_OPTIONS, allow_blank=True)
        dv_tester = DataValidation(type="list", formula1=TESTER_OPTIONS, allow_blank=True)
        dv_retest = DataValidation(type="list", formula1=STATUS_OPTIONS, allow_blank=True)
        dv_status.sqref = f"O2:O{max_r + 50}"
        dv_tester.sqref = f"M2:M{max_r + 50}"
        dv_retest.sqref = f"R2:R{max_r + 50}"
        ws.add_data_validation(dv_status)
        ws.add_data_validation(dv_tester)
        ws.add_data_validation(dv_retest)


# ═══════════════════════════════════════════════════════════════════
#  MAIN: BUILD / REFRESH MASTER FILE
# ═══════════════════════════════════════════════════════════════════
def refresh_master(filepath):
    """Open existing Master file, detect US sheets, rebuild Dashboard & Daily Tracking."""
    if not os.path.exists(filepath):
        print(f"❌ Không tìm thấy file: {filepath}")
        return

    print(f"--- Đang mở file: {filepath} ---")
    wb = load_workbook(filepath)

    # Detect US sheets (skip Dashboard & Daily Tracking)
    skip = {"📊 Dashboard", "📅 Daily Tracking"}
    us_sheets = []
    for sname in wb.sheetnames:
        if sname in skip:
            continue
        ws = wb[sname]
        fmt = _detect_format(ws)
        us_sheets.append((sname, fmt))
        print(f"  📋 {sname} — format: {fmt} — {ws.max_row - 1} TCs")
        format_us_sheet(ws, fmt)

    # Rebuild Dashboard & Daily Tracking
    build_dashboard(wb, us_sheets)
    build_daily_tracking(wb, us_sheets)

    # Reorder sheets: Dashboard first, then Daily Tracking, then US sheets
    desired_order = ["📊 Dashboard", "📅 Daily Tracking"] + [s for s, _ in us_sheets]
    for idx, name in enumerate(desired_order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))

    wb.save(filepath)
    print(f"\n✅ HOÀN THÀNH: {filepath}")
    print(f"   → {len(us_sheets)} US sheets | Dashboard + Daily Tracking đã cập nhật")


if __name__ == "__main__":
    target = sys.argv[1] if len(sys.argv) > 1 else TARGET_FILE
    refresh_master(target)

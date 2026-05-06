#!/usr/bin/env python3
"""US06 TC Batch 3: Boundary Value"""
import openpyxl
from openpyxl.styles import Alignment, Border, Side
import os

FPATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "US06_TestCases.xlsx")
wb = openpyxl.load_workbook(FPATH)
ws = wb["Test Cases"]
wrap = Alignment(wrap_text=True, vertical="top")
thin = Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
FEAT = "Khai báo Biểu phí"
PRE = "1. User đăng nhập quyền Maker tại Danh mục Biểu phí\n2. Đang ở màn hình Thêm mới Biểu phí"

data = [
    ("US06-TC-034","SC-27",'QTC-01.6 – "Mã max 50 ký tự"',FEAT,"Thêm mới – Boundary",
     "Mã Biểu phí: nhập đúng 50 ký tự (biên trên) → lưu thành công","Boundary Value","Medium",PRE,
     "1. Nhập Mã = 'A' * 50 (đúng 50 ký tự)\n2. Nhập đủ trường khác hợp lệ, gán Code phí\n3. Click 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Hệ thống chấp nhận Mã 50 ký tự.\n(ii) UI: Lưu thành công, toast thành công.","[Theo QTC-01.6]"),

    ("US06-TC-035","SC-28",'QTC-01.6 – "Mã max 50 ký tự"',FEAT,"Thêm mới – Boundary",
     "Mã Biểu phí: nhập 51 ký tự (vượt biên) → FE chặn","Boundary Value","Medium",PRE,
     "1. Nhập Mã = 'A' * 51 (51 ký tự)\n2. Kiểm tra phản hồi",
     "(i) Nghiệp vụ/Logic: Hệ thống không cho nhập quá 50 ký tự.\n(ii) UI: FE chặn không cho nhập thêm ký tự thứ 51.","[Theo QTC-01.6]"),

    ("US06-TC-036","SC-64",'QTC-01.6 – "Mã biên dưới"',FEAT,"Thêm mới – Boundary",
     "Mã Biểu phí: nhập 1 ký tự (biên dưới) → chấp nhận","Boundary Value","Low",PRE,
     "1. Nhập Mã = 'A' (1 ký tự)\n2. Nhập đủ trường khác, gán Code phí\n3. Click 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Hệ thống chấp nhận Mã 1 ký tự.\n(ii) UI: Lưu thành công.","[Theo QTC-01.6]"),

    ("US06-TC-037","SC-29",'QTC-01.6 – "Tên max 50 ký tự"',FEAT,"Thêm mới – Boundary",
     "Tên Biểu phí: nhập đúng 50 ký tự → lưu thành công","Boundary Value","Medium",PRE,
     "1. Nhập Tên = 'B' * 50\n2. Nhập đủ trường khác\n3. Click 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Hệ thống chấp nhận.\n(ii) UI: Lưu thành công.","[Theo QTC-01.6]"),

    ("US06-TC-038","SC-65",'QTC-01.6 – "Tên max 50 ký tự"',FEAT,"Thêm mới – Boundary",
     "Tên Biểu phí: nhập 51 ký tự → FE chặn","Boundary Value","Medium",PRE,
     "1. Nhập Tên = 'B' * 51\n2. Kiểm tra",
     "(i) Nghiệp vụ/Logic: Không cho nhập quá 50.\n(ii) UI: FE chặn ký tự thứ 51.","[Theo QTC-01.6]"),

    ("US06-TC-039","SC-30",'BA QA-02.2 – "Link iDoc max 500 ký tự"',FEAT,"Thêm mới – Boundary",
     "Link iDoc: nhập đúng 500 ký tự → lưu thành công","Boundary Value","Low",PRE,
     "1. Nhập Link iDoc = 'https://idoc.bank.vn/' + 'x'*479 (tổng 500 ký tự)\n2. Click 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Chấp nhận URL 500 ký tự.\n(ii) UI: Lưu thành công.","[BA QA-02.2]"),

    ("US06-TC-040","SC-31",'BA QA-02.2 – "Link iDoc max 500 ký tự"',FEAT,"Thêm mới – Boundary",
     "Link iDoc: nhập 501 ký tự → FE chặn","Boundary Value","Low",PRE,
     "1. Nhập Link iDoc = 'https://idoc.bank.vn/' + 'x'*480 (tổng 501)\n2. Kiểm tra",
     "(i) Nghiệp vụ/Logic: Không cho nhập quá 500.\n(ii) UI: FE chặn ký tự thứ 501.","[BA QA-02.2]"),

    ("US06-TC-041","SC-66",'QTC-01.6 – "Số VB max 50 ký tự"',FEAT,"Thêm mới – Boundary",
     "Số văn bản: nhập đúng 50 ký tự → chấp nhận","Boundary Value","Low",PRE,
     "1. Nhập Số văn bản = 'QD-' + 'X'*47 (50 ký tự)\n2. Click 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Chấp nhận.\n(ii) UI: Lưu thành công.","[Theo QTC-01.6]"),

    ("US06-TC-042","SC-67",'QTC-01.6 – "Tên VB max 50 ký tự"',FEAT,"Thêm mới – Boundary",
     "Tên văn bản: nhập đúng 50 ký tự → chấp nhận","Boundary Value","Low",PRE,
     "1. Nhập Tên văn bản = 'T' * 50\n2. Click 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Chấp nhận.\n(ii) UI: Lưu thành công.","[Theo QTC-01.6]"),

    ("US06-TC-043","SC-32",'Bảng mô tả trường R8 – ">= Ngày hệ thống"',FEAT,"Thêm mới – Boundary",
     "Ngày hiệu lực = Ngày hệ thống (biên dưới) → chấp nhận","Boundary Value","Medium",
     PRE+"\n3. Ngày HT = 06/05/2026",
     "1. Chọn Ngày hiệu lực = 06/05/2026 (= Ngày HT)\n2. Nhập đủ trường khác\n3. Click 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Chấp nhận vì >= Ngày HT.\n(ii) UI: Lưu thành công.",""),

    ("US06-TC-044","SC-33",'Bảng mô tả trường R9 – ">= Ngày hiệu lực"',FEAT,"Thêm mới – Boundary",
     "Ngày HHL = Ngày hiệu lực (biên dưới) → chấp nhận","Boundary Value","Medium",PRE,
     "1. Chọn Ngày HL = 01/07/2026\n2. Chọn Ngày HHL = 01/07/2026 (= Ngày HL)\n3. Click 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Chấp nhận vì HHL >= HL.\n(ii) UI: Lưu thành công.",""),

    ("US06-TC-045","SC-34",'Bảng mô tả trường R7 – "<= Ngày hệ thống"',FEAT,"Thêm mới – Boundary",
     "Ngày ban hành = Ngày hệ thống (biên trên) → chấp nhận","Boundary Value","Medium",
     PRE+"\n3. Ngày HT = 06/05/2026",
     "1. Chọn Ngày ban hành = 06/05/2026 (= Ngày HT)\n2. Click 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Chấp nhận vì <= Ngày HT.\n(ii) UI: Lưu thành công.",""),

    ("US06-TC-046","SC-35",'BA QA-01.8 – "ít nhất 1 Code phí"',FEAT,"Thêm mới – Boundary",
     "Gán đúng 1 Code phí (min) → lưu thành công","Boundary Value","Medium",PRE,
     "1. Nhập đủ thông tin chung\n2. Gán đúng 1 Code phí CP001\n3. Click 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Chấp nhận Biểu phí với 1 Code phí.\n(ii) UI: Lưu thành công.","[BA QA-01.8]"),

    ("US06-TC-047","SC-74",'Mục "Chỉnh sửa" P62 – "> Ngày HT, biên ="',FEAT,"Chỉnh sửa – Boundary",
     "Đang hiệu lực: Ngày HHL mới = Ngày HT (biên =, ràng buộc >) → FE chặn","Boundary Value","High",
     "1. User đăng nhập quyền Maker\n2. BP-013 Đang hiệu lực (HL=01/01/2026)\n3. Ngày HT = 06/05/2026",
     "1. Click 'Chỉnh sửa' tại BP-013\n2. Sửa Ngày HHL = 06/05/2026 (= Ngày HT)\n3. Click 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Không cho phép vì ràng buộc > (strict), = không đủ.\n(ii) UI: Hiển thị lỗi validation.",""),

    ("US06-TC-048","SC-75",'Mục "Chỉnh sửa" P62 – "> Ngày HL, biên ="',FEAT,"Chỉnh sửa – Boundary",
     "Đang hiệu lực: Ngày HHL mới = Ngày HL (biên =, ràng buộc >) → FE chặn","Boundary Value","High",
     "1. User đăng nhập quyền Maker\n2. BP-014 Đang hiệu lực (HL=01/03/2026)\n3. Ngày HT = 06/05/2026",
     "1. Click 'Chỉnh sửa' tại BP-014\n2. Sửa Ngày HHL = 01/03/2026 (= Ngày HL)\n3. Click 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Không cho phép vì ràng buộc > (strict).\n(ii) UI: Hiển thị lỗi validation.",""),
]

start = ws.max_row + 1
for i, row in enumerate(data):
    for j, val in enumerate(row):
        c = ws.cell(row=start+i, column=j+1, value=val)
        c.alignment = wrap; c.border = thin
wb.save(FPATH)
print(f"✅ Batch 3 (Boundary): {len(data)} TC → {FPATH}")

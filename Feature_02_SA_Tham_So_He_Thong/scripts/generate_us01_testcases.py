"""
Script sinh Test Case cho US01: Khai báo Danh mục Sản phẩm, Dịch vụ (Nghiệp vụ & SPDV chi tiết)
Module: SA01 | Chuẩn: Enterprise Level B2 | Workflow: generate_testcases_from_requirements
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
import os

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'output')
os.makedirs(OUTPUT_DIR, exist_ok=True)
OUTPUT_FILE = os.path.join(OUTPUT_DIR, 'US01_Test_Cases.xlsx')

# ==============================================================================
# DATA: 21 cột chuẩn Enterprise
# A=TC_ID | B=BR_Ref | C=URD_Ref | D=Module | E=Feature | F=Title
# G=Type | H=Category | I=Priority | J=Precondition | K=Steps | L=Expected
# M=Trace_ID | N=Note | O-U: Cột thực thi (để trống)
# ==============================================================================

TC = []

# ─────────────────────────────────────────────────────────────────────────────
# NHÓM A: TC-BR – Nghiệp vụ (Khai báo Nghiệp vụ cấp 1)
# ─────────────────────────────────────────────────────────────────────────────

TC.append({
    'TC_ID': 'SA01-BR-HAP-001',
    'BR_Ref': 'LOG-NV-THEMMOI-HAPPY',
    'URD_Ref': 'Mục "Yêu cầu nghiệp vụ > Khai báo Nghiệp vụ" – Luồng Thêm mới; Bảng Lưu đồ Thêm mới Nghiệp vụ',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Thêm mới Nghiệp vụ] Happy Path – Khai báo thành công với đầy đủ thông tin hợp lệ',
    'Type': 'TC-BR',
    'Category': 'Positive',
    'Priority': 'High',
    'Precondition': (
        '1. Người dùng có vai trò Maker, đã đăng nhập hệ thống.\n'
        '2. Người dùng có quyền truy cập chức năng: Tham số >> Danh mục sản phẩm dịch vụ và Code phí.\n'
        '3. Số lượng Nghiệp vụ hiện có trong hệ thống < 99 (chưa tràn mã tự tăng).'
    ),
    'Steps': (
        '1. Truy cập: Tham số >> Danh mục sản phẩm dịch vụ và Code phí.\n'
        '2. Nhấn nút "+ Thêm mới" trên thanh header.\n'
        '3. Hệ thống hiển thị màn hình Thêm mới Nghiệp vụ; kiểm tra trường "Cấp sản phẩm dịch vụ" tự động hiển thị = 1.\n'
        '4. Nhập "Tên Nghiệp vụ" = "Tài khoản tiết kiệm" (hợp lệ).\n'
        '5. Nhập "Mô tả" = "Nghiệp vụ quản lý tài khoản tiết kiệm".\n'
        '6. Chọn "Ngày hiệu lực" = ngày T+1 (tương lai).\n'
        '7. Chọn "Ngày hết hiệu lực" = ngày T+365.\n'
        '8. Nhấn nút "Xác nhận".'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Hệ thống tạo thành công bản ghi Nghiệp vụ mới; Mã Nghiệp vụ được hệ thống tự động sinh (số 2 ký tự tự tăng, duy nhất, không trùng).\n'
        '(ii) UI: Hiển thị thông báo "Thêm mới thành công Nghiệp vụ"; form đóng lại và quay về màn hình Danh sách.\n'
        '(iii) Trạng thái bản ghi: Bản ghi Nghiệp vụ vừa tạo có Trạng thái duyệt = "Chờ duyệt"; Trạng thái hoạt động = "Không hoạt động" (vì Ngày hiệu lực = T+1 > Ngày T).\n'
        '(iv) Output: Bản ghi xuất hiện tại màn hình "Tác vụ Pending của tôi" của Maker và màn hình "Tác vụ chờ duyệt" của Checker.'
    ),
    'Trace_ID': 'SA01-NV-THEMMOI-HAP',
    'Note': '',
})

TC.append({
    'TC_ID': 'SA01-BR-HAP-002',
    'BR_Ref': 'LOG-NV-STATUS-ACTIVE',
    'URD_Ref': 'Mục "Yêu cầu nghiệp vụ > Khai báo Nghiệp vụ" – đoạn "Nếu Ngày hiện tại = Ngày hiệu lực => Trạng thái = Hoạt động"',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Thêm mới Nghiệp vụ] Trạng thái hoạt động = "Hoạt động" khi Ngày hiệu lực = Ngày T (ngày hiện tại)',
    'Type': 'TC-BR',
    'Category': 'Positive',
    'Priority': 'High',
    'Precondition': (
        '1. Người dùng có vai trò Maker, đã đăng nhập hệ thống.\n'
        '2. Ngày hiện tại của hệ thống là ngày T.\n'
        '3. Có quyền truy cập chức năng: Tham số >> Danh mục sản phẩm dịch vụ và Code phí.'
    ),
    'Steps': (
        '1. Truy cập chức năng Thêm mới Nghiệp vụ.\n'
        '2. Nhập "Tên Nghiệp vụ" và "Mô tả" hợp lệ.\n'
        '3. Chọn "Ngày hiệu lực" = Ngày T (ngày hôm nay).\n'
        '4. Chọn "Ngày hết hiệu lực" = T+365.\n'
        '5. Nhấn "Xác nhận" → Bản ghi được Checker duyệt thành công.\n'
        '6. Kiểm tra Trạng thái hoạt động của bản ghi Nghiệp vụ vừa được duyệt.'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Sau khi duyệt thành công, vì Ngày T = Ngày hiệu lực, hệ thống xác định Trạng thái hoạt động = "Hoạt động".\n'
        '(ii) UI: Cột "Trạng thái hoạt động" trên lưới hiển thị = "Hoạt động".\n'
        '(iii) Trạng thái bản ghi: Trạng thái duyệt = "Đã duyệt"; Trạng thái hoạt động = "Hoạt động".\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-NV-STATUS-ACTIVE',
    'Note': '',
})

TC.append({
    'TC_ID': 'SA01-BR-HAP-003',
    'BR_Ref': 'LOG-NV-STATUS-INACTIVE',
    'URD_Ref': 'Mục "Yêu cầu nghiệp vụ > Khai báo Nghiệp vụ" – đoạn "Nếu Ngày T < Ngày hiệu lực => Trạng thái = Không hoạt động"',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Thêm mới Nghiệp vụ] Trạng thái hoạt động = "Không hoạt động" khi Ngày hiệu lực > Ngày T',
    'Type': 'TC-BR',
    'Category': 'Positive',
    'Priority': 'Medium',
    'Precondition': (
        '1. Người dùng có vai trò Maker, đã đăng nhập hệ thống.\n'
        '2. Ngày hiện tại của hệ thống là ngày T.\n'
        '3. Có quyền truy cập chức năng: Tham số >> Danh mục sản phẩm dịch vụ và Code phí.'
    ),
    'Steps': (
        '1. Truy cập chức năng Thêm mới Nghiệp vụ.\n'
        '2. Nhập "Tên Nghiệp vụ" và "Mô tả" hợp lệ.\n'
        '3. Chọn "Ngày hiệu lực" = T+10 (tương lai).\n'
        '4. Chọn "Ngày hết hiệu lực" = T+365.\n'
        '5. Nhấn "Xác nhận" → Bản ghi được Checker duyệt thành công.\n'
        '6. Kiểm tra Trạng thái hoạt động của bản ghi.'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Sau khi duyệt thành công, vì Ngày T < Ngày hiệu lực (T+10), hệ thống xác định Trạng thái hoạt động = "Không hoạt động".\n'
        '(ii) UI: Cột "Trạng thái hoạt động" trên lưới hiển thị = "Không hoạt động".\n'
        '(iii) Trạng thái bản ghi: Trạng thái duyệt = "Đã duyệt"; Trạng thái hoạt động = "Không hoạt động".\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-NV-STATUS-INACTIVE',
    'Note': '',
})

TC.append({
    'TC_ID': 'SA01-BR-HAP-004',
    'BR_Ref': 'LOG-NV-BATCH-AUTO-ACTIVE',
    'URD_Ref': 'Mục "Yêu cầu nghiệp vụ > Khai báo Nghiệp vụ" – đoạn "Hằng ngày, tại thời điểm đầu ngày, hệ thống tự động kiểm tra... Trường hợp ngày T = Ngày hiệu lực và trạng thái Không hoạt động => cập nhật = Hoạt động"',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Batch Job Nghiệp vụ] Hệ thống tự động cập nhật Trạng thái = "Hoạt động" khi đến Ngày hiệu lực',
    'Type': 'TC-BR',
    'Category': 'Positive',
    'Priority': 'High',
    'Precondition': (
        '1. Tồn tại bản ghi Nghiệp vụ đã được duyệt, Trạng thái hoạt động = "Không hoạt động", Ngày hiệu lực = ngày mai (T+1).\n'
        '2. Batch Job tự động của hệ thống được cấu hình chạy đầu ngày.\n'
        '3. Quyền kiểm tra: QA/Admin có thể xem Trạng thái bản ghi trên lưới.'
    ),
    'Steps': (
        '1. Xác nhận bản ghi Nghiệp vụ có Ngày hiệu lực = ngày mai và Trạng thái hoạt động = "Không hoạt động".\n'
        '2. Chờ hoặc trigger Batch Job kiểm tra ngày hiệu lực vào đầu ngày hôm sau (ngày T = Ngày hiệu lực).\n'
        '3. Sau khi Batch Job chạy xong, vào màn hình Danh sách Sản phẩm dịch vụ.\n'
        '4. Kiểm tra Trạng thái hoạt động của bản ghi Nghiệp vụ vừa đến Ngày hiệu lực.'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Batch Job cập nhật Trạng thái hoạt động của bản ghi từ "Không hoạt động" → "Hoạt động".\n'
        '(ii) UI: Cột "Trạng thái hoạt động" trên lưới hiển thị = "Hoạt động".\n'
        '(iii) Trạng thái bản ghi: Trạng thái hoạt động = "Hoạt động"; Ngày hiệu lực và các trường khác không thay đổi.\n'
        '(iv) Output: Không có output bổ sung. (Lưu ý: Thời điểm chạy Batch Job tùy cấu hình triển khai thực tế – xem Q07 đã Drop)'
    ),
    'Trace_ID': 'SA01-NV-BATCH-ACTIVE',
    'Note': 'Q07 của BA đã Drop – thời điểm cụ thể Batch Job phụ thuộc triển khai thực tế. QA cần trigger thủ công trong môi trường test.',
})

TC.append({
    'TC_ID': 'SA01-BR-HAP-005',
    'BR_Ref': 'LOG-NV-BATCH-AUTO-INACTIVE',
    'URD_Ref': 'Mục "Yêu cầu nghiệp vụ > Khai báo Nghiệp vụ" – đoạn "Trường hợp ngày T > Ngày hết hiệu lực và trạng thái Hoạt động => cập nhật = Không hoạt động"',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Batch Job Nghiệp vụ] Hệ thống tự động cập nhật Trạng thái = "Không hoạt động" khi quá Ngày hết hiệu lực',
    'Type': 'TC-BR',
    'Category': 'Positive',
    'Priority': 'High',
    'Precondition': (
        '1. Tồn tại bản ghi Nghiệp vụ đã được duyệt, Trạng thái hoạt động = "Hoạt động", Ngày hết hiệu lực = hôm qua (T-1).\n'
        '2. Batch Job tự động của hệ thống được cấu hình chạy đầu ngày.'
    ),
    'Steps': (
        '1. Xác nhận bản ghi Nghiệp vụ có Ngày hết hiệu lực = (T-1) và Trạng thái hoạt động = "Hoạt động".\n'
        '2. Trigger Batch Job kiểm tra ngày hiệu lực vào đầu ngày T.\n'
        '3. Sau khi Batch Job chạy xong, vào màn hình Danh sách Sản phẩm dịch vụ.\n'
        '4. Kiểm tra Trạng thái hoạt động của bản ghi.'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Batch Job phát hiện ngày T > Ngày hết hiệu lực (T-1) → cập nhật Trạng thái hoạt động = "Không hoạt động".\n'
        '(ii) UI: Cột "Trạng thái hoạt động" trên lưới hiển thị = "Không hoạt động".\n'
        '(iii) Trạng thái bản ghi: Trạng thái hoạt động = "Không hoạt động"; các trường khác không thay đổi.\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-NV-BATCH-INACTIVE',
    'Note': '',
})

TC.append({
    'TC_ID': 'SA01-BR-HAP-006',
    'BR_Ref': 'LOG-NV-AUTOCODE',
    'URD_Ref': 'Mục "Yêu cầu nghiệp vụ > Khai báo Nghiệp vụ" – đoạn "Mã nghiệp vụ do hệ thống tự động sinh, là số duy nhất 2 ký tự số tự tăng"',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Thêm mới Nghiệp vụ] Mã Nghiệp vụ được hệ thống tự động sinh – số thứ tự tự tăng, không trùng',
    'Type': 'TC-BR',
    'Category': 'Positive',
    'Priority': 'High',
    'Precondition': (
        '1. Người dùng có vai trò Maker, đã đăng nhập hệ thống.\n'
        '2. Có ít nhất 1 Nghiệp vụ đã tồn tại trong hệ thống (ví dụ: Mã = "01").\n'
        '3. Trường "Mã" không cho phép nhập liệu (hệ thống tự sinh).'
    ),
    'Steps': (
        '1. Truy cập chức năng Thêm mới Nghiệp vụ.\n'
        '2. Kiểm tra trường "Mã" trên form: quan sát trạng thái hiển thị (read-only / disabled).\n'
        '3. Nhập đầy đủ thông tin hợp lệ (Tên, Mô tả, Ngày hiệu lực, Ngày hết hiệu lực).\n'
        '4. Nhấn "Xác nhận".\n'
        '5. Sau khi được duyệt, quan sát giá trị Mã Nghiệp vụ được sinh ra.'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Mã Nghiệp vụ được hệ thống tự sinh = số kế tiếp của Mã lớn nhất hiện có (ví dụ: Mã hiện lớn nhất = "01" → Mã mới = "02"); Mã phải là duy nhất, không trùng với bất kỳ bản ghi nào đã tồn tại.\n'
        '(ii) UI: Trường "Mã" trên form ở trạng thái không cho phép nhập (read-only/disabled, theo Q01 BA đã xác nhận không ẩn trường do màn hình dùng chung); Mã hiển thị sau khi lưu thành công.\n'
        '(iii) Trạng thái bản ghi: Trạng thái = "Chờ duyệt".\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-NV-AUTOCODE',
    'Note': 'Q01: BA xác nhận trường Mã không ẩn, hiển thị read-only vì màn hình dùng chung nhiều mode.',
})

TC.append({
    'TC_ID': 'SA01-BR-NEG-001',
    'BR_Ref': 'LOG-NV-VALIDATE-REQUIRED',
    'URD_Ref': 'Bảng "Mô tả chi tiết các trường" (Table 2) – cột "Bắt buộc" ký hiệu ★ cho Tên Nghiệp vụ, Mô tả, Ngày hiệu lực, Ngày hết hiệu lực',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Thêm mới Nghiệp vụ] Negative – Nhấn Xác nhận khi bỏ trống trường bắt buộc "Tên Nghiệp vụ"',
    'Type': 'TC-BR',
    'Category': 'Negative',
    'Priority': 'High',
    'Precondition': (
        '1. Người dùng có vai trò Maker, đã đăng nhập hệ thống.\n'
        '2. Màn hình Thêm mới Nghiệp vụ đang mở.'
    ),
    'Steps': (
        '1. Mở màn hình Thêm mới Nghiệp vụ.\n'
        '2. Để trống trường "Tên Nghiệp vụ" (★ bắt buộc).\n'
        '3. Nhập đầy đủ các trường còn lại (Mô tả, Ngày hiệu lực, Ngày hết hiệu lực) với giá trị hợp lệ.\n'
        '4. Nhấn nút "Xác nhận".'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Hệ thống từ chối lưu dữ liệu; bản ghi không được tạo.\n'
        '(ii) UI: Hiển thị thông báo lỗi validation tại trường "Tên Nghiệp vụ" (ví dụ: "Trường này là bắt buộc"); form vẫn mở để user chỉnh sửa (theo Q09 BA sẽ update flowchart bổ sung mũi tên quay lại).\n'
        '(iii) Trạng thái bản ghi: Không có bản ghi nào được tạo trong DB.\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-NV-VALIDATE-REQ',
    'Note': '',
})

TC.append({
    'TC_ID': 'SA01-BR-NEG-002',
    'BR_Ref': 'LOG-NV-VALIDATE-REQUIRED',
    'URD_Ref': 'Bảng "Mô tả chi tiết các trường" (Table 2) – cột "Bắt buộc" ký hiệu ★ cho Mô tả, Ngày hiệu lực, Ngày hết hiệu lực',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Thêm mới Nghiệp vụ] Negative – Nhấn Xác nhận khi bỏ trống trường bắt buộc "Mô tả"',
    'Type': 'TC-BR',
    'Category': 'Negative',
    'Priority': 'Medium',
    'Precondition': (
        '1. Người dùng có vai trò Maker, đã đăng nhập hệ thống.\n'
        '2. Màn hình Thêm mới Nghiệp vụ đang mở.'
    ),
    'Steps': (
        '1. Mở màn hình Thêm mới Nghiệp vụ.\n'
        '2. Nhập "Tên Nghiệp vụ" hợp lệ.\n'
        '3. Để trống trường "Mô tả" (★ bắt buộc).\n'
        '4. Chọn Ngày hiệu lực, Ngày hết hiệu lực hợp lệ.\n'
        '5. Nhấn nút "Xác nhận".'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Hệ thống từ chối lưu dữ liệu; bản ghi không được tạo.\n'
        '(ii) UI: Hiển thị thông báo lỗi validation tại trường "Mô tả"; form vẫn mở.\n'
        '(iii) Trạng thái bản ghi: Không có bản ghi nào được tạo trong DB.\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-NV-VALIDATE-REQ',
    'Note': 'Đã cover logic tương tự ở SA01-BR-NEG-001 nhưng tách riêng để phủ từng trường bắt buộc.',
})

TC.append({
    'TC_ID': 'SA01-BR-NEG-003',
    'BR_Ref': 'LOG-NV-VALIDATE-DATE-PAST',
    'URD_Ref': 'Bảng "Mô tả chi tiết các trường" (Table 2) – dòng "Ngày hiệu lực", ràng buộc "Không chọn ngày quá khứ"',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Thêm mới Nghiệp vụ] Negative – Chọn Ngày hiệu lực là ngày quá khứ (T-1)',
    'Type': 'TC-BR',
    'Category': 'Negative',
    'Priority': 'High',
    'Precondition': (
        '1. Người dùng có vai trò Maker, đã đăng nhập hệ thống.\n'
        '2. Màn hình Thêm mới Nghiệp vụ đang mở.\n'
        '3. Ngày hiện tại của hệ thống là ngày T.'
    ),
    'Steps': (
        '1. Mở màn hình Thêm mới Nghiệp vụ.\n'
        '2. Nhập đầy đủ Tên Nghiệp vụ, Mô tả hợp lệ.\n'
        '3. Nhấn vào trường "Ngày hiệu lực" → Calendar mở ra.\n'
        '4. Cố gắng chọn ngày T-1 (ngày hôm qua) trên calendar.\n'
        '5. Nhấn nút "Xác nhận".'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Hệ thống không cho phép chọn ngày quá khứ; bản ghi không được tạo.\n'
        '(ii) UI: Calendar disabled (grayout) các ngày quá khứ; nếu user bypass được, hệ thống hiển thị thông báo lỗi validation "Ngày hiệu lực không được là ngày quá khứ".\n'
        '(iii) Trạng thái bản ghi: Không có bản ghi nào được tạo.\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-NV-DATE-PAST',
    'Note': '',
})

TC.append({
    'TC_ID': 'SA01-BR-NEG-004',
    'BR_Ref': 'LOG-NV-VALIDATE-DATE-ORDER',
    'URD_Ref': 'Bảng "Mô tả chi tiết các trường" (Table 2) – dòng "Ngày hết hiệu lực", ràng buộc "Ngày hết hiệu lực >= Ngày hiệu lực"',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Thêm mới Nghiệp vụ] Negative – Ngày hết hiệu lực nhỏ hơn Ngày hiệu lực',
    'Type': 'TC-BR',
    'Category': 'Negative',
    'Priority': 'High',
    'Precondition': (
        '1. Người dùng có vai trò Maker, đã đăng nhập hệ thống.\n'
        '2. Màn hình Thêm mới Nghiệp vụ đang mở.'
    ),
    'Steps': (
        '1. Mở màn hình Thêm mới Nghiệp vụ.\n'
        '2. Nhập đầy đủ Tên Nghiệp vụ, Mô tả hợp lệ.\n'
        '3. Chọn "Ngày hiệu lực" = T+10.\n'
        '4. Chọn "Ngày hết hiệu lực" = T+5 (nhỏ hơn Ngày hiệu lực).\n'
        '5. Nhấn nút "Xác nhận".'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Hệ thống từ chối lưu; bản ghi không được tạo.\n'
        '(ii) UI: Hiển thị thông báo lỗi validation "Ngày hết hiệu lực phải lớn hơn hoặc bằng Ngày hiệu lực"; form vẫn mở.\n'
        '(iii) Trạng thái bản ghi: Không có bản ghi nào được tạo.\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-NV-DATE-ORDER',
    'Note': '',
})

TC.append({
    'TC_ID': 'SA01-BR-NEG-005',
    'BR_Ref': 'LOG-NV-AUTOCODE-OVERFLOW',
    'URD_Ref': 'Mục "Yêu cầu nghiệp vụ > Khai báo Nghiệp vụ" – đoạn "Trường hợp số tự tăng >99 hệ thống báo lỗi và không cho phép lưu"',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Thêm mới Nghiệp vụ] Boundary – Tràn Mã Nghiệp vụ khi đã có đủ 99 bản ghi Nghiệp vụ',
    'Type': 'TC-BR',
    'Category': 'Negative',
    'Priority': 'Medium',
    'Precondition': (
        '1. Người dùng có vai trò Maker, đã đăng nhập hệ thống.\n'
        '2. Hệ thống đã có đủ 99 bản ghi Nghiệp vụ (số tự tăng đã đạt 99).\n'
        '3. Có quyền truy cập chức năng Thêm mới Nghiệp vụ.'
    ),
    'Steps': (
        '1. Truy cập chức năng Thêm mới Nghiệp vụ.\n'
        '2. Nhập đầy đủ thông tin hợp lệ (Tên, Mô tả, Ngày hiệu lực, Ngày hết hiệu lực).\n'
        '3. Nhấn nút "Xác nhận".'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Hệ thống phát hiện số tự tăng đã đạt giới hạn 99; từ chối lưu bản ghi mới.\n'
        '(ii) UI: Hiển thị thông báo lỗi (nội dung cụ thể tham chiếu Phụ lục mã lỗi – theo Q05 BA sẽ cập nhật lại).\n'
        '(iii) Trạng thái bản ghi: Không có bản ghi mới được tạo.\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-NV-AUTOCODE-OVF',
    'Note': 'Q05: BA xác nhận BA sẽ cập nhật lại lưu đồ không tham chiếu đến mã lỗi cụ thể. QA cần xác nhận thông báo lỗi khi có version mới.',
})

TC.append({
    'TC_ID': 'SA01-BR-NEG-006',
    'BR_Ref': 'LOG-NV-CONCURRENCY',
    'URD_Ref': 'Mục "Yêu cầu nghiệp vụ > Khai báo Nghiệp vụ" – Q08 BA confirm: Tên nghiệp vụ được phép trùng do đã có mã nghiệp vụ unique',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Thêm mới Nghiệp vụ] Concurrent – Hai Maker tạo Nghiệp vụ cùng thời điểm; tên trùng nhau vẫn được phép',
    'Type': 'TC-BR',
    'Category': 'Positive',
    'Priority': 'Medium',
    'Precondition': (
        '1. Có 2 tài khoản Maker A và Maker B đã đăng nhập hệ thống.\n'
        '2. Cả 2 Maker cùng mở màn hình Thêm mới Nghiệp vụ tại cùng thời điểm.'
    ),
    'Steps': (
        '1. Maker A nhập Tên Nghiệp vụ = "Nghiệp vụ TEST TRÙNG", Mô tả, Ngày hiệu lực, Ngày hết hiệu lực hợp lệ.\n'
        '2. Maker B nhập Tên Nghiệp vụ = "Nghiệp vụ TEST TRÙNG" (cùng tên với Maker A), Mô tả, Ngày hiệu lực, Ngày hết hiệu lực hợp lệ.\n'
        '3. Maker A nhấn "Xác nhận" trước.\n'
        '4. Maker B nhấn "Xác nhận" ngay sau.'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Cả 2 bản ghi đều được tạo thành công (tên trùng là hợp lệ vì Mã hệ thống tự sinh là unique); Maker B nhận được Mã khác với Maker A.\n'
        '(ii) UI: Cả 2 Maker nhận thông báo "Thêm mới thành công".\n'
        '(iii) Trạng thái bản ghi: 2 bản ghi với Mã khác nhau, cùng Trạng thái = "Chờ duyệt".\n'
        '(iv) Output: Cả 2 bản ghi xuất hiện ở màn hình "Tác vụ Pending của tôi" tương ứng.'
    ),
    'Trace_ID': 'SA01-NV-CONCURRENT',
    'Note': 'Q08: BA confirm tên nghiệp vụ được phép trùng do mã là unique.',
})

# ─────────────────────────────────────────────────────────────────────────────
# NHÓM A: TC-BR – Chỉnh sửa Nghiệp vụ
# ─────────────────────────────────────────────────────────────────────────────

TC.append({
    'TC_ID': 'SA01-BR-HAP-007',
    'BR_Ref': 'LOG-NV-EDIT-HAPPY',
    'URD_Ref': 'Mục "Yêu cầu nghiệp vụ > Khai báo Nghiệp vụ" – đoạn "Bản ghi Nghiệp vụ sau khi Duyệt thành công... khi người dùng chỉnh sửa, hệ thống chỉ cho phép chỉnh sửa Ngày hiệu lực, Ngày hết hiệu lực"',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Chỉnh sửa Nghiệp vụ] Happy Path – Cập nhật Ngày hiệu lực và Ngày hết hiệu lực thành công',
    'Type': 'TC-BR',
    'Category': 'Positive',
    'Priority': 'High',
    'Precondition': (
        '1. Người dùng có vai trò Maker, đã đăng nhập hệ thống.\n'
        '2. Tồn tại bản ghi Nghiệp vụ đã được duyệt (Trạng thái = "Đã duyệt").\n'
        '3. Bản ghi Nghiệp vụ không có yêu cầu chỉnh sửa nào đang ở trạng thái "Chờ duyệt".'
    ),
    'Steps': (
        '1. Truy cập màn hình Danh sách sản phẩm dịch vụ.\n'
        '2. Tìm bản ghi Nghiệp vụ đã duyệt; nhấn nút "Chỉnh sửa".\n'
        '3. Kiểm tra các trường Tên Nghiệp vụ, Mã, Mô tả ở trạng thái read-only/disabled.\n'
        '4. Thay đổi "Ngày hiệu lực" sang giá trị hợp lệ mới.\n'
        '5. Thay đổi "Ngày hết hiệu lực" sang giá trị hợp lệ mới (>= Ngày hiệu lực mới).\n'
        '6. Nhấn "Xác nhận".'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Hệ thống tạo bản ghi yêu cầu sửa Nghiệp vụ mới ở trạng thái "Chờ duyệt" với thông tin ngày vừa cập nhật; bản ghi gốc vẫn có hiệu lực cho đến khi yêu cầu được duyệt.\n'
        '(ii) UI: Hiển thị thông báo "Tạo yêu cầu Sửa thành công Nghiệp vụ"; quay về màn hình Danh sách.\n'
        '(iii) Trạng thái bản ghi: Yêu cầu chỉnh sửa có Trạng thái duyệt = "Chờ duyệt"; các trường Tên, Mã, Mô tả không thay đổi.\n'
        '(iv) Output: Bản ghi yêu cầu sửa xuất hiện tại màn hình "Tác vụ Pending của tôi" của Maker và "Tác vụ chờ duyệt" của Checker.'
    ),
    'Trace_ID': 'SA01-NV-EDIT-HAP',
    'Note': 'Q14: BA xác nhận màn hình chỉnh sửa hiển thị tương tự màn hình xem chi tiết, các trường không được sửa hiển thị read-only.',
})

TC.append({
    'TC_ID': 'SA01-BR-NEG-007',
    'BR_Ref': 'LOG-NV-EDIT-NOCHANGE',
    'URD_Ref': 'Lưu đồ Chỉnh sửa Nghiệp vụ (image2) – điều kiện Backend "Ngày hết hiệu lực/Ngày hiệu lực có thay đổi không?" nhánh "No" → báo lỗi (PR.01.03)',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Chỉnh sửa Nghiệp vụ] Negative – Nhấn Xác nhận mà không thay đổi giá trị Ngày hiệu lực / Ngày hết hiệu lực',
    'Type': 'TC-BR',
    'Category': 'Negative',
    'Priority': 'High',
    'Precondition': (
        '1. Người dùng có vai trò Maker, đã đăng nhập hệ thống.\n'
        '2. Tồn tại bản ghi Nghiệp vụ đã được duyệt (Trạng thái = "Đã duyệt"), ví dụ: Ngày hiệu lực = T+5, Ngày hết hiệu lực = T+365.'
    ),
    'Steps': (
        '1. Mở màn hình Chỉnh sửa bản ghi Nghiệp vụ đã duyệt.\n'
        '2. KHÔNG thay đổi bất kỳ giá trị nào (Ngày hiệu lực và Ngày hết hiệu lực giữ nguyên).\n'
        '3. Nhấn nút "Xác nhận".'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Hệ thống phát hiện không có thay đổi; từ chối tạo yêu cầu chỉnh sửa mới để tránh tạo bản ghi Chờ duyệt không cần thiết (theo Q04 BA Drop – xác nhận logic này là đúng).\n'
        '(ii) UI: Hiển thị thông báo (theo mã tham số hệ thống – Q04 Drop) ngăn user tạo yêu cầu khi không có thay đổi.\n'
        '(iii) Trạng thái bản ghi: Không có bản ghi yêu cầu sửa nào được tạo.\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-NV-EDIT-NOCHANGE',
    'Note': 'Q04 BA Drop – xác nhận behavior: không thay đổi = không tạo yêu cầu phê duyệt mới, tránh luồng phê duyệt thừa.',
})

# ─────────────────────────────────────────────────────────────────────────────
# NHÓM A: TC-BR – SPDV chi tiết (Thêm mới)
# ─────────────────────────────────────────────────────────────────────────────

TC.append({
    'TC_ID': 'SA01-BR-HAP-008',
    'BR_Ref': 'LOG-SPDV-THEMMOI-HAPPY',
    'URD_Ref': 'Mục "Yêu cầu nghiệp vụ > Khai báo SPDV chi tiết" – Luồng Thêm mới; Bảng Lưu đồ Thêm mới SPDV chi tiết',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Thêm mới SPDV chi tiết] Happy Path – Khai báo thành công SPDV chi tiết thuộc Nghiệp vụ cha',
    'Type': 'TC-BR',
    'Category': 'Positive',
    'Priority': 'High',
    'Precondition': (
        '1. Người dùng có vai trò Maker, đã đăng nhập hệ thống.\n'
        '2. Tồn tại bản ghi Nghiệp vụ cha đã được duyệt, Trạng thái = "Đã duyệt" (ví dụ: Mã "06", Tên "Tài khoản", Ngày hiệu lực = T+1, Ngày hết hiệu lực = T+365).\n'
        '3. Số lượng SPDV chi tiết cấp con trực tiếp của Nghiệp vụ cha < 99.'
    ),
    'Steps': (
        '1. Truy cập màn hình Danh sách sản phẩm dịch vụ.\n'
        '2. Nhấn nút "+" tại bản ghi Nghiệp vụ cha "Tài khoản" (Mã "06").\n'
        '3. Hệ thống hiển thị màn hình Thêm mới SPDV chi tiết; kiểm tra trường "Cấp SPDV liền trước" tự động = "06 - Tài khoản".\n'
        '4. Nhập "Tên" = "Tài khoản thanh toán".\n'
        '5. Nhập "Mô tả" = "Tài khoản sử dụng cho giao dịch thanh toán hàng ngày".\n'
        '6. Chọn "Ngày hiệu lực" = T+1 (>= Ngày hiệu lực của Nghiệp vụ cha T+1).\n'
        '7. Chọn "Ngày hết hiệu lực" = T+365 (<= Ngày hết hiệu lực cha T+365).\n'
        '8. Nhấn "Xác nhận".'
    ),
    'Expected': (
        '(i) Nghiệp vụ: SPDV chi tiết được tạo thành công; Mã SPDV chi tiết do hệ thống tự sinh = "0601" (Mã nghiệp vụ cha "06" + "01") và là duy nhất.\n'
        '(ii) UI: Hiển thị thông báo "Thêm mới thành công SPDV chi tiết"; form đóng lại.\n'
        '(iii) Trạng thái bản ghi: Bản ghi SPDV chi tiết có Trạng thái duyệt = "Chờ duyệt"; Trạng thái hoạt động = "Không hoạt động" (Ngày hiệu lực T+1 > Ngày T).\n'
        '(iv) Output: Bản ghi xuất hiện tại "Tác vụ Pending của tôi" của Maker và "Tác vụ chờ duyệt" của Checker.'
    ),
    'Trace_ID': 'SA01-SPDV-THEMMOI-HAP',
    'Note': '',
})

TC.append({
    'TC_ID': 'SA01-BR-HAP-009',
    'BR_Ref': 'LOG-SPDV-AUTOCODE',
    'URD_Ref': 'Bảng "Khai báo Sản phẩm, dịch vụ chi tiết" (Table 3) – dòng "Mã", mô tả "Mã nghiệp vụ/SPDV cấp cha liền trước + 02 ký tự số tự tăng từ 01 đến 99"',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Thêm mới SPDV chi tiết] Mã SPDV chi tiết tự sinh đúng quy tắc: Mã cha + 2 ký tự tự tăng',
    'Type': 'TC-BR',
    'Category': 'Positive',
    'Priority': 'High',
    'Precondition': (
        '1. Người dùng có vai trò Maker, đã đăng nhập hệ thống.\n'
        '2. Nghiệp vụ cha: Mã = "06" (Tài khoản), đã có 1 SPDV chi tiết con: Mã = "0601".\n'
        '3. Màn hình Thêm mới SPDV chi tiết của Nghiệp vụ cha "06" đã mở.'
    ),
    'Steps': (
        '1. Nhập thông tin hợp lệ: Tên, Mô tả, Ngày hiệu lực >= T, Ngày hết hiệu lực hợp lệ.\n'
        '2. Nhấn "Xác nhận".\n'
        '3. Sau khi duyệt, quan sát Mã SPDV chi tiết vừa được sinh.'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Mã SPDV chi tiết mới = "0602" (= "06" (Mã cha) + "02" (số tự tăng kế tiếp sau "01")); Mã là duy nhất, không trùng bất kỳ bản ghi nào.\n'
        '(ii) UI: Trường "Mã" hiển thị giá trị "0602" sau khi lưu thành công.\n'
        '(iii) Trạng thái bản ghi: Trạng thái duyệt = "Chờ duyệt".\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-SPDV-AUTOCODE',
    'Note': '',
})

TC.append({
    'TC_ID': 'SA01-BR-NEG-008',
    'BR_Ref': 'LOG-SPDV-DATE-PARENT-CONSTRAINT',
    'URD_Ref': 'Bảng "Khai báo Sản phẩm, dịch vụ chi tiết" (Table 3) – dòng "Ngày hiệu lực", ràng buộc ">= Ngày hiệu lực của các cấp cha trong cây SPDV"',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Thêm mới SPDV chi tiết] Negative – Ngày hiệu lực SPDV con nhỏ hơn Ngày hiệu lực Nghiệp vụ cha',
    'Type': 'TC-BR',
    'Category': 'Negative',
    'Priority': 'High',
    'Precondition': (
        '1. Người dùng có vai trò Maker, đã đăng nhập hệ thống.\n'
        '2. Nghiệp vụ cha: Mã "06", Ngày hiệu lực = T+10, Ngày hết hiệu lực = T+365, Trạng thái = "Đã duyệt".\n'
        '3. Màn hình Thêm mới SPDV chi tiết thuộc Nghiệp vụ "06" đang mở.'
    ),
    'Steps': (
        '1. Nhập Tên, Mô tả hợp lệ cho SPDV chi tiết.\n'
        '2. Chọn "Ngày hiệu lực" = T+5 (nhỏ hơn Ngày hiệu lực cha T+10).\n'
        '3. Chọn "Ngày hết hiệu lực" = T+365.\n'
        '4. Nhấn "Xác nhận".'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Hệ thống từ chối lưu; bản ghi SPDV chi tiết không được tạo.\n'
        '(ii) UI: Hiển thị thông báo lỗi "Ngày hiệu lực của SPDV chi tiết phải >= Ngày hiệu lực của cấp cha (T+10)".\n'
        '(iii) Trạng thái bản ghi: Không có bản ghi nào được tạo.\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-SPDV-DATE-PARENT-NEG',
    'Note': '',
})

TC.append({
    'TC_ID': 'SA01-BR-NEG-009',
    'BR_Ref': 'LOG-SPDV-DATE-ENDDATE-PARENT',
    'URD_Ref': 'Bảng "Khai báo Sản phẩm, dịch vụ chi tiết" (Table 3) – dòng "Ngày hết hiệu lực", ràng buộc "<= Ngày hết hiệu lực của các cấp cha trong cây SPDV"',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Thêm mới SPDV chi tiết] Negative – Ngày hết hiệu lực SPDV con lớn hơn Ngày hết hiệu lực Nghiệp vụ cha',
    'Type': 'TC-BR',
    'Category': 'Negative',
    'Priority': 'High',
    'Precondition': (
        '1. Người dùng có vai trò Maker, đã đăng nhập hệ thống.\n'
        '2. Nghiệp vụ cha: Mã "06", Ngày hiệu lực = T+1, Ngày hết hiệu lực = T+100, Trạng thái = "Đã duyệt".\n'
        '3. Màn hình Thêm mới SPDV chi tiết thuộc Nghiệp vụ "06" đang mở.'
    ),
    'Steps': (
        '1. Nhập Tên, Mô tả hợp lệ.\n'
        '2. Chọn "Ngày hiệu lực" = T+1.\n'
        '3. Chọn "Ngày hết hiệu lực" = T+200 (lớn hơn Ngày hết hiệu lực cha T+100).\n'
        '4. Nhấn "Xác nhận".'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Hệ thống từ chối lưu; bản ghi không được tạo.\n'
        '(ii) UI: Hiển thị thông báo lỗi "Ngày hết hiệu lực của SPDV chi tiết phải <= Ngày hết hiệu lực của cấp cha (T+100)".\n'
        '(iii) Trạng thái bản ghi: Không có bản ghi nào được tạo.\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-SPDV-DATE-ENDDATE-NEG',
    'Note': '',
})

TC.append({
    'TC_ID': 'SA01-BR-NEG-010',
    'BR_Ref': 'LOG-SPDV-AUTOCODE-OVERFLOW',
    'URD_Ref': 'Bảng "Khai báo Sản phẩm, dịch vụ chi tiết" (Table 3) – dòng "Mã", đoạn "Trường hợp số tự tăng >99 hệ thống báo lỗi và không cho phép lưu"',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Thêm mới SPDV chi tiết] Boundary – Tràn Mã SPDV chi tiết khi đã có đủ 99 SPDV con của cùng 1 cấp cha',
    'Type': 'TC-BR',
    'Category': 'Negative',
    'Priority': 'Medium',
    'Precondition': (
        '1. Người dùng có vai trò Maker, đã đăng nhập hệ thống.\n'
        '2. Nghiệp vụ cha (Mã "06") đã có đủ 99 SPDV chi tiết cấp con trực tiếp (Mã từ "0601" đến "0699").'
    ),
    'Steps': (
        '1. Nhấn nút "+" tại bản ghi Nghiệp vụ cha "06".\n'
        '2. Nhập thông tin hợp lệ (Tên, Mô tả, Ngày hiệu lực, Ngày hết hiệu lực).\n'
        '3. Nhấn "Xác nhận".'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Hệ thống phát hiện số tự tăng đã đạt giới hạn 99; từ chối tạo bản ghi mới.\n'
        '(ii) UI: Hiển thị thông báo lỗi (nội dung theo Phụ lục mã lỗi / tham số hệ thống – theo Q05 BA sẽ cập nhật lại).\n'
        '(iii) Trạng thái bản ghi: Không có bản ghi mới được tạo.\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-SPDV-AUTOCODE-OVF',
    'Note': '',
})

# ─────────────────────────────────────────────────────────────────────────────
# NHÓM A: TC-BR – Chỉnh sửa SPDV chi tiết
# ─────────────────────────────────────────────────────────────────────────────

TC.append({
    'TC_ID': 'SA01-BR-HAP-010',
    'BR_Ref': 'LOG-SPDV-EDIT-HAPPY',
    'URD_Ref': 'Mục "Yêu cầu nghiệp vụ > Khai báo SPDV chi tiết" – đoạn "Bản ghi SPDV chi tiết sau khi duyệt thành công... khi chỉnh sửa, hệ thống chỉ cho phép chỉnh sửa Ngày hiệu lực, Ngày hết hiệu lực"',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Chỉnh sửa SPDV chi tiết] Happy Path – Cập nhật Ngày hiệu lực và Ngày hết hiệu lực thành công, thỏa ràng buộc cấp cha-con',
    'Type': 'TC-BR',
    'Category': 'Positive',
    'Priority': 'High',
    'Precondition': (
        '1. Người dùng có vai trò Maker, đã đăng nhập hệ thống.\n'
        '2. Tồn tại bản ghi SPDV chi tiết đã được duyệt (Trạng thái = "Đã duyệt"), thuộc Nghiệp vụ cha có Ngày hiệu lực = T, Ngày hết hiệu lực = T+365.\n'
        '3. SPDV chi tiết hiện có: Ngày hiệu lực = T+1, Ngày hết hiệu lực = T+100.'
    ),
    'Steps': (
        '1. Nhấn nút "Chỉnh sửa" tại bản ghi SPDV chi tiết đã duyệt.\n'
        '2. Kiểm tra các trường Tên, Mã, Mô tả, Cấp SPDV liền trước ở trạng thái read-only.\n'
        '3. Thay đổi "Ngày hiệu lực" = T+5 (>= Ngày hiệu lực cha T).\n'
        '4. Thay đổi "Ngày hết hiệu lực" = T+200 (<= Ngày hết hiệu lực cha T+365).\n'
        '5. Nhấn "Xác nhận".'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Hệ thống tạo yêu cầu chỉnh sửa SPDV chi tiết mới ở trạng thái "Chờ duyệt"; ràng buộc ngày với cấp cha hợp lệ.\n'
        '(ii) UI: Hiển thị thông báo "Tạo yêu cầu Sửa thành công SPDV chi tiết".\n'
        '(iii) Trạng thái bản ghi: Yêu cầu chỉnh sửa = "Chờ duyệt"; bản ghi gốc vẫn hiệu lực cho đến khi được duyệt.\n'
        '(iv) Output: Yêu cầu sửa xuất hiện tại "Tác vụ Pending của tôi" của Maker và "Tác vụ chờ duyệt" của Checker.'
    ),
    'Trace_ID': 'SA01-SPDV-EDIT-HAP',
    'Note': '',
})

TC.append({
    'TC_ID': 'SA01-BR-NEG-011',
    'BR_Ref': 'LOG-SPDV-EDIT-CHILD-CONSTRAINT',
    'URD_Ref': 'Mục "Yêu cầu nghiệp vụ > Khai báo SPDV chi tiết" – đoạn "Ngày hiệu lực của cấp cha <= Ngày hiệu lực <= Ngày hiệu lực của các SPDV chi tiết cấp con"',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Chỉnh sửa SPDV chi tiết] Negative – Ngày hiệu lực mới lớn hơn Ngày hiệu lực của SPDV cấp con',
    'Type': 'TC-BR',
    'Category': 'Negative',
    'Priority': 'High',
    'Precondition': (
        '1. Người dùng có vai trò Maker, đã đăng nhập hệ thống.\n'
        '2. Tồn tại chuỗi SPDV: Nghiệp vụ cha "06" (Ngày HLực T) > SPDV con "0601" (Ngày HLực T+5) > SPDV cháu "060101" (Ngày HLực T+10).\n'
        '3. Đang chỉnh sửa bản ghi SPDV "0601".'
    ),
    'Steps': (
        '1. Mở màn hình Chỉnh sửa bản ghi SPDV "0601".\n'
        '2. Thay đổi "Ngày hiệu lực" = T+15 (lớn hơn Ngày hiệu lực của cháu "060101" = T+10).\n'
        '3. Nhấn "Xác nhận".'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Hệ thống phát hiện ràng buộc vi phạm với cấp con; từ chối lưu.\n'
        '(ii) UI: Hiển thị thông báo lỗi cascade (chi tiết thông báo theo BA sẽ cập nhật US – Q12).\n'
        '(iii) Trạng thái bản ghi: Không có yêu cầu chỉnh sửa nào được tạo.\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-SPDV-EDIT-CHILD-NEG',
    'Note': 'Q12: BA sẽ update lại US mô tả behavior. QA cần confirm lại expected result khi có version cập nhật.',
})

# ─────────────────────────────────────────────────────────────────────────────
# NHÓM A: TC-BR – Tham số số cấp SPDV (n)
# ─────────────────────────────────────────────────────────────────────────────

TC.append({
    'TC_ID': 'SA01-BR-HAP-011',
    'BR_Ref': 'LOG-SPDV-NCAP-INCREASE',
    'URD_Ref': 'Mục "Yêu cầu nghiệp vụ > Khai báo SPDV chi tiết" – đoạn "Số n tăng: Trạng thái các SPDV không thay đổi và người dùng có thể khai báo thêm SPDV cấp con"',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Tham số n - Số cấp SPDV] Khi n tăng – các SPDV hiện có không bị ảnh hưởng, được phép khai báo thêm cấp con',
    'Type': 'TC-BR',
    'Category': 'Positive',
    'Priority': 'Medium',
    'Precondition': (
        '1. Hệ thống đang cấu hình số cấp SPDV tối đa n = 3.\n'
        '2. Tồn tại SPDV cấp 3 (cấp cuối cùng), Trạng thái = "Hoạt động".\n'
        '3. Admin/Người có quyền thay đổi tham số hệ thống tăng n = 4.'
    ),
    'Steps': (
        '1. Admin thay đổi tham số số cấp SPDV tối đa từ n=3 lên n=4.\n'
        '2. Kiểm tra Trạng thái hoạt động của các SPDV hiện có (cấp 1, 2, 3).\n'
        '3. Thử thêm mới SPDV chi tiết cấp 4 (cấp con của SPDV cấp 3).'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Sau khi n tăng lên 4: Trạng thái hoạt động của tất cả SPDV hiện có (cấp 1, 2, 3) KHÔNG thay đổi; người dùng có thể khai báo thêm SPDV cấp 4.\n'
        '(ii) UI: Nút "+" xuất hiện tại các bản ghi SPDV cấp 3 (vốn trước đây là cấp cuối); thêm mới SPDV cấp 4 thành công.\n'
        '(iii) Trạng thái bản ghi: Không có bản ghi nào bị thay đổi trạng thái tự động.\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-SPDV-NCAP-UP',
    'Note': '',
})

TC.append({
    'TC_ID': 'SA01-BR-NEG-012',
    'BR_Ref': 'LOG-SPDV-NCAP-DECREASE',
    'URD_Ref': 'Mục "Yêu cầu nghiệp vụ > Khai báo SPDV chi tiết" – đoạn "Số n giảm: Hệ thống tự động cập nhật trạng thái SPDV có cấp > n thành Không hoạt động"',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Tham số n - Số cấp SPDV] Khi n giảm – SPDV có cấp > n tự động chuyển sang "Không hoạt động"',
    'Type': 'TC-BR',
    'Category': 'Negative',
    'Priority': 'High',
    'Precondition': (
        '1. Hệ thống đang cấu hình số cấp SPDV tối đa n = 4.\n'
        '2. Tồn tại SPDV cấp 4, Trạng thái = "Hoạt động".\n'
        '3. Admin thay đổi tham số số cấp SPDV từ n=4 xuống n=3.'
    ),
    'Steps': (
        '1. Admin thay đổi tham số số cấp SPDV tối đa từ n=4 xuống n=3.\n'
        '2. Kiểm tra Trạng thái hoạt động của SPDV cấp 4 (cấp > n mới).\n'
        '3. Kiểm tra Trạng thái hoạt động của SPDV cấp 1, 2, 3 (cấp <= n mới).'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Hệ thống tự động cập nhật Trạng thái hoạt động của tất cả SPDV cấp 4 (> n=3) = "Không hoạt động"; SPDV cấp 1, 2, 3 (<= n) KHÔNG bị ảnh hưởng.\n'
        '(ii) UI: Cột "Trạng thái hoạt động" của SPDV cấp 4 hiển thị = "Không hoạt động".\n'
        '(iii) Trạng thái bản ghi: SPDV cấp 4 = "Không hoạt động"; SPDV cấp 1-3 giữ nguyên trạng thái cũ.\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-SPDV-NCAP-DOWN',
    'Note': '',
})

# ─────────────────────────────────────────────────────────────────────────────
# NHÓM A: TC-BR – Xóa bản ghi (Chờ duyệt / Từ chối duyệt)
# ─────────────────────────────────────────────────────────────────────────────

TC.append({
    'TC_ID': 'SA01-BR-HAP-012',
    'BR_Ref': 'LOG-NV-DELETE-PENDING',
    'URD_Ref': 'Q13: Câu trả lời của BA "Chỉ được xóa khi bản ghi có trạng thái Từ chối duyệt/Chờ duyệt (Nút xóa nằm ở màn hình Tác vụ pending của tôi)"',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Xóa Nghiệp vụ / SPDV chi tiết] Maker xóa thành công bản ghi đang ở trạng thái "Chờ duyệt"',
    'Type': 'TC-BR',
    'Category': 'Positive',
    'Priority': 'High',
    'Precondition': (
        '1. Người dùng có vai trò Maker, đã đăng nhập hệ thống.\n'
        '2. Tồn tại bản ghi Nghiệp vụ hoặc SPDV chi tiết có Trạng thái = "Chờ duyệt" tại màn hình "Tác vụ Pending của tôi" của Maker.'
    ),
    'Steps': (
        '1. Truy cập màn hình "Tác vụ Pending của tôi".\n'
        '2. Tìm bản ghi Nghiệp vụ có Trạng thái = "Chờ duyệt".\n'
        '3. Nhấn nút "Xóa" tại bản ghi đó.\n'
        '4. Xác nhận thao tác xóa (nếu có popup xác nhận).'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Bản ghi "Chờ duyệt" bị xóa khỏi hệ thống.\n'
        '(ii) UI: Bản ghi biến mất khỏi danh sách "Tác vụ Pending của tôi" và "Tác vụ chờ duyệt" của Checker.\n'
        '(iii) Trạng thái bản ghi: Bản ghi không còn tồn tại trong DB.\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-NV-DELETE-PENDING',
    'Note': 'Q13: BA confirm chỉ được xóa khi trạng thái Từ chối duyệt/Chờ duyệt. Chức năng Xóa nằm ở màn hình "Tác vụ pending của tôi".',
})

TC.append({
    'TC_ID': 'SA01-BR-HAP-013',
    'BR_Ref': 'LOG-NV-DELETE-REJECTED',
    'URD_Ref': 'Q13: Câu trả lời của BA "Chỉ được xóa khi bản ghi có trạng thái Từ chối duyệt/Chờ duyệt"',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[Xóa Nghiệp vụ / SPDV chi tiết] Maker xóa thành công bản ghi đang ở trạng thái "Từ chối duyệt"',
    'Type': 'TC-BR',
    'Category': 'Positive',
    'Priority': 'Medium',
    'Precondition': (
        '1. Người dùng có vai trò Maker, đã đăng nhập hệ thống.\n'
        '2. Tồn tại bản ghi Nghiệp vụ hoặc SPDV chi tiết có Trạng thái = "Từ chối duyệt" tại màn hình "Tác vụ Pending của tôi" của Maker.'
    ),
    'Steps': (
        '1. Truy cập màn hình "Tác vụ Pending của tôi".\n'
        '2. Tìm bản ghi có Trạng thái = "Từ chối duyệt".\n'
        '3. Nhấn nút "Xóa" tại bản ghi đó.\n'
        '4. Xác nhận thao tác xóa (nếu có popup).'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Bản ghi "Từ chối duyệt" bị xóa khỏi hệ thống.\n'
        '(ii) UI: Bản ghi biến mất khỏi danh sách.\n'
        '(iii) Trạng thái bản ghi: Bản ghi không còn tồn tại trong DB.\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-NV-DELETE-REJECTED',
    'Note': '',
})

# ─────────────────────────────────────────────────────────────────────────────
# NHÓM B: TC-UI
# ─────────────────────────────────────────────────────────────────────────────

TC.append({
    'TC_ID': 'SA01-UI-001',
    'BR_Ref': 'UI-NV-FORM-OPEN',
    'URD_Ref': 'Lưu đồ Thêm mới Nghiệp vụ – Bước 1: Chọn nút "+ Thêm mới" trên thanh header; Bước 2: Hiển thị màn hình Thêm mới Nghiệp vụ',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[UI - Thêm mới Nghiệp vụ] Nhấn nút "+ Thêm mới" → Form Thêm mới Nghiệp vụ mở ra, trường "Cấp sản phẩm dịch vụ" mặc định = 1',
    'Type': 'TC-UI',
    'Category': 'Positive',
    'Priority': 'High',
    'Precondition': (
        '1. Người dùng đã đăng nhập, có quyền truy cập chức năng.\n'
        '2. Đang ở màn hình Danh sách sản phẩm dịch vụ.'
    ),
    'Steps': (
        '1. Nhấn nút "+ Thêm mới" trên thanh header của màn hình Danh sách sản phẩm dịch vụ.\n'
        '2. Quan sát form Thêm mới Nghiệp vụ.'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Form mở ra đúng chức năng Thêm mới Nghiệp vụ.\n'
        '(ii) UI: Form Thêm mới Nghiệp vụ hiển thị; trường "Cấp sản phẩm dịch vụ" tự động điền giá trị = 1 và không cho phép thay đổi (disabled/read-only); các trường còn lại trống, trường "Mã" ở dạng read-only (theo Q01).\n'
        '(iii) Trạng thái bản ghi: Không có bản ghi nào được tạo tại bước này.\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-UI-NV-FORM-OPEN',
    'Note': '',
})

TC.append({
    'TC_ID': 'SA01-UI-002',
    'BR_Ref': 'UI-NV-FORM-CLOSE',
    'URD_Ref': 'Lưu đồ Thêm mới Nghiệp vụ – Bước 2.1 (Đóng) → Bước 2.2 (Hiển thị màn hình quản lý cây SPDV)',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[UI - Thêm mới Nghiệp vụ] Nhấn nút "Đóng" → thoát form và quay về màn hình Danh sách, không lưu dữ liệu',
    'Type': 'TC-UI',
    'Category': 'Positive',
    'Priority': 'Medium',
    'Precondition': (
        '1. Người dùng đã đăng nhập, có quyền truy cập.\n'
        '2. Form Thêm mới Nghiệp vụ đang mở, đã nhập một số dữ liệu chưa lưu.'
    ),
    'Steps': (
        '1. Nhập một phần thông tin vào form (ví dụ: chỉ nhập Tên Nghiệp vụ).\n'
        '2. Nhấn nút "Đóng".'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Dữ liệu chưa lưu bị hủy; không có bản ghi mới được tạo.\n'
        '(ii) UI: Form đóng lại; màn hình trở về Danh sách sản phẩm dịch vụ (theo Q03 BA confirm: "Đóng" là logic chung, không có popup xác nhận).\n'
        '(iii) Trạng thái bản ghi: Không có bản ghi nào được tạo.\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-UI-NV-CLOSE',
    'Note': 'Q03: BA Drop – "Đóng" là logic chung của hệ thống, không có dialog xác nhận.',
})

TC.append({
    'TC_ID': 'SA01-UI-003',
    'BR_Ref': 'UI-SPDV-EDITSCREEN-READONLY',
    'URD_Ref': 'Mục "Yêu cầu nghiệp vụ > Khai báo Nghiệp vụ" – đoạn "hệ thống chỉ cho phép chỉnh sửa Ngày hiệu lực, Ngày hết hiệu lực"; Q14 BA: màn hình chỉnh sửa tương tự màn hình xem chi tiết',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[UI - Chỉnh sửa Nghiệp vụ] Màn hình Chỉnh sửa: các trường Tên, Mã, Mô tả hiển thị read-only; chỉ Ngày hiệu lực và Ngày hết hiệu lực được phép chỉnh sửa',
    'Type': 'TC-UI',
    'Category': 'Positive',
    'Priority': 'High',
    'Precondition': (
        '1. Người dùng có vai trò Maker, đã đăng nhập.\n'
        '2. Tồn tại bản ghi Nghiệp vụ đã được duyệt trên màn hình Danh sách.'
    ),
    'Steps': (
        '1. Tìm bản ghi Nghiệp vụ đã được duyệt trên lưới danh sách.\n'
        '2. Nhấn nút "Chỉnh sửa" tại bản ghi.\n'
        '3. Quan sát trạng thái của từng trường trên màn hình Chỉnh sửa.'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Màn hình Chỉnh sửa hiển thị đầy đủ thông tin bản ghi.\n'
        '(ii) UI: Các trường "Tên Nghiệp vụ", "Mã Nghiệp vụ", "Mô tả", "Cấp sản phẩm dịch vụ" hiển thị read-only (không có input box / disabled); chỉ trường "Ngày hiệu lực" và "Ngày hết hiệu lực" có input để chỉnh sửa.\n'
        '(iii) Trạng thái bản ghi: Không thay đổi tại bước này (chưa nhấn Xác nhận).\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-UI-EDIT-READONLY',
    'Note': 'Q14: BA Drop – logic chung màn hình chỉnh sửa tương tự xem chi tiết, hiển thị read-only các trường không được sửa.',
})

TC.append({
    'TC_ID': 'SA01-UI-004',
    'BR_Ref': 'UI-SPDV-ADDCHILD-BTN',
    'URD_Ref': 'Lưu đồ Thêm mới SPDV chi tiết – Bước 1: Chọn nút "+" tại bản ghi SPDV cấp cha liền trước',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[UI - Thêm mới SPDV chi tiết] Nhấn nút "+" tại bản ghi cha → Form Thêm mới SPDV chi tiết mở, trường "Cấp SPDV liền trước" tự động hiển thị thông tin cha',
    'Type': 'TC-UI',
    'Category': 'Positive',
    'Priority': 'High',
    'Precondition': (
        '1. Người dùng có vai trò Maker, đã đăng nhập.\n'
        '2. Tồn tại bản ghi Nghiệp vụ cha đã duyệt (ví dụ: Mã = "06", Tên = "Tài khoản") trên màn hình Danh sách.'
    ),
    'Steps': (
        '1. Trên màn hình Danh sách sản phẩm dịch vụ, tìm bản ghi Nghiệp vụ "06 - Tài khoản".\n'
        '2. Nhấn nút "+" tại bản ghi "06 - Tài khoản".\n'
        '3. Quan sát form Thêm mới SPDV chi tiết mở ra.'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Form Thêm mới SPDV chi tiết được mở đúng ngữ cảnh của Nghiệp vụ cha.\n'
        '(ii) UI: Trường "Cấp SPDV liền trước" tự động hiển thị = "06 - Tài khoản" (read-only); các trường khác trống; trường "Mã SPDV" read-only (hệ thống tự sinh).\n'
        '(iii) Trạng thái bản ghi: Không có bản ghi nào được tạo tại bước này.\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-UI-SPDV-ADD-BTN',
    'Note': '',
})

TC.append({
    'TC_ID': 'SA01-UI-005',
    'BR_Ref': 'UI-SPDV-LEAFLABEL',
    'URD_Ref': 'Bảng "Khai báo Sản phẩm, dịch vụ chi tiết" (Table 3) – dòng "Đây là cấp cuối cùng trong cây sơ đồ sản phẩm dịch vụ", Mô tả: "Chỉ hiển thị khi cấp khai báo là cấp cuối"',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[UI - Thêm mới SPDV chi tiết] Label "Đây là cấp cuối cùng" chỉ hiển thị khi cấp khai báo là cấp cuối (= tham số n)',
    'Type': 'TC-UI',
    'Category': 'Positive',
    'Priority': 'Medium',
    'Precondition': (
        '1. Người dùng có vai trò Maker, đã đăng nhập.\n'
        '2. Tham số hệ thống số cấp SPDV tối đa = n (ví dụ n=3).\n'
        '3. Đang mở form Thêm mới SPDV chi tiết cho một SPDV cấp 3 (cấp cuối).'
    ),
    'Steps': (
        '1. Mở form Thêm mới SPDV chi tiết cho một bản ghi ở cấp cuối (cấp n=3).\n'
        '2. Quan sát form xem có label "Đây là cấp cuối cùng trong cây sơ đồ sản phẩm dịch vụ" không.\n'
        '3. Mở form Thêm mới SPDV chi tiết cho một bản ghi ở cấp trung gian (cấp 2 < n=3).\n'
        '4. Quan sát form xem label có hiển thị không.'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Label chỉ hiển thị ở form thêm mới SPDV có cấp = cấp cuối; không hiển thị ở cấp trung gian.\n'
        '(ii) UI: Tại form cấp 3 (cấp n): label "Đây là cấp cuối cùng trong cây sơ đồ sản phẩm dịch vụ" hiển thị trên form; tại form cấp 2: label KHÔNG hiển thị.\n'
        '(iii) Trạng thái bản ghi: Không thay đổi.\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-UI-SPDV-LEAFLABEL',
    'Note': 'Q11: BA sẽ trao đổi với FE để update UI về vị trí và kiểu hiển thị label này.',
})

TC.append({
    'TC_ID': 'SA01-UI-006',
    'BR_Ref': 'UI-NV-CALENDAR-DISABLE-PAST',
    'URD_Ref': 'Bảng "Mô tả chi tiết các trường" (Table 2) – dòng "Ngày hiệu lực", ràng buộc "Không chọn ngày quá khứ"; tương tự Table 3',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[UI - Calendar] Ngày quá khứ bị disabled (grayout) trên calendar khi chọn Ngày hiệu lực / Ngày hết hiệu lực',
    'Type': 'TC-UI',
    'Category': 'Positive',
    'Priority': 'Medium',
    'Precondition': (
        '1. Người dùng đã đăng nhập, form Thêm mới Nghiệp vụ hoặc SPDV chi tiết đang mở.\n'
        '2. Ngày hiện tại là ngày T.'
    ),
    'Steps': (
        '1. Click vào trường "Ngày hiệu lực" để mở calendar.\n'
        '2. Quan sát trạng thái các ngày trên calendar (trước ngày T và từ ngày T trở đi).'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Calendar không cho phép chọn ngày quá khứ.\n'
        '(ii) UI: Các ngày trước ngày T bị grayout/disabled (không click được); từ ngày T trở đi có thể chọn.\n'
        '(iii) Trạng thái bản ghi: Không thay đổi.\n'
        '(iv) Output: Không có output bổ sung.'
    ),
    'Trace_ID': 'SA01-UI-CAL-DISABLE-PAST',
    'Note': '',
})

# ─────────────────────────────────────────────────────────────────────────────
# NHÓM C: TC – Luồng E2E
# ─────────────────────────────────────────────────────────────────────────────

TC.append({
    'TC_ID': 'SA01-E2E-001',
    'BR_Ref': 'LOG-E2E-FULL-FLOW-NV',
    'URD_Ref': 'Mục "Yêu cầu nghiệp vụ > Khai báo Nghiệp vụ" – toàn bộ luồng Maker khai báo → Checker duyệt → Trạng thái cập nhật',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[E2E - Nghiệp vụ] Maker tạo Nghiệp vụ → Checker duyệt → Trạng thái tự động Hoạt động khi đến Ngày hiệu lực',
    'Type': 'TC-E2E',
    'Category': 'Positive',
    'Priority': 'High',
    'Precondition': (
        '1. Có 2 tài khoản: Maker và Checker đã đăng nhập on hệ thống.\n'
        '2. Cả 2 đều có quyền truy cập chức năng Tham số >> Danh mục sản phẩm dịch vụ và Code phí.\n'
        '3. Số lượng Nghiệp vụ hiện có < 99.'
    ),
    'Steps': (
        '1. Maker: Truy cập Tham số >> Danh mục sản phẩm dịch vụ và Code phí >> Nhấn "+ Thêm mới".\n'
        '2. Maker: Nhập Tên Nghiệp vụ = "E2E Test Nghiệp vụ", Mô tả = "Test luồng E2E", Ngày hiệu lực = T (hôm nay), Ngày hết hiệu lực = T+365.\n'
        '3. Maker: Nhấn "Xác nhận" → nhận thông báo thành công.\n'
        '4. Checker: Truy cập màn hình "Tác vụ chờ duyệt" → tìm bản ghi "E2E Test Nghiệp vụ".\n'
        '5. Checker: Mở bản ghi, kiểm tra thông tin, nhấn "Duyệt" để phê duyệt.\n'
        '6. Sau khi duyệt, kiểm tra bản ghi trên màn hình Danh sách sản phẩm dịch vụ.'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Toàn bộ luồng hoàn thành thành công; bản ghi Nghiệp vụ được tạo và phê duyệt; vì Ngày T = Ngày hiệu lực nên Trạng thái hoạt động = "Hoạt động" ngay sau khi duyệt.\n'
        '(ii) UI: Maker nhận thông báo "Thêm mới thành công"; Checker nhận thông báo "Duyệt thành công"; bản ghi xuất hiện trên lưới Danh sách với Trạng thái = "Đã duyệt" & "Hoạt động".\n'
        '(iii) Trạng thái bản ghi: Trạng thái duyệt = "Đã duyệt"; Trạng thái hoạt động = "Hoạt động"; Mã Nghiệp vụ được hệ thống tự sinh thành công.\n'
        '(iv) Output: Bản ghi xuất hiện trên màn hình Danh sách sản phẩm dịch vụ và sẵn sàng để khai báo SPDV chi tiết con.'
    ),
    'Trace_ID': 'SA01-E2E-NV-FULL',
    'Note': '',
})

TC.append({
    'TC_ID': 'SA01-E2E-002',
    'BR_Ref': 'LOG-E2E-FULL-FLOW-SPDV',
    'URD_Ref': 'Mục "Yêu cầu nghiệp vụ > Khai báo SPDV chi tiết" – toàn bộ luồng khai báo cây SPDV 2 cấp',
    'Module': 'SA01',
    'Feature': 'Khai báo Danh mục Sản phẩm, Dịch vụ',
    'Title': '[E2E - Cây SPDV] Maker tạo Nghiệp vụ cha → Checker duyệt → Maker tạo SPDV chi tiết con → Checker duyệt → Cây SPDV 2 cấp hoàn chỉnh',
    'Type': 'TC-E2E',
    'Category': 'Positive',
    'Priority': 'High',
    'Precondition': (
        '1. Có 2 tài khoản: Maker và Checker đã đăng nhập hệ thống.\n'
        '2. Tham số số cấp SPDV tối đa n >= 2.'
    ),
    'Steps': (
        '1. Maker: Tạo Nghiệp vụ cha (Tên = "Nghiệp vụ E2E Cha", Ngày HLực = T, Ngày HHLực = T+365) → Nhấn "Xác nhận".\n'
        '2. Checker: Duyệt Nghiệp vụ cha trên "Tác vụ chờ duyệt".\n'
        '3. Maker: Trên màn hình Danh sách, nhấn "+" tại Nghiệp vụ cha vừa được duyệt.\n'
        '4. Maker: Nhập thông tin SPDV chi tiết con (Tên = "SPDV E2E Con", Mô tả, Ngày HLực = T, Ngày HHLực = T+300 <= T+365) → Nhấn "Xác nhận".\n'
        '5. Checker: Duyệt SPDV chi tiết con.\n'
        '6. Kiểm tra cây SPDV trên màn hình Danh sách: Nghiệp vụ cha và SPDV chi tiết con.'
    ),
    'Expected': (
        '(i) Nghiệp vụ: Toàn bộ luồng hoàn thành; Cây SPDV 2 cấp được thiết lập thành công; ràng buộc ngày giữa cha-con hợp lệ.\n'
        '(ii) UI: Trên màn hình Danh sách, Nghiệp vụ cha và SPDV chi tiết con đều hiển thị Trạng thái = "Đã duyệt" & "Hoạt động"; Mã SPDV chi tiết = Mã Nghiệp vụ cha + "01".\n'
        '(iii) Trạng thái bản ghi: Cả 2 bản ghi = "Đã duyệt" & "Hoạt động".\n'
        '(iv) Output: Cây SPDV hoàn chỉnh, sẵn sàng để sử dụng trong các chức năng tiếp theo của hệ thống ProfiX.'
    ),
    'Trace_ID': 'SA01-E2E-SPDV-TREE',
    'Note': '',
})

# ==============================================================================
# BUILD DATAFRAME & EXPORT EXCEL
# ==============================================================================

COLUMNS_EXEC = ['Người thực hiện', 'Ngày thực hiện', 'Môi trường', 'Kết quả', 'Ghi chú thực thi', 'Lần chạy', 'Bug ID']
COLUMNS_ALL = ['TC_ID', 'BR_Ref', 'URD_Ref', 'Module', 'Feature', 'Title',
               'Type', 'Category', 'Priority', 'Precondition', 'Steps', 'Expected',
               'Trace_ID', 'Note'] + COLUMNS_EXEC

df = pd.DataFrame(TC, columns=['TC_ID', 'BR_Ref', 'URD_Ref', 'Module', 'Feature', 'Title',
                                'Type', 'Category', 'Priority', 'Precondition', 'Steps', 'Expected',
                                'Trace_ID', 'Note'])

for col in COLUMNS_EXEC:
    df[col] = ''

df = df[COLUMNS_ALL]

# Export to Excel
df.to_excel(OUTPUT_FILE, sheet_name='Test Cases', index=False)

# ── Styling ──────────────────────────────────────────────────────────────────
wb = load_workbook(OUTPUT_FILE)
ws = wb['Test Cases']

HEADER_FILL  = PatternFill('solid', fgColor='1F3864')  # dark navy
HEADER_FONT  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
HAP_FILL     = PatternFill('solid', fgColor='E7F4E4')  # light green
NEG_FILL     = PatternFill('solid', fgColor='FDE9E9')  # light red
E2E_FILL     = PatternFill('solid', fgColor='EBF5FB')  # light blue
UI_FILL      = PatternFill('solid', fgColor='FEF9E7')  # light yellow
NORMAL_FONT  = Font(name='Calibri', size=9)
HEADER_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CELL_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Column widths (A-U)
COL_WIDTHS = {
    'A': 22, 'B': 28, 'C': 38, 'D': 10, 'E': 32, 'F': 55,
    'G': 10, 'H': 12, 'I': 10, 'J': 45, 'K': 50, 'L': 60,
    'M': 22, 'N': 35, 'O': 15, 'P': 15, 'Q': 14, 'R': 14, 'S': 25, 'T': 12, 'U': 15
}
for col_letter, width in COL_WIDTHS.items():
    ws.column_dimensions[col_letter].width = width

# Style header row
for cell in ws[1]:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = HEADER_BORDER

ws.row_dimensions[1].height = 30

# Style data rows
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    tc_type   = row[6].value or ''   # G = Type
    tc_cat    = row[7].value or ''   # H = Category
    row_fill = None
    if tc_type == 'TC-E2E':
        row_fill = E2E_FILL
    elif tc_type == 'TC-UI':
        row_fill = UI_FILL
    elif tc_cat == 'Positive':
        row_fill = HAP_FILL
    elif tc_cat == 'Negative':
        row_fill = NEG_FILL

    for cell in row:
        if row_fill:
            cell.fill = row_fill
        cell.font = NORMAL_FONT
        cell.border = CELL_BORDER
        cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Row height
    ws.row_dimensions[row[0].row].height = 85

# Freeze header
ws.freeze_panes = 'A2'

# Auto filter
ws.auto_filter.ref = ws.dimensions

wb.save(OUTPUT_FILE)
print(f'\n✅ Đã xuất thành công: {OUTPUT_FILE}')
print(f'   Tổng số Test Case: {len(TC)}')
tc_br = sum(1 for t in TC if t["Type"] == "TC-BR")
tc_ui = sum(1 for t in TC if t["Type"] == "TC-UI")
tc_e2e = sum(1 for t in TC if t["Type"] == "TC-E2E")
print(f'   TC-BR: {tc_br} | TC-UI: {tc_ui} | TC-E2E: {tc_e2e}')

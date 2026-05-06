#!/usr/bin/env python3
"""US06 TC Batch 2: Negative Path"""
import openpyxl
from openpyxl.styles import Alignment, Border, Side
import os

FPATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "US06_TestCases.xlsx")
wb = openpyxl.load_workbook(FPATH)
ws = wb["Test Cases"]
wrap = Alignment(wrap_text=True, vertical="top")
thin = Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))

FEAT = "Khai báo Biểu phí"
PRE = "1. User đăng nhập quyền Maker tại Danh mục Biểu phí\n2. Đang ở màn hình Thêm mới Biểu phí"

data = [
    ("US06-TC-015","SC-12",'Bảng mô tả trường, tất cả trường ★',FEAT,"Thêm mới – Validation FE",
     "Bỏ trống từng trường bắt buộc (Mã, Tên, Ngày BH, Ngày HL, Ngày HHL, Số VB, Tên VB) → FE chặn",
     "Negative Path","High",PRE,
     "1. Bỏ trống trường Mã Biểu phí, nhập đủ các trường còn lại → click 'Xác nhận'\n"
     "2. Lặp lại cho từng trường: Tên Biểu phí, Ngày ban hành, Ngày hiệu lực, Ngày hết hiệu lực, Số văn bản, Tên văn bản",
     "(i) Nghiệp vụ/Logic: Hệ thống không cho phép lưu khi thiếu bất kỳ trường bắt buộc nào.\n"
     "(ii) UI: FE highlight trường trống, hiển thị thông báo trường bắt buộc. Nút Xác nhận bị chặn hoặc toast lỗi.",""),

    ("US06-TC-016","SC-13",'Bảng mô tả trường R5 – "Mã Biểu phí là duy nhất"',FEAT,"Thêm mới – Validation FE",
     "Mã Biểu phí trùng với Mã đã tồn tại → BE trả lỗi unique","Negative Path","High",
     PRE+"\n3. Đã tồn tại Biểu phí mã 'BP-001' trong hệ thống",
     "1. Nhập Mã = 'BP-001' (trùng)\n2. Nhập đầy đủ các trường khác hợp lệ\n3. Gán Code phí CP001\n4. Click 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: BE kiểm tra trùng mã, không cho phép lưu.\n(ii) UI: Hiển thị thông báo lỗi trùng mã Biểu phí.",""),

    ("US06-TC-017","SC-14",'Bảng mô tả trường R7 – "Không chọn ngày tương lai"',FEAT,"Thêm mới – Validation FE",
     "Ngày ban hành > Ngày hệ thống → FE chặn","Negative Path","High",
     PRE+"\n3. Ngày hệ thống = 06/05/2026",
     "1. Chọn Ngày ban hành = 07/05/2026 (ngày mai)\n2. Nhập đủ các trường khác\n3. Click 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Hệ thống không cho phép chọn ngày tương lai cho Ngày ban hành.\n(ii) UI: FE chặn không cho chọn ngày tương lai trên calendar hoặc hiển thị lỗi.",""),

    ("US06-TC-018","SC-15",'Bảng mô tả trường R8 – "Không chọn ngày quá khứ"',FEAT,"Thêm mới – Validation FE",
     "Ngày hiệu lực < Ngày hệ thống → FE chặn","Negative Path","High",
     PRE+"\n3. Ngày hệ thống = 06/05/2026",
     "1. Chọn Ngày hiệu lực = 05/05/2026 (hôm qua)\n2. Nhập đủ các trường khác\n3. Click 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Hệ thống không cho phép Ngày hiệu lực < Ngày hệ thống.\n(ii) UI: FE chặn hoặc disable ngày quá khứ trên calendar.",""),

    ("US06-TC-019","SC-16",'Bảng mô tả trường R9 – "Ngày hết hiệu lực >= Ngày hiệu lực"',FEAT,"Thêm mới – Validation FE",
     "Ngày hết hiệu lực < Ngày hiệu lực → FE chặn","Negative Path","High",PRE,
     "1. Chọn Ngày hiệu lực = 01/07/2026\n2. Chọn Ngày hết hiệu lực = 30/06/2026 (< Ngày HL)\n3. Nhập đủ trường khác\n4. Click 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Hệ thống không cho phép Ngày HHL < Ngày HL.\n(ii) UI: Hiển thị lỗi validation ngày.",""),

    ("US06-TC-020","SC-17",'BA xác nhận QA-01.8 – "ít nhất 1 Code phí"',FEAT,"Thêm mới – Validation FE",
     "Không gán Code phí nào → FE disable nút Xác nhận","Negative Path","High",PRE,
     "1. Nhập đầy đủ thông tin chung hợp lệ\n2. KHÔNG gán bất kỳ Code phí nào (lưới Thông tin chi tiết trống)\n3. Kiểm tra nút 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Hệ thống yêu cầu ít nhất 1 Code phí.\n(ii) UI: Nút 'Xác nhận' ở trạng thái disabled.","[BA QA-01.8]"),

    ("US06-TC-021","SC-18",'QTC-07 Upload file – "Validate format/dung lượng"',FEAT,"Thêm mới – Upload",
     "Upload file sai định dạng hoặc vượt dung lượng → toast lỗi","Negative Path","Medium",PRE,
     "1. Click nút 'Chọn tệp'\n2. Chọn file 'data.pdf' (không phải .xlsx)\n3. Kiểm tra phản hồi hệ thống",
     "(i) Nghiệp vụ/Logic: Hệ thống từ chối file không đúng định dạng .xlsx.\n(ii) UI: Toast lỗi thông báo file không hợp lệ.","[Theo QTC-07, US07]"),

    ("US06-TC-022","SC-20",'BA xác nhận QA-02.2 – "validate URL http/https"',FEAT,"Thêm mới – Link iDoc",
     "Link iDoc không đúng format URL → FE chặn","Negative Path","Medium",PRE,
     "1. Nhập Link iDoc = 'ftp://idoc.bank.vn/doc1' (không phải http/https)\n2. Nhập đủ trường khác\n3. Click 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Hệ thống validate Link iDoc phải bắt đầu bằng http:// hoặc https://.\n(ii) UI: Hiển thị lỗi format URL không hợp lệ.","[BA QA-02.2]"),

    ("US06-TC-023","SC-21",'Mục "Chỉnh sửa" P52-P53 – "chặn nếu có tác vụ Chờ duyệt"',FEAT,"Chỉnh sửa – Chặn",
     "Chỉnh sửa khi có tác vụ đang Chờ duyệt → hệ thống chặn","Negative Path","High",
     "1. User đăng nhập quyền Maker\n2. Biểu phí BP-005 có tác vụ 'Chỉnh sửa' đang Chờ duyệt",
     "1. Tại lưới Danh mục, click 'Chỉnh sửa' tại BP-005",
     "(i) Nghiệp vụ/Logic: Hệ thống chặn không cho chỉnh sửa vì đã tồn tại tác vụ Chờ duyệt.\n(ii) UI: Hiển thị thông báo không cho phép chỉnh sửa.",""),

    ("US06-TC-024","SC-22",'Mục "Chỉnh sửa" P62 – "> Ngày hệ thống"',FEAT,"Chỉnh sửa – Đang hiệu lực",
     "Đang hiệu lực: Ngày HHL mới <= Ngày hệ thống → FE chặn","Negative Path","High",
     "1. User đăng nhập quyền Maker\n2. BP-006 Đang hiệu lực (HL=01/01/2026, HHL=30/06/2026)\n3. Ngày HT = 06/05/2026",
     "1. Click 'Chỉnh sửa' tại BP-006\n2. Sửa Ngày HHL = 06/05/2026 (= Ngày HT, không > Ngày HT)\n3. Click 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Hệ thống không cho phép vì Ngày HHL mới phải > Ngày HT.\n(ii) UI: Hiển thị lỗi validation.",""),

    ("US06-TC-025","SC-23",'Mục "Chỉnh sửa" P62 – "> Ngày hiệu lực"',FEAT,"Chỉnh sửa – Đang hiệu lực",
     "Đang hiệu lực: Ngày HHL mới <= Ngày hiệu lực → FE chặn","Negative Path","High",
     "1. User đăng nhập quyền Maker\n2. BP-007 Đang hiệu lực (HL=01/03/2026)\n3. Ngày HT = 06/05/2026",
     "1. Click 'Chỉnh sửa' tại BP-007\n2. Sửa Ngày HHL = 01/03/2026 (= Ngày HL, không > Ngày HL)\n3. Click 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Hệ thống không cho phép vì Ngày HHL mới phải > Ngày HL.\n(ii) UI: Hiển thị lỗi validation.",""),

    ("US06-TC-026","SC-24",'BA xác nhận QA-01.5 – "Không cho phép Thêm mới từng Code phí lẻ"',FEAT,"Chỉnh sửa – Chưa hiệu lực",
     "Chỉnh sửa Chưa hiệu lực: cố thêm Code phí lẻ → không cho phép","Negative Path","Medium",
     "1. User đăng nhập quyền Maker\n2. BP-008 trạng thái Chưa hiệu lực",
     "1. Click 'Chỉnh sửa' tại BP-008\n2. Cố gắng thêm Code phí lẻ qua cây SPDV (không qua Upload)\n3. Kiểm tra phản hồi",
     "(i) Nghiệp vụ/Logic: Hệ thống không cho phép gắn thêm Code phí lẻ. Phải dùng Upload Excel (US07).\n(ii) UI: Chức năng tích chọn thêm Code phí bị disable hoặc ẩn.","[BA QA-01.5]"),

    ("US06-TC-027","SC-25",'QTC-15 – "Nút Đóng: discard, không popup"',FEAT,"Thêm mới – Đóng",
     "Nhấn Đóng khi đang nhập dữ liệu → discard, quay về Danh mục","Negative Path","Medium",PRE,
     "1. Nhập Mã = 'BP-TEST', Tên = 'Test discard'\n2. Click nút 'Đóng'",
     "(i) Nghiệp vụ/Logic: Hệ thống discard toàn bộ dữ liệu đã nhập. Không lưu bản nháp.\n(ii) UI: Không hiển thị popup xác nhận. Quay về màn hình Danh mục Biểu phí.","[Theo QTC-15]"),

    ("US06-TC-028","SC-26",'QTC-11 – "FE-First Error Handling"',FEAT,"Chỉnh sửa – BE Exception",
     "BE validate thất bại (edge case) → hệ thống xử lý an toàn","Negative Path","Medium",
     "1. User đăng nhập quyền Maker\n2. BP-009 Chưa hiệu lực\n3. Giả lập bypass FE validation",
     "1. Click 'Chỉnh sửa' tại BP-009\n2. Bypass FE gửi request với dữ liệu không hợp lệ (VD: Ngày HHL < Ngày HL)\n3. Kiểm tra phản hồi BE",
     "(i) Nghiệp vụ/Logic: BE validate thất bại, không lưu dữ liệu sai. Hệ thống không crash.\n(ii) UI: Hiển thị thông báo lỗi từ BE.","[Theo QTC-11]"),

    ("US06-TC-029","SC-62",'Mục "Chỉnh sửa" P56 – "trừ Mã Biểu phí"',FEAT,"Chỉnh sửa – Chưa hiệu lực",
     "Cố sửa Mã Biểu phí khi Chưa hiệu lực → FE disable trường Mã","Negative Path","High",
     "1. User đăng nhập quyền Maker\n2. BP-010 trạng thái Chưa hiệu lực",
     "1. Click 'Chỉnh sửa' tại BP-010\n2. Kiểm tra trường Mã Biểu phí",
     "(i) Nghiệp vụ/Logic: Mã Biểu phí không được phép thay đổi sau khi tạo.\n(ii) UI: Trường Mã ở trạng thái disabled/readonly, không cho nhập.",""),

    ("US06-TC-030","SC-63",'QTC-12 + Mục "Chỉnh sửa" – "Checker từ chối Chỉnh sửa"',FEAT,"Chỉnh sửa – Maker-Checker",
     "Checker từ chối tác vụ Chỉnh sửa → giữ nguyên thông tin cũ","Negative Path","High",
     "1. User đăng nhập quyền Checker\n2. Tác vụ Chỉnh sửa BP-004 đang Chờ duyệt",
     "1. Tại Tác vụ chờ duyệt, mở tác vụ Chỉnh sửa BP-004\n2. Click nút 'Từ chối'\n3. Nhập lý do = 'Thông tin chưa chính xác'",
     "(i) Nghiệp vụ/Logic: Hệ thống giữ nguyên thông tin Biểu phí cũ. Bản ghi ở trạng thái Từ chối duyệt.\n"
     "(ii) UI: Toast thông báo từ chối. Bản ghi tại Tác vụ chờ duyệt cập nhật trạng thái Từ chối duyệt.","[Theo QTC-12]"),

    ("US06-TC-031","SC-76",'Bảng mô tả trường R9 – "Không chọn ngày quá khứ"',FEAT,"Thêm mới – Validation FE",
     "Thêm mới: Ngày hết hiệu lực < Ngày hệ thống → FE chặn","Negative Path","High",
     PRE+"\n3. Ngày HT = 06/05/2026",
     "1. Chọn Ngày hiệu lực = 06/05/2026\n2. Chọn Ngày hết hiệu lực = 05/05/2026 (< Ngày HT)\n3. Click 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Hệ thống không cho phép Ngày HHL < Ngày HT.\n(ii) UI: FE chặn hoặc disable ngày quá khứ trên calendar.",""),

    ("US06-TC-032","SC-77",'BA QA-01.8 + Mục "Chỉnh sửa" P61',FEAT,"Chỉnh sửa – Chưa hiệu lực",
     "Chỉnh sửa Chưa hiệu lực: bỏ gắn hết Code phí → FE disable Xác nhận","Negative Path","High",
     "1. User đăng nhập quyền Maker\n2. BP-011 Chưa hiệu lực, có 2 Code phí CP001, CP002",
     "1. Click 'Chỉnh sửa' tại BP-011\n2. Click 'Xóa' tại CP001\n3. Click 'Xóa' tại CP002\n4. Lưới trống\n5. Kiểm tra nút 'Xác nhận'",
     "(i) Nghiệp vụ/Logic: Biểu phí phải có ít nhất 1 Code phí.\n(ii) UI: Nút 'Xác nhận' disabled khi lưới Thông tin chi tiết trống.","[BA QA-01.8]"),

    ("US06-TC-033","SC-80",'QTC-15 – "Nút Đóng khi Chỉnh sửa"',FEAT,"Chỉnh sửa – Đóng",
     "Nhấn Đóng tại Chỉnh sửa → discard, quay về Danh mục","Negative Path","Medium",
     "1. User đăng nhập quyền Maker\n2. BP-012 Chưa hiệu lực",
     "1. Click 'Chỉnh sửa' tại BP-012\n2. Sửa Tên = 'Updated name'\n3. Click nút 'Đóng'",
     "(i) Nghiệp vụ/Logic: Hệ thống discard thay đổi. Tên Biểu phí giữ nguyên giá trị cũ.\n(ii) UI: Không popup xác nhận. Quay về Danh mục Biểu phí.","[Theo QTC-15]"),
]

start = ws.max_row + 1
for i, row in enumerate(data):
    for j, val in enumerate(row):
        c = ws.cell(row=start+i, column=j+1, value=val)
        c.alignment = wrap
        c.border = thin

wb.save(FPATH)
print(f"✅ Batch 2 (Negative): {len(data)} TC appended → {FPATH}")

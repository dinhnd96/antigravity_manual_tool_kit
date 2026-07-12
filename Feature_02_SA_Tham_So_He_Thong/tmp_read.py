from openpyxl import load_workbook

path = '/Users/mac/antigravity-testing-kit/Template upload Danh sách Code phí trong Biểu phí.xlsx'
wb = load_workbook(path, read_only=True)
print(f'Sheets: {wb.sheetnames}')

for sn in wb.sheetnames:
    ws = wb[sn]
    print(f'\n=== Sheet: {sn} (rows={ws.max_row}, cols={ws.max_column}) ===')
    # Print all rows
    for r in range(1, ws.max_row + 1):
        row_data = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            row_data.append(str(v)[:60] if v is not None else '')
        print(f'  R{r}: {" | ".join(row_data)}')
wb.close()

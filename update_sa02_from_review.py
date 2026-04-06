import openpyxl
import shutil
import warnings
import re

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# Đường dẫn file
source_file = 'Feature_02_SA_Tham_So_He_Thong/test case/SA02_Đăng xuất.xlsx'
dest_file = 'Feature_02_SA_Tham_So_He_Thong/test case/SA02_Đăng xuất - updated_final.xlsx'

# Copy file trước khi thao tác
shutil.copy(source_file, dest_file)

wb = openpyxl.load_workbook(dest_file)
ws = wb.active

# Lấy index của header thông qua array thay vì dict để an toàn hơn
headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
col_id = headers.index('TC_ID') + 1
col_title = headers.index('Title') + 1

# Tìm cột Expected (Có lúc tên là Expected hoặc Expected Result)
col_exp = None
for name in ['Expected', 'Expected Result']:
    if name in headers:
        col_exp = headers.index(name) + 1
        break

for r in range(2, ws.max_row + 1):
    tc_id = ws.cell(r, col_id).value
    if not tc_id: continue
    
    # 1. Update TC_ID Format
    new_id = str(tc_id)
    new_id = new_id.replace('SA02-HAPPY-', 'SA02-BR-HAP-')
    new_id = new_id.replace('SA02-SECURITY-', 'SA02-BR-SEC-')
    new_id = new_id.replace('SA02-NEGATIVE-', 'SA02-BR-NEG-')
    new_id = new_id.replace('SA02-NEG-', 'SA02-BR-NEG-')
    new_id = new_id.replace('SA02-BOUNDARY-', 'SA02-BR-BOU-')
    ws.cell(r, col_id).value = new_id
    
    title = str(ws.cell(r, col_title).value) if ws.cell(r, col_title).value else ""
    exp = str(ws.cell(r, col_exp).value).strip() if ws.cell(r, col_exp).value else ""
    
    # 2. Đảm bảo cấu trúc 4 Lớp (i-ii-iii-iv)
    if '(i)' in exp and '(ii)' in exp:
        if '(iii)' not in exp:
            exp += '\n(iii) Trạng thái/Audit: Bất hoạt token phiên hiện tại, xóa trạng thái đăng nhập.'
        if '(iv)' not in exp:
            exp += '\n(iv) Output: Không có.'
            
    # 3. Phủ cấu trúc UI Mismatch vào SA02-BR-HAP-001 và SA02-BR-HAP-005
    if new_id in ['SA02-BR-HAP-001', 'SA02-BR-HAP-005']:
        exp = re.sub(r'\(ii\) UI:.*?(?=\n\(iii|\n\(iv|$)', 
            '(ii) UI: Hệ thống chuyển hướng sang màn hình trung gian báo "Đăng xuất thành công" (Có Logo PVcomBank và biểu tượng check xanh). Click nút "Đăng nhập" trên màn này mới quay về Form Login.', 
            exp, flags=re.DOTALL)
            
    # 4. Gắn cờ cảnh báo BA cho các luồng Tối nghĩa
    if new_id == 'SA02-BR-SEC-002':
        title = "[CẦN BA CONFIRM UI AUTO-LOGOUT] " + title
        
    if new_id in ['SA02-BR-NEG-003', 'SA02-UI-008']:
        if not title.startswith("[CẦN BA"):
            title = "[CẦN BA CONFIRM POPUP WARNING] " + title

    ws.cell(r, col_title).value = title
    ws.cell(r, col_exp).value = exp

wb.save(dest_file)
print(f"Updated Test Cases successfully. Saved to: {dest_file}")

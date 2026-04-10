import openpyxl
import sys
import re

def process_testcases(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb["TestCases"]

    col_map = {cell.value: cell.column for cell in ws[1]}
    col_pre = col_map.get("Precondition")
    col_exp = col_map.get("Expected")
    col_steps = col_map.get("Steps")
    col_feature = col_map.get("Feature")
    col_note = col_map.get("Note")
    col_comment = col_map.get("Comment")
    col_title = col_map.get("Title")
    col_tcid = col_map.get("TC_ID")

    tc_ids_seen = set()

    for row in range(2, ws.max_row + 1):
        tc_id = ws.cell(row=row, column=col_tcid).value
        if not tc_id:
            continue

        comment = ws.cell(row=row, column=col_comment).value or ""
        steps = ws.cell(row=row, column=col_steps).value or ""
        expected = ws.cell(row=row, column=col_exp).value or ""
        precondition = ws.cell(row=row, column=col_pre).value or ""
        feature = ws.cell(row=row, column=col_feature).value or ""
        title = ws.cell(row=row, column=col_title).value or ""
        note = ws.cell(row=row, column=col_note).value or ""

        new_notes = []

        # -- FIX BA COMMENTS --
        # 1. Trùng case
        if "Trùng case" in comment:
            new_notes.append("📝 REVIEW (BA Comment): LỖI TRÙNG LẶP - TC này bị trùng lặp nội dung hoàn toàn. Yêu cầu xóa bỏ.")
        elif tc_id in tc_ids_seen:
            new_notes.append(f"📝 LỖI LẶP TC_ID: ID '{tc_id}' bị trùng lặp. Cần cập nhật TC_ID duy nhất.")
            
        tc_ids_seen.add(tc_id)

        # 2. Đã bỏ trường Tần suất, Ngày thu
        if "Đã bỏ trường" in comment or "Tần suất" in title or "Ngày thu" in title:
            new_notes.append("📝 REVIEW (BA Comment): SAI LOGIC (GAP) - URD đã gỡ bỏ trường Tần suất/Ngày thu. Vui lòng xem xét bỏ TC này.")
            # Xóa các text tương ứng trong Steps
            steps = steps.replace(", Tần suất, Ngày thu", "")
            steps = steps.replace("Tần suất, Ngày thu", "")
            steps = steps.replace(", Tần suất", "")
            steps = steps.replace("Tần suất", "")
            steps = steps.replace(", Ngày thu", "")
            ws.cell(row=row, column=col_steps).value = steps
            
            # Text UI Format in Expected
            expected = re.sub(r'Tần suất.*?\(radio\), ', '', expected)

        # -- FIX FORMATTING --
        # 3. Sửa Precondition
        if precondition:
            new_pre = []
            for line in str(precondition).split('\n'):
                if "Đăng nhập vào hệ thống" in line or "Đăng nhập" in line:
                    num_match = re.match(r'^(\d+\.\s*)', line)
                    prefix = num_match.group(1) if num_match else "1. "
                    feat_name = feature if feature else "hệ thống"
                    line = f"{prefix}Được phân quyền vào chức năng {feat_name}."
                    
                    if "LỖI FORMAT: Pre-condition" not in note:
                         new_notes.append("📝 REVIEW: Đã cập nhật Pre-condition từ 'Đăng nhập chung chung' sang chức năng chính xác.")
                new_pre.append(line)
            ws.cell(row=row, column=col_pre).value = "\n".join(new_pre)

        # 4. Sửa Expected - Trạng thái/Audit -> Trạng thái bản ghi, remove audit log text
        if expected:
            old_exp = expected
            expected = expected.replace("(iii) Trạng thái/Audit:", "(iii) Trạng thái bản ghi:")
            
            new_exp = []
            fixed_audit = False
            for line in expected.split('\n'):
                if "(iii) Trạng thái bản ghi:" in line:
                    if "Không tạo thay đổi" in line or "Không thay đổi" in line:
                        line = "(iii) Trạng thái bản ghi: Không thay đổi trạng thái của bản ghi."
                        fixed_audit = True
                    elif "Audit Log" in line or "Audit log" in line or "Audit" in line:
                        parts = re.split(r'\.\s*Audit\s*log|\.\s*Audit\s*Log|\.\s*Audit|\.\s*Ghi\s*log', line)
                        if len(parts) > 1:
                            line = parts[0].strip() + "."
                            if line.endswith(".."): line = line[:-1]
                            fixed_audit = True
                new_exp.append(line)
                
            if fixed_audit and "LỖI FORMAT: Mục (iii)" not in note:
                new_notes.append("📝 REVIEW: Đã cập nhật Expected item (iii) bỏ kiểm tra Audit Log.")
                
            ws.cell(row=row, column=col_exp).value = "\n".join(new_exp)
            
        # Update Notes
        if new_notes:
            # clean existing note from our previous script runs to avoid spam
            clean_note = re.sub(r'📝.*?\n', '', note)
            clean_note = re.sub(r'📝.*', '', clean_note).strip()
            
            final_note = "\n".join(new_notes)
            if clean_note:
                final_note = clean_note + "\n\n" + final_note
            ws.cell(row=row, column=col_note).value = final_note

    wb.save(file_path)
    print("Done applying fixes and reviews!")

if __name__ == "__main__":
    process_testcases(sys.argv[1])

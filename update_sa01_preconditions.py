import openpyxl

file_path = '/Users/mac/antigravity-testing-kit/Feature_02_SA_Tham_So_He_Thong/test case/SA01_Đăng nhập.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.worksheets[0]

headers = [cell.value for cell in ws[1]]
col_map = {name: idx for idx, name in enumerate(headers)}

updates = {
    'SA01-NEG-004': 'Truy cập thành công vào màn hình Đăng nhập.',
    'SA01-HAPPY-007': 'Thiết bị trình duyệt cho phép lưu Cookie/Cache. Tài khoản đăng nhập hợp lệ.',
    'SA01-UI-008': 'Truy cập thành công vào màn hình Đăng nhập.',
    'SA01-NEG-010': 'Hệ thống đã kết nối và cấu hình thành công với dịch vụ EntraID.',
    'SA01-NEG-013': 'Hệ thống đã kết nối và cấu hình thành công với dịch vụ EntraID.',
    'SA01-NEG-014': 'Truy cập thành công vào màn hình Đăng nhập.'
}

for row in ws.iter_rows(min_row=2):
    tc_id_val = row[col_map['TC_ID']].value
    if tc_id_val and tc_id_val in updates:
        row[col_map['Precondition']].value = updates[tc_id_val]
        print(f"Updated Precondition for {tc_id_val}")

wb.save(file_path)
print("\nEmpty preconditions updated successfully.")
